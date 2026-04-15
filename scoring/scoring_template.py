"""Generate scoring_template.xlsx per prompt.md §7.3.

Sheets:
  1. Scorecard — one row per TC, columns for each dimension (1-3), plus notes
  2. Grader Instructions — detailed per-dimension per-TC guidance from rubrics.yaml
  3. Inter-Rater Agreement — two raters' scores with Cohen's kappa formula
  4. Aggregate Dashboard — pivot tables by service line × capability × difficulty
"""

from __future__ import annotations

import datetime
import io
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import ExcelWriter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

DIMENSIONS = [
    "correctness",
    "completeness",
    "format_compliance",
    "robustness",
    "communication",
]

DIMENSION_LABELS = {
    "correctness": "Correctness",
    "completeness": "Completeness",
    "format_compliance": "Format Compliance",
    "robustness": "Robustness",
    "communication": "Communication",
}

# Pass/fail thresholds (prompt.md §7.4)
PASS_THRESHOLD = 2.4
CONDITIONAL_PASS_THRESHOLD = 2.0

# Fixed timestamps for deterministic output (matches generator convention)
_FIXED_DATETIME = datetime.datetime(2025, 3, 15, 9, 0, 0)
_FIXED_ZIP_DT = (2025, 3, 15, 9, 0, 0)

_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_SUBHEADER_FONT = Font(name="Calibri", bold=True, size=11)
_BODY_FONT = Font(name="Calibri", size=11)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_WRAP = Alignment(wrap_text=True, vertical="top")


def _style_header_row(ws, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _WRAP
        cell.border = _THIN_BORDER


def _style_cell(ws, row: int, col: int, *, font=None, fill=None) -> None:
    cell = ws.cell(row=row, column=col)
    cell.font = font or _BODY_FONT
    cell.alignment = _WRAP
    cell.border = _THIN_BORDER
    if fill:
        cell.fill = fill


# ---------------------------------------------------------------------------
# Sheet 1: Scorecard
# ---------------------------------------------------------------------------

def _build_scorecard(wb: Workbook, test_cases: dict) -> None:
    ws = wb.active
    ws.title = "Scorecard"

    headers = [
        "TC ID", "Title", "Service Line", "Difficulty",
        "Correctness (1-3)", "Completeness (1-3)", "Format Compliance (1-3)",
        "Robustness (1-3)", "Communication (1-3)",
        "Average", "Result", "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    sorted_tcs = sorted(test_cases.keys(), key=lambda k: int(k.split("-")[1]))
    for r, tc_id in enumerate(sorted_tcs, 2):
        tc = test_cases[tc_id]
        ws.cell(row=r, column=1, value=tc_id)
        ws.cell(row=r, column=2, value=tc.get("title", ""))
        ws.cell(row=r, column=3, value=tc.get("service_line", ""))
        ws.cell(row=r, column=4, value=tc.get("difficulty", ""))

        # Columns 5-9: dimension score entry (left blank for grader to fill)
        for col in range(5, 10):
            _style_cell(ws, r, col)

        # Average formula (cols 5-9)
        avg_formula = f"=IF(COUNTA(E{r}:I{r})=5,AVERAGE(E{r}:I{r}),\"\")"
        ws.cell(row=r, column=10, value=avg_formula)

        # Result formula based on §7.4 thresholds
        result_formula = (
            f'=IF(J{r}="","",IF(AND(J{r}>={PASS_THRESHOLD}),"Pass",'
            f'IF(AND(J{r}>={CONDITIONAL_PASS_THRESHOLD},'
            f'MIN(E{r}:I{r})>1),"Conditional Pass","Fail")))'
        )
        ws.cell(row=r, column=11, value=result_formula)

        # Notes column
        _style_cell(ws, r, 12)

        # Style all cells in the row
        for col in range(1, len(headers) + 1):
            _style_cell(ws, r, col)

    # Column widths
    widths = [8, 40, 15, 12, 16, 16, 20, 16, 18, 10, 18, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Sheet 2: Grader Instructions
# ---------------------------------------------------------------------------

def _build_grader_instructions(wb: Workbook, test_cases: dict) -> None:
    ws = wb.create_sheet("Grader Instructions")

    headers = ["TC ID", "Title", "Dimension", "Score 3", "Score 2", "Score 1"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    row = 2
    sorted_tcs = sorted(test_cases.keys(), key=lambda k: int(k.split("-")[1]))
    for tc_id in sorted_tcs:
        tc = test_cases[tc_id]
        first_row_for_tc = row
        for dim in DIMENSIONS:
            anchors = tc.get(dim, {})
            ws.cell(row=row, column=1, value=tc_id)
            ws.cell(row=row, column=2, value=tc.get("title", ""))
            ws.cell(row=row, column=3, value=DIMENSION_LABELS[dim])
            ws.cell(row=row, column=4, value=anchors.get(3, ""))
            ws.cell(row=row, column=5, value=anchors.get(2, ""))
            ws.cell(row=row, column=6, value=anchors.get(1, ""))
            for col in range(1, 7):
                _style_cell(ws, row, col)
            row += 1

        # Shade the TC block header row
        for col in range(1, 7):
            _style_cell(ws, first_row_for_tc, col, fill=_SUBHEADER_FILL)

    widths = [8, 40, 20, 60, 60, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Sheet 3: Inter-Rater Agreement
# ---------------------------------------------------------------------------

def _build_inter_rater(wb: Workbook, test_cases: dict) -> None:
    ws = wb.create_sheet("Inter-Rater Agreement")

    # Header row: TC ID, then for each dimension: Rater 1, Rater 2, Diff
    headers = ["TC ID"]
    for dim in DIMENSIONS:
        label = DIMENSION_LABELS[dim]
        headers.extend([f"{label} R1", f"{label} R2", f"{label} Diff"])
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    sorted_tcs = sorted(test_cases.keys(), key=lambda k: int(k.split("-")[1]))
    for r, tc_id in enumerate(sorted_tcs, 2):
        ws.cell(row=r, column=1, value=tc_id)
        _style_cell(ws, r, 1)
        col = 2
        for _ in DIMENSIONS:
            # R1 and R2 left blank for entry
            _style_cell(ws, r, col)
            _style_cell(ws, r, col + 1)
            # Diff = |R1 - R2|
            r1_ref = f"{get_column_letter(col)}{r}"
            r2_ref = f"{get_column_letter(col + 1)}{r}"
            ws.cell(
                row=r, column=col + 2,
                value=f'=IF(AND({r1_ref}<>"",{r2_ref}<>""),ABS({r1_ref}-{r2_ref}),"")',
            )
            _style_cell(ws, r, col + 2)
            col += 3

    # Cohen's kappa section below the data
    data_end_row = len(sorted_tcs) + 1
    kappa_start = data_end_row + 3
    n_tcs = len(sorted_tcs)

    ws.cell(row=kappa_start, column=1, value="Cohen's Kappa Calculation")
    ws.cell(row=kappa_start, column=1).font = Font(name="Calibri", bold=True, size=12)

    ws.cell(row=kappa_start + 1, column=1, value="Dimension")
    ws.cell(row=kappa_start + 1, column=2, value="Exact Agreement Count")
    ws.cell(row=kappa_start + 1, column=3, value="Observed Agreement (Po)")
    ws.cell(row=kappa_start + 1, column=4, value="Expected Agreement (Pe)")
    ws.cell(row=kappa_start + 1, column=5, value="Cohen's Kappa (κ)")
    ws.cell(row=kappa_start + 1, column=6, value="Interpretation")
    _style_header_row(ws, kappa_start + 1, 6)

    for d_idx, dim in enumerate(DIMENSIONS):
        row = kappa_start + 2 + d_idx
        r1_col = 2 + d_idx * 3      # R1 column for this dimension
        r2_col = r1_col + 1          # R2 column

        r1_letter = get_column_letter(r1_col)
        r2_letter = get_column_letter(r2_col)

        ws.cell(row=row, column=1, value=DIMENSION_LABELS[dim])

        # Exact agreement count
        ws.cell(
            row=row, column=2,
            value=f"=SUMPRODUCT(({r1_letter}2:{r1_letter}{data_end_row}={r2_letter}2:{r2_letter}{data_end_row})*({r1_letter}2:{r1_letter}{data_end_row}<>\"\"))",
        )

        # Po = agreement / N
        ws.cell(
            row=row, column=3,
            value=f"=IF(B{row}=\"\",\"\",B{row}/{n_tcs})",
        )

        # Pe (expected agreement by chance for 3-point scale)
        # Pe = sum over k in {1,2,3} of (proportion_R1_k * proportion_R2_k)
        # Simplified: for a 3-point scale, Pe = sum_k (n1k/N * n2k/N)
        pe_parts = []
        for score in [1, 2, 3]:
            pe_parts.append(
                f"(COUNTIF({r1_letter}2:{r1_letter}{data_end_row},{score})/{n_tcs})"
                f"*(COUNTIF({r2_letter}2:{r2_letter}{data_end_row},{score})/{n_tcs})"
            )
        pe_formula = f'=IF(C{row}="","",{"+".join(pe_parts)})'
        ws.cell(row=row, column=4, value=pe_formula)

        # Kappa = (Po - Pe) / (1 - Pe)
        ws.cell(
            row=row, column=5,
            value=f'=IF(OR(C{row}="",D{row}=""),"",IF(D{row}=1,1,(C{row}-D{row})/(1-D{row})))',
        )

        # Interpretation
        ws.cell(
            row=row, column=6,
            value=(
                f'=IF(E{row}="","",IF(E{row}>=0.81,"Almost Perfect",'
                f'IF(E{row}>=0.61,"Substantial",'
                f'IF(E{row}>=0.41,"Moderate",'
                f'IF(E{row}>=0.21,"Fair","Poor")))))'
            ),
        )

        for col in range(1, 7):
            _style_cell(ws, row, col)

    # Column widths
    ws.column_dimensions["A"].width = 10
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    # Kappa section columns
    for c in range(1, 7):
        ws.column_dimensions[get_column_letter(c)].width = max(
            ws.column_dimensions[get_column_letter(c)].width or 0, 22
        )


# ---------------------------------------------------------------------------
# Sheet 4: Aggregate Dashboard
# ---------------------------------------------------------------------------

def _build_dashboard(wb: Workbook, test_cases: dict) -> None:
    ws = wb.create_sheet("Aggregate Dashboard")

    sorted_tcs = sorted(test_cases.keys(), key=lambda k: int(k.split("-")[1]))
    tc_list = [(tc_id, test_cases[tc_id]) for tc_id in sorted_tcs]
    data_end_row = len(sorted_tcs) + 1  # row 1 is header on Scorecard

    # --- Section 1: By Service Line ---
    service_lines = sorted({tc.get("service_line", "") for _, tc in tc_list})
    row = 1
    ws.cell(row=row, column=1, value="Pass Rates by Service Line")
    ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=12)
    row += 1

    sl_headers = ["Service Line", "# TCs", "Avg Correctness", "Avg Completeness",
                   "Avg Format Compliance", "Avg Robustness", "Avg Communication",
                   "Overall Avg"]
    for col, h in enumerate(sl_headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, len(sl_headers))
    row += 1

    for sl in service_lines:
        tcs_in_sl = [tc_id for tc_id, tc in tc_list if tc.get("service_line") == sl]
        ws.cell(row=row, column=1, value=sl)
        ws.cell(row=row, column=2, value=len(tcs_in_sl))
        # Reference scorecard sheet for averages
        for d_idx, _ in enumerate(DIMENSIONS):
            sc_col = get_column_letter(5 + d_idx)  # E through I on Scorecard
            # Build AVERAGEIFS referencing Scorecard
            ws.cell(
                row=row, column=3 + d_idx,
                value=f'=IFERROR(AVERAGEIFS(Scorecard!{sc_col}2:{sc_col}{data_end_row},Scorecard!C2:C{data_end_row},A{row}),"")',
            )
        # Overall avg
        ws.cell(
            row=row, column=8,
            value=f'=IFERROR(AVERAGE(C{row}:G{row}),"")',
        )
        for col in range(1, len(sl_headers) + 1):
            _style_cell(ws, row, col)
        row += 1

    # --- Section 2: By Difficulty Tier ---
    row += 2
    ws.cell(row=row, column=1, value="Pass Rates by Difficulty Tier")
    ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=12)
    row += 1

    diff_headers = ["Difficulty", "# TCs", "Avg Correctness", "Avg Completeness",
                    "Avg Format Compliance", "Avg Robustness", "Avg Communication",
                    "Overall Avg"]
    for col, h in enumerate(diff_headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, len(diff_headers))
    row += 1

    difficulties = ["Routine", "Complex", "Adversarial"]
    for diff in difficulties:
        tcs_in_diff = [tc_id for tc_id, tc in tc_list if tc.get("difficulty") == diff]
        ws.cell(row=row, column=1, value=diff)
        ws.cell(row=row, column=2, value=len(tcs_in_diff))
        for d_idx, _ in enumerate(DIMENSIONS):
            sc_col = get_column_letter(5 + d_idx)
            ws.cell(
                row=row, column=3 + d_idx,
                value=f'=IFERROR(AVERAGEIFS(Scorecard!{sc_col}2:{sc_col}{data_end_row},Scorecard!D2:D{data_end_row},A{row}),"")',
            )
        ws.cell(
            row=row, column=8,
            value=f'=IFERROR(AVERAGE(C{row}:G{row}),"")',
        )
        for col in range(1, len(diff_headers) + 1):
            _style_cell(ws, row, col)
        row += 1

    # --- Section 3: By Capability Axis ---
    row += 2
    ws.cell(row=row, column=1, value="Pass Rates by Capability Axis")
    ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=12)
    row += 1

    # Capability matrix: each TC maps to one or more capabilities
    all_capabilities = sorted({
        cap
        for _, tc in tc_list
        for cap in tc.get("capabilities", [])
    })

    cap_headers = ["Capability", "# TCs", "Test Cases"]
    for col, h in enumerate(cap_headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, len(cap_headers))
    row += 1

    for cap in all_capabilities:
        tcs_with_cap = [tc_id for tc_id, tc in tc_list if cap in tc.get("capabilities", [])]
        ws.cell(row=row, column=1, value=cap)
        ws.cell(row=row, column=2, value=len(tcs_with_cap))
        ws.cell(row=row, column=3, value=", ".join(tcs_with_cap))
        for col in range(1, 4):
            _style_cell(ws, row, col)
        row += 1

    # --- Section 4: Full 3D Matrix (Service Line × Difficulty × Capability) ---
    row += 2
    ws.cell(row=row, column=1, value="3D Capability Matrix (§7.5)")
    ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=12)
    row += 1

    matrix_headers = ["Service Line", "Capability", "Difficulty", "Test Cases"]
    for col, h in enumerate(matrix_headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, len(matrix_headers))
    row += 1

    for sl in service_lines:
        for cap in all_capabilities:
            for diff in difficulties:
                matching = [
                    tc_id for tc_id, tc in tc_list
                    if tc.get("service_line") == sl
                    and tc.get("difficulty") == diff
                    and cap in tc.get("capabilities", [])
                ]
                if matching:
                    ws.cell(row=row, column=1, value=sl)
                    ws.cell(row=row, column=2, value=cap)
                    ws.cell(row=row, column=3, value=diff)
                    ws.cell(row=row, column=4, value=", ".join(matching))
                    for col in range(1, 5):
                        _style_cell(ws, row, col)
                    row += 1

    # Column widths
    widths = [20, 20, 20, 20, 20, 20, 20, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_scoring_template(rubrics_path: Path, output_path: Path) -> None:
    """Generate scoring_template.xlsx from rubrics.yaml.

    Args:
        rubrics_path: Path to scoring/rubrics.yaml.
        output_path: Path to write scoring_template.xlsx.
    """
    with open(rubrics_path) as f:
        rubrics = yaml.safe_load(f)

    test_cases = rubrics.get("test_cases", {})

    wb = Workbook()
    _build_scorecard(wb, test_cases)
    _build_grader_instructions(wb, test_cases)
    _build_inter_rater(wb, test_cases)
    _build_dashboard(wb, test_cases)
    _save_deterministic(wb, output_path)


def _save_deterministic(wb: Workbook, path: Path) -> None:
    """Save workbook with pinned timestamps for byte-identical reruns."""
    wb.properties.created = _FIXED_DATETIME
    wb.properties.modified = _FIXED_DATETIME

    buf = io.BytesIO()
    archive = ZipFile(buf, "w", ZIP_DEFLATED, allowZip64=True)
    writer = ExcelWriter(wb, archive)
    writer.save()

    buf.seek(0)
    with ZipFile(buf, "r") as src, ZipFile(str(path), "w", ZIP_DEFLATED) as dst:
        for item in src.infolist():
            info = ZipInfo(item.filename, date_time=_FIXED_ZIP_DT)
            info.compress_type = item.compress_type
            dst.writestr(info, src.read(item.filename))
