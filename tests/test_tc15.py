"""Tests for TC-15 — DCF Valuation formatter.

Verifies:
- 3-year historical financials (income statement, balance sheet, cash flow, key ratios)
- 5-year management projections with growth/margin/capex assumptions
- 10 comparable companies with trading multiples
- 15-page industry overview PDF
- Planted error ERR-022 (missing_data: blank EV/EBITDA for one comp)
- Planted error ERR-025 (mismatched_total: FY2024 revenue doesn't match sum)
- DCF gold standard values: WACC ~10.8%, EV ranges, equity midpoint ~$255M
- Canary embedding in all 4 input files
- Prompt and expected behavior markdown files
- Gold standard registration with margin expansion flag
"""

from __future__ import annotations

import json
import tempfile
from decimal import Decimal
from pathlib import Path

import openpyxl

from generator.canaries import CanaryRegistry
from generator.canaries import build_registry as build_canary_registry
from generator.errors import ErrorRegistry
from generator.formatters.tc15 import (
    _COMPARABLE_COMPANIES,
    _EQUITY_RISK_PREMIUM,
    _PERPETUITY_GROWTH,
    _PROJ_EBITDA_MARGIN_BASE,
    _PROJ_EBITDA_MARGIN_TARGET,
    _PROJ_REVENUE_GROWTH,
    _PROJECTION_YEARS,
    _RISK_FREE_RATE,
    _SIZE_PREMIUM,
    emit_tc15,
)
from generator.golds.framework import emit_gold
from generator.manifest import Manifest
from generator.model.build import CascadeModel, build_model

# ---------------------------------------------------------------------------
# Module-level fixture: build the model and run emit_tc15 once
# ---------------------------------------------------------------------------

_MODEL: CascadeModel | None = None
_OUTPUT: Path | None = None
_CANARIES: CanaryRegistry | None = None
_ERRORS: ErrorRegistry | None = None
_MANIFEST: Manifest | None = None

_CANARY_KEYS = sorted([
    "tc15_historical_financials",
    "tc15_management_projections",
    "tc15_comparable_companies",
    "tc15_industry_overview",
])


def _ensure_emitted() -> tuple[CascadeModel, Path, CanaryRegistry, ErrorRegistry, Manifest]:
    global _MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST  # noqa: PLW0603
    if _MODEL is None:
        _MODEL = build_model(seed=42)
        _OUTPUT = Path(tempfile.mkdtemp(prefix="tc15_test_"))
        _CANARIES = build_canary_registry(_CANARY_KEYS, seed=42)
        _ERRORS = ErrorRegistry()
        _MANIFEST = Manifest(_OUTPUT)
        _MANIFEST.__enter__()

        emit_tc15(_MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST)

        # Emit gold standard (registered via @register_gold)
        emit_gold("TC-15", _CANARIES, _ERRORS, _OUTPUT / "gold_standards", model=_MODEL)

        _MANIFEST.__exit__(None, None, None)
    return _MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST


# ---------------------------------------------------------------------------
# Input directory
# ---------------------------------------------------------------------------

_INPUT_DIR = "test_cases/TC-15/input_files"


# ---------------------------------------------------------------------------
# Constants verification (from prompt.md §5 TC-15)
# ---------------------------------------------------------------------------


class TestDCFConstants:
    """Verify formatter constants match prompt.md gold standard."""

    def test_risk_free_rate(self) -> None:
        assert _RISK_FREE_RATE == Decimal("0.042")

    def test_equity_risk_premium(self) -> None:
        assert _EQUITY_RISK_PREMIUM == Decimal("0.055")

    def test_size_premium(self) -> None:
        assert _SIZE_PREMIUM == Decimal("0.020")

    def test_perpetuity_growth(self) -> None:
        assert _PERPETUITY_GROWTH == Decimal("0.025")

    def test_ten_comparable_companies(self) -> None:
        assert len(_COMPARABLE_COMPANIES) == 10

    def test_five_projection_years(self) -> None:
        assert len(_PROJECTION_YEARS) == 5
        assert _PROJECTION_YEARS == [2026, 2027, 2028, 2029, 2030]

    def test_margin_expansion_200bps(self) -> None:
        """Management assumes 200bps EBITDA margin expansion — the aggressive flag."""
        expansion = _PROJ_EBITDA_MARGIN_TARGET - _PROJ_EBITDA_MARGIN_BASE
        assert expansion == Decimal("0.020")

    def test_revenue_growth_declining(self) -> None:
        """Revenue growth should decline from 8% to 5%."""
        assert _PROJ_REVENUE_GROWTH[0] == Decimal("0.080")
        assert _PROJ_REVENUE_GROWTH[-1] == Decimal("0.050")
        # Monotonically decreasing
        for i in range(len(_PROJ_REVENUE_GROWTH) - 1):
            assert _PROJ_REVENUE_GROWTH[i] >= _PROJ_REVENUE_GROWTH[i + 1]


# ---------------------------------------------------------------------------
# Historical financials xlsx
# ---------------------------------------------------------------------------


class TestHistoricalFinancials:
    """Verify historical_financials_3yr.xlsx."""

    def test_file_exists(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        path = out / _INPUT_DIR / "historical_financials_3yr.xlsx"
        assert path.exists()

    def test_has_four_sheets(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "historical_financials_3yr.xlsx")
        expected = {"Income Statement", "Balance Sheet", "Cash Flow", "Key Ratios"}
        assert set(wb.sheetnames) == expected

    def test_income_statement_three_years(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "historical_financials_3yr.xlsx")
        ws = wb["Income Statement"]
        # Headers in row 1: blank, FY2023, FY2024, FY2025
        assert ws.cell(row=1, column=2).value == "FY2023"
        assert ws.cell(row=1, column=3).value == "FY2024"
        assert ws.cell(row=1, column=4).value == "FY2025"

    def test_revenue_row_present(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "historical_financials_3yr.xlsx")
        ws = wb["Income Statement"]
        assert ws.cell(row=2, column=1).value == "Revenue"
        # Revenue values should be positive integers
        for col in [2, 3, 4]:
            val = ws.cell(row=2, column=col).value
            assert isinstance(val, (int, float))
            assert val > 0

    def test_ebitda_row_present(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "historical_financials_3yr.xlsx")
        ws = wb["Income Statement"]
        # EBITDA is the last data row (row 13)
        assert ws.cell(row=13, column=1).value == "EBITDA"

    def test_balance_sheet_key_items(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "historical_financials_3yr.xlsx")
        ws = wb["Balance Sheet"]
        labels = [ws.cell(row=r, column=1).value for r in range(2, 10)]
        assert "Total Assets" in labels
        assert "Total Debt" in labels
        assert "Net Debt" in labels

    def test_cash_flow_capex(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "historical_financials_3yr.xlsx")
        ws = wb["Cash Flow"]
        assert ws.cell(row=3, column=1).value == "Capital Expenditures"

    def test_key_ratios_sheet(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "historical_financials_3yr.xlsx")
        ws = wb["Key Ratios"]
        labels = [ws.cell(row=r, column=1).value for r in range(2, 9)]
        assert "EBITDA Margin" in labels
        assert "Effective Tax Rate" in labels
        assert "NWC / Revenue" in labels


# ---------------------------------------------------------------------------
# Management projections xlsx
# ---------------------------------------------------------------------------


class TestManagementProjections:
    """Verify management_projections.xlsx."""

    def test_file_exists(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        path = out / _INPUT_DIR / "management_projections.xlsx"
        assert path.exists()

    def test_has_three_sheets(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "management_projections.xlsx")
        expected = {"Revenue Projections", "EBITDA Projections", "Assumptions"}
        assert set(wb.sheetnames) == expected

    def test_five_projection_years_in_headers(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "management_projections.xlsx")
        ws = wb["Revenue Projections"]
        for c, year in enumerate(_PROJECTION_YEARS, 2):
            assert ws.cell(row=1, column=c).value == str(year)

    def test_assumptions_include_margin_note(self) -> None:
        """The 200bps expansion note must be present in the Assumptions sheet."""
        _, out, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "management_projections.xlsx")
        ws = wb["Assumptions"]
        # Search all cells for the margin expansion note
        found = False
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell and "200bps" in str(cell):
                    found = True
                    break
        assert found, "200bps margin expansion note not found in Assumptions sheet"


# ---------------------------------------------------------------------------
# Comparable companies xlsx
# ---------------------------------------------------------------------------


class TestComparableCompanies:
    """Verify comparable_companies_trading.xlsx."""

    def test_file_exists(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        path = out / _INPUT_DIR / "comparable_companies_trading.xlsx"
        assert path.exists()

    def test_has_two_sheets(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "comparable_companies_trading.xlsx")
        assert "Trading Comparables" in wb.sheetnames
        assert "Capital Structure" in wb.sheetnames

    def test_ten_companies_listed(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "comparable_companies_trading.xlsx")
        ws = wb["Trading Comparables"]
        # Companies in rows 2-11
        companies = []
        for r in range(2, 12):
            name = ws.cell(row=r, column=1).value
            if name:
                companies.append(name)
        assert len(companies) == 10

    def test_headers_include_multiples(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "comparable_companies_trading.xlsx")
        ws = wb["Trading Comparables"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 12)]
        assert "EV/Revenue" in headers
        assert "EV/EBITDA" in headers
        assert "Beta" in headers

    def test_summary_statistics_present(self) -> None:
        """Median and Mean rows should appear below the company data."""
        _, out, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "comparable_companies_trading.xlsx")
        ws = wb["Trading Comparables"]
        # Summary should be at row 13 (10 companies + header + blank)
        summary_row = len(_COMPARABLE_COMPANIES) + 3  # row 13
        assert ws.cell(row=summary_row, column=1).value == "Median"
        assert ws.cell(row=summary_row + 1, column=1).value == "Mean"

    def test_capital_structure_de_ratios(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "comparable_companies_trading.xlsx")
        ws = wb["Capital Structure"]
        # Should have Debt/Equity column header
        headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
        assert "Debt/Equity" in headers


# ---------------------------------------------------------------------------
# Industry overview PDF
# ---------------------------------------------------------------------------


class TestIndustryOverviewPDF:
    """Verify industry_overview.pdf."""

    def test_file_exists(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        path = out / _INPUT_DIR / "industry_overview.pdf"
        assert path.exists()

    def test_file_not_empty(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        path = out / _INPUT_DIR / "industry_overview.pdf"
        assert path.stat().st_size > 0


# ---------------------------------------------------------------------------
# Planted errors
# ---------------------------------------------------------------------------


class TestPlantedErrors:
    """Verify ERR-022 and ERR-025 are registered correctly."""

    def test_err022_registered(self) -> None:
        """ERR-022: missing_data — blank EV/EBITDA for one comparable."""
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.get("ERR-022")
        assert err is not None
        assert err.type == "missing_data"
        assert "EV/EBITDA" in err.description
        assert "MidWest Materials Group" in err.description

    def test_err022_blank_cell_in_xlsx(self) -> None:
        """The actual cell should be blank (None) in the workbook."""
        _, out, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "comparable_companies_trading.xlsx")
        ws = wb["Trading Comparables"]
        # MidWest Materials Group is index 2 → row 4, EV/EBITDA is column 9
        val = ws.cell(row=4, column=9).value
        assert val is None, f"ERR-022: expected blank EV/EBITDA, got {val}"

    def test_err025_registered(self) -> None:
        """ERR-025: mismatched_total — FY2024 revenue doesn't match sum."""
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.get("ERR-025")
        assert err is not None
        assert err.type == "mismatched_total"
        assert "FY2024" in err.description or "2024" in err.description

    def test_err025_corrupted_revenue(self) -> None:
        """FY2024 Revenue in the IS sheet should differ from other years' pattern."""
        _, out, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "historical_financials_3yr.xlsx")
        ws = wb["Income Statement"]
        # Row 2 is Revenue; col 3 is FY2024
        revenue_2024 = ws.cell(row=2, column=3).value
        assert revenue_2024 is not None
        # The error description includes the wrong value with comma formatting
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.get("ERR-025")
        formatted = f"${revenue_2024:,}"
        assert formatted in err.description, (
            f"Cell value {formatted} not found in error description: {err.description}"
        )


# ---------------------------------------------------------------------------
# Canary embedding
# ---------------------------------------------------------------------------


class TestCanaries:
    """Verify canary codes are embedded in all 4 input files."""

    def test_historical_financials_canary(self) -> None:
        _, out, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("tc15_historical_financials")
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "historical_financials_3yr.xlsx")
        assert canary in (wb.properties.description or ""), (
            f"Canary {canary} not in document properties"
        )

    def test_management_projections_canary(self) -> None:
        _, out, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("tc15_management_projections")
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "management_projections.xlsx")
        assert canary in (wb.properties.description or ""), (
            f"Canary {canary} not in document properties"
        )

    def test_comparable_companies_canary(self) -> None:
        _, out, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("tc15_comparable_companies")
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "comparable_companies_trading.xlsx")
        assert canary in (wb.properties.description or ""), (
            f"Canary {canary} not in document properties"
        )

    def test_industry_overview_canary(self) -> None:
        """Canary is in the PDF metadata (subject field)."""
        _, out, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("tc15_industry_overview")
        pdf_path = out / _INPUT_DIR / "industry_overview.pdf"
        content = pdf_path.read_bytes()
        assert canary.encode() in content, (
            f"Canary {canary} not found in industry_overview.pdf"
        )


# ---------------------------------------------------------------------------
# Prompt and expected behavior files
# ---------------------------------------------------------------------------


class TestMarkdownFiles:
    """Verify prompt.md and expected_behavior.md."""

    def test_prompt_exists(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        path = out / "test_cases/TC-15/prompt.md"
        assert path.exists()

    def test_prompt_contains_dcf_instructions(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        text = (out / "test_cases/TC-15/prompt.md").read_text()
        assert "DCF valuation" in text
        assert "WACC" in text
        assert "Gordon Growth" in text
        assert "sensitivity analysis" in text

    def test_expected_behavior_exists(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        path = out / "test_cases/TC-15/expected_behavior.md"
        assert path.exists()

    def test_expected_behavior_contains_gold_values(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        text = (out / "test_cases/TC-15/expected_behavior.md").read_text()
        assert "10.8%" in text
        assert "$245M" in text
        assert "$305M" in text
        assert "$255M" in text
        assert "200bps" in text


# ---------------------------------------------------------------------------
# Gold standard
# ---------------------------------------------------------------------------


class TestGoldStandard:
    """Verify the TC-15 gold standard JSON."""

    def test_gold_file_exists(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        path = out / "gold_standards/TC-15_gold.json"
        assert path.exists()

    def _load_gold(self) -> dict:
        _, out, _, _, _ = _ensure_emitted()
        return json.loads((out / "gold_standards/TC-15_gold.json").read_text())

    def test_wacc_approximately_10_8(self) -> None:
        gold = self._load_gold()
        wacc_pct = gold["expected_outputs"]["wacc"]["value_pct"]
        assert 9.5 <= wacc_pct <= 12.0, f"WACC {wacc_pct}% outside expected range"

    def test_ev_gordon_growth_range(self) -> None:
        """Enterprise value (Gordon Growth) should be $245M-$305M."""
        gold = self._load_gold()
        ev_gg = gold["expected_outputs"]["enterprise_value_gordon_growth"]
        assert ev_gg["range_low_M"] == 245
        assert ev_gg["range_high_M"] == 305
        computed = ev_gg["computed_M"]
        assert 200 <= computed <= 400, f"Computed EV (GG) ${computed}M looks wrong"

    def test_ev_exit_multiple_range(self) -> None:
        """Enterprise value (Exit Multiple) should be $260M-$320M."""
        gold = self._load_gold()
        ev_em = gold["expected_outputs"]["enterprise_value_exit_multiple"]
        assert ev_em["range_low_M"] == 260
        assert ev_em["range_high_M"] == 320
        computed = ev_em["computed_M"]
        assert 200 <= computed <= 400, f"Computed EV (EM) ${computed}M looks wrong"

    def test_implied_equity_value(self) -> None:
        gold = self._load_gold()
        equity = gold["expected_outputs"]["implied_equity_value_M"]
        assert 150 <= equity <= 400, f"Equity midpoint ${equity}M looks wrong"

    def test_margin_expansion_flag(self) -> None:
        gold = self._load_gold()
        flag = gold["expected_outputs"]["margin_expansion_flag"]
        assert flag["assumption_bps"] == 200
        assert flag["should_flag"] is True

    def test_projected_ufcf_five_years(self) -> None:
        gold = self._load_gold()
        ufcf = gold["expected_outputs"]["projected_ufcf_M"]
        assert len(ufcf) == 5
        for year_str in ["2026", "2027", "2028", "2029", "2030"]:
            assert year_str in ufcf

    def test_terminal_value_both_methods(self) -> None:
        gold = self._load_gold()
        tv = gold["expected_outputs"]["terminal_value"]
        assert "gordon_growth_M" in tv
        assert "exit_multiple_M" in tv
        assert tv["perpetuity_growth_rate"] == 0.025

    def test_canary_verification(self) -> None:
        gold = self._load_gold()
        cv = gold["canary_verification"]
        assert "read_historical_financials" in cv
        assert "read_management_projections" in cv
        assert "read_comparable_companies" in cv
        assert "read_industry_overview" in cv

    def test_error_detection(self) -> None:
        gold = self._load_gold()
        ed = gold["error_detection"]
        assert "ERR-022" in ed
        assert "ERR-025" in ed

    def test_scoring_hints_present(self) -> None:
        gold = self._load_gold()
        sh = gold["scoring_hints"]
        assert "correctness" in sh
        assert "completeness" in sh
        assert "robustness" in sh
