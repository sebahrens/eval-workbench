"""Tests for TC-02 — Bank Reconciliation & Confirmation Matching (Audit, Complex) formatter.

Verifies:
- Bank statement CSV with 340 transactions, cryptic descriptions, running balance
- GL cash detail xlsx with company-style descriptions, debit/credit/balance
- Bank confirmation PDF with confirmation balance of $4,312,117
- ERR-002 planted error (transposed digits in bank transaction amount)
- ERR-004 planted error (mismatched total in GL ending balance)
- Canary embedding in all files
- Gold standard structure (reconciliation, adjusted balances at $4,287,331)
- Prompt and expected behavior markdown files
"""

from __future__ import annotations

import csv
import json
import tempfile
from io import StringIO
from pathlib import Path

import openpyxl
from pypdf import PdfReader

from generator.canaries import CanaryRegistry
from generator.canaries import build_registry as build_canary_registry
from generator.errors import ErrorRegistry
from generator.formatters.tc02 import emit_tc02
from generator.golds.framework import emit_gold
from generator.manifest import Manifest
from generator.model.bank import (
    ADJUSTED_BALANCE,
    BANK_INTEREST,
    BANK_SERVICE_CHARGES,
    CONFIRMATION_BALANCE,
    GL_ENDING_BALANCE,
    TOTAL_DEPOSITS_IN_TRANSIT,
    TOTAL_OUTSTANDING_CHECKS,
)
from generator.model.build import CascadeModel, build_model

# ---------------------------------------------------------------------------
# Module-level fixture: build the model and run emit_tc02 once
# ---------------------------------------------------------------------------

_MODEL: CascadeModel | None = None
_OUTPUT: Path | None = None
_CANARIES: CanaryRegistry | None = None
_ERRORS: ErrorRegistry | None = None
_MANIFEST: Manifest | None = None

# All canary keys used by TC-02
_CANARY_KEYS = sorted([
    "bank_statement_dec2025",
    "cascade_gl_cash_dec2025",
    "bank_confirmation_fy2025",
])


def _ensure_emitted() -> tuple[CascadeModel, Path, CanaryRegistry, ErrorRegistry, Manifest]:
    global _MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST  # noqa: PLW0603
    if _MODEL is None:
        _MODEL = build_model(seed=42)
        _OUTPUT = Path(tempfile.mkdtemp(prefix="tc02_test_"))
        _CANARIES = build_canary_registry(_CANARY_KEYS, seed=42)
        _ERRORS = ErrorRegistry()
        _MANIFEST = Manifest(_OUTPUT)
        _MANIFEST.__enter__()

        emit_tc02(_MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST)

        # Emit gold standard
        emit_gold("TC-02", _CANARIES, _ERRORS, _OUTPUT / "gold_standards", model=_MODEL)

        _MANIFEST.__exit__(None, None, None)
    return _MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST


_INPUT_DIR = "test_cases/TC-02/input_files"


# ---------------------------------------------------------------------------
# Bank Statement CSV — bank_statement_dec2025.csv
# ---------------------------------------------------------------------------


def _read_csv_rows(output: Path) -> list[dict[str, str]]:
    """Read the bank CSV, skipping comment lines."""
    path = output / f"{_INPUT_DIR}/bank_statement_dec2025.csv"
    text = path.read_text()
    # Filter out comment lines (starting with #)
    data_lines = [line for line in text.splitlines() if not line.startswith("#")]
    reader = csv.DictReader(StringIO("\n".join(data_lines)))
    return list(reader)


class TestBankStatementCSV:
    """Verify bank_statement_dec2025.csv structure and content."""

    def test_file_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/bank_statement_dec2025.csv"
        assert path.exists()

    def test_has_340_transactions(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        rows = _read_csv_rows(output)
        assert len(rows) == 340, f"Expected 340 transactions, got {len(rows)}"

    def test_has_correct_columns(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        rows = _read_csv_rows(output)
        assert len(rows) > 0
        expected_cols = {"Date", "Description", "Amount", "Running Balance"}
        assert set(rows[0].keys()) == expected_cols

    def test_has_header_comments(self) -> None:
        """CSV should have comment lines with bank info."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/bank_statement_dec2025.csv"
        text = path.read_text()
        comment_lines = [line for line in text.splitlines() if line.startswith("#")]
        assert len(comment_lines) >= 3, "Expected header comment lines with bank info"

    def test_mentions_first_national_bank(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/bank_statement_dec2025.csv"
        text = path.read_text()
        assert "First National Bank" in text

    def test_mentions_account_number(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/bank_statement_dec2025.csv"
        text = path.read_text()
        assert "4782-0091" in text

    def test_amounts_are_numeric(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        rows = _read_csv_rows(output)
        for i, row in enumerate(rows[:10]):
            try:
                float(row["Amount"])
                float(row["Running Balance"])
            except ValueError:
                raise AssertionError(f"Row {i} has non-numeric Amount or Running Balance")


# ---------------------------------------------------------------------------
# GL Cash Detail — cascade_gl_cash_dec2025.xlsx
# ---------------------------------------------------------------------------


class TestGLCashDetail:
    """Verify cascade_gl_cash_dec2025.xlsx structure and content."""

    def test_file_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/cascade_gl_cash_dec2025.xlsx"
        assert path.exists()

    def test_has_cash_detail_sheet(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/cascade_gl_cash_dec2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        assert "Cash Detail - 1010" in wb.sheetnames

    def test_has_header_row_with_columns(self) -> None:
        """GL should have Date, Reference, Description, Debit, Credit, Balance."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/cascade_gl_cash_dec2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Cash Detail - 1010"]
        # Find row with headers
        header_found = False
        for row in range(1, 15):
            vals = []
            for col in range(1, 7):
                val = ws.cell(row=row, column=col).value
                if isinstance(val, str):
                    vals.append(val.strip())
            if "Date" in vals and "Debit" in vals and "Credit" in vals:
                header_found = True
                break
        assert header_found, "Expected header row with Date, Debit, Credit columns"

    def test_has_beginning_balance(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/cascade_gl_cash_dec2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Cash Detail - 1010"]
        found = False
        for row in range(1, 10):
            val = ws.cell(row=row, column=1).value
            if isinstance(val, str) and "beginning balance" in val.lower():
                found = True
                break
        assert found, "Expected 'Beginning Balance' label"

    def test_has_ending_balance(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/cascade_gl_cash_dec2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Cash Detail - 1010"]
        found = False
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if isinstance(val, str) and "ending balance" in val.lower():
                found = True
                break
        assert found, "Expected 'Ending Balance' label"

    def test_mentions_cascade_industries(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/cascade_gl_cash_dec2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Cash Detail - 1010"]
        found = False
        for row in range(1, 5):
            val = ws.cell(row=row, column=1).value
            if isinstance(val, str) and "cascade" in val.lower():
                found = True
                break
        assert found, "Expected company name in header area"

    def test_has_substantial_entries(self) -> None:
        """GL should have many journal entries (matching ~340 bank transactions)."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/cascade_gl_cash_dec2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Cash Detail - 1010"]
        data_rows = 0
        for row in range(8, ws.max_row + 1):
            date_val = ws.cell(row=row, column=1).value
            if date_val is not None and not isinstance(date_val, str):
                data_rows += 1
            elif isinstance(date_val, str) and "/" in date_val:
                data_rows += 1
        assert data_rows >= 100, f"Expected ≥100 GL entries, got {data_rows}"


# ---------------------------------------------------------------------------
# Bank Confirmation PDF — bank_confirmation_fy2025.pdf
# ---------------------------------------------------------------------------


class TestBankConfirmationPDF:
    """Verify bank_confirmation_fy2025.pdf structure and content."""

    def test_file_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/bank_confirmation_fy2025.pdf"
        assert path.exists()

    def test_is_text_native(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/bank_confirmation_fy2025.pdf"
        reader = PdfReader(str(path))
        text = reader.pages[0].extract_text()
        assert len(text) > 50, "PDF should be text-native with extractable text"

    def test_contains_confirmation_balance(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/bank_confirmation_fy2025.pdf"
        reader = PdfReader(str(path))
        all_text = "\n".join(p.extract_text() for p in reader.pages)
        confirmation_str = f"${int(CONFIRMATION_BALANCE):,}"
        assert confirmation_str in all_text, (
            f"Expected confirmation balance {confirmation_str} in PDF"
        )

    def test_contains_cascade_industries(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/bank_confirmation_fy2025.pdf"
        reader = PdfReader(str(path))
        all_text = "\n".join(p.extract_text() for p in reader.pages)
        assert "Cascade Industries" in all_text

    def test_contains_first_national_bank(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/bank_confirmation_fy2025.pdf"
        reader = PdfReader(str(path))
        all_text = "\n".join(p.extract_text() for p in reader.pages)
        assert "First National Bank" in all_text.replace("\n", " ")

    def test_contains_account_number(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/bank_confirmation_fy2025.pdf"
        reader = PdfReader(str(path))
        all_text = "\n".join(p.extract_text() for p in reader.pages)
        assert "4782-0091" in all_text

    def test_mentions_december_2025(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/bank_confirmation_fy2025.pdf"
        reader = PdfReader(str(path))
        all_text = "\n".join(p.extract_text() for p in reader.pages)
        assert "December 31, 2025" in all_text

    def test_canary_in_author_metadata(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("bank_confirmation_fy2025")
        path = output / f"{_INPUT_DIR}/bank_confirmation_fy2025.pdf"
        reader = PdfReader(str(path))
        author = reader.metadata.author or ""
        assert canary in author, f"Canary {canary} not in PDF author: {author}"


# ---------------------------------------------------------------------------
# ERR-002 — Transposed digits in bank transaction amount
# ---------------------------------------------------------------------------


class TestERR002PlantedError:
    """Verify ERR-002: transposed digits in the 5th bank transaction."""

    def test_err002_registered(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        assert "ERR-002" in errors.entries

    def test_err002_is_transposed_digits(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-002"]
        assert err.type == "transposed_digits"

    def test_err002_references_bank_statement(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-002"]
        assert "bank_statement_dec2025.csv" in err.file

    def test_err002_references_row_6(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-002"]
        assert "Row 6" in err.location or "transaction 5" in err.location.lower()


# ---------------------------------------------------------------------------
# ERR-004 — Mismatched total in GL ending balance
# ---------------------------------------------------------------------------


class TestERR004PlantedError:
    """Verify ERR-004: mismatched total in GL ending balance."""

    def test_err004_registered(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        assert "ERR-004" in errors.entries

    def test_err004_is_mismatched_total(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-004"]
        assert err.type == "mismatched_total"

    def test_err004_references_gl_file(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-004"]
        assert "cascade_gl_cash_dec2025.xlsx" in err.file

    def test_err004_references_ending_balance(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-004"]
        assert "ending balance" in err.location.lower() or "ending balance" in err.description.lower()


# ---------------------------------------------------------------------------
# Canary embedding
# ---------------------------------------------------------------------------


class TestCanaryEmbedding:
    """Verify canary codes are embedded in all files."""

    def test_all_canary_keys_assigned(self) -> None:
        _, _, canaries, _, _ = _ensure_emitted()
        for key in _CANARY_KEYS:
            code = canaries.canary_for(key)
            assert len(code) == 8, f"Canary for {key} should be 8 chars"

    def test_bank_csv_canary_in_comment(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("bank_statement_dec2025")
        path = output / f"{_INPUT_DIR}/bank_statement_dec2025.csv"
        text = path.read_text()
        assert canary in text, f"Canary {canary} not in bank statement CSV"

    def test_gl_cash_canary_in_xlsx(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("cascade_gl_cash_dec2025")
        path = output / f"{_INPUT_DIR}/cascade_gl_cash_dec2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        desc = wb.properties.description or ""
        assert canary in desc, f"Canary {canary} not in GL cash xlsx properties"

    def test_pdf_canary_in_metadata(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("bank_confirmation_fy2025")
        path = output / f"{_INPUT_DIR}/bank_confirmation_fy2025.pdf"
        reader = PdfReader(str(path))
        author = reader.metadata.author or ""
        assert canary in author, f"Canary {canary} not in PDF author metadata"


# ---------------------------------------------------------------------------
# Gold standard
# ---------------------------------------------------------------------------


class TestGoldStandard:
    """Verify gold standard JSON structure and content."""

    def test_gold_json_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / "gold_standards" / "TC-02_gold.json"
        assert path.exists()

    def test_gold_has_expected_outputs(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-02_gold.json").read_text())
        eo = gold["expected_outputs"]
        assert eo["file_type"] == "xlsx"
        assert "reconciliation" in eo
        assert "outstanding_checks_detail" in eo
        assert "deposits_in_transit_detail" in eo

    def test_gold_adjusted_balances_agree(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-02_gold.json").read_text())
        recon = gold["expected_outputs"]["reconciliation"]
        assert recon["adjusted_bank_balance"] == int(ADJUSTED_BALANCE)
        assert recon["adjusted_book_balance"] == int(ADJUSTED_BALANCE)
        assert recon["balances_agree"] is True

    def test_gold_confirmation_balance(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-02_gold.json").read_text())
        assert gold["expected_outputs"]["confirmation_balance"] == int(CONFIRMATION_BALANCE)
        assert gold["expected_outputs"]["confirmation_ties_to_bank_ending"] is True

    def test_gold_reconciliation_amounts(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-02_gold.json").read_text())
        recon = gold["expected_outputs"]["reconciliation"]
        assert recon["bank_ending_balance"] == int(CONFIRMATION_BALANCE)
        assert recon["deposits_in_transit"] == int(TOTAL_DEPOSITS_IN_TRANSIT)
        assert recon["outstanding_checks"] == int(TOTAL_OUTSTANDING_CHECKS)
        assert recon["gl_ending_balance"] == int(GL_ENDING_BALANCE)
        assert recon["bank_interest"] == int(BANK_INTEREST)
        assert recon["bank_service_charges"] == int(BANK_SERVICE_CHARGES)

    def test_gold_outstanding_checks_count(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-02_gold.json").read_text())
        assert len(gold["expected_outputs"]["outstanding_checks_detail"]) == 4

    def test_gold_deposits_in_transit_count(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-02_gold.json").read_text())
        assert len(gold["expected_outputs"]["deposits_in_transit_detail"]) == 2

    def test_gold_transaction_counts(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-02_gold.json").read_text())
        tc = gold["expected_outputs"]["transaction_count"]
        assert tc["bank_statement_rows"] == 340
        assert tc["outstanding_checks"] == 4
        assert tc["deposits_in_transit"] == 2
        assert tc["bank_only_items"] == 2

    def test_gold_canary_verification(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-02_gold.json").read_text())
        cv = gold["canary_verification"]
        for key in ["read_bank_statement", "read_gl_cash_detail", "read_bank_confirmation"]:
            assert key in cv, f"Missing canary verification key: {key}"

    def test_gold_scoring_hints(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-02_gold.json").read_text())
        hints = gold["scoring_hints"]
        for key in ["correctness", "completeness", "format_compliance", "robustness", "communication"]:
            assert key in hints


# ---------------------------------------------------------------------------
# Prompt and expected behavior
# ---------------------------------------------------------------------------


class TestPromptAndExpectedBehavior:
    """Verify prompt and expected behavior files are generated."""

    def test_prompt_md_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / "test_cases/TC-02/prompt.md"
        assert path.exists()

    def test_prompt_mentions_bank_reconciliation(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-02/prompt.md").read_text()
        assert "bank reconciliation" in text.lower()

    def test_prompt_mentions_key_tasks(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-02/prompt.md").read_text().lower()
        assert "outstanding checks" in text
        assert "deposits in transit" in text
        assert "fuzzy matching" in text or "fuzzy" in text

    def test_prompt_mentions_confirmation(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-02/prompt.md").read_text().lower()
        assert "confirmation" in text

    def test_expected_behavior_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / "test_cases/TC-02/expected_behavior.md"
        assert path.exists()

    def test_expected_behavior_mentions_adjusted_balance(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-02/expected_behavior.md").read_text()
        assert "$4,287,331" in text

    def test_expected_behavior_mentions_confirmation_balance(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-02/expected_behavior.md").read_text()
        assert "$4,312,117" in text

    def test_expected_behavior_mentions_outstanding_checks(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-02/expected_behavior.md").read_text()
        assert "outstanding checks" in text.lower()
        assert "4" in text  # 4 outstanding checks

    def test_expected_behavior_mentions_deposits_in_transit(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-02/expected_behavior.md").read_text()
        assert "deposits in transit" in text.lower()


# ---------------------------------------------------------------------------
# Manifest registration
# ---------------------------------------------------------------------------


class TestManifest:
    """Verify files are registered in the manifest."""

    def test_manifest_has_tc02_entries(self) -> None:
        _, _, _, _, manifest = _ensure_emitted()
        tc02_count = sum(
            1 for v in manifest.entries.values()
            if "TC-02" in (v.test_cases or [])
        )
        # bank_csv + gl_cash_xlsx + confirmation_pdf = 3
        assert tc02_count >= 3, f"Expected ≥3 TC-02 manifest entries, got {tc02_count}"
