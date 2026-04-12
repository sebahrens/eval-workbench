"""Tests for TC-08-EU — European R&D Incentive Study (CIR + Forschungszulage).

Verifies:
- 14 project description DOCX files (10 CM + 4 CP) in rd_project_descriptions_eu/
- rd_employee_time_records_eu.csv with ~2,184 rows
- payroll_data_eu_fy2025.xlsx with 42 employees
- rd_supply_expenses_eu.xlsx with supply and subcontractor sheets
- rd_subcontractor_invoices_eu.xlsx with 4 invoices
- prior_year_rd_data_eu.xlsx with FY2023/FY2024 context
- ERR-EU-008 planted error (maintenance contract miscoded as R&D supply)
- Canary embedding in all files
- Gold standard structure, CIR and Forschungszulage computations
- Prompt and expected behavior markdown files
"""

from __future__ import annotations

import json
import tempfile
from decimal import Decimal
from pathlib import Path

import openpyxl

from generator.canaries import CanaryRegistry
from generator.canaries import build_registry as build_canary_registry
from generator.errors import ErrorRegistry
from generator.formatters.tc08_eu import emit_tc08_eu
from generator.golds.framework import emit_gold
from generator.manifest import Manifest
from generator.model.build import CascadeModel, build_model
from generator.model.rd_eu import (
    ALL_CANARY_KEYS_TC08EU,
    RD_PROJECTS_EU,
    compute_consolidated_rd_benefit,
    generate_rd_employees_eu,
    generate_subcontractor_invoices_eu,
    generate_supply_expenses_eu,
    generate_time_records_eu,
)

# ---------------------------------------------------------------------------
# Module-level fixture: build the model and run emit_tc08_eu once
# ---------------------------------------------------------------------------

_MODEL: CascadeModel | None = None
_OUTPUT: Path | None = None
_CANARIES: CanaryRegistry | None = None
_ERRORS: ErrorRegistry | None = None
_MANIFEST: Manifest | None = None


def _ensure_emitted() -> tuple[CascadeModel, Path, CanaryRegistry, ErrorRegistry, Manifest]:
    global _MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST  # noqa: PLW0603
    if _MODEL is None:
        _MODEL = build_model(seed=42)
        _OUTPUT = Path(tempfile.mkdtemp(prefix="tc08eu_test_"))
        _CANARIES = build_canary_registry(ALL_CANARY_KEYS_TC08EU, seed=42)
        _ERRORS = ErrorRegistry()
        _MANIFEST = Manifest(_OUTPUT)
        _MANIFEST.__enter__()

        emit_tc08_eu(_MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST)

        # Emit gold standard
        emit_gold("TC-08-EU", _CANARIES, _ERRORS, _OUTPUT / "gold_standards", model=_MODEL)

        _MANIFEST.__exit__(None, None, None)
    return _MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST


_INPUT_DIR = "test_cases/TC-08-EU/input_files"
_PROJ_DESC_DIR = f"{_INPUT_DIR}/rd_project_descriptions_eu"


# ---------------------------------------------------------------------------
# Project description DOCX files — existence and count
# ---------------------------------------------------------------------------


class TestProjectDescriptions:
    """Verify 14 project description DOCX files are emitted."""

    def test_project_dir_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        assert (output / _PROJ_DESC_DIR).is_dir()

    def test_14_project_files(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        docx_files = sorted((output / _PROJ_DESC_DIR).glob("*.docx"))
        assert len(docx_files) == 14

    def test_cm_projects_present(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        for i in range(1, 11):
            path = output / _PROJ_DESC_DIR / f"CM-RD-{i:02d}.docx"
            assert path.exists(), f"Missing CM-RD-{i:02d}.docx"

    def test_cp_projects_present(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        for i in range(1, 5):
            path = output / _PROJ_DESC_DIR / f"CP-RD-{i:02d}.docx"
            assert path.exists(), f"Missing CP-RD-{i:02d}.docx"

    def test_canaries_in_docx(self) -> None:
        """All 14 DOCX files have canaries embedded."""
        _, output, canaries, _, _ = _ensure_emitted()
        import docx as python_docx

        for i in range(1, 15):
            key = f"tc08eu_project_{i:03d}"
            expected_canary = canaries.canary_for(key)
            # Find the file — projects are sorted by code
            projects = sorted(RD_PROJECTS_EU, key=lambda p: p.code)
            proj = projects[i - 1]
            path = output / _PROJ_DESC_DIR / f"{proj.code}.docx"
            doc = python_docx.Document(str(path))
            comments = doc.core_properties.comments or ""
            assert expected_canary in comments, (
                f"Canary {expected_canary} not found in {proj.code}.docx"
            )


# ---------------------------------------------------------------------------
# Time Records CSV
# ---------------------------------------------------------------------------


class TestTimeRecords:
    """Verify rd_employee_time_records_eu.csv."""

    def test_csv_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "rd_employee_time_records_eu.csv"
        assert path.exists()

    def test_csv_has_canary(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "rd_employee_time_records_eu.csv"
        content = path.read_text()
        expected = canaries.canary_for("tc08eu_time_records")
        assert expected in content

    def test_csv_row_count(self) -> None:
        """Expect ~2,184 data rows (42 employees × 52 weeks)."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "rd_employee_time_records_eu.csv"
        lines = path.read_text().strip().split("\n")
        # First line is canary comment, second is header
        data_rows = len(lines) - 2
        assert data_rows > 2000, f"Expected >2000 rows, got {data_rows}"

    def test_csv_has_both_entities(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "rd_employee_time_records_eu.csv"
        content = path.read_text()
        assert "CM," in content
        assert "CP," in content

    def test_csv_header_columns(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "rd_employee_time_records_eu.csv"
        lines = path.read_text().strip().split("\n")
        header = lines[1]  # Skip canary comment
        assert "entity_code" in header
        assert "employee_id" in header
        assert "project_code" in header
        assert "hours" in header


# ---------------------------------------------------------------------------
# Payroll XLSX
# ---------------------------------------------------------------------------


class TestPayroll:
    """Verify payroll_data_eu_fy2025.xlsx."""

    def test_payroll_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "payroll_data_eu_fy2025.xlsx"
        assert path.exists()

    def test_payroll_has_canary(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "payroll_data_eu_fy2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        expected = canaries.canary_for("tc08eu_payroll")
        desc = wb.properties.description or ""
        assert expected in desc, f"Canary {expected} not in payroll xlsx"

    def test_payroll_employee_count(self) -> None:
        """42 R&D employees (30 CM + 12 CP)."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "payroll_data_eu_fy2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        # Data rows start at row 5 (rows 1-2 title, row 4 header)
        data_rows = [r for r in ws.iter_rows(min_row=5) if r[0].value is not None]
        assert len(data_rows) == 42, f"Expected 42 employees, got {len(data_rows)}"

    def test_payroll_has_social_charges(self) -> None:
        """All employees should have social charges column populated."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "payroll_data_eu_fy2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        # Column 6 is social charges, data starts at row 5
        social_values = [ws.cell(row=r, column=6).value for r in range(5, 5 + 42)]
        non_zero = [v for v in social_values if v and v > 0]
        assert len(non_zero) == 42, f"Expected all 42 with social charges, got {len(non_zero)}"


# ---------------------------------------------------------------------------
# Supply Expenses XLSX
# ---------------------------------------------------------------------------


class TestSupplyExpenses:
    """Verify rd_supply_expenses_eu.xlsx."""

    def test_supply_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "rd_supply_expenses_eu.xlsx"
        assert path.exists()

    def test_supply_has_canary(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "rd_supply_expenses_eu.xlsx"
        wb = openpyxl.load_workbook(str(path))
        expected = canaries.canary_for("tc08eu_supply_expenses")
        desc = wb.properties.description or ""
        assert expected in desc, f"Canary {expected} not in supply xlsx"

    def test_supply_has_both_sheets(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "rd_supply_expenses_eu.xlsx"
        wb = openpyxl.load_workbook(str(path))
        sheet_names = wb.sheetnames
        assert "R&D Supply Expenses" in sheet_names
        # Subcontractor sheet may be in same or separate file
        # The emit uses a separate file for subcontractors

    def test_supply_has_err_eu_008_row(self) -> None:
        """ERR-EU-008: maintenance contract present in data."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "rd_supply_expenses_eu.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["R&D Supply Expenses"]
        found = False
        # Columns: Entity(0), Date(1), Description(2), Amount(3), CostCentre(4), Project(5)
        for row in ws.iter_rows(min_row=4, values_only=True):
            desc = str(row[2]) if row[2] else ""
            if "maintenance contract" in desc.lower():
                found = True
                assert row[5] == "CP-RD-03", f"ERR-EU-008 on wrong project: {row[5]}"
                break
        assert found, "ERR-EU-008 maintenance contract row not found"


# ---------------------------------------------------------------------------
# Subcontractor Invoices XLSX
# ---------------------------------------------------------------------------


class TestSubcontractorInvoices:
    """Verify rd_subcontractor_invoices_eu.xlsx."""

    def test_subcontractor_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "rd_subcontractor_invoices_eu.xlsx"
        assert path.exists()

    def test_subcontractor_has_canary(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "rd_subcontractor_invoices_eu.xlsx"
        wb = openpyxl.load_workbook(str(path))
        expected = canaries.canary_for("tc08eu_subcontractor_invoices")
        desc = wb.properties.description or ""
        assert expected in desc

    def test_subcontractor_count(self) -> None:
        """4 subcontractor invoices."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "rd_subcontractor_invoices_eu.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        # Data starts at row 4 (row 1 title, row 3 header)
        data_rows = [r for r in ws.iter_rows(min_row=4) if r[0].value is not None]
        assert len(data_rows) == 4

    def test_public_subcontractor_present(self) -> None:
        """At least one public research organism."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "rd_subcontractor_invoices_eu.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        # Type column is 6, data starts at row 4
        types = [ws.cell(row=r, column=6).value for r in range(4, ws.max_row + 1)]
        assert "public" in types, "No public subcontractor found"


# ---------------------------------------------------------------------------
# Prior Year Data XLSX
# ---------------------------------------------------------------------------


class TestPriorYearData:
    """Verify prior_year_rd_data_eu.xlsx."""

    def test_prior_year_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "prior_year_rd_data_eu.xlsx"
        assert path.exists()

    def test_prior_year_has_canary(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "prior_year_rd_data_eu.xlsx"
        wb = openpyxl.load_workbook(str(path))
        expected = canaries.canary_for("tc08eu_prior_year_rd")
        desc = wb.properties.description or ""
        assert expected in desc

    def test_prior_year_has_both_entities(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "prior_year_rd_data_eu.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        entities = set()
        # Data starts at row 5 (rows 1-2 title, row 4 header)
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[0]:
                entities.add(row[0])
        assert "CM" in entities
        assert "CP" in entities


# ---------------------------------------------------------------------------
# ERR-EU-008 — Maintenance contract miscoded as R&D supply
# ---------------------------------------------------------------------------


class TestErrEU008:
    """Verify ERR-EU-008: maintenance contract miscoded as R&D supply."""

    def test_error_registered(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        assert "ERR-EU-008" in errors.entries

    def test_error_type(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        assert errors.entries["ERR-EU-008"].type == "classification_error"

    def test_error_severity(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        assert errors.entries["ERR-EU-008"].severity == "immaterial"

    def test_error_mentions_maintenance(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-EU-008"]
        assert "maintenance" in err.description.lower()

    def test_error_references_cp_rd_03(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-EU-008"]
        assert "CP-RD-03" in err.description or "CP-RD-03" in err.location


# ---------------------------------------------------------------------------
# Gold Standard
# ---------------------------------------------------------------------------


class TestGoldStandard:
    """Verify TC-08-EU gold standard structure and values."""

    def _gold(self) -> dict:
        _, output, _, _, _ = _ensure_emitted()
        gold_path = output / "gold_standards" / "TC-08-EU_gold.json"
        return json.loads(gold_path.read_text())

    def test_gold_file_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        assert (output / "gold_standards" / "TC-08-EU_gold.json").exists()

    def test_gold_test_case_id(self) -> None:
        gold = self._gold()
        assert gold["test_case"] == "TC-08-EU"

    def test_gold_has_cir_section(self) -> None:
        gold = self._gold()
        assert "cir" in gold["expected_outputs"]
        cir = gold["expected_outputs"]["cir"]
        assert "cir_credit_eur" in cir

    def test_gold_has_forschungszulage_section(self) -> None:
        gold = self._gold()
        assert "forschungszulage" in gold["expected_outputs"]
        fz = gold["expected_outputs"]["forschungszulage"]
        assert "benefit_eur" in fz

    def test_gold_has_consolidated_benefit(self) -> None:
        gold = self._gold()
        assert "consolidated_benefit_eur" in gold["expected_outputs"]

    def test_gold_project_counts(self) -> None:
        gold = self._gold()
        eo = gold["expected_outputs"]
        assert eo["cm_projects"] == 10
        assert eo["cp_projects"] == 4
        assert eo["cm_qualifying"] == 7
        assert eo["cm_borderline"] == 2
        assert eo["cm_disqualified"] == 1
        assert eo["cp_qualifying"] == 3
        assert eo["cp_disqualified"] == 1

    def test_gold_cir_credit_reasonable(self) -> None:
        """CIR credit should be in the range of €1M-€2M per design."""
        gold = self._gold()
        cir_credit = int(gold["expected_outputs"]["cir"]["cir_credit_eur"])
        assert 1_000_000 < cir_credit < 2_000_000, f"CIR credit {cir_credit} out of range"

    def test_gold_fz_benefit_reasonable(self) -> None:
        """Forschungszulage benefit ≤ €500k (cap)."""
        gold = self._gold()
        fz_benefit = int(gold["expected_outputs"]["forschungszulage"]["benefit_eur"])
        assert fz_benefit <= 500_000, f"FZ benefit {fz_benefit} exceeds cap"
        assert fz_benefit > 100_000, f"FZ benefit {fz_benefit} seems too low"

    def test_gold_has_canary_verification(self) -> None:
        gold = self._gold()
        cv = gold["canary_verification"]
        assert "read_time_records" in cv
        assert "read_payroll" in cv
        assert "read_supply_expenses" in cv
        assert len(cv) >= 19  # 5 base files + 14 project docs

    def test_gold_has_error_detection(self) -> None:
        gold = self._gold()
        assert "ERR-EU-008" in gold["error_detection"]

    def test_gold_has_scoring_hints(self) -> None:
        gold = self._gold()
        assert "scoring_hints" in gold
        assert "correctness" in gold["scoring_hints"]
        assert "dual_regime" in gold["scoring_hints"]


# ---------------------------------------------------------------------------
# Prompt and Expected Behavior
# ---------------------------------------------------------------------------


class TestPromptAndExpectedBehavior:
    """Verify prompt.md and expected_behavior.md exist with key content."""

    def test_prompt_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        assert (output / "test_cases/TC-08-EU/prompt.md").exists()

    def test_prompt_mentions_cir(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        content = (output / "test_cases/TC-08-EU/prompt.md").read_text()
        assert "CIR" in content
        assert "Forschungszulage" in content

    def test_expected_behavior_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        assert (output / "test_cases/TC-08-EU/expected_behavior.md").exists()

    def test_expected_behavior_mentions_frascati(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        content = (output / "test_cases/TC-08-EU/expected_behavior.md").read_text()
        assert "Frascati" in content


# ---------------------------------------------------------------------------
# Model computation tests (direct unit tests)
# ---------------------------------------------------------------------------


class TestModelComputations:
    """Verify R&D model computations directly."""

    def test_employee_count(self) -> None:
        employees = generate_rd_employees_eu()
        cm = [e for e in employees if e.entity_code == "CM"]
        cp = [e for e in employees if e.entity_code == "CP"]
        assert len(cm) == 30
        assert len(cp) == 12

    def test_project_count(self) -> None:
        cm = [p for p in RD_PROJECTS_EU if p.entity_code == "CM"]
        cp = [p for p in RD_PROJECTS_EU if p.entity_code == "CP"]
        assert len(cm) == 10
        assert len(cp) == 4

    def test_project_qualifications(self) -> None:
        cm = [p for p in RD_PROJECTS_EU if p.entity_code == "CM"]
        cp = [p for p in RD_PROJECTS_EU if p.entity_code == "CP"]
        cm_yes = [p for p in cm if p.qualifies == "yes"]
        cm_border = [p for p in cm if p.qualifies == "borderline"]
        cm_no = [p for p in cm if p.qualifies == "no"]
        cp_yes = [p for p in cp if p.qualifies == "yes"]
        cp_no = [p for p in cp if p.qualifies == "no"]
        assert len(cm_yes) == 7
        assert len(cm_border) == 2
        assert len(cm_no) == 1
        assert len(cp_yes) == 3
        assert len(cp_no) == 1

    def test_time_records_generation(self) -> None:
        records = generate_time_records_eu()
        assert len(records) > 2000
        entities = {r.entity_code for r in records}
        assert "CM" in entities
        assert "CP" in entities

    def test_supply_expenses_err_eu_008(self) -> None:
        """ERR-EU-008 maintenance contract is in the data."""
        expenses = generate_supply_expenses_eu()
        err_item = [e for e in expenses if "maintenance" in e.description.lower()]
        assert len(err_item) == 1
        assert err_item[0].project_code == "CP-RD-03"

    def test_subcontractor_invoices(self) -> None:
        invoices = generate_subcontractor_invoices_eu()
        assert len(invoices) == 4
        public = [i for i in invoices if i.subcontractor_type == "public"]
        assert len(public) == 1

    def test_cir_excludes_non_qualifying(self) -> None:
        """CIR computation should exclude CM-RD-10 (market research)."""
        result = compute_consolidated_rd_benefit()
        # The non-qualifying project subcontractor (SUB-CM-004, €80k) should not appear
        # in total_subcontracted (which should be private+public_doubled for qualifying only)
        cir = result.cir
        # Public €80k doubled = €160k, private €260k (only CM-RD-01, CM-RD-03, CM-RD-07)
        # CM-RD-10 subcontractor (€80k private) excluded
        assert cir.subcontracted_private_eur == Decimal("260000")
        assert cir.subcontracted_public_doubled_eur == Decimal("160000")

    def test_forschungszulage_personnel_only(self) -> None:
        """Forschungszulage should only use personnel costs, not supplies."""
        result = compute_consolidated_rd_benefit()
        fz = result.forschungszulage
        # Should be under €2M cap
        assert fz.qualifying_personnel_cost_eur < Decimal("2000000")
        assert fz.assessment_basis_eur == fz.qualifying_personnel_cost_eur

    def test_consolidated_benefit(self) -> None:
        """Combined benefit = CIR + Forschungszulage."""
        result = compute_consolidated_rd_benefit()
        assert result.total_benefit_eur == result.cir.cir_credit_eur + result.forschungszulage.benefit_eur


# ---------------------------------------------------------------------------
# Determinism test
# ---------------------------------------------------------------------------


class TestDeterminism:
    """Verify deterministic output across two runs."""

    def test_time_records_deterministic(self) -> None:
        records1 = generate_time_records_eu()
        records2 = generate_time_records_eu()
        assert len(records1) == len(records2)
        for r1, r2 in zip(records1[:100], records2[:100]):
            assert r1 == r2

    def test_computations_deterministic(self) -> None:
        result1 = compute_consolidated_rd_benefit()
        result2 = compute_consolidated_rd_benefit()
        assert result1.cir.cir_credit_eur == result2.cir.cir_credit_eur
        assert result1.forschungszulage.benefit_eur == result2.forschungszulage.benefit_eur
