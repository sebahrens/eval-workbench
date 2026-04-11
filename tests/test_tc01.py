"""Tests for TC-01 — Trial Balance Reconciliation (Audit, Complex) formatter.

Verifies:
- Messy client TB with merged header, inconsistent names, abbreviations
- Clean prior year workpaper with standardized names and lead schedule mappings
- 8-page signed FY2024 financial statements PDF
- ERR-001 planted error (transposed digit in A/R FY2025 balance)
- Canary embedding in all files
- Gold standard structure (mapping counts, variance flags, tie-out discrepancy)
- Prompt and expected behavior markdown files
"""

from __future__ import annotations

import json
import tempfile
from pathlib import Path

import openpyxl
from pypdf import PdfReader

from generator.canaries import CanaryRegistry
from generator.canaries import build_registry as build_canary_registry
from generator.errors import ErrorRegistry
from generator.formatters.tc01 import (
    _ABBREVIATIONS,
    _RENAMED_ACCOUNTS,
    emit_tc01,
)
from generator.golds.framework import emit_gold
from generator.manifest import Manifest
from generator.model.build import CascadeModel, build_model

# ---------------------------------------------------------------------------
# Module-level fixture: build the model and run emit_tc01 once
# ---------------------------------------------------------------------------

_MODEL: CascadeModel | None = None
_OUTPUT: Path | None = None
_CANARIES: CanaryRegistry | None = None
_ERRORS: ErrorRegistry | None = None
_MANIFEST: Manifest | None = None

# All canary keys used by TC-01
_CANARY_KEYS = sorted([
    "cascade_tb_fy2025",
    "cascade_tb_fy2024_workpaper",
    "cascade_financials_fy2024_signed",
])


def _ensure_emitted() -> tuple[CascadeModel, Path, CanaryRegistry, ErrorRegistry, Manifest]:
    global _MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST  # noqa: PLW0603
    if _MODEL is None:
        _MODEL = build_model(seed=42)
        _OUTPUT = Path(tempfile.mkdtemp(prefix="tc01_test_"))
        _CANARIES = build_canary_registry(_CANARY_KEYS, seed=42)
        _ERRORS = ErrorRegistry()
        _MANIFEST = Manifest(_OUTPUT)
        _MANIFEST.__enter__()

        emit_tc01(_MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST)

        # Emit gold standard
        emit_gold("TC-01", _CANARIES, _ERRORS, _OUTPUT / "gold_standards", model=_MODEL)

        _MANIFEST.__exit__(None, None, None)
    return _MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST


_INPUT_DIR = "test_cases/TC-01/input_files"


# ---------------------------------------------------------------------------
# Messy client TB — cascade_tb_fy2025.xlsx
# ---------------------------------------------------------------------------


class TestMessyClientTB:
    """Verify cascade_tb_fy2025.xlsx structure and messiness."""

    def test_file_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/cascade_tb_fy2025.xlsx"
        assert path.exists()

    def test_has_trial_balance_sheet(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/cascade_tb_fy2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        assert "Trial Balance" in wb.sheetnames

    def test_has_merged_cells_in_header(self) -> None:
        """The messy TB should have merged cells in the header area."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/cascade_tb_fy2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Trial Balance"]
        assert len(ws.merged_cells.ranges) > 0, "Expected merged cells in header"

    def test_has_abbreviations(self) -> None:
        """Some account names should use abbreviations from the spec."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/cascade_tb_fy2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Trial Balance"]
        names = set()
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=2).value
            if isinstance(val, str):
                names.add(val.strip())
        # Check at least one abbreviation is present
        abbrev_found = any(abbr in names for abbr in _ABBREVIATIONS.values())
        assert abbrev_found, "Expected abbreviated account names in messy TB"

    def test_has_debit_credit_columns(self) -> None:
        """TB should have debit and credit columns with numeric values."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/cascade_tb_fy2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Trial Balance"]
        # Find header row with Debit/Credit
        header_found = False
        for row in range(1, min(10, ws.max_row + 1)):
            vals = [ws.cell(row=row, column=c).value for c in range(1, ws.max_column + 1)]
            str_vals = [str(v).strip().lower() for v in vals if v is not None]
            if "debit" in str_vals and "credit" in str_vals:
                header_found = True
                break
        assert header_found, "Expected Debit and Credit column headers"

    def test_account_count(self) -> None:
        """TB should have a substantial number of accounts."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/cascade_tb_fy2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Trial Balance"]
        acct_count = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if isinstance(val, (int, str)) and str(val).strip().isdigit():
                acct_count += 1
        assert acct_count >= 50, f"Expected ≥50 accounts, got {acct_count}"


# ---------------------------------------------------------------------------
# Prior year workpaper — cascade_tb_fy2024_workpaper.xlsx
# ---------------------------------------------------------------------------


class TestPriorYearWorkpaper:
    """Verify cascade_tb_fy2024_workpaper.xlsx structure."""

    def test_file_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/cascade_tb_fy2024_workpaper.xlsx"
        assert path.exists()

    def test_has_clean_standardized_names(self) -> None:
        """Prior year workpaper should have clean, standardized account names."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/cascade_tb_fy2024_workpaper.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        names = []
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=2).value
            if isinstance(val, str) and val.strip():
                names.append(val.strip())
        # Clean names should not have leading spaces
        leading_space_count = sum(1 for n in names if n != n.lstrip())
        assert leading_space_count == 0, "Prior year WP should have clean names (no leading spaces)"

    def test_has_lead_schedule_mappings(self) -> None:
        """Workpaper should have a lead schedule column."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/cascade_tb_fy2024_workpaper.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        # Header row is 5 (rows 1-3 are title, row 4 blank)
        headers = []
        for row in range(1, min(10, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=row, column=col).value
                if isinstance(val, str) and "lead" in val.lower():
                    headers.append(val)
        assert len(headers) > 0, "Expected a 'Lead Schedule' column header"


# ---------------------------------------------------------------------------
# Signed financials PDF — cascade_financials_fy2024_signed.pdf
# ---------------------------------------------------------------------------


class TestSignedFinancialsPDF:
    """Verify cascade_financials_fy2024_signed.pdf structure."""

    def test_file_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/cascade_financials_fy2024_signed.pdf"
        assert path.exists()

    def test_has_multiple_pages(self) -> None:
        """PDF should have multiple pages (cover, BS, IS, CF, notes)."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/cascade_financials_fy2024_signed.pdf"
        reader = PdfReader(str(path))
        assert len(reader.pages) >= 5, f"Expected ≥5 pages, got {len(reader.pages)}"

    def test_is_text_native(self) -> None:
        """PDF should be text-native (extractable text)."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/cascade_financials_fy2024_signed.pdf"
        reader = PdfReader(str(path))
        text = reader.pages[0].extract_text()
        assert len(text) > 50, "Page 1 should have extractable text"

    def test_contains_cascade_industries(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/cascade_financials_fy2024_signed.pdf"
        reader = PdfReader(str(path))
        all_text = "\n".join(p.extract_text() for p in reader.pages)
        assert "Cascade Industries" in all_text

    def test_contains_balance_sheet_and_income_statement(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/cascade_financials_fy2024_signed.pdf"
        reader = PdfReader(str(path))
        all_text = "\n".join(p.extract_text() for p in reader.pages)
        assert "Balance Sheet" in all_text
        assert "Income Statement" in all_text or "Statement of Income" in all_text

    def test_canary_in_author_metadata(self) -> None:
        """Canary should be in PDF author field."""
        _, output, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("cascade_financials_fy2024_signed")
        path = output / f"{_INPUT_DIR}/cascade_financials_fy2024_signed.pdf"
        reader = PdfReader(str(path))
        author = reader.metadata.author or ""
        assert canary in author, f"Canary {canary} not in PDF author: {author}"


# ---------------------------------------------------------------------------
# ERR-001 — Transposed digits in A/R FY2025 balance
# ---------------------------------------------------------------------------


class TestERR001PlantedError:
    """Verify ERR-001: transposed digit in Accounts Receivable balance."""

    def test_err001_registered(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        assert "ERR-001" in errors.entries

    def test_err001_is_transposed_digits(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-001"]
        assert err.type == "transposed_digits"

    def test_err001_references_ar(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-001"]
        assert "1100" in err.location or "receivable" in err.description.lower()

    def test_err001_in_fy2025_tb_file(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-001"]
        assert "cascade_tb_fy2025.xlsx" in err.file


# ---------------------------------------------------------------------------
# Canary embedding
# ---------------------------------------------------------------------------


class TestCanaryEmbedding:
    """Verify canary codes are embedded in all files."""

    def test_all_canary_keys_assigned(self) -> None:
        _, _, canaries, _, _ = _ensure_emitted()
        for key in _CANARY_KEYS:
            code = canaries.canary_for(key)
            assert len(code) == 8, f"Canary for {key} should be 8 chars"

    def test_tb_fy2025_canary_in_xlsx(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("cascade_tb_fy2025")
        path = output / f"{_INPUT_DIR}/cascade_tb_fy2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        desc = wb.properties.description or ""
        assert canary in desc, f"Canary {canary} not in FY2025 TB properties"

    def test_workpaper_canary_in_xlsx(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("cascade_tb_fy2024_workpaper")
        path = output / f"{_INPUT_DIR}/cascade_tb_fy2024_workpaper.xlsx"
        wb = openpyxl.load_workbook(str(path))
        desc = wb.properties.description or ""
        assert canary in desc, f"Canary {canary} not in FY2024 workpaper properties"

    def test_pdf_canary_in_metadata(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("cascade_financials_fy2024_signed")
        path = output / f"{_INPUT_DIR}/cascade_financials_fy2024_signed.pdf"
        reader = PdfReader(str(path))
        author = reader.metadata.author or ""
        assert canary in author, f"Canary {canary} not in PDF author metadata"


# ---------------------------------------------------------------------------
# Gold standard
# ---------------------------------------------------------------------------


class TestGoldStandard:
    """Verify gold standard JSON structure and content."""

    def test_gold_json_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / "gold_standards" / "TC-01_gold.json"
        assert path.exists()

    def test_gold_has_expected_outputs(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-01_gold.json").read_text())
        eo = gold["expected_outputs"]
        assert eo["file_type"] == "xlsx"
        assert "mapping" in eo
        assert "variance_analysis" in eo
        assert "tie_out" in eo

    def test_gold_required_sheets(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-01_gold.json").read_text())
        sheets = gold["expected_outputs"]["required_sheets"]
        assert set(sheets) == {"Mapping", "Variance Analysis", "Exceptions", "Tie-Out"}

    def test_gold_mapping_counts(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-01_gold.json").read_text())
        m = gold["expected_outputs"]["mapping"]
        assert isinstance(m["new_accounts_flagged"], int)
        assert isinstance(m["missing_accounts_flagged"], int)
        assert m["renamed_accounts_correctly_identified"] == len(_RENAMED_ACCOUNTS)
        assert m["total_accounts_mapped"] > 50

    def test_gold_variance_analysis(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-01_gold.json").read_text())
        va = gold["expected_outputs"]["variance_analysis"]
        assert va["flagged_accounts_count"] > 0
        assert len(va["flagged_accounts"]) == va["flagged_accounts_count"]
        assert va["largest_variance_account"] != ""
        assert va["largest_variance_pct"] > 0

    def test_gold_tie_out_shows_err001(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-01_gold.json").read_text())
        to = gold["expected_outputs"]["tie_out"]
        assert to["discrepancies_found"] == 1
        assert "ERR-001" in to["discrepancy_details"]

    def test_gold_canary_verification(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-01_gold.json").read_text())
        cv = gold["canary_verification"]
        for key in ["read_correct_tb", "read_correct_prior_wp", "read_correct_pdf"]:
            assert key in cv, f"Missing canary verification key: {key}"

    def test_gold_error_detection(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-01_gold.json").read_text())
        assert "ERR-001" in gold["error_detection"]

    def test_gold_scoring_hints(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-01_gold.json").read_text())
        hints = gold["scoring_hints"]
        for key in ["correctness", "completeness", "format_compliance", "robustness", "communication"]:
            assert key in hints


# ---------------------------------------------------------------------------
# Prompt and expected behavior
# ---------------------------------------------------------------------------


class TestPromptAndExpectedBehavior:
    """Verify prompt and expected behavior files are generated."""

    def test_prompt_md_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / "test_cases/TC-01/prompt.md"
        assert path.exists()

    def test_prompt_mentions_trial_balance(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-01/prompt.md").read_text()
        assert "trial balance" in text.lower()

    def test_prompt_mentions_key_tasks(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-01/prompt.md").read_text().lower()
        assert "map" in text
        assert "variance" in text
        assert "tie" in text or "verify" in text

    def test_prompt_mentions_required_sheets(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-01/prompt.md").read_text()
        assert "Mapping" in text
        assert "Variance Analysis" in text
        assert "Exceptions" in text
        assert "Tie-Out" in text

    def test_expected_behavior_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / "test_cases/TC-01/expected_behavior.md"
        assert path.exists()

    def test_expected_behavior_mentions_merged_cells(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-01/expected_behavior.md").read_text()
        assert "merged" in text.lower()

    def test_expected_behavior_mentions_err001(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-01/expected_behavior.md").read_text()
        assert "ERR-001" in text

    def test_expected_behavior_mentions_abbreviations(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-01/expected_behavior.md").read_text()
        assert "abbreviat" in text.lower() or "Accts Recv" in text


# ---------------------------------------------------------------------------
# Manifest registration
# ---------------------------------------------------------------------------


class TestManifest:
    """Verify files are registered in the manifest."""

    def test_manifest_has_tc01_entries(self) -> None:
        _, _, _, _, manifest = _ensure_emitted()
        tc01_count = sum(
            1 for v in manifest.entries.values()
            if "TC-01" in (v.test_cases or [])
        )
        # tb_fy2025 + workpaper + pdf = 3
        assert tc01_count >= 3, f"Expected ≥3 TC-01 manifest entries, got {tc01_count}"
