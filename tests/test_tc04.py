"""Tests for TC-04 — Lease Extraction & ASC 842 Schedule Population (Audit, Adversarial).

Verifies:
- 15 lease PDFs in leases/ (10 text-native, 3 scanned-style, 2 with amendments)
- lease_schedule_partial.xlsx (8 of 15 leases pre-filled, 3 with blank fields)
- ERR-007 planted error (LS-002 commencement date off by one month)
- 2 leases qualify for short-term exemption (LS-009, LS-013)
- Canary embedding in all files
- Gold standard structure and scoring hints
- Prompt and expected behavior markdown files
"""

from __future__ import annotations

import json
import tempfile
from pathlib import Path

import openpyxl

from generator.canaries import CanaryRegistry
from generator.canaries import build_registry as build_canary_registry
from generator.errors import ErrorRegistry
from generator.formatters.tc04 import (
    _AMENDMENT_PDF_IDS,
    _BLANK_FIELDS_IDS,
    _PARTIAL_LEASE_IDS,
    _SCANNED_IDS,
    _SCHEDULE_KEY,
    emit_tc04,
)
from generator.golds.framework import emit_gold
from generator.manifest import Manifest
from generator.model.build import CascadeModel, build_model

# ---------------------------------------------------------------------------
# Module-level fixture: build the model and run emit_tc04 once
# ---------------------------------------------------------------------------

_MODEL: CascadeModel | None = None
_OUTPUT: Path | None = None
_CANARIES: CanaryRegistry | None = None
_ERRORS: ErrorRegistry | None = None
_MANIFEST: Manifest | None = None

# All canary keys used by TC-04 (15 lease PDFs + 1 schedule)
_CANARY_KEYS = sorted(
    [f"tc04_lease_{i:03d}" for i in range(1, 16)]
    + [_SCHEDULE_KEY]
)


def _ensure_emitted() -> tuple[CascadeModel, Path, CanaryRegistry, ErrorRegistry, Manifest]:
    global _MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST  # noqa: PLW0603
    if _MODEL is None:
        _MODEL = build_model(seed=42)
        _OUTPUT = Path(tempfile.mkdtemp(prefix="tc04_test_"))
        _CANARIES = build_canary_registry(_CANARY_KEYS, seed=42)
        _ERRORS = ErrorRegistry()
        _MANIFEST = Manifest(_OUTPUT)
        _MANIFEST.__enter__()

        emit_tc04(_MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST)

        # Emit gold standard
        emit_gold("TC-04", _CANARIES, _ERRORS, _OUTPUT / "gold_standards", model=_MODEL)

        _MANIFEST.__exit__(None, None, None)
    return _MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST


_INPUT_DIR = "test_cases/TC-04/input_files"


# ---------------------------------------------------------------------------
# Lease PDFs — existence and counts
# ---------------------------------------------------------------------------


class TestLeasePDFs:
    """Verify 15 lease PDFs exist in leases/ directory."""

    def test_leases_dir_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/leases"
        assert path.is_dir()

    def test_15_lease_pdfs(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        leases_dir = output / f"{_INPUT_DIR}/leases"
        pdfs = sorted(leases_dir.glob("LS-*.pdf"))
        assert len(pdfs) == 15, f"Expected 15 lease PDFs, got {len(pdfs)}"

    def test_lease_ids_ls001_through_ls015(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        leases_dir = output / f"{_INPUT_DIR}/leases"
        expected = {f"LS-{i:03d}.pdf" for i in range(1, 16)}
        actual = {p.name for p in leases_dir.glob("LS-*.pdf")}
        assert actual == expected

    def test_10_text_native_pdfs(self) -> None:
        """10 leases should be text-native (not scanned, not amendment)."""
        all_ids = {f"LS-{i:03d}" for i in range(1, 16)}
        text_native = all_ids - _SCANNED_IDS - _AMENDMENT_PDF_IDS
        assert len(text_native) == 10

    def test_3_scanned_style_pdfs(self) -> None:
        assert _SCANNED_IDS == {"LS-003", "LS-008", "LS-013"}

    def test_2_amendment_pdfs(self) -> None:
        assert _AMENDMENT_PDF_IDS == {"LS-002", "LS-007"}

    def test_pdfs_are_non_empty(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        leases_dir = output / f"{_INPUT_DIR}/leases"
        for pdf in leases_dir.glob("LS-*.pdf"):
            assert pdf.stat().st_size > 0, f"{pdf.name} is empty"


# ---------------------------------------------------------------------------
# Partial lease schedule xlsx
# ---------------------------------------------------------------------------


class TestPartialSchedule:
    """Verify lease_schedule_partial.xlsx structure and content."""

    def test_file_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/lease_schedule_partial.xlsx"
        assert path.exists()

    def test_has_lease_schedule_sheet(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/lease_schedule_partial.xlsx"
        wb = openpyxl.load_workbook(str(path))
        assert "Lease Schedule" in wb.sheetnames

    def test_title_row(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/lease_schedule_partial.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Lease Schedule"]
        assert "Cascade Industries" in (ws["A1"].value or "")
        assert "ASC 842" in (ws["A1"].value or "")

    def test_header_columns(self) -> None:
        """Row 3 should have the 13 schedule columns."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/lease_schedule_partial.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Lease Schedule"]
        headers = []
        for col in range(1, 14):
            val = ws.cell(row=3, column=col).value
            if val:
                headers.append(val)
        assert len(headers) == 13
        assert "Lease ID" in headers
        assert "Commencement Date" in headers
        assert "Monthly Base Rent" in headers
        assert "Short-Term Exempt?" in headers

    def test_8_leases_populated(self) -> None:
        """First 8 leases (LS-001..LS-008) should have data rows."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/lease_schedule_partial.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Lease Schedule"]
        lease_ids = []
        for row in range(4, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if isinstance(val, str) and val.startswith("LS-"):
                lease_ids.append(val)
        assert len(lease_ids) == 8
        assert set(lease_ids) == _PARTIAL_LEASE_IDS

    def test_blank_fields_for_3_leases(self) -> None:
        """LS-003, LS-005, LS-007 should have blank rent (col 7) and escalation (col 8)."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/lease_schedule_partial.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Lease Schedule"]
        for row in range(4, 12):
            lease_id = ws.cell(row=row, column=1).value
            if lease_id in _BLANK_FIELDS_IDS:
                rent = ws.cell(row=row, column=7).value
                escalation = ws.cell(row=row, column=8).value
                assert rent is None, f"{lease_id} rent should be blank, got {rent}"
                assert escalation is None, f"{lease_id} escalation should be blank, got {escalation}"

    def test_non_blank_leases_have_rent(self) -> None:
        """Pre-populated leases that aren't in the blank set should have rent values."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/lease_schedule_partial.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Lease Schedule"]
        for row in range(4, 12):
            lease_id = ws.cell(row=row, column=1).value
            if isinstance(lease_id, str) and lease_id not in _BLANK_FIELDS_IDS:
                rent = ws.cell(row=row, column=7).value
                assert rent is not None and rent > 0, (
                    f"{lease_id} should have a positive rent value, got {rent}"
                )

    def test_7_blank_rows_for_remaining_leases(self) -> None:
        """Rows after the 8 populated leases should be blank (for agent to fill)."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/lease_schedule_partial.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Lease Schedule"]
        # Row 12 should be the start of blank rows (row 4 + 8 data rows = 12)
        blank_count = 0
        for row in range(12, 19):
            val = ws.cell(row=row, column=1).value
            if val is None:
                blank_count += 1
        assert blank_count == 7, f"Expected 7 blank rows, got {blank_count}"


# ---------------------------------------------------------------------------
# ERR-007 — Date inconsistency in LS-002 commencement date
# ---------------------------------------------------------------------------


class TestERR007PlantedError:
    """Verify ERR-007: LS-002 commencement date in schedule differs from PDF."""

    def test_err007_registered(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        assert "ERR-007" in errors.entries

    def test_err007_is_date_inconsistency(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-007"]
        assert err.type == "date_inconsistency"

    def test_err007_references_ls002(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-007"]
        assert "LS-002" in err.location or "LS-002" in err.description

    def test_err007_in_schedule_file(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-007"]
        assert "lease_schedule_partial.xlsx" in err.file

    def test_err007_one_month_shift(self) -> None:
        """The schedule date for LS-002 should be exactly one month later than the model."""
        model, output, _, _, _ = _ensure_emitted()
        ls002 = next(ls for ls in model.leases if ls.lease_id == "LS-002")
        correct = ls002.commencement_date

        path = output / f"{_INPUT_DIR}/lease_schedule_partial.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Lease Schedule"]
        for row in range(4, 12):
            if ws.cell(row=row, column=1).value == "LS-002":
                cell_date = ws.cell(row=row, column=5).value
                # openpyxl returns datetime; compare month
                if hasattr(cell_date, "date"):
                    cell_date = cell_date.date()
                expected_wrong_month = correct.month % 12 + 1
                assert cell_date.month == expected_wrong_month, (
                    f"LS-002 schedule date month should be {expected_wrong_month}, "
                    f"got {cell_date.month}"
                )
                assert cell_date.day == correct.day
                break
        else:
            raise AssertionError("LS-002 not found in schedule")  # noqa: TRY003


# ---------------------------------------------------------------------------
# Short-term exemption
# ---------------------------------------------------------------------------


class TestShortTermExemption:
    """Verify exactly 2 leases qualify for short-term exemption."""

    def test_two_short_term_leases(self) -> None:
        model, _, _, _, _ = _ensure_emitted()
        exempt = [ls for ls in model.leases if ls.short_term_exempt]
        assert len(exempt) == 2, f"Expected 2 short-term exempt leases, got {len(exempt)}"

    def test_ls009_is_short_term(self) -> None:
        model, _, _, _, _ = _ensure_emitted()
        ls009 = next(ls for ls in model.leases if ls.lease_id == "LS-009")
        assert ls009.short_term_exempt

    def test_ls013_is_short_term(self) -> None:
        model, _, _, _, _ = _ensure_emitted()
        ls013 = next(ls for ls in model.leases if ls.lease_id == "LS-013")
        assert ls013.short_term_exempt

    def test_short_term_in_schedule(self) -> None:
        """LS-009 is in partial schedule (LS-001..008) — check its flag."""
        # LS-009 is NOT in the partial set (001..008), so we only check
        # that LS-003 and LS-008 are NOT short-term (they're in the schedule).
        model, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/lease_schedule_partial.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Lease Schedule"]
        for row in range(4, 12):
            lease_id = ws.cell(row=row, column=1).value
            exempt_val = ws.cell(row=row, column=12).value
            if lease_id in ("LS-003", "LS-008"):
                assert exempt_val == "No", f"{lease_id} should not be short-term exempt"


# ---------------------------------------------------------------------------
# Canary embedding
# ---------------------------------------------------------------------------


class TestCanaryEmbedding:
    """Verify canary codes are embedded in files."""

    def test_all_canary_keys_assigned(self) -> None:
        _, _, canaries, _, _ = _ensure_emitted()
        for key in _CANARY_KEYS:
            code = canaries.canary_for(key)
            assert len(code) == 8, f"Canary for {key} should be 8 chars"

    def test_schedule_canary_in_xlsx(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for(_SCHEDULE_KEY)
        path = output / f"{_INPUT_DIR}/lease_schedule_partial.xlsx"
        wb = openpyxl.load_workbook(str(path))
        desc = wb.properties.description or ""
        assert canary in desc, f"Canary {canary} not in schedule properties"


# ---------------------------------------------------------------------------
# Gold standard
# ---------------------------------------------------------------------------


class TestGoldStandard:
    """Verify gold standard JSON structure."""

    def test_gold_json_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / "gold_standards" / "TC-04_gold.json"
        assert path.exists()

    def test_gold_has_15_leases(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-04_gold.json").read_text())
        assert gold["expected_outputs"]["total_leases"] == 15
        assert len(gold["expected_outputs"]["leases"]) == 15

    def test_gold_short_term_exempt_ids(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-04_gold.json").read_text())
        assert gold["expected_outputs"]["short_term_exempt_count"] == 2
        assert set(gold["expected_outputs"]["short_term_exempt_ids"]) == {"LS-009", "LS-013"}

    def test_gold_amended_lease_ids(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-04_gold.json").read_text())
        amended = set(gold["expected_outputs"]["amended_lease_ids"])
        # LS-002, LS-007, LS-010 all have amendments
        assert "LS-002" in amended
        assert "LS-007" in amended
        assert "LS-010" in amended

    def test_gold_scanned_lease_ids(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-04_gold.json").read_text())
        assert set(gold["expected_outputs"]["scanned_lease_ids"]) == {"LS-003", "LS-008", "LS-013"}

    def test_gold_ocr_traps(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-04_gold.json").read_text())
        traps = gold["expected_outputs"]["ocr_traps"]
        assert "LS-003" in traps
        assert "LS-008" in traps
        assert "LS-013" in traps

    def test_gold_schedule_pre_populated(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-04_gold.json").read_text())
        assert gold["expected_outputs"]["schedule_pre_populated"] == 8

    def test_gold_canary_verification(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-04_gold.json").read_text())
        cv = gold["canary_verification"]
        # Should have read_LS-001..LS-015 + read_lease_schedule
        assert "read_LS-001" in cv
        assert "read_LS-015" in cv
        assert "read_lease_schedule" in cv

    def test_gold_error_detection(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-04_gold.json").read_text())
        assert "ERR-007" in gold["error_detection"]

    def test_gold_scoring_hints(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-04_gold.json").read_text())
        hints = gold["scoring_hints"]
        for key in ["correctness", "completeness", "format_compliance", "robustness", "communication"]:
            assert key in hints


# ---------------------------------------------------------------------------
# Prompt and expected behavior
# ---------------------------------------------------------------------------


class TestPromptAndExpectedBehavior:
    """Verify prompt and expected behavior files are generated."""

    def test_prompt_md_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / "test_cases/TC-04/prompt.md"
        assert path.exists()

    def test_prompt_mentions_asc842(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-04/prompt.md").read_text()
        assert "ASC 842" in text

    def test_prompt_mentions_lease_extraction(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-04/prompt.md").read_text().lower()
        assert "extract" in text
        assert "lease" in text

    def test_prompt_mentions_amendments(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-04/prompt.md").read_text().lower()
        assert "amendment" in text

    def test_prompt_mentions_short_term(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-04/prompt.md").read_text().lower()
        assert "short-term" in text

    def test_expected_behavior_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / "test_cases/TC-04/expected_behavior.md"
        assert path.exists()

    def test_expected_behavior_mentions_ocr_traps(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-04/expected_behavior.md").read_text()
        assert "OCR" in text
        assert "LS-003" in text
        assert "LS-008" in text
        assert "LS-013" in text

    def test_expected_behavior_mentions_err007(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-04/expected_behavior.md").read_text()
        assert "ERR-007" in text

    def test_expected_behavior_mentions_2_short_term(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-04/expected_behavior.md").read_text()
        assert "2 short-term" in text or "2 leases" in text.lower()


# ---------------------------------------------------------------------------
# Manifest registration
# ---------------------------------------------------------------------------


class TestManifest:
    """Verify files are registered in the manifest."""

    def test_manifest_has_tc04_entries(self) -> None:
        _, _, _, _, manifest = _ensure_emitted()
        tc04_count = sum(
            1 for v in manifest.entries.values()
            if "TC-04" in (v.test_cases or [])
        )
        # 15 lease PDFs + 1 schedule = 16
        assert tc04_count >= 16, f"Expected ≥16 TC-04 manifest entries, got {tc04_count}"
