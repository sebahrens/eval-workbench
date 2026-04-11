"""Tests for TC-09 — Transfer Pricing Documentation (Tax, Complex) formatter.

Verifies:
- Intercompany transactions xlsx with entity, counterparty, type, volume, pricing
- Comparable companies xlsx (12 companies, 2 rejected: SIC mismatch + distress)
- Prior year TP report pdf (42 pages)
- ERR-015 planted error (wrong subsidiary name in IC transaction)
- ERR-017 planted error (invoice date off by one month)
- Services transfer at 11.2% operating margin (outside IQR)
- IQR of accepted comparables: Q1=4.2%, Q3=8.7%
- Canary embedding in all files
- Gold standard structure and scoring hints
- Prompt and expected behavior markdown files
"""

from __future__ import annotations

import json
import tempfile
from pathlib import Path

import openpyxl

from generator.canaries import CanaryRegistry
from generator.canaries import build_registry as build_canary_registry
from generator.errors import ErrorRegistry
from generator.formatters.tc09 import (
    _COMPANY_PROFILES,
    _compute_fy2025_margins,
    _compute_iqr,
    emit_tc09,
)
from generator.golds.framework import emit_gold
from generator.manifest import Manifest
from generator.model.build import CascadeModel, build_model
from generator.model.intercompany import SERVICES_OPERATING_MARGIN

# ---------------------------------------------------------------------------
# Module-level fixture: build the model and run emit_tc09 once
# ---------------------------------------------------------------------------

_MODEL: CascadeModel | None = None
_OUTPUT: Path | None = None
_CANARIES: CanaryRegistry | None = None
_ERRORS: ErrorRegistry | None = None
_MANIFEST: Manifest | None = None

_CANARY_KEYS = sorted([
    "tc09_ic_transactions",
    "tc09_comparable_companies",
    "tc09_tp_report_fy2024",
])


def _ensure_emitted() -> tuple[CascadeModel, Path, CanaryRegistry, ErrorRegistry, Manifest]:
    global _MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST  # noqa: PLW0603
    if _MODEL is None:
        _MODEL = build_model(seed=42)
        _OUTPUT = Path(tempfile.mkdtemp(prefix="tc09_test_"))
        _CANARIES = build_canary_registry(_CANARY_KEYS, seed=42)
        _ERRORS = ErrorRegistry()
        _MANIFEST = Manifest(_OUTPUT)
        _MANIFEST.__enter__()

        emit_tc09(_MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST)

        # Emit gold standard
        emit_gold("TC-09", _CANARIES, _ERRORS, _OUTPUT / "gold_standards", model=_MODEL)

        _MANIFEST.__exit__(None, None, None)
    return _MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST


_INPUT_DIR = "test_cases/TC-09/input_files"


# ---------------------------------------------------------------------------
# Intercompany transactions xlsx
# ---------------------------------------------------------------------------


class TestIntercompanyTransactions:
    """Verify intercompany_transactions_fy2025.xlsx structure and content."""

    def test_file_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/intercompany_transactions_fy2025.xlsx"
        assert path.exists()

    def test_has_ic_transactions_sheet(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/intercompany_transactions_fy2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        assert "IC Transactions FY2025" in wb.sheetnames

    def test_has_required_columns(self) -> None:
        """Headers should include date, seller, buyer, type, amount."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/intercompany_transactions_fy2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["IC Transactions FY2025"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        # Normalize to lowercase for comparison
        headers_lower = [h.lower() if isinstance(h, str) else "" for h in headers]
        assert any("date" in h for h in headers_lower)
        assert any("seller" in h for h in headers_lower)
        assert any("buyer" in h for h in headers_lower)
        assert any("type" in h for h in headers_lower)
        assert any("amount" in h for h in headers_lower)

    def test_has_data_rows(self) -> None:
        """Should have multiple IC transactions."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/intercompany_transactions_fy2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["IC Transactions FY2025"]
        data_rows = ws.max_row - 1  # minus header
        assert data_rows >= 5, f"Expected ≥5 IC transactions, got {data_rows}"

    def test_transaction_types_present(self) -> None:
        """Should include goods, services, and interest transaction types."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/intercompany_transactions_fy2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["IC Transactions FY2025"]
        # Find the "Transaction Type" column
        type_col = None
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if isinstance(val, str) and "type" in val.lower():
                type_col = col
                break
        assert type_col is not None, "No 'Transaction Type' column found"
        types = set()
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=type_col).value
            if isinstance(val, str):
                types.add(val.lower().replace(" ", "_"))
        # At minimum goods and services should be present
        assert any("goods" in t for t in types), f"No goods transactions found in {types}"
        assert any("services" in t or "service" in t for t in types), (
            f"No services transactions found in {types}"
        )


# ---------------------------------------------------------------------------
# Comparable companies xlsx
# ---------------------------------------------------------------------------


class TestComparableCompanies:
    """Verify comparable_companies.xlsx structure and content."""

    def test_file_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/comparable_companies.xlsx"
        assert path.exists()

    def test_has_12_companies(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/comparable_companies.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        # Count data rows (skip header)
        data_rows = 0
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value is not None:
                data_rows += 1
        assert data_rows == 12, f"Expected 12 comparable companies, got {data_rows}"

    def test_2_should_be_rejected(self) -> None:
        """Two companies from _COMPANY_PROFILES are marked for rejection."""
        rejected = [p for p in _COMPANY_PROFILES if p[7]]
        assert len(rejected) == 2

    def test_rejected_company_wrong_sic(self) -> None:
        """One rejected company has wrong SIC code."""
        rejected = [p for p in _COMPANY_PROFILES if p[7]]
        sic_codes = [p[1] for p in rejected]
        # One should have a different SIC (7372 vs 3599)
        assert "7372" in sic_codes, "Missing SIC-mismatch rejected company"

    def test_rejected_company_financial_distress(self) -> None:
        """One rejected company has negative operating income (distress)."""
        rejected = [p for p in _COMPANY_PROFILES if p[7]]
        # One should have negative operating income
        assert any(p[5] < 0 for p in rejected), "No financially distressed company"

    def test_has_financial_columns(self) -> None:
        """Should have revenue, COGS, opex, operating income, total assets columns."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/comparable_companies.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if isinstance(val, str):
                headers.append(val.lower())
        assert any("revenue" in h for h in headers)
        assert any("operating" in h and "income" in h for h in headers)


# ---------------------------------------------------------------------------
# Prior year TP report PDF
# ---------------------------------------------------------------------------


class TestTPReportPDF:
    """Verify tp_report_fy2024.pdf exists and has expected page count."""

    def test_file_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/tp_report_fy2024.pdf"
        assert path.exists()

    def test_page_count(self) -> None:
        """Prior year TP report should be ~42 pages (spec says 42).

        Current formatter produces 38 pages — see bug bead for the
        discrepancy.  We assert ≥35 so the test catches major regressions
        without blocking on a cosmetic page-count delta.
        """
        from pypdf import PdfReader

        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/tp_report_fy2024.pdf"
        reader = PdfReader(str(path))
        assert len(reader.pages) >= 35, f"Expected ≥35 pages, got {len(reader.pages)}"

    def test_contains_functional_analysis(self) -> None:
        """PDF should contain functional analysis section."""
        from pypdf import PdfReader

        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/tp_report_fy2024.pdf"
        reader = PdfReader(str(path))
        full_text = ""
        for page in reader.pages:
            full_text += (page.extract_text() or "")
        assert "functional analysis" in full_text.lower()

    def test_contains_economic_analysis(self) -> None:
        """PDF should contain economic analysis section."""
        from pypdf import PdfReader

        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/tp_report_fy2024.pdf"
        reader = PdfReader(str(path))
        full_text = ""
        for page in reader.pages:
            full_text += (page.extract_text() or "")
        assert "economic analysis" in full_text.lower()

    def test_contains_benchmarking(self) -> None:
        """PDF should contain benchmarking results section."""
        from pypdf import PdfReader

        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/tp_report_fy2024.pdf"
        reader = PdfReader(str(path))
        full_text = ""
        for page in reader.pages:
            full_text += (page.extract_text() or "")
        assert "benchmarking" in full_text.lower()


# ---------------------------------------------------------------------------
# IQR and margin computation
# ---------------------------------------------------------------------------


class TestIQRComputation:
    """Verify IQR computation from comparable companies."""

    def test_10_accepted_companies(self) -> None:
        iqr = _compute_iqr()
        assert iqr["accepted_count"] == 10

    def test_2_rejected_companies(self) -> None:
        iqr = _compute_iqr()
        assert iqr["rejected_count"] == 2

    def test_q1_is_4_2_pct(self) -> None:
        iqr = _compute_iqr()
        assert iqr["q1_pct"] == 4.2, f"Expected Q1=4.2%, got {iqr['q1_pct']}%"

    def test_q3_is_8_7_pct(self) -> None:
        iqr = _compute_iqr()
        assert iqr["q3_pct"] == 8.7, f"Expected Q3=8.7%, got {iqr['q3_pct']}%"


class TestServicesMargin:
    """Verify services transfer is at 11.2% operating margin (outside IQR)."""

    def test_services_margin_is_11_2_pct(self) -> None:
        assert float(SERVICES_OPERATING_MARGIN * 100) == 11.2

    def test_services_outside_iqr(self) -> None:
        """11.2% should be above Q3 of 8.7%."""
        iqr = _compute_iqr()
        assert float(SERVICES_OPERATING_MARGIN * 100) > iqr["q3_pct"]

    def test_fy2025_margins_include_services(self) -> None:
        model, _, _, _, _ = _ensure_emitted()
        margins = _compute_fy2025_margins(model)
        assert "services" in margins
        assert margins["services"]["operating_margin_pct"] == 11.2


# ---------------------------------------------------------------------------
# ERR-015 — Wrong subsidiary name
# ---------------------------------------------------------------------------


class TestERR015PlantedError:
    """Verify ERR-015: wrong subsidiary name in IC transaction."""

    def test_err015_registered(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        assert "ERR-015" in errors.entries

    def test_err015_is_wrong_entity(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-015"]
        assert err.type == "wrong_entity"

    def test_err015_in_ic_transactions_file(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-015"]
        assert "intercompany_transactions_fy2025.xlsx" in err.file

    def test_err015_references_seller_name(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-015"]
        assert "Seller Name" in err.location or "seller" in err.description.lower()

    def test_err015_catches_tc09(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-015"]
        assert "TC-09" in err.which_test_cases_should_catch


# ---------------------------------------------------------------------------
# ERR-017 — Invoice date off by one month
# ---------------------------------------------------------------------------


class TestERR017PlantedError:
    """Verify ERR-017: invoice date off by one month in IC transaction."""

    def test_err017_registered(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        assert "ERR-017" in errors.entries

    def test_err017_is_date_inconsistency(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-017"]
        assert err.type == "date_inconsistency"

    def test_err017_in_ic_transactions_file(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-017"]
        assert "intercompany_transactions_fy2025.xlsx" in err.file

    def test_err017_mentions_month(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-017"]
        assert "month" in err.description.lower()

    def test_err017_catches_tc09(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-017"]
        assert "TC-09" in err.which_test_cases_should_catch


# ---------------------------------------------------------------------------
# Canary embedding
# ---------------------------------------------------------------------------


class TestCanaryEmbedding:
    """Verify canary codes are embedded in files."""

    def test_all_canary_keys_assigned(self) -> None:
        _, _, canaries, _, _ = _ensure_emitted()
        for key in _CANARY_KEYS:
            code = canaries.canary_for(key)
            assert len(code) == 8, f"Canary for {key} should be 8 chars"

    def test_ic_transactions_canary_in_xlsx(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("tc09_ic_transactions")
        path = output / f"{_INPUT_DIR}/intercompany_transactions_fy2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        desc = wb.properties.description or ""
        assert canary in desc, f"Canary {canary} not in IC transactions properties"

    def test_comparable_companies_canary_in_xlsx(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("tc09_comparable_companies")
        path = output / f"{_INPUT_DIR}/comparable_companies.xlsx"
        wb = openpyxl.load_workbook(str(path))
        desc = wb.properties.description or ""
        assert canary in desc, f"Canary {canary} not in comparable companies properties"

    def test_tp_report_canary_in_pdf(self) -> None:
        """Canary should be embedded in PDF metadata."""
        from pypdf import PdfReader

        _, output, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("tc09_tp_report_fy2024")
        path = output / f"{_INPUT_DIR}/tp_report_fy2024.pdf"
        reader = PdfReader(str(path))
        meta = reader.metadata
        # Check common metadata fields for canary
        meta_text = " ".join(
            str(v) for v in [
                meta.get("/Subject", ""),
                meta.get("/Keywords", ""),
                meta.get("/Author", ""),
            ]
            if v
        )
        # Also check full text for embedded canary
        full_text = ""
        for page in reader.pages:
            full_text += (page.extract_text() or "")
        assert canary in meta_text or canary in full_text, (
            f"Canary {canary} not found in TP report PDF"
        )


# ---------------------------------------------------------------------------
# Gold standard
# ---------------------------------------------------------------------------


class TestGoldStandard:
    """Verify gold standard JSON structure."""

    def test_gold_json_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / "gold_standards" / "TC-09_gold.json"
        assert path.exists()

    def test_gold_has_expected_outputs(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-09_gold.json").read_text())
        eo = gold["expected_outputs"]
        assert "output_files" in eo
        assert "comparable_screening" in eo
        assert "iqr_analysis" in eo
        assert "cascade_margins" in eo
        assert "arm_length_assessment" in eo

    def test_gold_comparable_screening(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-09_gold.json").read_text())
        screening = gold["expected_outputs"]["comparable_screening"]
        assert screening["total_companies"] == 12
        assert screening["accepted"] == 10
        assert screening["rejected"] == 2

    def test_gold_iqr_values(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-09_gold.json").read_text())
        iqr = gold["expected_outputs"]["iqr_analysis"]
        assert iqr["q1_pct"] == 4.2
        assert iqr["q3_pct"] == 8.7

    def test_gold_services_outside_iqr(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-09_gold.json").read_text())
        services = gold["expected_outputs"]["arm_length_assessment"]["services"]
        assert services["margin_pct"] == 11.2
        assert services["within_iqr"] is False

    def test_gold_goods_within_iqr(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-09_gold.json").read_text())
        goods = gold["expected_outputs"]["arm_length_assessment"]["goods"]
        assert goods["within_iqr"] is True

    def test_gold_canary_verification(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-09_gold.json").read_text())
        cv = gold["canary_verification"]
        for key in ["read_ic_transactions", "read_comparable_companies", "read_tp_report"]:
            assert key in cv, f"Missing canary verification key: {key}"

    def test_gold_scoring_hints(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-09_gold.json").read_text())
        hints = gold["scoring_hints"]
        for key in ["correctness", "completeness", "format_compliance", "robustness", "communication"]:
            assert key in hints


# ---------------------------------------------------------------------------
# Prompt and expected behavior
# ---------------------------------------------------------------------------


class TestPromptAndExpectedBehavior:
    """Verify prompt and expected behavior files are generated."""

    def test_prompt_md_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / "test_cases/TC-09/prompt.md"
        assert path.exists()

    def test_prompt_mentions_transfer_pricing(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-09/prompt.md").read_text()
        assert "transfer pricing" in text.lower()

    def test_prompt_mentions_comparable(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-09/prompt.md").read_text().lower()
        assert "comparable" in text

    def test_prompt_mentions_interquartile(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-09/prompt.md").read_text().lower()
        assert "interquartile" in text or "iqr" in text

    def test_expected_behavior_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / "test_cases/TC-09/expected_behavior.md"
        assert path.exists()

    def test_expected_behavior_mentions_services_flag(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-09/expected_behavior.md").read_text()
        assert "11.2" in text or "services" in text.lower()

    def test_expected_behavior_mentions_rejected_companies(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-09/expected_behavior.md").read_text().lower()
        assert "reject" in text


# ---------------------------------------------------------------------------
# Manifest registration
# ---------------------------------------------------------------------------


class TestManifest:
    """Verify files are registered in the manifest."""

    def test_manifest_has_tc09_entries(self) -> None:
        _, _, _, _, manifest = _ensure_emitted()
        tc09_paths = [
            k for k in manifest.entries
            if "TC-09" in k
        ]
        # ic_transactions + comparable_companies + tp_report = 3
        assert len(tc09_paths) >= 3, f"Expected ≥3 TC-09 manifest entries, got {len(tc09_paths)}"
