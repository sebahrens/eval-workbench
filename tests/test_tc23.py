"""Tests for TC-23 — Swiss Multi-Currency Bank Reconciliation (CHF/EUR/USD).

Verifies:
- 6 input files (3 bank statements CSV, GL cash xlsx, bank confirmation PDF, SNB rates CSV)
- 185 CHF transactions, 62 EUR transactions, 28 USD transactions
- CHF reconciliation: 2 outstanding checks, 1 deposit in transit, bank interest/charges
- EUR reconciliation: 1 outstanding SEPA, FX revaluation with stale rate (ERR-CH-002)
- USD reconciliation: 1 outstanding wire, correct FX revaluation
- Gold standard values: CHF 2,817,905 / EUR adj CHF 1,264,142 / USD CHF 536,518
- EUR FX error: CHF 6,324 (stale rate on EUR_REVAL_BALANCE portion)
- Consolidated cash CHF 4,618,565
- Canary embedding in all 6 input files
- Deterministic output
"""

from __future__ import annotations

import json
import random
import tempfile
from decimal import Decimal
from pathlib import Path

import openpyxl

from generator.canaries import CanaryRegistry
from generator.canaries import build_registry as build_canary_registry
from generator.errors import ErrorRegistry
from generator.formatters.tc23 import emit_tc23
from generator.golds.framework import emit_gold
from generator.manifest import Manifest
from generator.model.bank_ch import (
    CHF_ADJUSTED_BANK,
    CHF_BANK_CHARGES,
    CHF_BANK_ENDING,
    CHF_BANK_INTEREST,
    CHF_DEPOSITS_IN_TRANSIT,
    CHF_DEPOSITS_IN_TRANSIT_TOTAL,
    CHF_GL_ENDING,
    CHF_OUTSTANDING_CHECKS,
    CHF_OUTSTANDING_CHECKS_TOTAL,
    CHF_TX_COUNT,
    CONSOLIDATED_CHF,
    CREDIT_FACILITY,
    ENTITY_NAME,
    EUR_ADJUSTED_BANK,
    EUR_ADJUSTED_BANK_CHF,
    EUR_BANK_ENDING,
    EUR_FX_ERROR,
    EUR_GL_ENDING_CHF,
    EUR_OUTSTANDING_SEPA,
    EUR_OUTSTANDING_SEPA_TOTAL,
    EUR_REVAL_BALANCE,
    EUR_TX_COUNT,
    FX_CHF_EUR_CLOSING,
    FX_CHF_EUR_STALE,
    FX_CHF_USD_CLOSING,
    USD_ADJUSTED_BANK,
    USD_ADJUSTED_BANK_CHF,
    USD_BANK_ENDING,
    USD_GL_ENDING_CHF,
    USD_OUTSTANDING_WIRE,
    USD_OUTSTANDING_WIRE_TOTAL,
    USD_TX_COUNT,
    generate_bank_ch_model,
    validate_reconciliation_ch,
)
from generator.model.build import CascadeModel, build_model

# ---------------------------------------------------------------------------
# Module-level fixture: build the model and run emit_tc23 once
# ---------------------------------------------------------------------------

_MODEL: CascadeModel | None = None
_OUTPUT: Path | None = None
_CANARIES: CanaryRegistry | None = None
_ERRORS: ErrorRegistry | None = None
_MANIFEST: Manifest | None = None

_CANARY_KEYS = sorted([
    "tc23_bank_chf",
    "tc23_bank_eur",
    "tc23_bank_usd",
    "tc23_gl_cash",
    "tc23_bank_confirm",
    "tc23_snb_fx_rates",
])


def _ensure_emitted() -> tuple[CascadeModel, Path, CanaryRegistry, ErrorRegistry, Manifest]:
    global _MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST  # noqa: PLW0603
    if _MODEL is None:
        _MODEL = build_model(seed=42)
        _OUTPUT = Path(tempfile.mkdtemp(prefix="tc23_test_"))
        _CANARIES = build_canary_registry(_CANARY_KEYS, seed=42)
        _ERRORS = ErrorRegistry()
        _MANIFEST = Manifest(_OUTPUT)
        _MANIFEST.__enter__()

        emit_tc23(_MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST)

        emit_gold("TC-23", _CANARIES, _ERRORS, _OUTPUT / "gold_standards", model=_MODEL)

        _MANIFEST.__exit__(None, None, None)
    return _MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST


_INPUT_DIR = "test_cases/TC-23/input_files"


# ---------------------------------------------------------------------------
# Model constants verification
# ---------------------------------------------------------------------------


class TestBankModelConstants:
    """Verify model constants match design spec."""

    def test_chf_adjusted_bank_equals_book(self) -> None:
        assert CHF_ADJUSTED_BANK == Decimal("2817905")

    def test_chf_outstanding_checks_count(self) -> None:
        assert len(CHF_OUTSTANDING_CHECKS) == 2

    def test_chf_outstanding_checks_total(self) -> None:
        assert CHF_OUTSTANDING_CHECKS_TOTAL == Decimal("47815")

    def test_chf_deposits_in_transit_count(self) -> None:
        assert len(CHF_DEPOSITS_IN_TRANSIT) == 1

    def test_chf_deposits_in_transit_total(self) -> None:
        assert CHF_DEPOSITS_IN_TRANSIT_TOTAL == Decimal("18430")

    def test_chf_bank_ending(self) -> None:
        assert CHF_BANK_ENDING == Decimal("2847290")

    def test_chf_gl_ending(self) -> None:
        assert CHF_GL_ENDING == Decimal("2815280")

    def test_chf_bank_interest(self) -> None:
        assert CHF_BANK_INTEREST == Decimal("4000")

    def test_chf_bank_charges(self) -> None:
        assert CHF_BANK_CHARGES == Decimal("1375")

    def test_eur_bank_ending(self) -> None:
        assert EUR_BANK_ENDING == Decimal("1215400")

    def test_eur_adjusted_bank(self) -> None:
        assert EUR_ADJUSTED_BANK == Decimal("1186650")

    def test_eur_adjusted_bank_chf(self) -> None:
        assert EUR_ADJUSTED_BANK_CHF == Decimal("1264142")

    def test_eur_gl_ending_chf(self) -> None:
        assert EUR_GL_ENDING_CHF == Decimal("1270466")

    def test_eur_fx_error_exact(self) -> None:
        """The FX error must be exactly CHF 6,324 (review change #1)."""
        assert EUR_FX_ERROR == Decimal("6324")

    def test_eur_reval_balance(self) -> None:
        """Opening EUR balance subject to stale-rate error."""
        assert EUR_REVAL_BALANCE == Decimal("540400")

    def test_eur_outstanding_sepa_count(self) -> None:
        assert len(EUR_OUTSTANDING_SEPA) == 1

    def test_eur_outstanding_sepa_total(self) -> None:
        assert EUR_OUTSTANDING_SEPA_TOTAL == Decimal("28750")

    def test_usd_bank_ending(self) -> None:
        assert USD_BANK_ENDING == Decimal("489750")

    def test_usd_adjusted_bank(self) -> None:
        assert USD_ADJUSTED_BANK == Decimal("474550")

    def test_usd_adjusted_bank_chf(self) -> None:
        assert USD_ADJUSTED_BANK_CHF == Decimal("536518")

    def test_usd_gl_ending_chf_matches_adjusted(self) -> None:
        """No FX error on USD account."""
        assert USD_GL_ENDING_CHF == USD_ADJUSTED_BANK_CHF

    def test_usd_outstanding_wire_count(self) -> None:
        assert len(USD_OUTSTANDING_WIRE) == 1

    def test_usd_outstanding_wire_total(self) -> None:
        assert USD_OUTSTANDING_WIRE_TOTAL == Decimal("15200")

    def test_consolidated_cash(self) -> None:
        assert CONSOLIDATED_CHF == Decimal("4618565")
        assert CONSOLIDATED_CHF == CHF_ADJUSTED_BANK + EUR_ADJUSTED_BANK_CHF + USD_ADJUSTED_BANK_CHF

    def test_fx_rates(self) -> None:
        assert FX_CHF_EUR_CLOSING == Decimal("0.9387")
        assert FX_CHF_EUR_STALE == Decimal("0.9285")
        assert FX_CHF_USD_CLOSING == Decimal("0.8845")

    def test_credit_facility(self) -> None:
        assert CREDIT_FACILITY == Decimal("5000000")


# ---------------------------------------------------------------------------
# Model generation and validation
# ---------------------------------------------------------------------------


class TestBankModelGeneration:
    """Verify BankModelCH generates and validates correctly."""

    def test_model_generates(self) -> None:
        rng = random.Random(42 + 23)
        model = generate_bank_ch_model(rng)
        assert model is not None

    def test_model_validates(self) -> None:
        rng = random.Random(42 + 23)
        model = generate_bank_ch_model(rng)
        errors = validate_reconciliation_ch(model)
        assert errors == [], f"Validation errors: {errors}"

    def test_chf_transaction_count(self) -> None:
        rng = random.Random(42 + 23)
        model = generate_bank_ch_model(rng)
        assert len(model.chf_bank_transactions) == CHF_TX_COUNT

    def test_eur_transaction_count(self) -> None:
        rng = random.Random(42 + 23)
        model = generate_bank_ch_model(rng)
        assert len(model.eur_bank_transactions) == EUR_TX_COUNT

    def test_usd_transaction_count(self) -> None:
        rng = random.Random(42 + 23)
        model = generate_bank_ch_model(rng)
        assert len(model.usd_bank_transactions) == USD_TX_COUNT

    def test_chf_bank_ending_balance(self) -> None:
        rng = random.Random(42 + 23)
        model = generate_bank_ch_model(rng)
        assert model.chf_bank_transactions[-1].running_balance == CHF_BANK_ENDING

    def test_eur_bank_ending_balance(self) -> None:
        rng = random.Random(42 + 23)
        model = generate_bank_ch_model(rng)
        assert model.eur_bank_transactions[-1].running_balance == EUR_BANK_ENDING

    def test_usd_bank_ending_balance(self) -> None:
        rng = random.Random(42 + 23)
        model = generate_bank_ch_model(rng)
        assert model.usd_bank_transactions[-1].running_balance == USD_BANK_ENDING

    def test_chf_gl_ending(self) -> None:
        rng = random.Random(42 + 23)
        model = generate_bank_ch_model(rng)
        assert model.chf_gl_entries[-1].running_balance == CHF_GL_ENDING

    def test_eur_gl_ending(self) -> None:
        rng = random.Random(42 + 23)
        model = generate_bank_ch_model(rng)
        assert model.eur_gl_entries[-1].running_balance == EUR_GL_ENDING_CHF

    def test_usd_gl_ending(self) -> None:
        rng = random.Random(42 + 23)
        model = generate_bank_ch_model(rng)
        assert model.usd_gl_entries[-1].running_balance == USD_GL_ENDING_CHF

    def test_deterministic_output(self) -> None:
        """Two runs with same seed must produce identical models."""
        rng1 = random.Random(42 + 23)
        m1 = generate_bank_ch_model(rng1)
        rng2 = random.Random(42 + 23)
        m2 = generate_bank_ch_model(rng2)

        assert len(m1.chf_bank_transactions) == len(m2.chf_bank_transactions)
        for a, b in zip(m1.chf_bank_transactions, m2.chf_bank_transactions):
            assert a.date == b.date
            assert a.amount == b.amount
            assert a.description == b.description

    def test_eur_revaluation_entry_uses_stale_rate(self) -> None:
        """The EUR GL must contain a revaluation entry referencing the stale rate."""
        rng = random.Random(42 + 23)
        model = generate_bank_ch_model(rng)
        reval_entries = [
            e for e in model.eur_gl_entries
            if e.category == "fx_revaluation"
        ]
        assert len(reval_entries) == 1
        assert reval_entries[0].fx_rate == FX_CHF_EUR_STALE
        assert "0.9285" in reval_entries[0].description

    def test_usd_revaluation_entry_uses_correct_rate(self) -> None:
        """The USD GL must contain a revaluation entry with the correct closing rate."""
        rng = random.Random(42 + 23)
        model = generate_bank_ch_model(rng)
        reval_entries = [
            e for e in model.usd_gl_entries
            if e.category == "fx_revaluation"
        ]
        assert len(reval_entries) == 1
        assert reval_entries[0].fx_rate == FX_CHF_USD_CLOSING


# ---------------------------------------------------------------------------
# Input file existence
# ---------------------------------------------------------------------------


class TestInputFiles:
    """Verify all 6 TC-23 input files are generated."""

    def test_bank_statement_chf_exists(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        assert (out / _INPUT_DIR / "cpi_bank_statement_chf_dec2025.csv").exists()

    def test_bank_statement_eur_exists(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        assert (out / _INPUT_DIR / "cpi_bank_statement_eur_dec2025.csv").exists()

    def test_bank_statement_usd_exists(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        assert (out / _INPUT_DIR / "cpi_bank_statement_usd_dec2025.csv").exists()

    def test_gl_cash_xlsx_exists(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        assert (out / _INPUT_DIR / "cpi_gl_cash_dec2025.xlsx").exists()

    def test_bank_confirmation_pdf_exists(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        assert (out / _INPUT_DIR / "cpi_bank_confirmations_fy2025.pdf").exists()

    def test_snb_fx_rates_csv_exists(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        assert (out / _INPUT_DIR / "snb_fx_rates_dec2025.csv").exists()


# ---------------------------------------------------------------------------
# Bank statement CSVs
# ---------------------------------------------------------------------------


class TestBankStatementCSVs:
    """Verify bank statement CSV content."""

    def test_chf_csv_has_header_row(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        text = (out / _INPUT_DIR / "cpi_bank_statement_chf_dec2025.csv").read_text()
        assert "Buchungsdatum" in text
        assert "Buchungstext" in text
        assert "Betrag" in text
        assert "Saldo" in text

    def test_chf_csv_has_ubs_header(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        text = (out / _INPUT_DIR / "cpi_bank_statement_chf_dec2025.csv").read_text()
        assert "UBS" in text
        assert ENTITY_NAME in text

    def test_eur_csv_has_transactions(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        text = (out / _INPUT_DIR / "cpi_bank_statement_eur_dec2025.csv").read_text()
        lines = [line for line in text.strip().split("\n") if not line.startswith("#")]
        # Header + data rows
        assert len(lines) >= EUR_TX_COUNT + 1

    def test_usd_csv_has_transactions(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        text = (out / _INPUT_DIR / "cpi_bank_statement_usd_dec2025.csv").read_text()
        lines = [line for line in text.strip().split("\n") if not line.startswith("#")]
        assert len(lines) >= USD_TX_COUNT + 1

    def test_chf_csv_mentions_dta(self) -> None:
        """Swiss DTA payment format should appear in CHF transactions."""
        _, out, _, _, _ = _ensure_emitted()
        text = (out / _INPUT_DIR / "cpi_bank_statement_chf_dec2025.csv").read_text()
        assert "DTA" in text

    def test_chf_csv_mentions_lsv(self) -> None:
        """Swiss LSV+ direct debit should appear in CHF transactions."""
        _, out, _, _, _ = _ensure_emitted()
        text = (out / _INPUT_DIR / "cpi_bank_statement_chf_dec2025.csv").read_text()
        assert "LSV+" in text

    def test_eur_csv_mentions_sepa(self) -> None:
        """SEPA transactions should appear in EUR statement."""
        _, out, _, _, _ = _ensure_emitted()
        text = (out / _INPUT_DIR / "cpi_bank_statement_eur_dec2025.csv").read_text()
        assert "SEPA" in text

    def test_usd_csv_mentions_swift(self) -> None:
        """SWIFT/MT103 references should appear in USD statement."""
        _, out, _, _, _ = _ensure_emitted()
        text = (out / _INPUT_DIR / "cpi_bank_statement_usd_dec2025.csv").read_text()
        assert "SWIFT" in text or "MT103" in text


# ---------------------------------------------------------------------------
# GL Cash xlsx
# ---------------------------------------------------------------------------


class TestGLCashXlsx:
    """Verify cpi_gl_cash_dec2025.xlsx."""

    def test_has_three_sheets(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "cpi_gl_cash_dec2025.xlsx")
        assert len(wb.sheetnames) == 3

    def test_sheet_names(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "cpi_gl_cash_dec2025.xlsx")
        assert "1020 - Bank CHF" in wb.sheetnames
        assert "1021 - Bank EUR" in wb.sheetnames
        assert "1022 - Bank USD" in wb.sheetnames

    def test_chf_sheet_has_entries(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "cpi_gl_cash_dec2025.xlsx")
        ws = wb["1020 - Bank CHF"]
        # Count data rows (after header row 7)
        data_rows = sum(1 for r in range(8, ws.max_row + 1)
                       if ws.cell(row=r, column=1).value is not None
                       and ws.cell(row=r, column=1).value != "Endsaldo")
        assert data_rows > 100, f"Expected 180+ CHF GL entries, got {data_rows}"

    def test_eur_sheet_has_revaluation_entry(self) -> None:
        """EUR sheet must have an FX revaluation entry mentioning rate 0.9285."""
        _, out, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "cpi_gl_cash_dec2025.xlsx")
        ws = wb["1021 - Bank EUR"]
        found_reval = False
        for row in ws.iter_rows(min_row=8, values_only=True):
            if row[2] and "Neubewertung" in str(row[2]) and "0.9285" in str(row[2]):
                found_reval = True
                break
        assert found_reval, "EUR FX revaluation entry with stale rate 0.9285 not found"

    def test_usd_sheet_has_revaluation_entry(self) -> None:
        """USD sheet must have an FX revaluation entry with correct rate."""
        _, out, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "cpi_gl_cash_dec2025.xlsx")
        ws = wb["1022 - Bank USD"]
        found_reval = False
        for row in ws.iter_rows(min_row=8, values_only=True):
            if row[2] and "Neubewertung" in str(row[2]) and "0.8845" in str(row[2]):
                found_reval = True
                break
        assert found_reval, "USD FX revaluation entry with rate 0.8845 not found"


# ---------------------------------------------------------------------------
# Bank confirmation PDF
# ---------------------------------------------------------------------------


class TestBankConfirmationPDF:
    """Verify cpi_bank_confirmations_fy2025.pdf."""

    def test_pdf_not_empty(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        path = out / _INPUT_DIR / "cpi_bank_confirmations_fy2025.pdf"
        assert path.stat().st_size > 2000, "PDF should be substantial"

    def test_pdf_contains_ubs(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        content = (out / _INPUT_DIR / "cpi_bank_confirmations_fy2025.pdf").read_bytes()
        assert b"UBS" in content

    def test_pdf_contains_entity_name(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        content = (out / _INPUT_DIR / "cpi_bank_confirmations_fy2025.pdf").read_bytes()
        assert ENTITY_NAME.encode() in content


# ---------------------------------------------------------------------------
# SNB FX rates CSV
# ---------------------------------------------------------------------------


class TestSNBFxRatesCSV:
    """Verify snb_fx_rates_dec2025.csv."""

    def test_csv_has_closing_rates(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        text = (out / _INPUT_DIR / "snb_fx_rates_dec2025.csv").read_text()
        assert "0.9387" in text, "EUR closing rate must appear"
        assert "0.8845" in text, "USD closing rate must appear"

    def test_csv_has_december_dates(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        text = (out / _INPUT_DIR / "snb_fx_rates_dec2025.csv").read_text()
        assert "2025-12-31" in text, "Dec 31 closing date must appear"


# ---------------------------------------------------------------------------
# Planted error
# ---------------------------------------------------------------------------


class TestPlantedError:
    """Verify ERR-CH-002 is registered correctly."""

    def test_err_ch_002_registered(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.get("ERR-CH-002")
        assert err is not None

    def test_err_ch_002_type(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.get("ERR-CH-002")
        assert err.type == "stale_data"

    def test_err_ch_002_mentions_stale_rate(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.get("ERR-CH-002")
        assert "0.9285" in err.description
        assert "0.9387" in err.description

    def test_err_ch_002_impact_amount(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.get("ERR-CH-002")
        assert "6,324" in err.description


# ---------------------------------------------------------------------------
# Canary embedding
# ---------------------------------------------------------------------------


class TestCanaries:
    """Verify canary codes are embedded in all 6 input files."""

    def test_bank_chf_canary(self) -> None:
        _, out, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("tc23_bank_chf")
        text = (out / _INPUT_DIR / "cpi_bank_statement_chf_dec2025.csv").read_text()
        assert canary in text

    def test_bank_eur_canary(self) -> None:
        _, out, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("tc23_bank_eur")
        text = (out / _INPUT_DIR / "cpi_bank_statement_eur_dec2025.csv").read_text()
        assert canary in text

    def test_bank_usd_canary(self) -> None:
        _, out, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("tc23_bank_usd")
        text = (out / _INPUT_DIR / "cpi_bank_statement_usd_dec2025.csv").read_text()
        assert canary in text

    def test_gl_cash_canary(self) -> None:
        _, out, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("tc23_gl_cash")
        wb = openpyxl.load_workbook(out / _INPUT_DIR / "cpi_gl_cash_dec2025.xlsx")
        assert canary in (wb.properties.description or "")

    def test_bank_confirm_canary(self) -> None:
        _, out, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("tc23_bank_confirm")
        content = (out / _INPUT_DIR / "cpi_bank_confirmations_fy2025.pdf").read_bytes()
        assert canary.encode() in content

    def test_snb_fx_rates_canary(self) -> None:
        _, out, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("tc23_snb_fx_rates")
        text = (out / _INPUT_DIR / "snb_fx_rates_dec2025.csv").read_text()
        assert canary in text


# ---------------------------------------------------------------------------
# Prompt and expected behavior files
# ---------------------------------------------------------------------------


class TestMarkdownFiles:
    """Verify prompt.md and expected_behavior.md."""

    def test_prompt_exists(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        assert (out / "test_cases/TC-23/prompt.md").exists()

    def test_prompt_mentions_multicurrency(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        text = (out / "test_cases/TC-23/prompt.md").read_text()
        assert "CHF" in text
        assert "EUR" in text
        assert "USD" in text

    def test_prompt_mentions_snb(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        text = (out / "test_cases/TC-23/prompt.md").read_text()
        assert "SNB" in text

    def test_expected_behavior_exists(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        assert (out / "test_cases/TC-23/expected_behavior.md").exists()

    def test_expected_behavior_mentions_error(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        text = (out / "test_cases/TC-23/expected_behavior.md").read_text()
        assert "ERR-CH-002" in text
        assert "6,324" in text


# ---------------------------------------------------------------------------
# Gold standard
# ---------------------------------------------------------------------------


class TestGoldStandard:
    """Verify the TC-23 gold standard JSON."""

    def test_gold_file_exists(self) -> None:
        _, out, _, _, _ = _ensure_emitted()
        assert (out / "gold_standards/TC-23_gold.json").exists()

    def _load_gold(self) -> dict:
        _, out, _, _, _ = _ensure_emitted()
        return json.loads((out / "gold_standards/TC-23_gold.json").read_text())

    def test_required_sheets(self) -> None:
        gold = self._load_gold()
        sheets = gold["expected_outputs"]["required_sheets"]
        assert "Recon CHF" in sheets
        assert "Recon EUR" in sheets
        assert "Recon USD" in sheets
        assert "FX Revaluation Check" in sheets
        assert "Confirmation Tie-Out" in sheets
        assert "Consolidated Cash" in sheets

    def test_chf_reconciliation(self) -> None:
        gold = self._load_gold()
        chf = gold["expected_outputs"]["recon_chf"]
        assert chf["bank_ending_balance_chf"] == 2847290
        assert chf["outstanding_checks_count"] == 2
        assert chf["outstanding_checks_total_chf"] == 47815
        assert chf["deposits_in_transit_count"] == 1
        assert chf["deposits_in_transit_total_chf"] == 18430
        assert chf["adjusted_bank_balance_chf"] == 2817905
        assert chf["adjusted_book_balance_chf"] == 2817905
        assert chf["reconciliation_difference"] == 0

    def test_eur_reconciliation(self) -> None:
        gold = self._load_gold()
        eur = gold["expected_outputs"]["recon_eur"]
        assert eur["bank_ending_balance_eur"] == 1215400
        assert eur["adjusted_bank_balance_eur"] == 1186650
        assert eur["adjusted_bank_balance_chf_at_closing"] == 1264142
        assert eur["gl_ending_balance_chf"] == 1270466
        assert eur["fx_revaluation_error_chf"] == 6324
        assert eur["outstanding_sepa_count"] == 1
        assert eur["outstanding_sepa_total_eur"] == 28750

    def test_usd_reconciliation(self) -> None:
        gold = self._load_gold()
        usd = gold["expected_outputs"]["recon_usd"]
        assert usd["bank_ending_balance_usd"] == 489750
        assert usd["adjusted_bank_balance_usd"] == 474550
        assert usd["adjusted_bank_balance_chf_at_closing"] == 536518
        assert usd["gl_ending_balance_chf"] == 536518
        assert usd["reconciliation_difference"] == 0

    def test_fx_revaluation_check(self) -> None:
        gold = self._load_gold()
        fx = gold["expected_outputs"]["fx_revaluation_check"]
        assert fx["eur_rate_discrepancy"] is True
        assert fx["eur_impact_chf"] == 6324
        assert fx["usd_rate_discrepancy"] is False

    def test_consolidated_cash(self) -> None:
        gold = self._load_gold()
        cash = gold["expected_outputs"]["consolidated_cash"]
        assert cash["chf_adjusted_chf"] == 2817905
        assert cash["eur_adjusted_chf"] == 1264142
        assert cash["usd_adjusted_chf"] == 536518
        assert cash["total_consolidated_chf"] == 4618565

    def test_confirmation_tie_out(self) -> None:
        gold = self._load_gold()
        conf = gold["expected_outputs"]["confirmation_tie_out"]
        assert conf["chf_confirmation_matches_statement"] is True
        assert conf["eur_confirmation_matches_statement"] is True
        assert conf["usd_confirmation_matches_statement"] is True

    def test_error_detection(self) -> None:
        gold = self._load_gold()
        assert "ERR-CH-002" in gold["error_detection"]
        assert "6,324" in gold["error_detection"]["ERR-CH-002"]

    def test_canary_verification(self) -> None:
        gold = self._load_gold()
        cv = gold["canary_verification"]
        assert "read_bank_statement_chf" in cv
        assert "read_bank_statement_eur" in cv
        assert "read_bank_statement_usd" in cv
        assert "read_gl_cash_detail" in cv
        assert "read_bank_confirmation" in cv
        assert "read_snb_fx_rates" in cv

    def test_scoring_hints_present(self) -> None:
        gold = self._load_gold()
        sh = gold["scoring_hints"]
        assert "correctness" in sh
        assert "completeness" in sh
        assert "format_compliance" in sh
        assert "robustness" in sh
        assert "communication" in sh
