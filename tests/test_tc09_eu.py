"""Tests for TC-09-EU — OECD Transfer Pricing Documentation (Master File / Local File).

Verifies:
- Intercompany transactions XLSX (120+ rows, correct columns, all entity pairs)
- Comparable companies XLSX (2 sheets: Manufacturing with 15 companies, Distribution with 10)
- Interest rate benchmarks XLSX (EURIBOR rates + credit spreads, ERR-EU-005 planted)
- Master file PDF (exists, has canary in metadata)
- Local file CP PDF (exists, has canary)
- ERR-EU-005 planted error (registered, correct type "rounding_discrepancy", severity "material")
- Canary embedding in all 5 files
- Gold standard JSON (structure, expected outputs, scoring hints)
- Prompt and expected behavior markdown files
- Model computation tests (IQR calculations, transaction generation)
"""

from __future__ import annotations

import json
import tempfile
from pathlib import Path

import openpyxl

from generator.canaries import CanaryRegistry
from generator.canaries import build_registry as build_canary_registry
from generator.errors import ErrorRegistry
from generator.formatters.tc09_eu import emit_tc09_eu
from generator.golds.framework import emit_gold
from generator.manifest import Manifest
from generator.model.build import CascadeModel, build_model
from generator.model.tp_eu import (
    ALL_CANARY_KEYS_TC09EU,
    DIST_COMPARABLES,
    MFG_COMPARABLES,
    compute_dist_iqr,
    compute_mfg_iqr,
    generate_credit_spread_data,
    generate_euribor_data,
    generate_ic_transactions_eu,
)

# ---------------------------------------------------------------------------
# Module-level fixture: build the model and run emit_tc09_eu once
# ---------------------------------------------------------------------------

_MODEL: CascadeModel | None = None
_OUTPUT: Path | None = None
_CANARIES: CanaryRegistry | None = None
_ERRORS: ErrorRegistry | None = None
_MANIFEST: Manifest | None = None


def _ensure_emitted() -> tuple[CascadeModel, Path, CanaryRegistry, ErrorRegistry, Manifest]:
    global _MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST  # noqa: PLW0603
    if _MODEL is None:
        _MODEL = build_model(seed=42)
        _OUTPUT = Path(tempfile.mkdtemp(prefix="tc09eu_test_"))
        _CANARIES = build_canary_registry(ALL_CANARY_KEYS_TC09EU, seed=42)
        _ERRORS = ErrorRegistry()
        _MANIFEST = Manifest(_OUTPUT)
        _MANIFEST.__enter__()

        emit_tc09_eu(_MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST)
        emit_gold("TC-09-EU", _CANARIES, _ERRORS, _OUTPUT / "gold_standards", model=_MODEL)

        _MANIFEST.__exit__(None, None, None)
    return _MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST


_INPUT_DIR = "test_cases/TC-09-EU/input_files"


# ---------------------------------------------------------------------------
# Intercompany Transactions XLSX
# ---------------------------------------------------------------------------


class TestIntercompanyTransactions:
    """Verify intercompany_transactions_eu_fy2025.xlsx."""

    def test_xlsx_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "intercompany_transactions_eu_fy2025.xlsx"
        assert path.exists()

    def test_xlsx_has_canary(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "intercompany_transactions_eu_fy2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        expected = canaries.canary_for("tc09eu_intercompany_transactions")
        desc = wb.properties.description or ""
        assert expected in desc, f"Canary {expected} not in IC transactions xlsx"

    def test_has_transactions_sheet(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "intercompany_transactions_eu_fy2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        assert "Transactions" in wb.sheetnames

    def test_transaction_count(self) -> None:
        """Expect 80+ transaction rows; title rows 1-2, headers row 3, data from row 4."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "intercompany_transactions_eu_fy2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Transactions"]
        data_rows = [r for r in ws.iter_rows(min_row=4) if r[0].value is not None]
        assert len(data_rows) >= 80, f"Expected >=80 rows, got {len(data_rows)}"

    def test_has_all_entity_pairs(self) -> None:
        """All expected intercompany flows are present."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "intercompany_transactions_eu_fy2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Transactions"]
        pairs = set()
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row[0] is not None:
                from_e, to_e, tx_type = row[1], row[2], row[3]
                pairs.add((from_e, to_e, tx_type))
        # CP->CM goods, CP->CD goods, CE->CP/CM/CD management_fee,
        # CE->CM interest, CM->CP royalty
        assert ("CP", "CM", "goods") in pairs
        assert ("CP", "CD", "goods") in pairs
        assert ("CE", "CP", "management_fee") in pairs
        assert ("CE", "CM", "management_fee") in pairs
        assert ("CE", "CD", "management_fee") in pairs
        assert ("CE", "CM", "interest") in pairs
        # Royalty flow
        royalty_found = any(tx_type == "royalty" for (_, _, tx_type) in pairs)
        assert royalty_found, "No royalty transactions found"

    def test_columns(self) -> None:
        """Header row has expected column names."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "intercompany_transactions_eu_fy2025.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Transactions"]
        headers = [ws.cell(row=3, column=c).value for c in range(1, 11)]
        assert "Transaction ID" in headers
        assert "From Entity" in headers
        assert "To Entity" in headers
        assert "Transaction Type" in headers
        assert "Description" in headers


# ---------------------------------------------------------------------------
# Comparable Companies XLSX
# ---------------------------------------------------------------------------


class TestComparableCompanies:
    """Verify comparable_companies_eu.xlsx."""

    def test_xlsx_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "comparable_companies_eu.xlsx"
        assert path.exists()

    def test_xlsx_has_canary(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "comparable_companies_eu.xlsx"
        wb = openpyxl.load_workbook(str(path))
        expected = canaries.canary_for("tc09eu_comparable_companies")
        desc = wb.properties.description or ""
        assert expected in desc, f"Canary {expected} not in comparables xlsx"

    def test_has_mfg_sheet(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "comparable_companies_eu.xlsx"
        wb = openpyxl.load_workbook(str(path))
        assert "Manufacturing Comparables" in wb.sheetnames

    def test_has_dist_sheet(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "comparable_companies_eu.xlsx"
        wb = openpyxl.load_workbook(str(path))
        assert "Distribution Comparables" in wb.sheetnames

    def test_mfg_company_count(self) -> None:
        """15 manufacturing companies (12 accepted + 3 rejected)."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "comparable_companies_eu.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Manufacturing Comparables"]
        data_rows = [r for r in ws.iter_rows(min_row=2) if r[0].value is not None]
        assert len(data_rows) == 15, f"Expected 15 mfg companies, got {len(data_rows)}"

    def test_dist_company_count(self) -> None:
        """10 distribution companies (8 accepted + 2 rejected)."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "comparable_companies_eu.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Distribution Comparables"]
        data_rows = [r for r in ws.iter_rows(min_row=2) if r[0].value is not None]
        assert len(data_rows) == 10, f"Expected 10 dist companies, got {len(data_rows)}"

    def test_mfg_has_rejected(self) -> None:
        """3 manufacturing companies should be rejected."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "comparable_companies_eu.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Manufacturing Comparables"]
        rejected = []
        # Col 10 (index 9) is Accepted/Rejected
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            if row[9] == "Rejected":
                rejected.append(row[0])
        assert len(rejected) == 3, f"Expected 3 rejected mfg, got {len(rejected)}: {rejected}"

    def test_dist_has_rejected(self) -> None:
        """2 distribution companies should be rejected."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "comparable_companies_eu.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Distribution Comparables"]
        rejected = []
        # Col 9 (index 8) is Accepted/Rejected for dist sheet
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            if row[8] == "Rejected":
                rejected.append(row[0])
        assert len(rejected) == 2, f"Expected 2 rejected dist, got {len(rejected)}: {rejected}"


# ---------------------------------------------------------------------------
# Interest Rate Benchmarks XLSX
# ---------------------------------------------------------------------------


class TestInterestRateBenchmarks:
    """Verify interest_rate_benchmarks_eu.xlsx."""

    def test_xlsx_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "interest_rate_benchmarks_eu.xlsx"
        assert path.exists()

    def test_xlsx_has_canary(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "interest_rate_benchmarks_eu.xlsx"
        wb = openpyxl.load_workbook(str(path))
        expected = canaries.canary_for("tc09eu_interest_rate_benchmarks")
        desc = wb.properties.description or ""
        assert expected in desc, f"Canary {expected} not in interest rate xlsx"

    def test_has_euribor_sheet(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "interest_rate_benchmarks_eu.xlsx"
        wb = openpyxl.load_workbook(str(path))
        assert "EURIBOR Rates" in wb.sheetnames

    def test_has_credit_spread_sheet(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "interest_rate_benchmarks_eu.xlsx"
        wb = openpyxl.load_workbook(str(path))
        assert "Credit Spreads" in wb.sheetnames

    def test_euribor_has_err_eu_005(self) -> None:
        """CRITICAL: Q3 FY2025 12M rate should show 0.38 (the wrong value)."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "interest_rate_benchmarks_eu.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["EURIBOR Rates"]
        found = False
        # Columns: Period(0), Year(1), Quarter(2), Tenor(3), Rate%(4)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == 2025 and row[2] == 3 and row[3] == "12M":
                rate = row[4]
                assert abs(rate - 0.38) < 0.01, (
                    f"Q3 FY2025 12M should be 0.38, got {rate}"
                )
                found = True
                break
        assert found, "Q3 FY2025 12M row not found in EURIBOR Rates sheet"

    def test_euribor_entry_count(self) -> None:
        """24 entries: 8 quarters × 3 tenors (3M, 6M, 12M)."""
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "interest_rate_benchmarks_eu.xlsx"
        wb = openpyxl.load_workbook(str(path))
        ws = wb["EURIBOR Rates"]
        data_rows = [r for r in ws.iter_rows(min_row=2) if r[0].value is not None]
        assert len(data_rows) == 24, f"Expected 24 EURIBOR rows, got {len(data_rows)}"


# ---------------------------------------------------------------------------
# Master File PDF
# ---------------------------------------------------------------------------


class TestMasterFilePDF:
    """Verify master_file_fy2024.pdf."""

    def test_pdf_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "master_file_fy2024.pdf"
        assert path.exists()

    def test_pdf_has_canary(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        from pypdf import PdfReader

        path = output / _INPUT_DIR / "master_file_fy2024.pdf"
        reader = PdfReader(str(path))
        author = reader.metadata.author or ""
        expected = canaries.canary_for("tc09eu_master_file_fy2024")
        assert expected in author, f"Canary {expected} not in master file PDF author"

    def test_pdf_page_count(self) -> None:
        """Master file should be >= 20 pages (target 28)."""
        _, output, _, _, _ = _ensure_emitted()
        from pypdf import PdfReader

        path = output / _INPUT_DIR / "master_file_fy2024.pdf"
        reader = PdfReader(str(path))
        assert len(reader.pages) >= 20, (
            f"Expected >= 20 pages, got {len(reader.pages)}"
        )


# ---------------------------------------------------------------------------
# Local File CP PDF
# ---------------------------------------------------------------------------


class TestLocalFileCPPDF:
    """Verify local_file_cp_fy2024.pdf."""

    def test_pdf_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / _INPUT_DIR / "local_file_cp_fy2024.pdf"
        assert path.exists()

    def test_pdf_has_canary(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        from pypdf import PdfReader

        path = output / _INPUT_DIR / "local_file_cp_fy2024.pdf"
        reader = PdfReader(str(path))
        author = reader.metadata.author or ""
        expected = canaries.canary_for("tc09eu_local_file_cp_fy2024")
        assert expected in author, f"Canary {expected} not in local file CP PDF author"

    def test_pdf_page_count(self) -> None:
        """Local file CP should be >= 25 pages (target 35)."""
        _, output, _, _, _ = _ensure_emitted()
        from pypdf import PdfReader

        path = output / _INPUT_DIR / "local_file_cp_fy2024.pdf"
        reader = PdfReader(str(path))
        assert len(reader.pages) >= 25, (
            f"Expected >= 25 pages, got {len(reader.pages)}"
        )


# ---------------------------------------------------------------------------
# ERR-EU-005 — EURIBOR decimal point error
# ---------------------------------------------------------------------------


class TestErrEU005:
    """Verify ERR-EU-005: EURIBOR 12M Q3 FY2025 decimal point error.

    Note: The error ID in the formatter is ERR-EU-005 (rounding_discrepancy).
    """

    def test_error_registered(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        assert "ERR-EU-005" in errors.entries

    def test_error_type(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        assert errors.entries["ERR-EU-005"].type == "rounding_discrepancy"

    def test_error_severity(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        assert errors.entries["ERR-EU-005"].severity == "material"

    def test_error_mentions_euribor(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-EU-005"]
        assert "EURIBOR" in err.description or "euribor" in err.description.lower()

    def test_error_references_decimal_point(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-EU-005"]
        assert "decimal" in err.description.lower() or "0.38" in err.description


# ---------------------------------------------------------------------------
# Gold Standard
# ---------------------------------------------------------------------------


class TestGoldStandard:
    """Verify TC-09-EU gold standard structure and values."""

    def _gold(self) -> dict:
        _, output, _, _, _ = _ensure_emitted()
        gold_path = output / "gold_standards" / "TC-09-EU_gold.json"
        return json.loads(gold_path.read_text())

    def test_gold_file_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        assert (output / "gold_standards" / "TC-09-EU_gold.json").exists()

    def test_gold_test_case_id(self) -> None:
        gold = self._gold()
        assert gold["test_case"] == "TC-09-EU"

    def test_gold_has_mfg_screening(self) -> None:
        gold = self._gold()
        assert "manufacturing_comparable_screening" in gold["expected_outputs"]
        mfg = gold["expected_outputs"]["manufacturing_comparable_screening"]
        assert mfg["total_companies"] == 15
        assert mfg["accepted"] == 12
        assert mfg["rejected"] == 3

    def test_gold_has_dist_screening(self) -> None:
        gold = self._gold()
        assert "distribution_comparable_screening" in gold["expected_outputs"]
        dist = gold["expected_outputs"]["distribution_comparable_screening"]
        assert dist["total_companies"] == 10
        assert dist["accepted"] == 8
        assert dist["rejected"] == 2

    def test_gold_has_arm_length_assessment(self) -> None:
        gold = self._gold()
        eo = gold["expected_outputs"]
        ala = eo["arm_length_assessment"]
        assert ala["cp_within_iqr"] is True
        assert ala["cd_within_iqr"] is True
        assert ala["loan_within_range"] is True

    def test_gold_has_interest_rate_benchmarking(self) -> None:
        gold = self._gold()
        irb = gold["expected_outputs"]["interest_rate_benchmarking"]
        assert irb["actual_rate_pct"] == "4.5"
        assert "arm_length_range_pct" in irb

    def test_gold_has_canary_verification(self) -> None:
        gold = self._gold()
        cv = gold["canary_verification"]
        assert len(cv) == 5
        assert "read_ic_transactions" in cv
        assert "read_comparable_companies" in cv
        assert "read_master_file" in cv
        assert "read_local_file_cp" in cv
        assert "read_interest_benchmarks" in cv

    def test_gold_has_error_detection(self) -> None:
        gold = self._gold()
        assert "ERR-EU-005" in gold["error_detection"]

    def test_gold_has_scoring_hints(self) -> None:
        gold = self._gold()
        assert "scoring_hints" in gold
        assert "correctness" in gold["scoring_hints"]
        assert "completeness" in gold["scoring_hints"]
        assert "robustness" in gold["scoring_hints"]


# ---------------------------------------------------------------------------
# Prompt and Expected Behavior
# ---------------------------------------------------------------------------


class TestPromptAndExpectedBehavior:
    """Verify prompt.md and expected_behavior.md exist with key content."""

    def test_prompt_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        assert (output / "test_cases/TC-09-EU/prompt.md").exists()

    def test_prompt_mentions_oecd(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        content = (output / "test_cases/TC-09-EU/prompt.md").read_text()
        assert "OECD" in content

    def test_prompt_mentions_master_file(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        content = (output / "test_cases/TC-09-EU/prompt.md").read_text()
        assert "master file" in content.lower() or "Master File" in content

    def test_expected_behavior_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        assert (output / "test_cases/TC-09-EU/expected_behavior.md").exists()

    def test_expected_behavior_mentions_iqr(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        content = (output / "test_cases/TC-09-EU/expected_behavior.md").read_text()
        assert "IQR" in content or "interquartile" in content.lower()

    def test_expected_behavior_mentions_dempe(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        content = (output / "test_cases/TC-09-EU/expected_behavior.md").read_text()
        assert "DEMPE" in content


# ---------------------------------------------------------------------------
# Model computation tests (direct unit tests)
# ---------------------------------------------------------------------------


class TestModelComputations:
    """Verify transfer pricing model computations directly."""

    def test_ic_transaction_count(self) -> None:
        """Model generates 84 transactions (12 months x 7 flows); formatter adds extras."""
        txns = generate_ic_transactions_eu()
        assert len(txns) >= 80, f"Expected >= 80 IC transactions, got {len(txns)}"

    def test_ic_transaction_entities(self) -> None:
        """All expected entity pairs present in model transactions."""
        txns = generate_ic_transactions_eu()
        pairs = {(t.from_entity, t.to_entity, t.transaction_type) for t in txns}
        assert ("CP", "CM", "goods") in pairs
        assert ("CP", "CD", "goods") in pairs
        assert ("CE", "CP", "management_fee") in pairs
        assert ("CE", "CM", "management_fee") in pairs
        assert ("CE", "CD", "management_fee") in pairs
        assert ("CE", "CM", "interest") in pairs
        assert ("CP", "CM", "royalty") in pairs

    def test_mfg_iqr_values(self) -> None:
        """Manufacturing IQR: Q1=3.8, median~5.6, Q3=7.9."""
        iqr = compute_mfg_iqr()
        assert iqr["q1_pct"] == 3.8, f"MFG Q1 expected 3.8, got {iqr['q1_pct']}"
        assert iqr["median_pct"] == 5.6, f"MFG median expected 5.6, got {iqr['median_pct']}"
        assert iqr["q3_pct"] == 7.9, f"MFG Q3 expected 7.9, got {iqr['q3_pct']}"

    def test_dist_iqr_values(self) -> None:
        """Distribution IQR: Q1=1.2, median~2.1, Q3=3.5."""
        iqr = compute_dist_iqr()
        assert iqr["q1_pct"] == 1.2, f"DIST Q1 expected 1.2, got {iqr['q1_pct']}"
        assert iqr["median_pct"] == 2.1, f"DIST median expected 2.1, got {iqr['median_pct']}"
        assert iqr["q3_pct"] == 3.5, f"DIST Q3 expected 3.5, got {iqr['q3_pct']}"

    def test_mfg_rejected_count(self) -> None:
        """3 manufacturing companies rejected."""
        rejected = [c for c in MFG_COMPARABLES if c[9]]
        assert len(rejected) == 3, f"Expected 3 rejected mfg, got {len(rejected)}"

    def test_dist_rejected_count(self) -> None:
        """2 distribution companies rejected."""
        rejected = [c for c in DIST_COMPARABLES if c[8]]
        assert len(rejected) == 2, f"Expected 2 rejected dist, got {len(rejected)}"

    def test_euribor_data_count(self) -> None:
        """24 EURIBOR entries (8 quarters x 3 tenors)."""
        data = generate_euribor_data()
        assert len(data) == 24, f"Expected 24 EURIBOR entries, got {len(data)}"

    def test_credit_spread_data_count(self) -> None:
        """8 credit spread entries."""
        data = generate_credit_spread_data()
        assert len(data) == 8, f"Expected 8 credit spread entries, got {len(data)}"


# ---------------------------------------------------------------------------
# Determinism test
# ---------------------------------------------------------------------------


class TestDeterminism:
    """Verify deterministic output across two runs."""

    def test_ic_transactions_deterministic(self) -> None:
        txns1 = generate_ic_transactions_eu()
        txns2 = generate_ic_transactions_eu()
        assert len(txns1) == len(txns2)
        for t1, t2 in zip(txns1[:50], txns2[:50]):
            assert t1 == t2

    def test_iqr_deterministic(self) -> None:
        mfg1 = compute_mfg_iqr()
        mfg2 = compute_mfg_iqr()
        assert mfg1 == mfg2
        dist1 = compute_dist_iqr()
        dist2 = compute_dist_iqr()
        assert dist1 == dist2
