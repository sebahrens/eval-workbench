"""Tests for TC-07-EU — European Partnership Investment Allocation Extraction.

Verifies:
- 8 allocation statement PDFs in allocation_statements/ (5 clean, 2 scanned, 1 amended)
- tc07eu_investment_register.xlsx
- tc07eu_withholding_tax_summary.xlsx
- ERR-EU-004 planted error (WHT rate mismatch Thames Valley 20% vs 15%)
- ERR-EU-005 planted error (partner share mismatch Capital Croissance 4.8% vs 5.0%)
- Amended statement detection (Beteiligungen München KG: €340K→€285K + €55K interest)
- Canary embedding in all files
- Gold standard structure, per-investment detail, consolidated totals, scoring hints
- Prompt and expected behavior markdown files
"""

from __future__ import annotations

import json
import tempfile
from decimal import Decimal
from pathlib import Path

import openpyxl

from generator.canaries import CanaryRegistry
from generator.canaries import build_registry as build_canary_registry
from generator.errors import ErrorRegistry
from generator.formatters.tc07_eu import emit_tc07_eu
from generator.golds.framework import emit_gold
from generator.manifest import Manifest
from generator.model.build import CascadeModel, build_model
from generator.model.k1_eu import (
    ALL_CANARY_KEYS,
    GBP_EUR_RATE,
    AllocLanguage,
    AllocPdfStyle,
    alloc_canary_key,
    generate_eu_investments,
)

# ---------------------------------------------------------------------------
# Module-level fixture: build the model and run emit_tc07_eu once
# ---------------------------------------------------------------------------

_MODEL: CascadeModel | None = None
_OUTPUT: Path | None = None
_CANARIES: CanaryRegistry | None = None
_ERRORS: ErrorRegistry | None = None
_MANIFEST: Manifest | None = None


def _ensure_emitted() -> tuple[CascadeModel, Path, CanaryRegistry, ErrorRegistry, Manifest]:
    global _MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST  # noqa: PLW0603
    if _MODEL is None:
        _MODEL = build_model(seed=42)
        _OUTPUT = Path(tempfile.mkdtemp(prefix="tc07eu_test_"))
        _CANARIES = build_canary_registry(ALL_CANARY_KEYS, seed=42)
        _ERRORS = ErrorRegistry()
        _MANIFEST = Manifest(_OUTPUT)
        _MANIFEST.__enter__()

        emit_tc07_eu(_MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST)

        # Emit gold standard
        emit_gold("TC-07-EU", _CANARIES, _ERRORS, _OUTPUT / "gold_standards", model=_MODEL)

        _MANIFEST.__exit__(None, None, None)
    return _MODEL, _OUTPUT, _CANARIES, _ERRORS, _MANIFEST


_INPUT_DIR = "test_cases/TC-07-EU/input_files"
_ALLOC_DIR = f"{_INPUT_DIR}/allocation_statements"


# ---------------------------------------------------------------------------
# Allocation statement PDFs — existence and count
# ---------------------------------------------------------------------------


class TestAllocationPDFs:
    """Verify 8 allocation statement PDFs are emitted."""

    def test_alloc_directory_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        assert (output / _ALLOC_DIR).is_dir()

    def test_8_alloc_pdfs_exist(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        investments = generate_eu_investments()
        for inv in investments:
            path = output / _ALLOC_DIR / inv.alloc_filename
            assert path.exists(), f"Missing {inv.alloc_filename}"

    def test_no_extra_pdfs(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        pdfs = list((output / _ALLOC_DIR).glob("*.pdf"))
        assert len(pdfs) == 8

    def test_5_clean_2_scanned_1_amended(self) -> None:
        """PDF styles: 6 clean text (incl amended), 2 scanned-style."""
        investments = generate_eu_investments()
        clean = [i for i in investments if i.pdf_style == AllocPdfStyle.CLEAN_TEXT]
        scanned = [i for i in investments if i.pdf_style == AllocPdfStyle.SCANNED]
        assert len(clean) == 6
        assert len(scanned) == 2

    def test_multilingual(self) -> None:
        """Verify 4 languages represented."""
        investments = generate_eu_investments()
        languages = {i.language for i in investments}
        assert AllocLanguage.GERMAN in languages
        assert AllocLanguage.FRENCH in languages
        assert AllocLanguage.DUTCH in languages
        assert AllocLanguage.ENGLISH in languages

    def test_jurisdictions(self) -> None:
        """Verify 5 jurisdictions: DE, FR, LU, NL, UK."""
        investments = generate_eu_investments()
        jurisdictions = {i.jurisdiction for i in investments}
        assert jurisdictions == {"DE", "FR", "LU", "NL", "UK"}


# ---------------------------------------------------------------------------
# Investment register XLSX
# ---------------------------------------------------------------------------


class TestInvestmentRegister:
    """Verify tc07eu_investment_register.xlsx."""

    def test_file_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/tc07eu_investment_register.xlsx"
        assert path.exists()

    def test_has_8_data_rows(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(output / f"{_INPUT_DIR}/tc07eu_investment_register.xlsx")
        ws = wb.active
        # Header at row 1, data starts at row 2
        data_rows = [r for r in ws.iter_rows(min_row=2, max_col=1, values_only=True) if r[0]]
        assert len(data_rows) == 8

    def test_capital_croissance_register_shows_5pct(self) -> None:
        """ERR-EU-005: register should show 5.0% for Capital Croissance (not 4.8%)."""
        _, output, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(output / f"{_INPUT_DIR}/tc07eu_investment_register.xlsx")
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == "ALLOC-004":
                assert row[4] == 5.0, f"Expected 5.0% for ALLOC-004, got {row[4]}"
                break
        else:
            raise AssertionError("ALLOC-004 not found in register")

    def test_benelux_ventures_ias28_note(self) -> None:
        """Investment register notes board-seat justification for IAS 28."""
        _, output, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(output / f"{_INPUT_DIR}/tc07eu_investment_register.xlsx")
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == "ALLOC-005":
                assert row[8] == "IAS 28 associate"
                assert "board seat" in str(row[10]).lower()
                break


# ---------------------------------------------------------------------------
# WHT summary XLSX
# ---------------------------------------------------------------------------


class TestWHTSummary:
    """Verify tc07eu_withholding_tax_summary.xlsx."""

    def test_file_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        path = output / f"{_INPUT_DIR}/tc07eu_withholding_tax_summary.xlsx"
        assert path.exists()

    def test_has_rate_entries(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(output / f"{_INPUT_DIR}/tc07eu_withholding_tax_summary.xlsx")
        ws = wb.active
        data_rows = [r for r in ws.iter_rows(min_row=2, max_col=1, values_only=True) if r[0]]
        assert len(data_rows) >= 10

    def test_uk_nrl_entry(self) -> None:
        """WHT summary includes UK NRL scheme entry with 20% domestic / 15% treaty."""
        _, output, _, _, _ = _ensure_emitted()
        wb = openpyxl.load_workbook(output / f"{_INPUT_DIR}/tc07eu_withholding_tax_summary.xlsx")
        ws = wb.active
        found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == "UK" and "NRL" in str(row[1]):
                assert row[2] == 20.0
                assert row[3] == 15.0
                found = True
                break
        assert found, "UK NRL entry not found in WHT summary"


# ---------------------------------------------------------------------------
# ERR-EU-004 — WHT rate mismatch Thames Valley
# ---------------------------------------------------------------------------


class TestERR_EU_004:
    """Verify ERR-EU-004: Thames Valley WHT 20% vs 15% treaty."""

    def test_registered(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        assert "ERR-EU-004" in errors.entries

    def test_type(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        assert errors.entries["ERR-EU-004"].type == "mismatched_total"

    def test_references_thames_valley(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-EU-004"]
        assert "Thames Valley" in err.description or "thames_valley" in err.file

    def test_mentions_rate_discrepancy(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-EU-004"]
        assert "20%" in err.description


# ---------------------------------------------------------------------------
# ERR-EU-005 — Partner share mismatch Capital Croissance
# ---------------------------------------------------------------------------


class TestERR_EU_005:
    """Verify ERR-EU-005: Capital Croissance 4.8% vs 5.0%."""

    def test_registered(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        assert "ERR-EU-005" in errors.entries

    def test_type(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        assert errors.entries["ERR-EU-005"].type == "mismatched_total"

    def test_mentions_share_discrepancy(self) -> None:
        _, _, _, errors, _ = _ensure_emitted()
        err = errors.entries["ERR-EU-005"]
        assert "4.8%" in err.description
        assert "5.0%" in err.description


# ---------------------------------------------------------------------------
# Amended statement detection
# ---------------------------------------------------------------------------


class TestAmendedStatement:
    """Verify the amended allocation (Beteiligungen München KG)."""

    def test_exactly_one_amended(self) -> None:
        investments = generate_eu_investments()
        amended = [i for i in investments if i.is_amended]
        assert len(amended) == 1

    def test_amended_is_alloc_008(self) -> None:
        investments = generate_eu_investments()
        amended = [i for i in investments if i.is_amended][0]
        assert amended.alloc_id == "ALLOC-008"

    def test_amended_profit_285k(self) -> None:
        investments = generate_eu_investments()
        amended = [i for i in investments if i.is_amended][0]
        profit_cats = [c for c in amended.categories if c.local_label == "Gewinnanteil"]
        assert len(profit_cats) == 1
        assert profit_cats[0].amount_eur == Decimal("285000")

    def test_amended_interest_55k(self) -> None:
        investments = generate_eu_investments()
        amended = [i for i in investments if i.is_amended][0]
        interest_cats = [c for c in amended.categories if "Zinserträge" in c.local_label]
        assert len(interest_cats) == 1
        assert interest_cats[0].amount_eur == Decimal("55000")

    def test_amendment_records(self) -> None:
        investments = generate_eu_investments()
        amended = [i for i in investments if i.is_amended][0]
        assert len(amended.amendments) == 2


# ---------------------------------------------------------------------------
# GBP conversion
# ---------------------------------------------------------------------------


class TestGBPConversion:
    """Verify Thames Valley GBP amounts convert correctly."""

    def test_gbp_eur_rate(self) -> None:
        assert GBP_EUR_RATE == Decimal("1.17")

    def test_thames_valley_income_eur(self) -> None:
        investments = generate_eu_investments()
        thames = next(i for i in investments if i.alloc_id == "ALLOC-006")
        expected = Decimal("192500") * GBP_EUR_RATE
        assert thames.total_income_eur == expected

    def test_thames_valley_wht_eur(self) -> None:
        investments = generate_eu_investments()
        thames = next(i for i in investments if i.alloc_id == "ALLOC-006")
        expected = Decimal("38500") * GBP_EUR_RATE
        assert thames.wht_amount_eur == expected


# ---------------------------------------------------------------------------
# Canary embedding
# ---------------------------------------------------------------------------


class TestCanaryEmbedding:
    """Verify canary codes are embedded in all TC-07-EU files."""

    def test_all_canary_keys_assigned(self) -> None:
        _, _, canaries, _, _ = _ensure_emitted()
        for key in ALL_CANARY_KEYS:
            code = canaries.canary_for(key)
            assert len(code) == 8

    def test_alloc_pdf_canaries_in_metadata(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        investments = generate_eu_investments()
        for inv in investments:
            key = alloc_canary_key(inv.alloc_id)
            canary = canaries.canary_for(key)
            path = output / _ALLOC_DIR / inv.alloc_filename
            content = path.read_bytes()
            assert canary.encode() in content, (
                f"Canary {canary} not found in {inv.alloc_filename}"
            )

    def test_register_canary(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("tc07eu_investment_register")
        path = output / f"{_INPUT_DIR}/tc07eu_investment_register.xlsx"
        wb = openpyxl.load_workbook(str(path))
        desc = wb.properties.description or ""
        assert canary in desc, f"Canary {canary} not in register properties"

    def test_wht_summary_canary(self) -> None:
        _, output, canaries, _, _ = _ensure_emitted()
        canary = canaries.canary_for("tc07eu_wht_summary")
        path = output / f"{_INPUT_DIR}/tc07eu_withholding_tax_summary.xlsx"
        wb = openpyxl.load_workbook(str(path))
        desc = wb.properties.description or ""
        assert canary in desc, f"Canary {canary} not in WHT summary properties"


# ---------------------------------------------------------------------------
# Gold standard
# ---------------------------------------------------------------------------


class TestGoldStandard:
    """Verify gold standard JSON structure."""

    def test_gold_json_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        assert (output / "gold_standards" / "TC-07-EU_gold.json").exists()

    def test_gold_has_8_allocations(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-07-EU_gold.json").read_text())
        assert gold["expected_outputs"]["allocation_count"] == 8

    def test_gold_has_allocation_details(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-07-EU_gold.json").read_text())
        details = gold["expected_outputs"]["allocation_details"]
        assert len(details) == 8

    def test_gold_has_consolidated_totals(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-07-EU_gold.json").read_text())
        assert len(gold["expected_outputs"]["consolidated_by_category"]) > 0
        assert len(gold["expected_outputs"]["consolidated_by_jurisdiction"]) > 0

    def test_gold_has_amended_info(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-07-EU_gold.json").read_text())
        amended = gold["expected_outputs"]["amended_statement"]
        assert amended["alloc_id"] == "ALLOC-008"

    def test_gold_has_participation_flags(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-07-EU_gold.json").read_text())
        flags = gold["expected_outputs"]["participation_exemption_flags"]
        assert len(flags) >= 4

    def test_gold_canary_verification_count(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-07-EU_gold.json").read_text())
        assert len(gold["canary_verification"]) == 10

    def test_gold_error_detection(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-07-EU_gold.json").read_text())
        assert "ERR-EU-004" in gold["error_detection"]
        assert "ERR-EU-005" in gold["error_detection"]

    def test_gold_scoring_hints(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        gold = json.loads((output / "gold_standards" / "TC-07-EU_gold.json").read_text())
        for key in ["correctness", "completeness", "format_compliance", "communication"]:
            assert key in gold["scoring_hints"]


# ---------------------------------------------------------------------------
# Prompt and expected behavior
# ---------------------------------------------------------------------------


class TestPromptAndExpectedBehavior:
    """Verify prompt and expected behavior files."""

    def test_prompt_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        assert (output / "test_cases/TC-07-EU/prompt.md").exists()

    def test_prompt_mentions_allocation_statements(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-07-EU/prompt.md").read_text()
        assert "allocation" in text.lower()

    def test_prompt_mentions_deelnemingsvrijstelling(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-07-EU/prompt.md").read_text()
        assert "deelnemingsvrijstelling" in text

    def test_prompt_mentions_gbp_conversion(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-07-EU/prompt.md").read_text()
        assert "1.17" in text

    def test_expected_behavior_exists(self) -> None:
        _, output, _, _, _ = _ensure_emitted()
        assert (output / "test_cases/TC-07-EU/expected_behavior.md").exists()

    def test_expected_behavior_no_us_references(self) -> None:
        """Rubric should not mention K-1/Form 1120/Section 199A."""
        _, output, _, _, _ = _ensure_emitted()
        text = (output / "test_cases/TC-07-EU/expected_behavior.md").read_text()
        assert "K-1" not in text
        assert "Form 1120" not in text
        assert "Section 199A" not in text


# ---------------------------------------------------------------------------
# Manifest registration
# ---------------------------------------------------------------------------


class TestManifest:
    """Verify files are registered in the manifest."""

    def test_manifest_has_tc07eu_entries(self) -> None:
        _, _, _, _, manifest = _ensure_emitted()
        tc07eu_count = sum(
            1 for v in manifest.entries.values()
            if "TC-07-EU" in (v.test_cases or [])
        )
        assert tc07eu_count >= 10
