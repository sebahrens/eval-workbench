"""Deterministic document writer helpers for the Cascade test suite.

Shared utilities for saving xlsx/docx/PDF files with pinned timestamps
so that output is byte-identical across runs.
"""

from __future__ import annotations

import datetime
import io
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side

# ── Fixed timestamps ────────────────────────────────────────────────────────

FIXED_DATETIME = datetime.datetime(2025, 3, 15, 9, 0, 0)
FIXED_ZIP_DT = (2025, 3, 15, 9, 0, 0)

# ── Common styles ───────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BOLD_FONT = Font(bold=True, size=10)
NORMAL_FONT = Font(size=10)
THIN_BORDER = Border(
    bottom=Side(style="thin"),
    top=Side(style="thin"),
    left=Side(style="thin"),
    right=Side(style="thin"),
)
NUMBER_FMT = "#,##0"
DOLLAR_FMT = '#,##0'


# ── Deterministic save helpers ──────────────────────────────────────────────


def pin_xlsx_dates(wb: openpyxl.Workbook) -> None:
    """Pin created/modified timestamps for determinism."""
    wb.properties.created = FIXED_DATETIME
    wb.properties.modified = FIXED_DATETIME


def save_xlsx_deterministic(wb: openpyxl.Workbook, path: str | Path) -> None:
    """Save workbook with pinned timestamps and fixed zip entry dates."""
    from openpyxl.writer.excel import ExcelWriter

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    buf = io.BytesIO()
    wb.properties.modified = FIXED_DATETIME
    archive = ZipFile(buf, "w", ZIP_DEFLATED, allowZip64=True)
    writer = ExcelWriter(wb, archive)
    writer.save()

    buf.seek(0)
    with ZipFile(buf, "r") as src, ZipFile(str(path), "w", ZIP_DEFLATED) as dst:
        for item in src.infolist():
            info = ZipInfo(item.filename, date_time=FIXED_ZIP_DT)
            info.compress_type = item.compress_type
            dst.writestr(info, src.read(item.filename))


def save_docx_deterministic(doc: Any, path: str | Path) -> None:
    """Save a python-docx Document with fixed zip entry timestamps."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    buf = io.BytesIO()
    doc.save(buf)

    buf.seek(0)
    with ZipFile(buf, "r") as src, ZipFile(str(path), "w", ZIP_DEFLATED) as dst:
        for item in src.infolist():
            info = ZipInfo(item.filename, date_time=FIXED_ZIP_DT)
            info.compress_type = item.compress_type
            dst.writestr(info, src.read(item.filename))


def new_workbook() -> openpyxl.Workbook:
    """Create a new Workbook with timestamps pinned for determinism."""
    wb = openpyxl.Workbook()
    pin_xlsx_dates(wb)
    return wb


def write_header_row(ws: Any, headers: list[str], row: int = 1) -> None:
    """Write a styled header row to a worksheet."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
