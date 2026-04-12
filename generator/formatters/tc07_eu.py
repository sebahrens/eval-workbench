"""Formatter: TC-07-EU — European Partnership Investment Allocation Extraction.

Emits:
- test_cases/TC-07-EU/input_files/allocation_statements/
  8 allocation statement PDFs from European partnerships (5 clean, 2 scanned, 1 amended)
- test_cases/TC-07-EU/input_files/tc07eu_investment_register.xlsx
  CE's investment register (8 investments)
- test_cases/TC-07-EU/input_files/tc07eu_withholding_tax_summary.xlsx
  WHT rates by jurisdiction and income type
- test_cases/TC-07-EU/prompt.md
- test_cases/TC-07-EU/expected_behavior.md
- gold_standards/TC-07-EU_gold.json

Planted errors:
  ERR-EU-004: mismatched_total — Thames Valley LLP WHT at 20% NRL vs 15% treaty
  ERR-EU-005: mismatched_total — Capital Croissance partner share 4.8% vs 5.0%

Uses deterministic European investment model — never hardcodes numbers that
should come from the model.
"""

from __future__ import annotations

import datetime
import io
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import openpyxl
from fpdf import FPDF
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from generator.canaries import (
    CanaryRegistry,
    embed_canary_pdf_fpdf2,
    embed_canary_xlsx,
)
from generator.errors import ErrorRegistry, PlantedError
from generator.golds.framework import GoldStandard, register_gold
from generator.manifest import Manifest
from generator.model.build import CascadeModel
from generator.model.k1_eu import (
    GBP_EUR_RATE,
    INVESTMENT_REGISTER_KEY,
    WHT_RATES,
    WHT_SUMMARY_KEY,
    EUPartnershipInvestment,
    alloc_canary_key,
    consolidated_by_jurisdiction,
    consolidated_totals_eu,
    generate_eu_investments,
    total_wht,
)

# ── Constants ────────────────────────────────────────────────────────────────

_TC = "TC-07-EU"
_INPUT_DIR = f"test_cases/{_TC}/input_files"
_ALLOC_DIR = f"{_INPUT_DIR}/allocation_statements"

_FIXED_DATE = "2025-03-15"
_CREATION_DATE = datetime.datetime(2025, 3, 15, 9, 0, 0)
_FIXED_DATETIME = datetime.datetime(2025, 3, 15, 9, 0, 0)
_FIXED_ZIP_DT = (2025, 3, 15, 9, 0, 0)


# ── Formatting helpers ──────────────────────────────────────────────────────

def _fmt_eur(d: Decimal | int) -> str:
    """Format EUR with European comma-decimal: 45.230,00 EUR."""
    if isinstance(d, int):
        val = d
    else:
        val = int(d.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    if val < 0:
        return f"-{abs(val):,.0f} EUR".replace(",", ".")
    return f"{val:,.0f} EUR".replace(",", ".")


def _fmt_eur_comma(d: Decimal | int) -> str:
    """Format EUR with comma decimal: 45.230,00 EUR.

    Uses 'EUR' instead of the euro sign to stay within Latin-1 encoding
    (fpdf2 core fonts do not support U+20AC).
    """
    if isinstance(d, int):
        val = d
    else:
        val = int(d.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    # Use period for thousands, comma for decimal
    formatted = f"{val:,}".replace(",", ".")
    return f"{formatted},00 EUR"


def _fmt_gbp(d: Decimal | int) -> str:
    """Format GBP: £38,500.00."""
    if isinstance(d, int):
        val = d
    else:
        val = int(d.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    return f"£{val:,.2f}"


def _fmt_pct(d: Decimal) -> str:
    """Format percentage: 8,0% (European style) or 8.0% (English)."""
    return f"{d}%"


def _fmt_int(d: Decimal | int | None) -> int | None:
    """Convert Decimal to int for gold standard."""
    if d is None:
        return None
    if isinstance(d, int):
        return d
    return int(d.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


# ── Deterministic save helpers ──────────────────────────────────────────────

def _save_xlsx_deterministic(wb: openpyxl.Workbook, path: str | Path) -> None:
    """Save workbook with pinned timestamps and fixed zip entry dates."""
    from openpyxl.writer.excel import ExcelWriter

    path = Path(path)
    wb.properties.modified = _FIXED_DATETIME

    buf = io.BytesIO()
    archive = ZipFile(buf, "w", ZIP_DEFLATED, allowZip64=True)
    writer = ExcelWriter(wb, archive)
    writer.save()

    buf.seek(0)
    with ZipFile(buf, "r") as src, ZipFile(str(path), "w", ZIP_DEFLATED) as dst:
        for item in src.infolist():
            info = ZipInfo(item.filename, date_time=_FIXED_ZIP_DT)
            info.compress_type = item.compress_type
            dst.writestr(info, src.read(item.filename))


def _pin_xlsx_dates(wb: openpyxl.Workbook) -> None:
    wb.properties.created = _FIXED_DATETIME


def _latin1_safe(text: str) -> str:
    """Replace characters unsupported by fpdf2 core fonts (latin-1 only).

    Replaces euro sign (U+20AC) with 'EUR' and em-dash (U+2014) with '-'.
    """
    return text.replace("\u20ac", "EUR").replace("\u2014", "-")


# ── PDF helpers ─────────────────────────────────────────────────────────────


def _latin1_safe(text: str) -> str:
    """Replace characters outside Latin-1 with ASCII equivalents for fpdf2 core fonts."""
    return text.replace("\u20ac", "EUR").replace("\u2014", "-").replace("\u2013", "-")


def _new_pdf(canary: str) -> FPDF:
    """Create an FPDF instance with deterministic metadata and embedded canary."""
    pdf = FPDF()
    pdf.set_creation_date(_CREATION_DATE)
    pdf.set_auto_page_break(auto=False)
    embed_canary_pdf_fpdf2(pdf, canary)
    return pdf


# ── PDF renderers by jurisdiction ──────────────────────────────────────────

def _render_german_kg(pdf: FPDF, inv: EUPartnershipInvestment) -> None:
    """Render a German KG/GmbH & Co. KG allocation statement (clean text)."""
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)

    if inv.is_amended:
        pdf.cell(0, 8, "KORRIGIERT / AMENDED", new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.ln(2)

    pdf.cell(0, 8, "Gewinnverteilungsmitteilung", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Geschäftsjahr {inv.fiscal_year}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(5)

    # Partnership info
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, "Gesellschaft:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, f"  {inv.fund_name}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, f"  Rechtsform: {inv.legal_form}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # Partner info
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, "Kommanditist:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, "  Cascade Europe Holdings B.V.", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, f"  Beteiligungsquote: {inv.partner_share_pct}%", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # Income allocation
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, "Ergebnisverteilung:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)

    for cat in inv.categories:
        pdf.cell(120, 5, f"  {cat.local_label}")
        pdf.cell(0, 5, _fmt_eur_comma(cat.amount_eur), new_x="LMARGIN", new_y="NEXT")

    pdf.ln(2)
    pdf.cell(120, 5, "  Gesamt")
    pdf.cell(0, 5, _fmt_eur_comma(inv.total_income_eur), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # WHT
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, "Kapitalertragsteuer einbehalten:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(120, 5, f"  Steuersatz: {inv.wht_rate_pct}%")
    pdf.cell(0, 5, _fmt_eur_comma(inv.wht_amount_eur), new_x="LMARGIN", new_y="NEXT")

    # Amendment details
    if inv.is_amended and inv.amendments:
        pdf.ln(5)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 6, "Korrekturhinweise:", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 9)
        for amend in inv.amendments:
            pdf.cell(0, 5, _latin1_safe(f"  - {amend.description}"), new_x="LMARGIN", new_y="NEXT")

    # Feststellungsbescheid reference
    pdf.ln(3)
    pdf.set_font("Helvetica", "I", 8)
    ref_num = inv.alloc_id.split("-")[1]
    pdf.cell(
        0, 4,
        f"Feststellungsbescheid Referenz: FA München {inv.fiscal_year}/KG/{ref_num}",
        new_x="LMARGIN", new_y="NEXT",
    )
    pdf.cell(0, 4, f"Erstellt am {_FIXED_DATE}", new_x="LMARGIN", new_y="NEXT")


def _render_french_sci(pdf: FPDF, inv: EUPartnershipInvestment) -> None:
    """Render a French SCI allocation statement (scanned-style, lower quality)."""
    pdf.add_page()

    # Scanned style: slightly different font, lower quality feel
    pdf.set_font("Courier", "B", 12)
    pdf.cell(0, 8, "ATTESTATION DE QUOTE-PART DES RÉSULTATS", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_font("Courier", "", 9)
    pdf.cell(0, 5, f"Exercice fiscal {inv.fiscal_year}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(5)

    # SCI info
    pdf.set_font("Courier", "B", 9)
    pdf.cell(0, 5, "Société:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Courier", "", 9)
    pdf.cell(0, 5, f"  {inv.fund_name}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, f"  Forme juridique: {inv.legal_form}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # Partner
    pdf.set_font("Courier", "B", 9)
    pdf.cell(0, 5, "Associé:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Courier", "", 9)
    pdf.cell(0, 5, "  Cascade Europe Holdings B.V.", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, f"  Quote-part: {inv.partner_share_pct}%", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # Results
    pdf.set_font("Courier", "B", 9)
    pdf.cell(0, 5, "Répartition des résultats:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Courier", "", 9)

    for cat in inv.categories:
        pdf.cell(110, 5, f"  {cat.local_label}")
        pdf.cell(0, 5, _fmt_eur_comma(cat.amount_eur), new_x="LMARGIN", new_y="NEXT")

    pdf.ln(2)
    pdf.cell(110, 5, "  Total")
    pdf.cell(0, 5, _fmt_eur_comma(inv.total_income_eur), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # WHT
    pdf.set_font("Courier", "B", 9)
    pdf.cell(0, 5, "Prélèvement à la source:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Courier", "", 9)
    pdf.cell(110, 5, f"  Taux applicable: {inv.wht_rate_pct}%")
    pdf.cell(0, 5, _fmt_eur_comma(inv.wht_amount_eur), new_x="LMARGIN", new_y="NEXT")

    # Crédit d'impôt note
    pdf.ln(3)
    pdf.set_font("Courier", "I", 8)
    pdf.cell(0, 4, "Crédit d'impôt: voir convention fiscale franco-néerlandaise", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 4, f"Document émis le {_FIXED_DATE}", new_x="LMARGIN", new_y="NEXT")


def _render_lux_scsp(pdf: FPDF, inv: EUPartnershipInvestment) -> None:
    """Render a Luxembourg SCSp allocation statement (clean, English/French mix)."""
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, "PARTNER ALLOCATION STATEMENT", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Fiscal Year {inv.fiscal_year}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(5)

    # Fund info
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, "Fund:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, f"  {inv.fund_name}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, f"  Legal form: {inv.legal_form}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, "  Jurisdiction: Grand Duchy of Luxembourg", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # Partner info
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, "Limited Partner:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, "  Cascade Europe Holdings B.V.", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, f"  Partnership interest: {inv.partner_share_pct}%", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # Allocation
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, "Allocated Income:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)

    for cat in inv.categories:
        pdf.cell(120, 5, f"  {cat.local_label}")
        pdf.cell(0, 5, _fmt_eur_comma(cat.amount_eur), new_x="LMARGIN", new_y="NEXT")

    pdf.ln(2)
    pdf.cell(120, 5, "  Total Allocated Income")
    pdf.cell(0, 5, _fmt_eur_comma(inv.total_income_eur), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # WHT (usually 0 for LU)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, "Withholding Tax:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    if inv.wht_amount_eur == Decimal(0):
        pdf.cell(0, 5, "  No withholding tax applicable (tax transparent vehicle)", new_x="LMARGIN", new_y="NEXT")
    else:
        pdf.cell(120, 5, f"  Rate: {inv.wht_rate_pct}%")
        pdf.cell(0, 5, _fmt_eur_comma(inv.wht_amount_eur), new_x="LMARGIN", new_y="NEXT")

    # Footer
    pdf.ln(5)
    pdf.set_font("Helvetica", "I", 8)
    pdf.cell(0, 4, f"Prepared {_FIXED_DATE} - {inv.fund_name}, Luxembourg", new_x="LMARGIN", new_y="NEXT")


def _render_dutch_cv(pdf: FPDF, inv: EUPartnershipInvestment) -> None:
    """Render a Dutch CV allocation statement (clean text, Dutch)."""
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, "WINSTVERDELINGSOPGAVE", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Boekjaar {inv.fiscal_year}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(5)

    # CV info
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, "Vennootschap:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, f"  {inv.fund_name}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, f"  Rechtsvorm: {inv.legal_form}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # Partner
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, "Commanditair vennoot:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, "  Cascade Europe Holdings B.V.", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, f"  Deelnemingspercentage: {inv.partner_share_pct}%", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # Winstverdeling
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, "Winstverdeling:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)

    for cat in inv.categories:
        pdf.cell(120, 5, f"  {cat.local_label}")
        pdf.cell(0, 5, _fmt_eur_comma(cat.amount_eur), new_x="LMARGIN", new_y="NEXT")

    pdf.ln(2)
    pdf.cell(120, 5, "  Totaal")
    pdf.cell(0, 5, _fmt_eur_comma(inv.total_income_eur), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # WHT
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, "Dividendbelasting ingehouden:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    if inv.wht_amount_eur == Decimal(0):
        pdf.cell(0, 5, "  Niet van toepassing (binnenlandse CV)", new_x="LMARGIN", new_y="NEXT")
    else:
        pdf.cell(120, 5, f"  Tarief: {inv.wht_rate_pct}%")
        pdf.cell(0, 5, _fmt_eur_comma(inv.wht_amount_eur), new_x="LMARGIN", new_y="NEXT")

    # Kapitaalrekening note
    pdf.ln(3)
    pdf.set_font("Helvetica", "I", 8)
    pdf.cell(0, 4, "Kapitaalrekening: zie bijlage", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 4, f"Opgesteld op {_FIXED_DATE}", new_x="LMARGIN", new_y="NEXT")


def _render_uk_llp(pdf: FPDF, inv: EUPartnershipInvestment) -> None:
    """Render a UK LLP allocation statement (scanned-style, English, GBP)."""
    pdf.add_page()

    # Scanned style
    pdf.set_font("Courier", "B", 12)
    pdf.cell(0, 8, "PARTNER PROFIT ALLOCATION", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_font("Courier", "", 9)
    pdf.cell(0, 5, f"Financial Year Ended 31 March {inv.fiscal_year}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(5)

    # LLP info
    pdf.set_font("Courier", "B", 9)
    pdf.cell(0, 5, "Limited Liability Partnership:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Courier", "", 9)
    pdf.cell(0, 5, f"  {inv.fund_name}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, f"  Legal form: {inv.legal_form}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, "  Registered: England and Wales", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # Partner
    pdf.set_font("Courier", "B", 9)
    pdf.cell(0, 5, "Member:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Courier", "", 9)
    pdf.cell(0, 5, "  Cascade Europe Holdings B.V.", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, f"  Profit share: {inv.partner_share_pct}%", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # Allocation in GBP
    pdf.set_font("Courier", "B", 9)
    pdf.cell(0, 5, "Profit Allocation:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Courier", "", 9)

    for cat in inv.categories:
        local_amount = cat.amount_local if cat.amount_local is not None else cat.amount_eur
        pdf.cell(110, 5, f"  {cat.local_label}")
        pdf.cell(0, 5, _fmt_gbp(local_amount), new_x="LMARGIN", new_y="NEXT")

    # Total in GBP
    total_gbp = sum(
        c.amount_local if c.amount_local is not None else c.amount_eur
        for c in inv.categories
    )
    pdf.ln(2)
    pdf.cell(110, 5, "  Total Allocation")
    pdf.cell(0, 5, _fmt_gbp(total_gbp), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # WHT (NRL scheme)
    pdf.set_font("Courier", "B", 9)
    pdf.cell(0, 5, "UK Income Tax Deducted at Source:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Courier", "", 9)
    wht_gbp = inv.wht_amount_local if inv.wht_amount_local is not None else inv.wht_amount_eur
    pdf.cell(110, 5, f"  Non-Resident Landlord Scheme ({inv.wht_rate_pct}%)")
    pdf.cell(0, 5, _fmt_gbp(wht_gbp), new_x="LMARGIN", new_y="NEXT")

    # Footer
    pdf.ln(5)
    pdf.set_font("Courier", "I", 8)
    pdf.cell(0, 4, "Note: All amounts in GBP. Non-resident partners should", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 4, "consult their tax adviser regarding treaty relief claims.", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 4, f"Prepared {_FIXED_DATE}", new_x="LMARGIN", new_y="NEXT")


# ── Allocation statement PDF writer ────────────────────────────────────────

def _render_allocation_pdf(pdf: FPDF, inv: EUPartnershipInvestment) -> None:
    """Dispatch to the appropriate jurisdiction renderer."""
    if inv.jurisdiction == "DE":
        _render_german_kg(pdf, inv)
    elif inv.jurisdiction == "FR":
        _render_french_sci(pdf, inv)
    elif inv.jurisdiction == "LU":
        _render_lux_scsp(pdf, inv)
    elif inv.jurisdiction == "NL":
        _render_dutch_cv(pdf, inv)
    elif inv.jurisdiction == "UK":
        _render_uk_llp(pdf, inv)


def _write_allocation_pdfs(
    investments: list[EUPartnershipInvestment],
    output_dir: Path,
    canaries: CanaryRegistry,
    manifest: Manifest,
) -> None:
    """Write 8 allocation statement PDFs."""
    alloc_path = output_dir / _ALLOC_DIR
    alloc_path.mkdir(parents=True, exist_ok=True)

    for inv in investments:
        key = alloc_canary_key(inv.alloc_id)
        canary_code = canaries.canary_for(key)
        pdf = _new_pdf(canary_code)

        _render_allocation_pdf(pdf, inv)

        filename = inv.alloc_filename
        path = alloc_path / filename
        pdf.output(str(path))

        canaries.set_location(
            key,
            f"{_ALLOC_DIR}/{filename}",
            "PDF metadata \u2192 Subject",
        )
        manifest.register(
            f"{_ALLOC_DIR}/{filename}",
            "pdf",
            canary=canary_code,
            test_cases=[_TC],
        )


# ── Investment register XLSX ───────────────────────────────────────────────

_REGISTER_HEADERS = [
    "Investment ID",
    "Fund Name",
    "Jurisdiction",
    "Legal Form",
    "Partner Share %",
    "Acquisition Date",
    "Cost Basis (EUR)",
    "Current Carrying Value (EUR)",
    "Classification",
    "Accounting Method",
    "Notes",
]

# Styles
_HEADER_FONT = Font(bold=True, size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _write_investment_register(
    investments: list[EUPartnershipInvestment],
    output_dir: Path,
    canaries: CanaryRegistry,
    manifest: Manifest,
) -> None:
    """Write tc07eu_investment_register.xlsx."""
    wb = openpyxl.Workbook()
    _pin_xlsx_dates(wb)
    ws = wb.active
    ws.title = "Investment Register"

    # Headers
    for col_idx, header in enumerate(_REGISTER_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT_WHITE
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows — use register share (5.0% for Capital Croissance, not 4.8%)
    for row_idx, inv in enumerate(investments, start=2):
        # ERR-EU-005: register shows 5.0% for Capital Croissance
        share_pct = inv.partner_share_pct
        if inv.alloc_id == "ALLOC-004":
            share_pct = Decimal("5.0")  # Register value differs from statement

        values = [
            inv.alloc_id,
            inv.fund_name,
            inv.jurisdiction,
            inv.legal_form,
            float(share_pct),
            inv.acquisition_date,
            _fmt_int(inv.cost_basis_eur),
            _fmt_int(inv.carrying_value_eur),
            inv.classification,
            inv.accounting_method,
            inv.classification_note,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _THIN_BORDER

    # Column widths
    col_widths = [14, 35, 14, 30, 14, 16, 18, 22, 22, 22, 50]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Embed canary and save
    canary_code = canaries.canary_for(INVESTMENT_REGISTER_KEY)
    embed_canary_xlsx(wb, canary_code)

    file_rel = f"{_INPUT_DIR}/tc07eu_investment_register.xlsx"
    path = output_dir / file_rel
    path.parent.mkdir(parents=True, exist_ok=True)
    _save_xlsx_deterministic(wb, path)

    canaries.set_location(
        INVESTMENT_REGISTER_KEY,
        file_rel,
        "Excel custom property \u2192 _canary",
    )
    manifest.register(file_rel, "xlsx", canary=canary_code, test_cases=[_TC])


# ── WHT summary XLSX ──────────────────────────────────────────────────────

_WHT_HEADERS = [
    "Jurisdiction",
    "Income Type",
    "Domestic WHT Rate (%)",
    "Treaty Rate (NL) (%)",
    "Tax Treaty Reference",
]


def _write_wht_summary(
    output_dir: Path,
    canaries: CanaryRegistry,
    manifest: Manifest,
) -> None:
    """Write tc07eu_withholding_tax_summary.xlsx."""
    wb = openpyxl.Workbook()
    _pin_xlsx_dates(wb)
    ws = wb.active
    ws.title = "WHT Summary"

    # Headers
    for col_idx, header in enumerate(_WHT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT_WHITE
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data
    for row_idx, rate in enumerate(WHT_RATES, start=2):
        values = [
            rate.jurisdiction,
            rate.income_type,
            float(rate.domestic_rate_pct),
            float(rate.treaty_rate_pct),
            rate.treaty_reference,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _THIN_BORDER

    # Column widths
    col_widths = [14, 30, 20, 20, 45]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Embed canary and save
    canary_code = canaries.canary_for(WHT_SUMMARY_KEY)
    embed_canary_xlsx(wb, canary_code)

    file_rel = f"{_INPUT_DIR}/tc07eu_withholding_tax_summary.xlsx"
    path = output_dir / file_rel
    path.parent.mkdir(parents=True, exist_ok=True)
    _save_xlsx_deterministic(wb, path)

    canaries.set_location(
        WHT_SUMMARY_KEY,
        file_rel,
        "Excel custom property \u2192 _canary",
    )
    manifest.register(file_rel, "xlsx", canary=canary_code, test_cases=[_TC])


# ── Prompt & expected behavior ─────────────────────────────────────────────

def _write_prompt(output_dir: Path) -> None:
    """Write the TC-07-EU prompt.md."""
    text = """\
# TC-07-EU: European Partnership Investment Allocation Extraction & Consolidation

## Input Files
- `allocation_statements/` — 8 allocation statement PDFs from European partnerships
  (Germany, France, Luxembourg, Netherlands, United Kingdom)
- `tc07eu_investment_register.xlsx` — CE's investment register showing 8 investments
- `tc07eu_withholding_tax_summary.xlsx` — WHT rates by jurisdiction

## Instructions

Cascade Europe Holdings B.V. holds minority interests in 8 European
partnership and fund vehicles.

1. Extract all income, loss, and withholding tax items from each
   allocation statement in the allocation_statements/ folder.
   - Map each item to a standardised category: rental/property income,
     interest income, dividend income, capital gains, trading/business
     income, management fees, carried interest, other income/losses.
   - Convert GBP amounts to EUR at the average FY2025 rate (1.17).
2. For the amended allocation statement, identify what changed from
   the original and use the amended figures.
3. Cross-reference each allocation statement against the investment
   register to verify partner share percentages match.
4. Consolidate all partnership allocations into a single summary schedule
   grouped by:
   - Investment (rows)
   - Income category (columns)
   - With subtotals by jurisdiction and a grand total
5. Prepare a withholding tax credit schedule:
   - List all withholding taxes by jurisdiction and income type.
   - Determine whether each withholding tax qualifies for a foreign
     tax credit under Dutch tax law (voorkoming dubbele belasting)
     or the applicable tax treaty.
   - Flag any items where the treaty rate differs from the amount
     actually withheld.
6. Identify items requiring special handling:
   - Participation exemption (deelnemingsvrijstelling): flag any
     investment where the Dutch participation exemption may apply
     (>=5% interest in qualifying entity), making the income exempt
     from Dutch CIT but also making the WHT non-creditable.
   - IAS 28 vs IFRS 9 classification: verify that equity-method
     investments have consistent treatment.
   - Anti-abuse considerations: flag if any partnership income could
     trigger Dutch CFC or anti-avoidance rules.

Export the consolidated schedule as an Excel workbook with sheets for:
- Allocation Detail (one row per fund per category)
- Consolidated Summary (by jurisdiction)
- Withholding Tax Credit Schedule
- Special Handling Flags
"""
    path = output_dir / f"test_cases/{_TC}/prompt.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text)


def _write_expected_behavior(output_dir: Path) -> None:
    """Write the TC-07-EU expected_behavior.md."""
    text = """\
# TC-07-EU: Expected Behavior

## Key Evaluation Criteria

### Multilingual Extraction
- Agent must correctly extract amounts and categories from German, French,
  Dutch, and English allocation statements.
- Misinterpreting "Gewinnanteil" (profit share), "Revenus fonciers" (property
  income), "Winstaandeel" (profit share), or "Mieteinnahmen" (rental income)
  is a failure.

### Amended Statement Handling
- Agent must identify Beteiligungen München KG as a corrected/amended statement
  (marked "KORRIGIERT / AMENDED").
- The amendment changed Gewinnanteil (profit share) from EUR 340,000 to EUR 285,000.
- A new line was added: Zinserträge (interest income) EUR 55,000.
- Agent must use the amended figures, not the originals.

### EUR/GBP Translation
- Thames Valley Property LLP amounts are in GBP and must be converted to EUR
  at the 1.17 average FY2025 rate.
- GBP 192,500 total allocation -> approximately EUR 225,225.
- GBP 38,500 WHT -> approximately EUR 45,045.

### Participation Exemption Identification
- Agent must identify which investments potentially qualify for the Dutch
  participation exemption (deelnemingsvrijstelling) at >=5% share:
  - Rheinland KG (8.0%) — qualifies, active business
  - Südbayern GmbH & Co. KG (6.0%) — borderline, passive real estate may fail
    the asset test
  - Benelux Ventures CV (12.0%) — qualifies if IAS 28 classification stands
  - Beteiligungen München KG (5.5%) — qualifies, active holding
- If income is exempt under participation exemption, associated foreign WHT
  is NOT creditable. Agent must NOT claim WHT credits on exempt income.

### WHT Treaty Rate Mismatch (ERR-EU-004)
- Thames Valley LLP withheld at 20% (UK NRL scheme) vs 15% treaty rate.
- Agent must flag the discrepancy and note the excess 5% is reclaimable.

### Partner Share Discrepancy (ERR-EU-005)
- Capital Croissance SLP allocation statement shows 4.8% partner share.
- Investment register shows 5.0%.
- Agent must flag the 0.2% discrepancy and recommend verification.
- This also creates a participation exemption boundary condition:
  at 5.0% it potentially qualifies, at 4.8% it does not.

### IAS 28 Classification Judgment
- Benelux Ventures CV is classified as IAS 28 associate at 12% share (below
  the 20% presumption threshold).
- Justification: "significant influence via board seat."
- Agent should note this justification rather than silently reclassifying.

### Consolidation Accuracy
- All 8 investments must be consolidated into standard categories.
- Totals must be arithmetically correct across all investments and categories.

### Anti-Abuse Awareness
- Agent should note whether any partnership structure could trigger Dutch CFC
  rules (ATAD Directive anti-avoidance), even if concluding it doesn't apply.
"""
    path = output_dir / f"test_cases/{_TC}/expected_behavior.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text)


# ── Planted errors ─────────────────────────────────────────────────────────

def _register_errors(
    investments: list[EUPartnershipInvestment],
    errors: ErrorRegistry,
) -> None:
    """Register ERR-EU-004 and ERR-EU-005."""
    # ERR-EU-004: WHT rate mismatch — Thames Valley at 20% vs 15% treaty
    thames = next(i for i in investments if i.alloc_id == "ALLOC-006")
    wht_gbp = thames.wht_amount_local
    errors.add(PlantedError(
        error_id="ERR-EU-004",
        file=f"{_ALLOC_DIR}/{thames.alloc_filename}",
        location="UK Income Tax Deducted at Source section",
        type="mismatched_total",
        description=(
            f"Thames Valley LLP withheld £{_fmt_int(wht_gbp):,} at 20% "
            f"(UK NRL scheme) on £{_fmt_int(thames.categories[0].amount_local):,} "
            "property income. NL-UK treaty rate is 15%. "
            "Excess 5% (£9,625) is reclaimable under the treaty."
        ),
        severity="material",
        which_test_cases_should_catch=[_TC],
    ))

    # ERR-EU-005: Partner share mismatch — Capital Croissance 4.8% vs 5.0%
    capital = next(i for i in investments if i.alloc_id == "ALLOC-004")
    errors.add(PlantedError(
        error_id="ERR-EU-005",
        file=f"{_INPUT_DIR}/tc07eu_investment_register.xlsx",
        location="Investment Register — Partner Share % column, ALLOC-004 row",
        type="mismatched_total",
        description=(
            f"Capital Croissance SLP allocation statement shows {capital.partner_share_pct}% "
            "partner share, but investment register shows 5.0%. "
            "Discrepancy creates a participation exemption boundary condition: "
            "at 5.0% potentially qualifies, at 4.8% does not."
        ),
        severity="material",
        which_test_cases_should_catch=[_TC],
    ))


# ── Gold Standard ──────────────────────────────────────────────────────────

# Dutch CIT category mapping for consolidated schedule
_DUTCH_CIT_MAPPING: dict[str, str] = {
    "Rental/property income": "Dutch CIT — taxable or exempt via deelnemingsvrijstelling",
    "Interest income": "Dutch CIT — taxable (voorkoming dubbele belasting for foreign WHT)",
    "Dividend income": "Dutch CIT — taxable or exempt via deelnemingsvrijstelling",
    "Capital gains": "Dutch CIT — taxable or exempt via deelnemingsvrijstelling",
    "Trading/business income": "Dutch CIT — taxable or exempt via deelnemingsvrijstelling",
    "Infrastructure income": "Dutch CIT — taxable",
    "Carried interest": "Dutch CIT — taxable",
    "Profit share": "Dutch CIT — taxable or exempt via deelnemingsvrijstelling",
}


@register_gold("TC-07-EU")
def _tc07_eu_gold(
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    **model_kwargs: Any,
) -> GoldStandard:
    """Build the TC-07-EU gold standard from the canonical model."""
    investments = generate_eu_investments()
    cat_totals = consolidated_totals_eu(investments)
    jur_totals = consolidated_by_jurisdiction(investments)
    grand_wht = total_wht(investments)

    # Per-investment detail
    alloc_details: dict[str, dict[str, Any]] = {}
    for inv in investments:
        detail: dict[str, Any] = {
            "fund_name": inv.fund_name,
            "jurisdiction": inv.jurisdiction,
            "legal_form": inv.legal_form,
            "partner_share_pct": float(inv.partner_share_pct),
            "language": inv.language.value,
            "pdf_style": inv.pdf_style.value,
            "is_amended": inv.is_amended,
            "currency": inv.currency,
        }

        # Categories
        categories: dict[str, dict[str, Any]] = {}
        for cat in inv.categories:
            cat_entry: dict[str, Any] = {
                "local_label": cat.local_label,
                "amount_eur": _fmt_int(cat.amount_eur),
            }
            if cat.amount_local is not None:
                cat_entry["amount_local"] = _fmt_int(cat.amount_local)
                cat_entry["local_currency"] = inv.currency
            categories[cat.english_label] = cat_entry
        detail["categories"] = categories
        detail["total_income_eur"] = _fmt_int(inv.total_income_eur)

        # WHT
        detail["wht"] = {
            "amount_eur": _fmt_int(inv.wht_amount_eur),
            "rate_pct": float(inv.wht_rate_pct),
        }
        if inv.wht_amount_local is not None:
            detail["wht"]["amount_local"] = _fmt_int(inv.wht_amount_local)
            detail["wht"]["local_currency"] = inv.currency

        # Classification
        detail["classification"] = inv.classification
        detail["accounting_method"] = inv.accounting_method
        detail["classification_note"] = inv.classification_note

        # Participation exemption
        detail["participation_exempt"] = inv.participation_exempt
        detail["participation_exempt_note"] = inv.participation_exempt_note

        # Amendments
        if inv.is_amended:
            detail["amendments"] = [
                {
                    "field": a.field_changed,
                    "original_label": a.original_label,
                    "original_value": _fmt_int(a.original_value),
                    "amended_label": a.amended_label,
                    "amended_value": _fmt_int(a.amended_value),
                    "description": a.description,
                }
                for a in inv.amendments
            ]

        alloc_details[inv.alloc_id] = detail

    # Consolidated totals as ints
    cat_totals_int = {k: _fmt_int(v) for k, v in sorted(cat_totals.items())}
    jur_totals_int = {k: _fmt_int(v) for k, v in sorted(jur_totals.items())}

    # Grand total income
    grand_income = sum(inv.total_income_eur for inv in investments)

    # Canary verification
    canary_verification: dict[str, str] = {}
    for inv in investments:
        key = alloc_canary_key(inv.alloc_id)
        canary_verification[f"read_{inv.alloc_id}"] = canaries.canary_for(key)
    canary_verification["read_investment_register"] = canaries.canary_for(INVESTMENT_REGISTER_KEY)
    canary_verification["read_wht_summary"] = canaries.canary_for(WHT_SUMMARY_KEY)

    return GoldStandard(
        test_case=_TC,
        expected_outputs={
            "file_type": "xlsx",
            "description": (
                "Consolidated European partnership allocation schedule with "
                "8 investments across 5 jurisdictions"
            ),
            "allocation_count": 8,
            "allocation_details": alloc_details,
            "consolidated_by_category": cat_totals_int,
            "consolidated_by_jurisdiction": jur_totals_int,
            "grand_total_income_eur": _fmt_int(grand_income),
            "grand_total_wht_eur": _fmt_int(grand_wht),
            "gbp_eur_rate": float(GBP_EUR_RATE),
            "dutch_cit_mapping": _DUTCH_CIT_MAPPING,
            "amended_statement": {
                "alloc_id": "ALLOC-008",
                "fund": "Beteiligungen München KG",
                "changes": [
                    {
                        "field": "Gewinnanteil (profit share)",
                        "original": 340000,
                        "amended": 285000,
                    },
                    {
                        "field": "Zinserträge (interest income)",
                        "original": 0,
                        "amended": 55000,
                    },
                ],
            },
            "participation_exemption_flags": [
                {
                    "alloc_id": "ALLOC-001",
                    "fund": "Rheinland Industriepark KG",
                    "share_pct": 8.0,
                    "qualifies": True,
                    "note": "Active business — qualifies",
                },
                {
                    "alloc_id": "ALLOC-002",
                    "fund": "Südbayern Gewerbe GmbH & Co. KG",
                    "share_pct": 6.0,
                    "qualifies": "borderline",
                    "note": "Passive real estate may fail asset test",
                },
                {
                    "alloc_id": "ALLOC-005",
                    "fund": "Benelux Ventures CV",
                    "share_pct": 12.0,
                    "qualifies": True,
                    "note": "Qualifies if IAS 28 classification stands (board seat justification)",
                },
                {
                    "alloc_id": "ALLOC-008",
                    "fund": "Beteiligungen München KG",
                    "share_pct": 5.5,
                    "qualifies": True,
                    "note": "Active industrial holding — qualifies",
                },
            ],
            "wht_treaty_rate_mismatch": {
                "alloc_id": "ALLOC-006",
                "fund": "Thames Valley Property LLP",
                "actual_rate_pct": 20.0,
                "treaty_rate_pct": 15.0,
                "excess_reclaimable_gbp": 9625,
                "note": "UK NRL scheme 20% vs NL-UK treaty 15%; excess reclaimable",
            },
            "partner_share_discrepancy": {
                "alloc_id": "ALLOC-004",
                "fund": "Capital Croissance SLP",
                "statement_pct": 4.8,
                "register_pct": 5.0,
                "note": "Boundary for participation exemption (5% threshold)",
            },
            "ias28_classification_judgment": {
                "alloc_id": "ALLOC-005",
                "fund": "Benelux Ventures CV",
                "share_pct": 12.0,
                "classification": "IAS 28 associate",
                "justification": "Significant influence via board seat despite <20% share",
            },
            "special_handling": [
                "Participation exemption — income may be exempt but WHT then non-creditable",
                "WHT treaty rate mismatch — Thames Valley 20% vs 15% treaty",
                "Partner share discrepancy — Capital Croissance 4.8% vs 5.0%",
                "IAS 28 classification judgment — Benelux Ventures at 12% with board seat",
                "GBP->EUR conversion required for Thames Valley",
                "Anti-abuse/CFC awareness — ATAD Directive considerations",
            ],
        },
        canary_verification=canary_verification,
        error_detection={
            "ERR-EU-004": (
                "Thames Valley LLP withheld at 20% (UK NRL scheme) vs 15% "
                "NL-UK treaty rate. Excess 5% is reclaimable."
            ),
            "ERR-EU-005": (
                "Capital Croissance SLP statement shows 4.8% partner share "
                "but investment register shows 5.0%. Creates participation "
                "exemption boundary condition."
            ),
        },
        scoring_hints={
            "correctness": (
                "All 8 allocation statements extracted correctly; consolidated "
                "totals match gold standard; amended values used (not originals); "
                "GBP amounts converted at 1.17"
            ),
            "completeness": (
                "Multilingual extraction successful; WHT credit schedule with "
                "treaty rate analysis; participation exemption flags; IAS 28 "
                "classification noted; anti-abuse awareness"
            ),
            "format_compliance": (
                "Valid xlsx with Allocation Detail, Consolidated Summary, "
                "WHT Credit Schedule, and Special Handling Flags sheets"
            ),
            "communication": (
                "Identified amended statement and described changes; flagged "
                "participation exemption interactions with WHT credits; noted "
                "treaty rate mismatch; flagged partner share discrepancy"
            ),
        },
    )


# ── Public entry point ─────────────────────────────────────────────────────

def emit_tc07_eu(
    model: CascadeModel,
    output_dir: Path,
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    manifest: Manifest,
    **kwargs: object,
) -> None:
    """Write all TC-07-EU files to *output_dir*."""
    investments = generate_eu_investments()
    _write_allocation_pdfs(investments, output_dir, canaries, manifest)
    _write_investment_register(investments, output_dir, canaries, manifest)
    _write_wht_summary(output_dir, canaries, manifest)
    _register_errors(investments, errors)
    _write_prompt(output_dir)
    _write_expected_behavior(output_dir)
