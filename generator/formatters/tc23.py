"""Formatter: TC-23 — Swiss Multi-Currency Bank Reconciliation (Audit, Complex).

Emits:
- test_cases/TC-23/input_files/cpi_bank_statement_chf_dec2025.csv  (185 txns)
- test_cases/TC-23/input_files/cpi_bank_statement_eur_dec2025.csv  (62 txns)
- test_cases/TC-23/input_files/cpi_bank_statement_usd_dec2025.csv  (28 txns)
- test_cases/TC-23/input_files/cpi_gl_cash_dec2025.xlsx            (3 sheets)
- test_cases/TC-23/input_files/cpi_bank_confirmations_fy2025.pdf   (UBS letter)
- test_cases/TC-23/input_files/snb_fx_rates_dec2025.csv            (daily rates)
- test_cases/TC-23/prompt.md
- test_cases/TC-23/expected_behavior.md
- gold_standards/TC-23_gold.json

Uses the canonical BankModelCH — never hardcodes numbers.
"""

from __future__ import annotations

import datetime
import io
import random
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.writer.excel import ExcelWriter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from generator.canaries import (
    CanaryRegistry,
    embed_canary_csv_comment,
    embed_canary_xlsx,
)
from generator.errors import ErrorRegistry, PlantedError
from generator.golds.framework import GoldStandard, register_gold
from generator.manifest import Manifest
from generator.model.bank_ch import (
    CHF_ADJUSTED_BANK,
    CHF_BANK_CHARGES,
    CHF_BANK_ENDING,
    CHF_BANK_INTEREST,
    CHF_DEPOSITS_IN_TRANSIT,
    CHF_DEPOSITS_IN_TRANSIT_TOTAL,
    CHF_GL_ENDING,
    CHF_IBAN,
    CHF_OUTSTANDING_CHECKS,
    CHF_OUTSTANDING_CHECKS_TOTAL,
    CONSOLIDATED_CHF,
    CREDIT_FACILITY,
    ENTITY_CITY,
    ENTITY_NAME,
    ENTITY_SHORT,
    EUR_ADJUSTED_BANK,
    EUR_ADJUSTED_BANK_CHF,
    EUR_BANK_ENDING,
    EUR_FX_ERROR,
    EUR_GL_ENDING_CHF,
    EUR_IBAN,
    EUR_OUTSTANDING_SEPA,
    EUR_OUTSTANDING_SEPA_TOTAL,
    FX_CHF_EUR_CLOSING,
    FX_CHF_EUR_STALE,
    FX_CHF_USD_CLOSING,
    USD_ADJUSTED_BANK,
    USD_ADJUSTED_BANK_CHF,
    USD_BANK_ENDING,
    USD_GL_ENDING_CHF,
    USD_IBAN,
    USD_OUTSTANDING_WIRE,
    USD_OUTSTANDING_WIRE_TOTAL,
    BankModelCH,
    _approximate_daily_rate,
    generate_bank_ch_model,
)
from generator.model.build import CascadeModel
from generator.noise import ExclusionZone, apply_csv_noise, make_noise_rng
from generator.scenario_context import ScenarioContext

# ── Constants ────────────────────────────────────────────────────────────────

_TC = "TC-23"
_INPUT_DIR = f"test_cases/{_TC}/input_files"

# Fixed datetime for xlsx metadata (determinism)
_FIXED_DATETIME = datetime.datetime(2025, 3, 15, 9, 0, 0)

# December 2025 business days for SNB rates
_DEC_START = datetime.date(2025, 12, 1)
_DEC_END = datetime.date(2025, 12, 31)

# EUR start rate for daily interpolation (same as model)
_EUR_START_SNB = FX_CHF_EUR_CLOSING * Decimal("1.004")
_USD_START_SNB = FX_CHF_USD_CLOSING * Decimal("1.003")

# FY2024 closing rates (for reference row in SNB CSV)
_FY2024_EUR = FX_CHF_EUR_STALE  # 0.9285
_FY2024_USD = Decimal("0.8902")  # Plausible FY2024 USD rate


# ── Helpers ──────────────────────────────────────────────────────────────────


def _save_xlsx_deterministic(wb: openpyxl.Workbook, path: str | Path) -> None:
    """Save workbook with pinned timestamps for byte-identical output."""
    path = Path(path)
    wb.properties.modified = _FIXED_DATETIME
    wb.properties.created = _FIXED_DATETIME

    buf = io.BytesIO()
    archive = ZipFile(buf, "w", ZIP_DEFLATED, allowZip64=True)
    writer = ExcelWriter(wb, archive)
    writer.save()

    fixed_date_time = (2025, 3, 15, 9, 0, 0)
    buf.seek(0)
    with ZipFile(buf, "r") as src, ZipFile(str(path), "w", ZIP_DEFLATED) as dst:
        for item in src.infolist():
            info = ZipInfo(item.filename, date_time=fixed_date_time)
            info.compress_type = item.compress_type
            dst.writestr(info, src.read(item.filename))


def _fmt_amount(d: Decimal) -> str:
    """Format a Decimal as a string with 2 decimal places for CSV."""
    return str(d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _fmt_date_ch(d: datetime.date) -> str:
    """Format a date as DD.MM.YYYY (Swiss/German convention)."""
    return d.strftime("%d.%m.%Y")


def _dec_business_days() -> list[datetime.date]:
    """Return all business days in December 2025."""
    days: list[datetime.date] = []
    d = _DEC_START
    while d <= _DEC_END:
        if d.weekday() < 5:
            days.append(d)
        d += datetime.timedelta(days=1)
    return days


# ── Bank Statement CSVs ─────────────────────────────────────────────────────


def _write_bank_csv_single(
    bank_ch: BankModelCH,
    currency: str,
    filename: str,
    file_key: str,
    starting_balance: Decimal,
    output_dir: Path,
    canaries: CanaryRegistry,
    manifest: Manifest,
) -> None:
    """Write a single bank statement CSV for one currency account."""
    canary_code = canaries.canary_for(file_key)
    canary_line = embed_canary_csv_comment(canary_code)

    path = output_dir / _INPUT_DIR / filename
    path.parent.mkdir(parents=True, exist_ok=True)

    # Select transactions by currency
    if currency == "CHF":
        transactions = bank_ch.chf_bank_transactions
        iban = CHF_IBAN
    elif currency == "EUR":
        transactions = bank_ch.eur_bank_transactions
        iban = EUR_IBAN
    else:
        transactions = bank_ch.usd_bank_transactions
        iban = USD_IBAN

    lines: list[str] = [canary_line]
    lines.append("# UBS Switzerland AG — Kontoauszug\n")
    lines.append(f"# Konto: {ENTITY_NAME} — IBAN {iban}\n")
    lines.append(f"# Währung: {currency}\n")
    lines.append("# Auszugszeitraum: 01.12.2025 - 31.12.2025\n")
    lines.append(
        f"# Anfangssaldo: {_fmt_amount(starting_balance)} {currency}\n"
    )
    lines.append("#\n")
    lines.append("Buchungsdatum,Buchungstext,Betrag,Saldo\n")

    for txn in transactions:
        date_str = _fmt_date_ch(txn.date)
        amt_str = _fmt_amount(txn.amount)
        bal_str = _fmt_amount(txn.running_balance)
        desc = txn.description
        if "," in desc:
            desc = f'"{desc}"'
        lines.append(f"{date_str},{desc},{amt_str},{bal_str}\n")

    # Controlled noise
    # Protect canary line (0), comment lines (1-6), header (7)
    excl = ExclusionZone(rows={0, 1, 2, 3, 4, 5, 6, 7})
    noise_rng = make_noise_rng(
        ScenarioContext(seed=42), _TC, file_key,
    )
    lines = apply_csv_noise(lines, noise_rng, excl)

    path.write_text("".join(lines), encoding="utf-8")

    canaries.set_location(
        file_key,
        f"{_INPUT_DIR}/{filename}",
        "First line comment",
    )
    manifest.register(
        f"{_INPUT_DIR}/{filename}",
        "csv",
        canary=canary_code,
        test_cases=[_TC],
    )


def _write_bank_csvs(
    bank_ch: BankModelCH,
    output_dir: Path,
    canaries: CanaryRegistry,
    manifest: Manifest,
) -> None:
    """Write all 3 bank statement CSVs."""
    _write_bank_csv_single(
        bank_ch, "CHF", "cpi_bank_statement_chf_dec2025.csv",
        "tc23_bank_chf", bank_ch.chf_bank_starting,
        output_dir, canaries, manifest,
    )
    _write_bank_csv_single(
        bank_ch, "EUR", "cpi_bank_statement_eur_dec2025.csv",
        "tc23_bank_eur", bank_ch.eur_bank_starting,
        output_dir, canaries, manifest,
    )
    _write_bank_csv_single(
        bank_ch, "USD", "cpi_bank_statement_usd_dec2025.csv",
        "tc23_bank_usd", bank_ch.usd_bank_starting,
        output_dir, canaries, manifest,
    )


# ── GL Cash xlsx (3 sheets) ─────────────────────────────────────────────────


def _write_gl_cash_xlsx(
    bank_ch: BankModelCH,
    output_dir: Path,
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    manifest: Manifest,
) -> None:
    """Write cpi_gl_cash_dec2025.xlsx with 3 sheets (CHF, EUR, USD)."""
    file_key = "tc23_gl_cash"
    canary_code = canaries.canary_for(file_key)

    wb = openpyxl.Workbook()
    loc = embed_canary_xlsx(wb, canary_code)

    # ── Sheet 1: 1020 - Bank CHF ────────────────────────────────────
    ws_chf = wb.active
    ws_chf.title = "1020 - Bank CHF"
    _write_gl_sheet(
        ws_chf,
        account_name="1020 - Bank CHF",
        currency_label="CHF",
        gl_entries=bank_ch.chf_gl_entries,
        starting_balance=bank_ch.chf_gl_starting,
        ending_balance=CHF_GL_ENDING,
    )

    # ── Sheet 2: 1021 - Bank EUR ────────────────────────────────────
    ws_eur = wb.create_sheet("1021 - Bank EUR")
    _write_gl_sheet(
        ws_eur,
        account_name="1021 - Bank EUR",
        currency_label="CHF (EUR-Konto, funktionale Währung CHF)",
        gl_entries=bank_ch.eur_gl_entries,
        starting_balance=bank_ch.eur_gl_starting_chf,
        ending_balance=EUR_GL_ENDING_CHF,
    )

    # Register ERR-CH-002
    errors.add(PlantedError(
        error_id="ERR-CH-002",
        file=f"{_INPUT_DIR}/cpi_gl_cash_dec2025.xlsx",
        location="Sheet '1021 - Bank EUR', revaluation entry dated 31.12.2025",
        type="stale_data",
        description=(
            "EUR account opening-balance FX revaluation uses stale FY2024 closing "
            "rate (CHF/EUR 0.9285, i.e. 1 EUR = 1.0770 CHF) instead of FY2025 "
            "closing rate (CHF/EUR 0.9387, i.e. 1 EUR = 1.0653 CHF); "
            f"CHF {int(EUR_FX_ERROR):,} overstatement of EUR cash balance in CHF terms"
        ),
        severity="immaterial",
        which_test_cases_should_catch=["TC-23"],
    ))

    # ── Sheet 3: 1022 - Bank USD ────────────────────────────────────
    ws_usd = wb.create_sheet("1022 - Bank USD")
    _write_gl_sheet(
        ws_usd,
        account_name="1022 - Bank USD",
        currency_label="CHF (USD-Konto, funktionale Währung CHF)",
        gl_entries=bank_ch.usd_gl_entries,
        starting_balance=bank_ch.usd_gl_starting_chf,
        ending_balance=USD_GL_ENDING_CHF,
    )

    # ── Save ────────────────────────────────────────────────────────
    path = output_dir / _INPUT_DIR / "cpi_gl_cash_dec2025.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    _save_xlsx_deterministic(wb, path)

    canaries.set_location(
        file_key,
        f"{_INPUT_DIR}/cpi_gl_cash_dec2025.xlsx",
        loc,
    )
    manifest.register(
        f"{_INPUT_DIR}/cpi_gl_cash_dec2025.xlsx",
        "xlsx",
        canary=canary_code,
        test_cases=[_TC],
    )


def _write_gl_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    account_name: str,
    currency_label: str,
    gl_entries: list,
    starting_balance: Decimal,
    ending_balance: Decimal,
) -> None:
    """Populate a single GL sheet with header, data rows, and ending balance."""
    # Header
    ws["A1"] = ENTITY_NAME
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = f"Hauptbuch — {account_name}"
    ws["A2"].font = Font(italic=True, size=11)
    ws["A3"] = f"Dezember 2025 — Beträge in {currency_label}"
    ws["A3"].font = Font(size=10)

    # Starting balance
    ws["A5"] = "Anfangssaldo"
    ws["A5"].font = Font(bold=True, size=10)
    ws["F5"] = int(starting_balance)
    ws["F5"].font = Font(bold=True, size=10)
    ws["F5"].number_format = "#,##0"

    # Column headers
    header_row = 7
    headers = ["Datum", "Beleg-Nr", "Buchungstext", "Soll", "Haben", "Saldo"]
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A3C6E")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    money_fmt = "#,##0"
    data_font = Font(size=10)
    row = header_row + 1

    for entry in gl_entries:
        ws.cell(
            row=row, column=1, value=_fmt_date_ch(entry.date)
        ).font = data_font
        ws.cell(row=row, column=2, value=entry.reference).font = data_font
        ws.cell(row=row, column=3, value=entry.description).font = data_font

        debit_cell = ws.cell(
            row=row, column=4,
            value=int(entry.debit) if entry.debit > 0 else None,
        )
        debit_cell.font = data_font
        debit_cell.number_format = money_fmt

        credit_cell = ws.cell(
            row=row, column=5,
            value=int(entry.credit) if entry.credit > 0 else None,
        )
        credit_cell.font = data_font
        credit_cell.number_format = money_fmt

        bal_cell = ws.cell(row=row, column=6, value=int(entry.running_balance))
        bal_cell.font = data_font
        bal_cell.number_format = money_fmt

        row += 1

    # Ending balance row
    row += 1
    ws.cell(row=row, column=1, value="Endsaldo").font = Font(bold=True, size=10)
    end_cell = ws.cell(row=row, column=6, value=int(ending_balance))
    end_cell.font = Font(bold=True, size=10)
    end_cell.number_format = money_fmt
    end_cell.border = Border(bottom=Side(style="double"))

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18


# ── Bank Confirmation PDF ────────────────────────────────────────────────────


def _write_bank_confirmation_pdf(
    bank_ch: BankModelCH,
    output_dir: Path,
    canaries: CanaryRegistry,
    manifest: Manifest,
) -> None:
    """Write cpi_bank_confirmations_fy2025.pdf — UBS confirmation letter in German."""
    file_key = "tc23_bank_confirm"
    canary_code = canaries.canary_for(file_key)

    path = output_dir / _INPUT_DIR / "cpi_bank_confirmations_fy2025.pdf"
    path.parent.mkdir(parents=True, exist_ok=True)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ConfTitle", parent=styles["Title"],
        fontSize=14, spaceAfter=6, textColor=colors.HexColor("#003366"),
    )
    heading_style = ParagraphStyle(
        "ConfHeading", parent=styles["Heading2"],
        fontSize=11, spaceAfter=6, textColor=colors.HexColor("#003366"),
    )
    body_style = ParagraphStyle(
        "ConfBody", parent=styles["Normal"],
        fontSize=10, spaceAfter=6, leading=14,
    )
    small_style = ParagraphStyle(
        "ConfSmall", parent=styles["Normal"],
        fontSize=8, spaceAfter=4, textColor=colors.HexColor("#666666"),
    )

    story: list = []

    # UBS letterhead
    story.append(Paragraph("UBS Switzerland AG", title_style))
    story.append(Paragraph(
        "Corporate &amp; Institutional Banking<br/>"
        "Bahnhofstrasse 45<br/>"
        "CH-8001 Z\u00fcrich",
        body_style,
    ))
    story.append(Spacer(1, 0.3 * inch))

    # Confirmation header
    story.append(Paragraph("BANKBEST\u00c4TIGUNG", heading_style))
    story.append(Paragraph(
        "(Standardbest\u00e4tigung f\u00fcr Revisionsgesellschaften)",
        body_style,
    ))
    story.append(Spacer(1, 0.2 * inch))

    # Date and addressee
    story.append(Paragraph("15. Januar 2026", body_style))
    story.append(Spacer(1, 0.15 * inch))
    story.append(Paragraph(
        "Treuhand &amp; Revision Z\u00fcrich AG<br/>"
        "z. Hd. Pr\u00fcfungsteam<br/>"
        "Seestrasse 42<br/>"
        "CH-8002 Z\u00fcrich",
        body_style,
    ))
    story.append(Spacer(1, 0.2 * inch))

    story.append(Paragraph(
        f"Betreff: {ENTITY_NAME}, {ENTITY_CITY}",
        body_style,
    ))
    story.append(Paragraph(
        "Sehr geehrte Damen und Herren,",
        body_style,
    ))
    story.append(Spacer(1, 0.1 * inch))

    # Body text
    story.append(Paragraph(
        f"Im Zusammenhang mit Ihrer Pr\u00fcfung des Jahresabschlusses der "
        f"<b>{ENTITY_NAME}</b> best\u00e4tigen wir nachstehend die "
        f"Kontoverh\u00e4ltnisse per <b>31. Dezember 2025</b> "
        f"(Gesch\u00e4ftsschluss):",
        body_style,
    ))
    story.append(Spacer(1, 0.15 * inch))

    # Account details table — all 3 accounts
    chf_bal = f"CHF {int(CHF_BANK_ENDING):,}"
    eur_bal = f"EUR {int(EUR_BANK_ENDING):,}"
    usd_bal = f"USD {int(USD_BANK_ENDING):,}"

    acct_data = [
        ["Kontobezeichnung", "IBAN", "Kontotyp", "Saldo"],
        [
            f"{ENTITY_SHORT} \u2014 Betriebskonto CHF",
            CHF_IBAN,
            "Kontokorrent CHF",
            chf_bal,
        ],
        [
            f"{ENTITY_SHORT} \u2014 Fremdw\u00e4hrungskonto EUR",
            EUR_IBAN,
            "Kontokorrent EUR",
            eur_bal,
        ],
        [
            f"{ENTITY_SHORT} \u2014 Fremdw\u00e4hrungskonto USD",
            USD_IBAN,
            "Kontokorrent USD",
            usd_bal,
        ],
    ]
    acct_table = Table(
        acct_data,
        colWidths=[2.0 * inch, 1.8 * inch, 1.3 * inch, 1.3 * inch],
    )
    acct_table.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 9),
        ("FONT", (0, 1), (-1, -1), "Helvetica", 9),
        ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#003366")),
        ("LINEBELOW", (0, -1), (-1, -1), 0.5, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E0E8F0")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(acct_table)
    story.append(Spacer(1, 0.2 * inch))

    # Credit facility
    credit_chf = f"CHF {int(CREDIT_FACILITY):,}"
    story.append(Paragraph(
        f"<b>Kreditlimiten:</b> Revolvierende Kreditlimite \u00fcber "
        f"{credit_chf}, per Stichtag nicht beansprucht.",
        body_style,
    ))
    story.append(Spacer(1, 0.1 * inch))

    # Pledges / fiduciary
    story.append(Paragraph(
        "<b>Verpf\u00e4ndungen:</b> Keine.",
        body_style,
    ))
    story.append(Paragraph(
        "<b>Treuhandanlagen:</b> Keine.",
        body_style,
    ))
    story.append(Spacer(1, 0.15 * inch))

    # Closing
    story.append(Paragraph(
        "Diese Best\u00e4tigung wird ausschliesslich f\u00fcr "
        "Revisionszwecke ausgestellt und darf nicht f\u00fcr andere "
        "Zwecke verwendet werden. Die vorstehenden Angaben sind per "
        "Stichtag korrekt.",
        body_style,
    ))
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph("Freundliche Gr\u00fcsse,", body_style))
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(
        "<b>Dr. Markus W. Brunner</b><br/>"
        "Managing Director, Corporate Banking<br/>"
        "UBS Switzerland AG",
        body_style,
    ))
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(
        f"Referenz: UBS-BEST-2025-{canary_code}",
        small_style,
    ))

    # Build PDF — A4 page size (Swiss standard)
    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
        leftMargin=inch,
        rightMargin=inch,
        title=f"Bankbest\u00e4tigung \u2014 {ENTITY_NAME}",
        author=f"CANARY: {canary_code}",
        creator="UBS Switzerland AG",
        invariant=True,
    )
    doc.build(story)

    canaries.set_location(
        file_key,
        f"{_INPUT_DIR}/cpi_bank_confirmations_fy2025.pdf",
        "PDF metadata \u2192 Author; also in Referenz footer",
    )
    manifest.register(
        f"{_INPUT_DIR}/cpi_bank_confirmations_fy2025.pdf",
        "pdf",
        canary=canary_code,
        test_cases=[_TC],
    )


# ── SNB FX Rates CSV ────────────────────────────────────────────────────────


def _write_snb_fx_rates_csv(
    output_dir: Path,
    canaries: CanaryRegistry,
    manifest: Manifest,
) -> None:
    """Write snb_fx_rates_dec2025.csv — daily SNB rates for December 2025."""
    file_key = "tc23_snb_fx_rates"
    canary_code = canaries.canary_for(file_key)
    canary_line = embed_canary_csv_comment(canary_code)

    path = output_dir / _INPUT_DIR / "snb_fx_rates_dec2025.csv"
    path.parent.mkdir(parents=True, exist_ok=True)

    lines: list[str] = [canary_line]
    lines.append("# Swiss National Bank \u2014 Foreign Exchange Rates\n")
    lines.append("# Convention: Foreign currency units per 1 CHF\n")
    lines.append("Date,Currency,Rate\n")

    # FY2024 closing rates as reference row
    lines.append(f"2024-12-31,EUR,{_FY2024_EUR}\n")
    lines.append(f"2024-12-31,USD,{_FY2024_USD}\n")

    # Daily rates for all business days of December 2025
    business_days = _dec_business_days()
    for bd in business_days:
        eur_rate = _approximate_daily_rate(
            bd.day, FX_CHF_EUR_CLOSING, start_snb=_EUR_START_SNB,
        )
        usd_rate = _approximate_daily_rate(
            bd.day, FX_CHF_USD_CLOSING, start_snb=_USD_START_SNB,
        )
        date_str = bd.strftime("%Y-%m-%d")
        lines.append(f"{date_str},EUR,{eur_rate}\n")
        lines.append(f"{date_str},USD,{usd_rate}\n")

    path.write_text("".join(lines), encoding="utf-8")

    canaries.set_location(
        file_key,
        f"{_INPUT_DIR}/snb_fx_rates_dec2025.csv",
        "First line comment",
    )
    manifest.register(
        f"{_INPUT_DIR}/snb_fx_rates_dec2025.csv",
        "csv",
        canary=canary_code,
        test_cases=[_TC],
    )


# ── Prompt & Expected Behavior ───────────────────────────────────────────────


def _write_prompt(output_dir: Path) -> None:
    """Write test_cases/TC-23/prompt.md per spec."""
    text = """\
Perform a multi-currency bank reconciliation for Cascade Precision Instruments AG
as of December 31, 2025. CPI maintains three bank accounts at UBS AG in CHF, EUR,
and USD.

1. For each currency account (CHF, EUR, USD):
   a. Match transactions between the bank statement and the general ledger.
      Use fuzzy matching on dates (allow \u00b12 business days) and amounts (exact
      match in original currency).
   b. Identify all outstanding items (checks, transfers, deposits in transit).
   c. Identify any bank charges, interest, or FX gains/losses not recorded in the GL.
   d. Prepare a standard bank reconciliation schedule showing:
      - Balance per bank statement (original currency)
      - Add/Less: outstanding items
      - Adjusted bank balance (original currency)
      - Balance per GL (in CHF, the functional currency)
      - Add/Less: book adjustments needed
      - Adjusted book balance (CHF)

2. Verify the month-end FX revaluation of the EUR and USD accounts:
   - The GL should reflect EUR and USD balances translated at the SNB closing
     rate as of 31.12.2025
   - Compare the GL revaluation entries to the correct SNB rates
   - Flag any rate discrepancies

3. Reconcile all three accounts to the Swiss bank confirmation letter
   (Bankbest\u00e4tigung):
   - Confirm that the bank statement ending balances match the confirmation
     balances in each original currency
   - Note any additional disclosures in the confirmation (credit facilities,
     pledges, fiduciary deposits)

4. Prepare a consolidated cash position summary in CHF showing:
   - CHF account adjusted balance
   - EUR account adjusted balance (translated to CHF at closing rate)
   - USD account adjusted balance (translated to CHF at closing rate)
   - Total consolidated cash position in CHF

Export as an Excel workbook with sheets:
- "Recon CHF": CHF account reconciliation
- "Recon EUR": EUR account reconciliation with FX translation
- "Recon USD": USD account reconciliation with FX translation
- "FX Revaluation Check": Verification of month-end revaluation rates
- "Confirmation Tie-Out": Bank confirmation agreement
- "Consolidated Cash": Summary cash position in CHF
"""
    path = output_dir / f"test_cases/{_TC}/prompt.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def _write_expected_behavior(output_dir: Path) -> None:
    """Write test_cases/TC-23/expected_behavior.md per spec."""
    text = """\
# TC-23: Swiss Multi-Currency Bank Reconciliation \u2014 Expected Behavior

## Data Challenges
- Three bank accounts in different currencies (CHF, EUR, USD) must be
  reconciled separately against a single GL that records everything in CHF.
- Bank statements use **German-language descriptions** and Swiss date format
  (DD.MM.YYYY) requiring parsing and fuzzy matching.
- The GL records foreign currency transactions at **intra-month spot rates**,
  requiring the agent to understand the difference between transaction-date
  rates and closing rates.
- **SNB FX rates** are provided in the Swiss convention (foreign currency
  per 1 CHF), which is the inverse of the more common EUR/CHF quotation.

## CHF Account Reconciliation
- **185 transactions** on the bank statement.
- **2 outstanding checks** totaling CHF 47,815 (in GL but not on bank).
- **1 deposit in transit** of CHF 18,430 (in GL but not on bank).
- **Bank interest** CHF 4,000 on statement, not in GL.
- **Bank charges** CHF 1,375 on statement, not in GL.
- Adjusted bank = adjusted book = CHF 2,817,905.

## EUR Account Reconciliation
- **62 transactions** on the bank statement.
- **1 outstanding SEPA** of EUR 28,750 (in GL but not on bank).
- Adjusted bank balance EUR 1,186,650.
- At closing rate (0.9387): CHF 1,264,142.
- GL shows CHF 1,270,466 (stale FY2024 rate on opening balance) \u2014 **ERR-CH-002**.

## USD Account Reconciliation
- **28 transactions** on the bank statement.
- **1 outstanding wire** of USD 15,200 (in GL but not on bank).
- Adjusted bank balance USD 474,550.
- At closing rate (0.8845): CHF 536,518.
- No FX error on USD account.

## FX Revaluation Check
- The EUR revaluation entry uses **stale rate 0.9285** (FY2024) instead of
  the correct **closing rate 0.9387** (FY2025). This is ERR-CH-002.
- Impact: CHF 6,324 overstatement of EUR cash balance.
- The USD revaluation uses the correct closing rate 0.8845.

## Bank Confirmation Tie-Out
- The UBS confirmation letter covers all 3 accounts.
- All ending balances must match between statements and confirmation.
- The confirmation notes a CHF 5M revolving credit facility (undrawn),
  no pledges, and no fiduciary deposits.

## Consolidated Cash Position
- CHF adjusted: CHF 2,817,905
- EUR adjusted (at closing rate): CHF 1,264,142
- USD adjusted (at closing rate): CHF 536,518
- **Total consolidated: CHF 4,618,565**

## Output Quality
- The workbook should contain 6 sheets covering all reconciliation aspects.
- FX rate discrepancies should be clearly flagged.
- The consolidated position should use correct closing rates throughout.
"""
    path = output_dir / f"test_cases/{_TC}/expected_behavior.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


# ── Gold Standard ────────────────────────────────────────────────────────────


@register_gold("TC-23")
def _tc23_gold(
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    **model_kwargs: Any,
) -> GoldStandard:
    """Build the TC-23 gold standard from the canonical BankModelCH."""
    # Compute derived values from model constants
    eur_closing_chf_per_eur = (Decimal("1") / FX_CHF_EUR_CLOSING).quantize(
        Decimal("0.0001"), rounding=ROUND_HALF_UP,
    )
    eur_stale_chf_per_eur = (Decimal("1") / FX_CHF_EUR_STALE).quantize(
        Decimal("0.0001"), rounding=ROUND_HALF_UP,
    )
    usd_closing_chf_per_usd = (Decimal("1") / FX_CHF_USD_CLOSING).quantize(
        Decimal("0.0001"), rounding=ROUND_HALF_UP,
    )

    return GoldStandard(
        test_case="TC-23",
        expected_outputs={
            "file_type": "xlsx",
            "required_sheets": [
                "Recon CHF", "Recon EUR", "Recon USD",
                "FX Revaluation Check", "Confirmation Tie-Out",
                "Consolidated Cash",
            ],
            "recon_chf": {
                "bank_ending_balance_chf": int(CHF_BANK_ENDING),
                "outstanding_checks_count": len(CHF_OUTSTANDING_CHECKS),
                "outstanding_checks_total_chf": int(CHF_OUTSTANDING_CHECKS_TOTAL),
                "deposits_in_transit_count": len(CHF_DEPOSITS_IN_TRANSIT),
                "deposits_in_transit_total_chf": int(CHF_DEPOSITS_IN_TRANSIT_TOTAL),
                "adjusted_bank_balance_chf": int(CHF_ADJUSTED_BANK),
                "gl_ending_balance_chf": int(CHF_GL_ENDING),
                "bank_charges_not_in_gl_chf": int(CHF_BANK_CHARGES),
                "interest_not_in_gl_chf": int(CHF_BANK_INTEREST),
                "adjusted_book_balance_chf": int(CHF_ADJUSTED_BANK),
                "reconciliation_difference": 0,
            },
            "recon_eur": {
                "bank_ending_balance_eur": int(EUR_BANK_ENDING),
                "outstanding_sepa_count": len(EUR_OUTSTANDING_SEPA),
                "outstanding_sepa_total_eur": int(EUR_OUTSTANDING_SEPA_TOTAL),
                "adjusted_bank_balance_eur": int(EUR_ADJUSTED_BANK),
                "adjusted_bank_balance_chf_at_closing": int(EUR_ADJUSTED_BANK_CHF),
                "gl_ending_balance_chf": int(EUR_GL_ENDING_CHF),
                "fx_revaluation_error_chf": int(EUR_FX_ERROR),
                "correct_closing_rate_chf_per_eur": float(eur_closing_chf_per_eur),
                "stale_rate_used_chf_per_eur": float(eur_stale_chf_per_eur),
            },
            "recon_usd": {
                "bank_ending_balance_usd": int(USD_BANK_ENDING),
                "outstanding_wire_count": len(USD_OUTSTANDING_WIRE),
                "outstanding_wire_total_usd": int(USD_OUTSTANDING_WIRE_TOTAL),
                "adjusted_bank_balance_usd": int(USD_ADJUSTED_BANK),
                "adjusted_bank_balance_chf_at_closing": int(USD_ADJUSTED_BANK_CHF),
                "gl_ending_balance_chf": int(USD_GL_ENDING_CHF),
                "reconciliation_difference": 0,
            },
            "fx_revaluation_check": {
                "eur_rate_per_gl": float(eur_stale_chf_per_eur),
                "eur_rate_per_snb": float(eur_closing_chf_per_eur),
                "eur_rate_discrepancy": True,
                "eur_impact_chf": int(EUR_FX_ERROR),
                "usd_rate_per_gl": float(usd_closing_chf_per_usd),
                "usd_rate_per_snb": float(usd_closing_chf_per_usd),
                "usd_rate_discrepancy": False,
            },
            "confirmation_tie_out": {
                "chf_confirmation_matches_statement": True,
                "eur_confirmation_matches_statement": True,
                "usd_confirmation_matches_statement": True,
                "credit_facility_noted": "CHF 5M revolving, undrawn",
                "pledges": "none",
                "fiduciary_deposits": "none",
            },
            "consolidated_cash": {
                "chf_adjusted_chf": int(CHF_ADJUSTED_BANK),
                "eur_adjusted_chf": int(EUR_ADJUSTED_BANK_CHF),
                "usd_adjusted_chf": int(USD_ADJUSTED_BANK_CHF),
                "total_consolidated_chf": int(CONSOLIDATED_CHF),
            },
        },
        canary_verification={
            "read_bank_statement_chf": canaries.canary_for("tc23_bank_chf"),
            "read_bank_statement_eur": canaries.canary_for("tc23_bank_eur"),
            "read_bank_statement_usd": canaries.canary_for("tc23_bank_usd"),
            "read_gl_cash_detail": canaries.canary_for("tc23_gl_cash"),
            "read_bank_confirmation": canaries.canary_for("tc23_bank_confirm"),
            "read_snb_fx_rates": canaries.canary_for("tc23_snb_fx_rates"),
        },
        error_detection={
            "ERR-CH-002": (
                "EUR account opening-balance FX revaluation uses stale "
                "FY2024 rate (CHF/EUR 0.9285) instead of FY2025 closing "
                f"rate (0.9387); CHF {int(EUR_FX_ERROR):,} overstatement"
            ),
        },
        scoring_hints={
            "correctness": (
                "CHF recon must balance at CHF 2,817,905; EUR and USD "
                "adjusted balances must convert correctly at SNB closing rates"
            ),
            "completeness": (
                "All outstanding items identified: 2 CHF checks, 1 EUR SEPA, "
                "1 USD wire; bank charges and interest on CHF account"
            ),
            "format_compliance": (
                "Valid xlsx with 6 sheets: Recon CHF/EUR/USD, FX Revaluation "
                "Check, Confirmation Tie-Out, Consolidated Cash"
            ),
            "robustness": (
                "Handled multi-currency FX conversion, SNB rate convention, "
                "German-language bank descriptions, and fuzzy date matching"
            ),
            "communication": (
                "ERR-CH-002 (stale EUR rate) identified and impact quantified; "
                "bank confirmation tied out for all 3 accounts"
            ),
        },
        scenario_pack="cascade_swiss",
        service_line="audit",
    )


# ── Public entry point ───────────────────────────────────────────────────────


def emit_tc23(
    model: CascadeModel,
    output_dir: Path,
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    manifest: Manifest,
    **kwargs: object,
) -> None:
    """Write all TC-23 files to *output_dir*."""
    # Generate the Swiss bank model with an isolated RNG
    bank_ch_rng = random.Random(42 + 23)
    bank_ch = generate_bank_ch_model(bank_ch_rng)

    _write_bank_csvs(bank_ch, output_dir, canaries, manifest)
    _write_gl_cash_xlsx(bank_ch, output_dir, canaries, errors, manifest)
    _write_bank_confirmation_pdf(bank_ch, output_dir, canaries, manifest)
    _write_snb_fx_rates_csv(output_dir, canaries, manifest)
    _write_prompt(output_dir)
    _write_expected_behavior(output_dir)
