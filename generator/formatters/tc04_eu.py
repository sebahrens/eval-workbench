"""Formatter: TC-04-EU — IFRS 16 Lease Extraction & Schedule Population (Audit, Adversarial).

European variant of TC-04 replacing ASC 842 with IFRS 16 single lessee model.

Emits:
- test_cases/TC-04-EU/input_files/leases/LS-EU-001.pdf through LS-EU-015.pdf
  15 lease PDFs: 10 text-native (reportlab), 3 scanned-style (fpdf2 low-res),
  2 with amendments appended as additional pages
- test_cases/TC-04-EU/input_files/lease_schedule_partial.xlsx
  Partially populated IFRS 16 schedule (8 of 15 leases filled)
- test_cases/TC-04-EU/prompt.md
- test_cases/TC-04-EU/expected_behavior.md
- gold_standards/TC-04-EU_gold.json

One planted error: ERR-EU-001 date_format_inconsistency — commencement date in
the partial schedule for LS-EU-002 uses MM/DD/YYYY while the source PDF uses
DD.MM.YYYY, resulting in a different date (05.03.2023 = March 5 → schedule
shows 05/03/2023 interpreted as May 3).

Uses the canonical European lease model — never hardcodes numbers.
"""

from __future__ import annotations

import datetime
import io
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from fpdf import FPDF
from openpyxl.writer.excel import ExcelWriter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from generator.canaries import (
    CanaryRegistry,
    embed_canary_pdf_fpdf2,
    embed_canary_xlsx,
)
from generator.errors import ErrorRegistry, PlantedError
from generator.golds.framework import GoldStandard, register_gold
from generator.manifest import Manifest
from generator.model.build import CascadeModel
from generator.model.leases_eu import (
    EscalationTypeEU,
    LeaseEU,
    generate_leases_eu,
)
from generator.scenario_context import ScenarioContext

# ── Constants ────────────────────────────────────────────────────────────────

_TC = "TC-04-EU"
_INPUT_DIR = f"test_cases/{_TC}/input_files"
_LEASES_DIR = f"{_INPUT_DIR}/leases"

# fpdf2 fixed creation date for determinism
_CREATION_DATE = datetime.datetime(2025, 3, 15, 9, 0, 0)

# A4 page size for European documents
_PAGE_SIZE = A4

# ── Lease-to-rendering-type assignment ───────────────────────────────────────
# Per design: 10 text-native, 3 scanned-style, 2 with amendments.
#
# Amendments:
#   LS-EU-001 (CE head office — termination exercised then reversed)
#   LS-EU-003 (CP manufacturing — rent increase + term extension)
#   LS-EU-010 (CM R&D lab — scope change, additional floor)
# LS-EU-001 and LS-EU-003 rendered with amendment pages.
# LS-EU-010's amendment is described in body (no separate page).
#
# Scanned-style:
#   LS-EU-002 (CE parking — old Amsterdam lease, scanned)
#   LS-EU-004 (CP warehouse — German-language, scanned)
#   LS-EU-014 (CM storage — French-language, scanned)

_SCANNED_IDS = {"LS-EU-002", "LS-EU-004", "LS-EU-014"}
_AMENDMENT_PDF_IDS = {"LS-EU-001", "LS-EU-003"}

# ── Which 8 leases go in the partial schedule ─────────────────────────────
_PARTIAL_LEASE_IDS = {f"LS-EU-{i:03d}" for i in range(1, 9)}
_BLANK_FIELDS_IDS = {"LS-EU-004", "LS-EU-006", "LS-EU-008"}


# ── EUR formatting helpers ──────────────────────────────────────────────────

def _fmt_eur(d: Decimal | int) -> str:
    """Format amount in European style: 12.500,00 EUR."""
    if isinstance(d, int):
        val = d
    else:
        val = int(d.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    if val < 0:
        # European negative: (12.500 EUR) or -12.500 EUR
        return f"({abs(val):,} EUR)".replace(",", ".")
    return f"{val:,} EUR".replace(",", ".")


def _fmt_eur_short(d: Decimal | int) -> str:
    """Short EUR format for PDF body text: EUR 12.500."""
    if isinstance(d, int):
        val = d
    else:
        val = int(d.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    return f"EUR {val:,}".replace(",", ".")


def _fmt_date_eu(d: datetime.date) -> str:
    """European date format: DD.MM.YYYY."""
    return d.strftime("%d.%m.%Y")


def _fmt_date_eu_long(d: datetime.date) -> str:
    """Long European date: 5 March 2023."""
    return f"{d.day} {d.strftime('%B')} {d.year}"


def _escalation_text_eu(lease: LeaseEU) -> str:
    if lease.escalation_type == EscalationTypeEU.FIXED_PCT:
        pct = float(lease.escalation_pct) * 100
        return f"Fixed annual escalation of {pct:.1f}%"
    elif lease.escalation_type == EscalationTypeEU.HICP:
        return "Annual adjustment based on Eurozone Harmonised Index of Consumer Prices (HICP)"
    elif lease.escalation_type == EscalationTypeEU.STEPPED:
        if lease.escalation_steps:
            steps = ", ".join(_fmt_eur_short(s) for s in lease.escalation_steps)
            return f"Stepped rent schedule: {steps} per month (years 1, 2, 3)"
        return "Stepped rent schedule per attached schedule"
    return "None"


def _renewal_text_eu(lease: LeaseEU) -> str:
    if lease.renewal_option_months <= 0:
        return "None"
    mos = lease.renewal_option_months
    increase = float(lease.renewal_rent_increase_pct) * 100
    return (
        f"Option to renew for {mos} months at {increase:.0f}% above "
        f"the then-current base rent, exercisable with 90-day prior written notice"
    )


def _purchase_text_eu(lease: LeaseEU) -> str:
    if not lease.purchase_option:
        return "None"
    price = _fmt_eur_short(lease.purchase_option_price) if lease.purchase_option_price else "fair market value"
    return f"Purchase option at {price} at end of lease term"


def _canary_key(lease_id: str) -> str:
    num = lease_id.split("-")[2]
    return f"tc04eu_lease_{num}"


_SCHEDULE_KEY = "tc04eu_lease_schedule_partial"


# ── Text-native PDF via reportlab (A4, European style) ──────────────────────

def _write_lease_pdf_reportlab(
    lease: LeaseEU,
    path: Path,
    canary: str,
) -> None:
    """Write a text-native lease agreement PDF using reportlab (A4, EUR)."""
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "LeaseTitle", parent=styles["Title"],
        fontSize=14, spaceAfter=6, textColor=colors.HexColor("#1A3C6E"),
    )
    heading_style = ParagraphStyle(
        "LeaseHeading", parent=styles["Heading2"],
        fontSize=11, spaceAfter=6, textColor=colors.HexColor("#1A3C6E"),
    )
    body_style = ParagraphStyle(
        "LeaseBody", parent=styles["Normal"],
        fontSize=10, spaceAfter=6, leading=14,
    )
    small_style = ParagraphStyle(
        "LeaseSmall", parent=styles["Normal"],
        fontSize=8, spaceAfter=4, textColor=colors.HexColor("#666666"),
    )

    story: list = []

    story.append(Paragraph("GEWERBEMIETVERTRAG / COMMERCIAL LEASE AGREEMENT", title_style))
    story.append(Spacer(1, 0.5 * cm))

    story.append(Paragraph("VERTRAGSPARTEIEN / PARTIES", heading_style))
    story.append(Paragraph(
        f"This Commercial Lease Agreement (the &quot;Agreement&quot;) is entered into as of "
        f"<b>{_fmt_date_eu(lease.commencement_date)}</b>, by and between:",
        body_style,
    ))
    story.append(Paragraph(
        f"<b>Vermieter / Lessor:</b> {lease.lessor}<br/>"
        f"<b>Mieter / Lessee:</b> {lease.lessee}",
        body_style,
    ))
    story.append(Spacer(1, 0.3 * cm))

    story.append(Paragraph("ABSCHNITT 1 -- MIETOBJEKT / PREMISES", heading_style))
    story.append(Paragraph(
        f"The Lessor hereby leases to the Lessee the following: "
        f"<b>{lease.description}</b>.",
        body_style,
    ))
    story.append(Spacer(1, 0.2 * cm))

    story.append(Paragraph("ABSCHNITT 2 -- MIETDAUER / LEASE TERM", heading_style))
    story.append(Paragraph(
        f"The lease term shall commence on <b>{_fmt_date_eu(lease.commencement_date)}</b> "
        f"and shall continue for a period of <b>{lease.term_months} months</b>, "
        f"terminating on <b>{_fmt_date_eu(lease.end_date)}</b>, unless earlier "
        f"terminated in accordance with the provisions of this Agreement.",
        body_style,
    ))
    story.append(Spacer(1, 0.2 * cm))

    story.append(Paragraph("ABSCHNITT 3 -- MIETZINS / BASE RENT", heading_style))
    story.append(Paragraph(
        f"The Lessee shall pay base rent of <b>{_fmt_eur_short(lease.monthly_base_rent)} per month</b>, "
        f"due on the first day of each calendar month during the lease term.",
        body_style,
    ))
    story.append(Spacer(1, 0.2 * cm))

    story.append(Paragraph("ABSCHNITT 4 -- MIETANPASSUNG / RENT ESCALATION", heading_style))
    story.append(Paragraph(_escalation_text_eu(lease), body_style))
    story.append(Spacer(1, 0.2 * cm))

    story.append(Paragraph("ABSCHNITT 5 -- VERLAENGERUNGSOPTION / RENEWAL OPTION", heading_style))
    story.append(Paragraph(_renewal_text_eu(lease), body_style))
    story.append(Spacer(1, 0.2 * cm))

    story.append(Paragraph("ABSCHNITT 6 -- KAUFOPTION / PURCHASE OPTION", heading_style))
    story.append(Paragraph(_purchase_text_eu(lease), body_style))
    story.append(Spacer(1, 0.2 * cm))

    story.append(Paragraph("ABSCHNITT 7 -- KUENDIGUNG / TERMINATION", heading_style))
    story.append(Paragraph(lease.termination_provision, body_style))
    story.append(Spacer(1, 0.2 * cm))

    # Low-value note if applicable
    if lease.low_value_exempt and lease.asset_value_when_new:
        story.append(Paragraph("ABSCHNITT 8 -- GERINGWERTIG / LOW-VALUE ASSET NOTE", heading_style))
        story.append(Paragraph(
            f"The underlying asset has a value when new of approximately "
            f"<b>{_fmt_eur_short(lease.asset_value_when_new)}</b>. "
            f"The lessee may elect the IFRS 16 low-value asset exemption.",
            body_style,
        ))
        story.append(Spacer(1, 0.2 * cm))

    # Signatures
    story.append(Paragraph("UNTERSCHRIFTEN / SIGNATURES", heading_style))
    story.append(Spacer(1, 0.6 * cm))

    sig_data = [
        ["Vermieter:", "________________________", "Datum:", "________________________"],
        ["", lease.lessor, "", _fmt_date_eu(lease.commencement_date)],
        ["", "", "", ""],
        ["Mieter:", "________________________", "Datum:", "________________________"],
        ["", lease.lessee, "", _fmt_date_eu(lease.commencement_date)],
    ]
    sig_table = Table(sig_data, colWidths=[2.0 * cm, 6.5 * cm, 1.5 * cm, 6.5 * cm])
    sig_table.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, -1), "Helvetica", 9),
        ("FONT", (0, 0), (0, 0), "Helvetica-Bold", 9),
        ("FONT", (0, 3), (0, 3), "Helvetica-Bold", 9),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(sig_table)

    story.append(Spacer(1, 0.6 * cm))
    story.append(Paragraph(
        f"Lease Reference: {lease.lease_id} / CANARY: {canary}",
        small_style,
    ))

    doc = SimpleDocTemplate(
        str(path),
        pagesize=_PAGE_SIZE,
        topMargin=2.0 * cm,
        bottomMargin=2.0 * cm,
        leftMargin=2.5 * cm,
        rightMargin=2.5 * cm,
        title=f"Lease Agreement -- {lease.lease_id}",
        author=f"CANARY: {canary}",
        creator=lease.lessor,
        invariant=True,
    )
    doc.build(story)


def _write_amendment_page_reportlab(
    lease: LeaseEU,
    path: Path,
    canary: str,
) -> None:
    """Write a lease PDF with original agreement + amendment page(s) (A4, EUR)."""
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "LeaseTitle", parent=styles["Title"],
        fontSize=14, spaceAfter=6, textColor=colors.HexColor("#1A3C6E"),
    )
    heading_style = ParagraphStyle(
        "LeaseHeading", parent=styles["Heading2"],
        fontSize=11, spaceAfter=6, textColor=colors.HexColor("#1A3C6E"),
    )
    body_style = ParagraphStyle(
        "LeaseBody", parent=styles["Normal"],
        fontSize=10, spaceAfter=6, leading=14,
    )
    small_style = ParagraphStyle(
        "LeaseSmall", parent=styles["Normal"],
        fontSize=8, spaceAfter=4, textColor=colors.HexColor("#666666"),
    )

    story: list = []

    # ── Original agreement ──
    story.append(Paragraph("GEWERBEMIETVERTRAG / COMMERCIAL LEASE AGREEMENT", title_style))
    story.append(Spacer(1, 0.5 * cm))

    story.append(Paragraph("VERTRAGSPARTEIEN / PARTIES", heading_style))
    story.append(Paragraph(
        f"This Commercial Lease Agreement is entered into as of "
        f"<b>{_fmt_date_eu(lease.commencement_date)}</b>, by and between:",
        body_style,
    ))
    story.append(Paragraph(
        f"<b>Vermieter / Lessor:</b> {lease.lessor}<br/>"
        f"<b>Mieter / Lessee:</b> {lease.lessee}",
        body_style,
    ))
    story.append(Spacer(1, 0.2 * cm))

    story.append(Paragraph("ABSCHNITT 1 -- MIETOBJEKT / PREMISES", heading_style))
    story.append(Paragraph(f"<b>{lease.description}</b>.", body_style))

    story.append(Paragraph("ABSCHNITT 2 -- MIETDAUER / LEASE TERM", heading_style))
    story.append(Paragraph(
        f"Commencement: <b>{_fmt_date_eu(lease.commencement_date)}</b>. "
        f"Original term: <b>{lease.term_months} months</b>.",
        body_style,
    ))

    story.append(Paragraph("ABSCHNITT 3 -- MIETZINS / BASE RENT", heading_style))
    story.append(Paragraph(
        f"Base rent: <b>{_fmt_eur_short(lease.monthly_base_rent)} per month</b>.",
        body_style,
    ))

    story.append(Paragraph("ABSCHNITT 4 -- MIETANPASSUNG / RENT ESCALATION", heading_style))
    story.append(Paragraph(_escalation_text_eu(lease), body_style))

    story.append(Paragraph("ABSCHNITT 5 -- VERLAENGERUNGSOPTION / RENEWAL OPTION", heading_style))
    story.append(Paragraph(_renewal_text_eu(lease), body_style))

    story.append(Paragraph("ABSCHNITT 6 -- KUENDIGUNG / TERMINATION", heading_style))
    story.append(Paragraph(lease.termination_provision, body_style))

    story.append(Spacer(1, 0.4 * cm))
    story.append(Paragraph(
        f"Lease Reference: {lease.lease_id} / CANARY: {canary}",
        small_style,
    ))

    # ── Amendment pages ──
    for i, amendment in enumerate(lease.amendments, start=1):
        story.append(PageBreak())
        story.append(Paragraph(
            f"NACHTRAG NR. {i} ZUM GEWERBEMIETVERTRAG / AMENDMENT NO. {i}",
            title_style,
        ))
        story.append(Spacer(1, 0.3 * cm))

        story.append(Paragraph(
            f"This Amendment No. {i} to the Commercial Lease Agreement dated "
            f"{_fmt_date_eu(lease.commencement_date)} is entered into effective as of "
            f"<b>{_fmt_date_eu(amendment.effective_date)}</b>, by and between "
            f"<b>{lease.lessor}</b> (Vermieter) and <b>{lease.lessee}</b> (Mieter).",
            body_style,
        ))
        story.append(Spacer(1, 0.2 * cm))

        story.append(Paragraph("PRAEAMBEL / RECITALS", heading_style))
        story.append(Paragraph(amendment.description, body_style))
        story.append(Spacer(1, 0.2 * cm))

        story.append(Paragraph("GEAENDERTE BEDINGUNGEN / AMENDED TERMS", heading_style))
        amended_items: list[str] = []
        if amendment.new_monthly_rent is not None:
            amended_items.append(
                f"Abschnitt 3 (Mietzins) is amended to <b>{_fmt_eur_short(amendment.new_monthly_rent)} "
                f"per month</b>, effective {_fmt_date_eu(amendment.effective_date)}."
            )
        if amendment.new_term_months is not None:
            amended_items.append(
                f"Abschnitt 2 (Mietdauer) is amended to a total term of "
                f"<b>{amendment.new_term_months} months</b> from the original "
                f"commencement date."
            )
        if amendment.new_escalation_pct is not None:
            pct = float(amendment.new_escalation_pct) * 100
            amended_items.append(
                f"Abschnitt 4 (Mietanpassung) is amended to {pct:.1f}% annually."
            )
        # For LS-EU-001 (termination reversed), the description carries the terms
        if not amended_items:
            amended_items.append(amendment.description)

        for item in amended_items:
            story.append(Paragraph(item, body_style))

        story.append(Spacer(1, 0.2 * cm))
        story.append(Paragraph(
            "All other terms and conditions of the Original Agreement remain in "
            "full force and effect. / Alle uebrigen Bedingungen des Ursprungsvertrags "
            "bleiben in vollem Umfang gueltig.",
            body_style,
        ))

        # Amendment signatures
        story.append(Spacer(1, 0.6 * cm))
        sig_data = [
            ["Vermieter:", "________________________", "Datum:", "________________________"],
            ["", lease.lessor, "", _fmt_date_eu(amendment.effective_date)],
            ["", "", "", ""],
            ["Mieter:", "________________________", "Datum:", "________________________"],
            ["", lease.lessee, "", _fmt_date_eu(amendment.effective_date)],
        ]
        sig_table = Table(sig_data, colWidths=[2.0 * cm, 6.5 * cm, 1.5 * cm, 6.5 * cm])
        sig_table.setStyle(TableStyle([
            ("FONT", (0, 0), (-1, -1), "Helvetica", 9),
            ("FONT", (0, 0), (0, 0), "Helvetica-Bold", 9),
            ("FONT", (0, 3), (0, 3), "Helvetica-Bold", 9),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(sig_table)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=_PAGE_SIZE,
        topMargin=2.0 * cm,
        bottomMargin=2.0 * cm,
        leftMargin=2.5 * cm,
        rightMargin=2.5 * cm,
        title=f"Lease Agreement -- {lease.lease_id}",
        author=f"CANARY: {canary}",
        creator=lease.lessor,
        invariant=True,
    )
    doc.build(story)


# ── Scanned-style PDF via fpdf2 ─────────────────────────────────────────────

# OCR confusion pairs for EUR context
_OCR_TRAPS_EU: dict[str, str] = {
    "EUR": "EUP",   # Euro sign/abbreviation misread
    "1": "l",       # 1/l confusion retained
}


def _latin1_safe(text: str) -> str:
    """Replace characters unsupported by fpdf2 core fonts (latin-1 only)."""
    return (
        text
        .replace("\u2014", "--")
        .replace("\u2013", "-")
        .replace("\u201c", '"')
        .replace("\u201d", '"')
        .replace("\u00e4", "ae")
        .replace("\u00f6", "oe")
        .replace("\u00fc", "ue")
        .replace("\u00c4", "Ae")
        .replace("\u00d6", "Oe")
        .replace("\u00dc", "Ue")
        .replace("\u00df", "ss")
        .replace("\u00e9", "e")
        .replace("\u00e8", "e")
        .replace("\u00ea", "e")
        .replace("\u00e0", "a")
        .replace("\u00e2", "a")
        .replace("\u00ee", "i")
        .replace("\u00f4", "o")
    )


def _ocr_mangle_eu(text: str, lease_id: str) -> str:
    """Introduce deliberate OCR-like confusion for European context."""
    # LS-EU-002 (CE parking): EUR → EUP in rent line
    if lease_id == "LS-EU-002" and "EUR 6.800" in text:
        text = text.replace("EUR 6.800", "EUP 6.800", 1)
    # LS-EU-004 (CP warehouse): "1" → "l" in term
    if lease_id == "LS-EU-004" and "84 months" in text:
        text = text.replace("84 months", "84 rnonths", 1)
    # LS-EU-014 (CM storage): mangle comma decimal "9.500" → "9.S00"
    if lease_id == "LS-EU-014" and "EUR 9.500" in text:
        text = text.replace("EUR 9.500", "EUR 9.S00", 1)
    return text


def _write_lease_pdf_fpdf2(
    lease: LeaseEU,
    path: Path,
    canary: str,
) -> None:
    """Write a 'scanned-style' lease PDF using fpdf2 with OCR traps."""
    pdf = FPDF()
    pdf.set_creation_date(_CREATION_DATE)
    pdf.set_auto_page_break(auto=True, margin=15)
    embed_canary_pdf_fpdf2(pdf, canary)

    pdf.add_page()

    pdf.set_font("Courier", "B", 13)
    pdf.cell(0, 8, "GEWERBEMIETVERTRAG / COMMERCIAL LEASE AGREEMENT", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(4)

    pdf.set_font("Courier", "", 10)

    def _line(text: str) -> None:
        mangled = _ocr_mangle_eu(_latin1_safe(text), lease.lease_id)
        pdf.multi_cell(0, 5, mangled, new_x="LMARGIN", new_y="NEXT")

    _line(f"Datum: {_fmt_date_eu(lease.commencement_date)}")
    _line(f"Vermieter / Lessor: {lease.lessor}")
    _line(f"Mieter / Lessee: {lease.lessee}")
    pdf.ln(3)

    _line(f"MIETOBJEKT / PREMISES: {lease.description}")
    pdf.ln(2)

    _line(f"MIETDAUER / LEASE TERM: {lease.term_months} months, commencing {_fmt_date_eu(lease.commencement_date)}")
    pdf.ln(2)

    _line(f"MIETZINS / BASE RENT: {_fmt_eur_short(lease.monthly_base_rent)} per month")
    pdf.ln(2)

    _line(f"MIETANPASSUNG / ESCALATION: {_escalation_text_eu(lease)}")
    pdf.ln(2)

    _line(f"VERLAENGERUNG / RENEWAL: {_renewal_text_eu(lease)}")
    pdf.ln(2)

    if lease.purchase_option:
        _line(f"KAUFOPTION / PURCHASE OPTION: {_purchase_text_eu(lease)}")
        pdf.ln(2)

    _line(f"KUENDIGUNG / TERMINATION: {lease.termination_provision}")
    pdf.ln(4)

    pdf.set_font("Courier", "", 8)
    _line(f"Ref: {lease.lease_id} / CANARY: {canary}")

    pdf.output(str(path))


# ── Lease PDFs ───────────────────────────────────────────────────────────────

def _write_lease_pdfs(
    leases: list[LeaseEU],
    output_dir: Path,
    canaries: CanaryRegistry,
    manifest: Manifest,
) -> None:
    """Write all 15 European lease PDFs."""
    leases_dir = output_dir / _LEASES_DIR
    leases_dir.mkdir(parents=True, exist_ok=True)

    for lease in leases:
        key = _canary_key(lease.lease_id)
        canary = canaries.canary_for(key)
        path = leases_dir / f"{lease.lease_id}.pdf"
        rel_path = f"{_LEASES_DIR}/{lease.lease_id}.pdf"

        if lease.lease_id in _SCANNED_IDS:
            _write_lease_pdf_fpdf2(lease, path, canary)
        elif lease.lease_id in _AMENDMENT_PDF_IDS:
            _write_amendment_page_reportlab(lease, path, canary)
        else:
            _write_lease_pdf_reportlab(lease, path, canary)

        canaries.set_location(
            key, rel_path,
            "PDF metadata -> Author (reportlab) or Subject (fpdf2); also in footer reference",
        )
        manifest.register(
            rel_path, "pdf",
            canary=canary,
            test_cases=[_TC],
        )


# ── Partial lease schedule xlsx (IFRS 16) ──────────────────────────────────

_FIXED_DATETIME = datetime.datetime(2025, 3, 15, 9, 0, 0)


def _save_xlsx_deterministic(wb: Any, path: str | Path) -> None:
    """Save workbook with pinned timestamps for byte-identical output."""
    path = Path(path)
    buf = io.BytesIO()
    wb.properties.modified = _FIXED_DATETIME
    archive = ZipFile(buf, "w", ZIP_DEFLATED, allowZip64=True)
    writer = ExcelWriter(wb, archive)
    writer.save()

    fixed_date_time = (2025, 3, 15, 9, 0, 0)
    buf.seek(0)
    with ZipFile(buf, "r") as src, ZipFile(str(path), "w", ZIP_DEFLATED) as dst:
        for item in src.infolist():
            info = ZipInfo(item.filename, date_time=fixed_date_time)
            info.compress_type = item.compress_type
            dst.writestr(info, src.read(item.filename))


# IFRS 16 schedule columns — no Lease Type column (single model)
# Adds Low-Value Exemption column (IFRS 16-specific)
_SCHEDULE_COLUMNS_EU = [
    "Lease ID",
    "Entity",
    "Lessor",
    "Description",
    "Commencement Date",
    "Term (Months)",
    "Monthly Base Rent (EUR)",
    "Escalation",
    "Renewal Option",
    "Purchase Option",
    "Short-Term Exemption",
    "Low-Value Exemption",
    "Incremental Borrowing Rate",
    "ROU Asset (Initial, EUR)",
    "Lease Liability (Initial, EUR)",
    "Notes",
]


def _write_partial_schedule(
    leases: list[LeaseEU],
    output_dir: Path,
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    manifest: Manifest,
) -> None:
    """Write IFRS 16 lease_schedule_partial.xlsx with 8 of 15 leases populated."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IFRS 16 Lease Schedule"

    wb.properties.created = _FIXED_DATETIME
    wb.properties.modified = _FIXED_DATETIME

    canary = canaries.canary_for(_SCHEDULE_KEY)
    embed_canary_xlsx(wb, canary)

    # Header styling
    header_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title row
    ws.merge_cells("A1:P1")
    ws["A1"] = "Cascade Europe -- IFRS 16 Lease Schedule (FY2025)"
    ws["A1"].font = Font(name="Calibri", bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Headers in row 3
    for col_idx, header in enumerate(_SCHEDULE_COLUMNS_EU, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Populate first 8 leases
    row = 4
    partial_leases = [ls for ls in leases if ls.lease_id in _PARTIAL_LEASE_IDS]
    partial_leases.sort(key=lambda ls: ls.lease_id)

    # ERR-EU-001: date_format_inconsistency — LS-EU-002 commencement date
    # Source PDF: DD.MM.YYYY (e.g. 05.03.2023 = March 5)
    # Schedule: shows 05/03/2023 which could be read as May 3 (MM/DD/YYYY)
    # We deliberately swap day and month to create the error
    err_lease_id = "LS-EU-002"
    err_correct_date: datetime.date | None = None
    err_wrong_date: datetime.date | None = None

    for lease in partial_leases:
        leave_blank = lease.lease_id in _BLANK_FIELDS_IDS

        ws.cell(row=row, column=1, value=lease.lease_id).border = thin_border
        ws.cell(row=row, column=2, value=lease.entity_code).border = thin_border
        ws.cell(row=row, column=3, value=lease.lessor).border = thin_border
        ws.cell(row=row, column=4, value=lease.description).border = thin_border

        # Commencement date — ERR-EU-001 for LS-EU-002
        if lease.lease_id == err_lease_id:
            correct = lease.commencement_date
            # Swap day and month to simulate DD/MM vs MM/DD confusion
            # Only works when day ≤ 12 so both interpretations are valid dates
            if correct.day <= 12 and correct.day != correct.month:
                wrong = datetime.date(correct.year, correct.day, correct.month)
            else:
                # Fallback: shift by two months if swap isn't possible
                wrong_month = (correct.month + 1) % 12 + 1
                wrong = datetime.date(correct.year, wrong_month, correct.day)
            err_correct_date = correct
            err_wrong_date = wrong
            ws.cell(row=row, column=5, value=wrong).border = thin_border
        else:
            ws.cell(row=row, column=5, value=lease.commencement_date).border = thin_border

        ws.cell(row=row, column=5).number_format = "DD/MM/YYYY"

        if leave_blank:
            ws.cell(row=row, column=6, value=lease.term_months).border = thin_border
            ws.cell(row=row, column=7).border = thin_border  # blank rent
            ws.cell(row=row, column=8).border = thin_border  # blank escalation
        else:
            ws.cell(row=row, column=6, value=lease.term_months).border = thin_border
            ws.cell(row=row, column=7, value=float(lease.monthly_base_rent)).border = thin_border
            ws.cell(row=row, column=7).number_format = '#.##0'
            ws.cell(row=row, column=8, value=_escalation_text_eu(lease)).border = thin_border

        ws.cell(row=row, column=9, value=_renewal_text_eu(lease)).border = thin_border
        ws.cell(row=row, column=10, value=_purchase_text_eu(lease)).border = thin_border
        ws.cell(row=row, column=11, value="Yes" if lease.short_term_exempt else "No").border = thin_border
        ws.cell(row=row, column=12, value="Yes" if lease.low_value_exempt else "No").border = thin_border

        # IBR and IFRS 16 computed values
        if leave_blank:
            ws.cell(row=row, column=13).border = thin_border
            ws.cell(row=row, column=14).border = thin_border
            ws.cell(row=row, column=15).border = thin_border
        else:
            ws.cell(row=row, column=13, value=f"{float(lease.discount_rate) * 100:.2f}%").border = thin_border
            ws.cell(row=row, column=14, value=float(lease.rou_asset_initial)).border = thin_border
            ws.cell(row=row, column=14).number_format = '#.##0'
            ws.cell(row=row, column=15, value=float(lease.lease_liability_initial)).border = thin_border
            ws.cell(row=row, column=15).number_format = '#.##0'

        ws.cell(row=row, column=16).border = thin_border  # notes
        row += 1

    # Blank rows for remaining 7 leases
    for i in range(7):
        for col in range(1, len(_SCHEDULE_COLUMNS_EU) + 1):
            ws.cell(row=row + i, column=col).border = thin_border

    # Column widths
    widths = [12, 6, 30, 40, 16, 10, 16, 28, 40, 20, 14, 14, 12, 16, 18, 20]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

    rel_path = f"{_INPUT_DIR}/lease_schedule_partial.xlsx"
    full_path = output_dir / rel_path
    full_path.parent.mkdir(parents=True, exist_ok=True)
    _save_xlsx_deterministic(wb, full_path)

    canaries.set_location(_SCHEDULE_KEY, rel_path, "Document properties -> description")
    manifest.register(rel_path, "xlsx", canary=canary, test_cases=[_TC])

    # Register ERR-EU-001
    assert err_correct_date is not None and err_wrong_date is not None
    errors.add(PlantedError(
        error_id="ERR-EU-001",
        file=rel_path,
        location=f"Sheet 'IFRS 16 Lease Schedule', Row 5 ({err_lease_id}), Column E (Commencement Date)",
        type="date_inconsistency",
        description=(
            f"Commencement date for {err_lease_id} shows {err_wrong_date.strftime('%d/%m/%Y')} "
            f"in the schedule but the source PDF uses {err_correct_date.strftime('%d.%m.%Y')} -- "
            f"day and month are swapped (DD/MM vs MM/DD ambiguity)"
        ),
        severity="significant",
        which_test_cases_should_catch=["TC-04-EU"],
    ))


# ── Prompt & expected behavior ───────────────────────────────────────────────

def _write_prompt(output_dir: Path) -> None:
    text = """\
The audit team needs to complete the IFRS 16 lease schedule for the Cascade Europe group.

1. Extract key terms from each lease agreement in the leases/ folder:
   - Lessee and lessor names
   - Commencement date
   - Lease term (in months)
   - Monthly/annual base rent (EUR)
   - Escalation terms (fixed %, index-linked, or stepped)
   - Renewal options (term and rent)
   - Purchase options
   - Termination provisions
2. For leases with amendments, use the amended terms (not the original).
3. Populate the lease schedule, matching the format of the existing entries.
4. Flag any leases that qualify for the short-term lease exemption (<=12 months
   remaining at reporting date with no purchase option reasonably certain
   to be exercised).
5. Flag any leases that may qualify for the low-value asset exemption
   (underlying asset value when new <= USD 5,000 / ~EUR 4,500).
6. Flag any leases where extracted terms are uncertain (e.g., from scanned documents
   where OCR may be unreliable).
7. Note any leases that require judgment calls (e.g., whether a renewal is
   "reasonably certain" to be exercised, or whether a variable lease payment
   is in-substance fixed).

Export the completed schedule as an Excel file.
"""
    path = output_dir / f"test_cases/{_TC}/prompt.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text)


def _write_expected_behavior(output_dir: Path) -> None:
    text = """\
# TC-04-EU: IFRS 16 Lease Extraction & Schedule Population -- Expected Behavior

## Data Challenges
- **15 lease PDFs** across three Cascade Europe entities (CE, CP, CM):
  - 10 text-native PDFs (A4 format, European date/currency conventions)
  - 3 scanned-style PDFs with deliberate **OCR traps**:
    - LS-EU-002: "EUR" rendered as "EUP" in rent amount
    - LS-EU-004: "months" rendered as "rnonths" in lease term
    - LS-EU-014: "9.500" rendered as "9.S00" (digit/letter confusion)
  - 2 PDFs with **amendment pages** modifying original terms (LS-EU-001, LS-EU-003)
- European formatting: dates as DD.MM.YYYY, amounts in EUR with period thousands
  separator (12.500 EUR = twelve thousand five hundred euros)
- The partial schedule has 8 of 15 leases pre-populated, but 3 of those have
  **blank fields** (rent, escalation, IBR, ROU/liability) that need to be filled.

## IFRS 16 Single Model (Judgment Trap)
- IFRS 16 does NOT classify leases as operating vs. finance for the lessee.
- All leases go on the balance sheet (ROU asset + lease liability) except those
  qualifying for short-term or low-value exemptions.
- If the agent produces an operating/finance classification, this indicates
  application of **ASC 842 logic** and should be flagged as incorrect.

## Amendment Handling
- **LS-EU-001** (CE head office): Early termination clause exercised then reversed.
  Agent must use the reinstated original terms, not the terminated state.
- **LS-EU-003** (CP manufacturing facility): Post-COVID renegotiation increased
  rent by EUR 8,000/month and extended term by 24 months.
- **LS-EU-010** (CM R&D lab): Scope change adding additional floor; rent increased
  by EUR 12,000/month. Amendment is in the lease body, not a separate page.

## Short-Term Exemptions
- **LS-EU-008** (CP vehicle fleet): 10-month lease, no renewal, no purchase
  option -> qualifies for short-term exemption.
- **LS-EU-015** (CM IT server rack): 8-month lease, no renewal, no purchase
  option -> qualifies for short-term exemption.
- The agent should identify exactly **2 short-term exempt leases**.

## Low-Value Asset Exemption (IFRS 16-Specific)
- **LS-EU-007** (CP office printer): Asset value when new EUR 3,800 (below the
  ~EUR 4,500 / USD 5,000 threshold). The agent should identify this as
  potentially qualifying for the low-value asset exemption and state the threshold.
- This exemption does NOT exist under ASC 842 -- it tests IFRS 16-specific
  knowledge.

## Planted Errors
- **ERR-EU-001**: The partial schedule shows LS-EU-002's commencement date with
  day and month swapped compared to the source PDF (DD.MM.YYYY in PDF vs
  MM/DD/YYYY interpretation in schedule). The agent should flag this date
  format inconsistency.

## OCR Uncertainty Flagging
- The agent should **flag uncertainty** on at least the 3 scanned-style leases
  rather than silently inserting potentially wrong values.

## Date Format Handling
- European dates use DD.MM.YYYY or DD/MM/YYYY format.
- At least one lease has an ambiguous date where DD/MM could be read as MM/DD
  (e.g., 05/03/2023 -- is it March 5 or May 3?). The agent should flag this.

## EUR Formatting
- Amounts use comma decimal separator in formal documents (12.500,00 EUR) but
  the schedule uses period-based grouping (12.500 EUR).
- The agent must correctly parse these as twelve thousand five hundred, not 12.5.

## IBR Selection
- IFRS 16 emphasises the incremental borrowing rate (EURIBOR-based for European
  entities). The rate implicit in the lease is rarely determinable by the lessee.
- If the agent uses "rate implicit in the lease" without justification, this
  should be flagged.

## Output Quality
- Completed Excel schedule with all 15 leases populated.
- Amendment terms correctly applied (3 leases).
- Short-term exemptions flagged (2 leases).
- Low-value exemption flagged (1 lease).
- No operating/finance classification (IFRS 16 single model).
- OCR uncertainties flagged (3 leases).
- Date format inconsistency identified (1 instance).
"""
    path = output_dir / f"test_cases/{_TC}/expected_behavior.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text)


# ── Gold Standard ────────────────────────────────────────────────────────────

@register_gold("TC-04-EU")
def _tc04_eu_gold(
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    **model_kwargs: Any,
) -> GoldStandard:
    """Build the TC-04-EU gold standard from the European lease model."""
    # EU leases are generated by the formatter, not stored on CascadeModel,
    # so we regenerate them deterministically from the same seed.
    ctx: ScenarioContext = model_kwargs["ctx"]
    rng_obj = ctx.named_rng("tc04_eu_leases")
    # Convert numpy Generator to stdlib Random for the model
    import random as _random
    seed_bytes = rng_obj.bytes(8)
    seed_int = int.from_bytes(seed_bytes, "big")
    rng = _random.Random(seed_int)
    leases = generate_leases_eu(rng)

    lease_details = []
    for ls in sorted(leases, key=lambda x: x.lease_id):
        detail: dict[str, Any] = {
            "lease_id": ls.lease_id,
            "entity_code": ls.entity_code,
            "lessee": ls.lessee,
            "lessor": ls.lessor,
            "description": ls.description,
            "commencement_date": ls.commencement_date.isoformat(),
            "original_term_months": ls.term_months,
            "effective_term_months": ls.effective_term_months,
            "original_monthly_rent_eur": int(ls.monthly_base_rent),
            "effective_monthly_rent_eur": int(ls.effective_monthly_rent),
            "escalation_type": ls.escalation_type.value,
            "short_term_exempt": ls.short_term_exempt,
            "low_value_exempt": ls.low_value_exempt,
            "asset_value_when_new_eur": int(ls.asset_value_when_new) if ls.asset_value_when_new else None,
            "has_amendment": len(ls.amendments) > 0,
            "purchase_option": ls.purchase_option,
            "on_balance_sheet": ls.is_on_balance_sheet,
            "discount_rate": str(ls.discount_rate),
            "rou_asset_initial_eur": int(ls.rou_asset_initial),
            "lease_liability_initial_eur": int(ls.lease_liability_initial),
        }
        if ls.amendments:
            detail["amendments"] = [
                {
                    "effective_date": a.effective_date.isoformat(),
                    "description": a.description,
                    "new_monthly_rent_eur": int(a.new_monthly_rent) if a.new_monthly_rent else None,
                    "new_term_months": a.new_term_months,
                }
                for a in ls.amendments
            ]
        lease_details.append(detail)

    short_term_ids = sorted(ls.lease_id for ls in leases if ls.short_term_exempt)
    low_value_ids = sorted(ls.lease_id for ls in leases if ls.low_value_exempt)
    amended_ids = sorted(ls.lease_id for ls in leases if ls.amendments)
    scanned_ids = sorted(_SCANNED_IDS)
    on_bs_ids = sorted(ls.lease_id for ls in leases if ls.is_on_balance_sheet)

    canary_verif = {}
    for ls in sorted(leases, key=lambda x: x.lease_id):
        key = _canary_key(ls.lease_id)
        canary_verif[f"read_{ls.lease_id}"] = canaries.canary_for(key)
    canary_verif["read_lease_schedule"] = canaries.canary_for(_SCHEDULE_KEY)

    return GoldStandard(
        test_case=_TC,
        expected_outputs={
            "file_type": "xlsx",
            "total_leases": 15,
            "leases": lease_details,
            "short_term_exempt_ids": short_term_ids,
            "short_term_exempt_count": len(short_term_ids),
            "low_value_exempt_ids": low_value_ids,
            "low_value_exempt_count": len(low_value_ids),
            "amended_lease_ids": amended_ids,
            "scanned_lease_ids": scanned_ids,
            "on_balance_sheet_ids": on_bs_ids,
            "on_balance_sheet_count": len(on_bs_ids),
            "ocr_traps": {
                "LS-EU-002": "EUR rendered as EUP in rent amount",
                "LS-EU-004": "months rendered as rnonths in lease term",
                "LS-EU-014": "9.500 rendered as 9.S00 (digit confusion)",
            },
            "schedule_pre_populated": 8,
            "schedule_blank_fields": sorted(_BLANK_FIELDS_IDS),
            "ifrs16_single_model": True,
            "no_operating_finance_classification": True,
            "currency": "EUR",
            "date_format": "DD.MM.YYYY",
        },
        canary_verification=canary_verif,
        error_detection={
            "ERR-EU-001": (
                "Commencement date for LS-EU-002 in the partial schedule has "
                "day and month swapped vs the source PDF -- date_format_inconsistency"
            ),
        },
        scoring_hints={
            "correctness": (
                "All 15 leases extracted with correct terms; 2 short-term exemptions identified; "
                "1 low-value exemption identified; 3 amendments applied; no operating/finance "
                "classification (IFRS 16 single model)"
            ),
            "completeness": (
                "Complete IFRS 16 schedule matching existing format; all 15 leases populated with "
                "all required fields including IBR, ROU asset, and lease liability"
            ),
            "format_compliance": "Valid xlsx matching partial schedule format; EUR formatting; DD.MM.YYYY dates",
            "robustness": (
                "Handled European date formats; parsed EUR amounts correctly; flagged OCR "
                "uncertainties on 3 scanned leases; date format inconsistency identified"
            ),
            "communication": (
                "Flagged uncertain extractions from scanned PDFs; noted IFRS 16 vs ASC 842 "
                "considerations; highlighted date format ambiguity; stated low-value threshold"
            ),
        },
        scenario_pack="cascade_europe_ifrs",
        service_line="audit",
        judgment_traps=[
            {
                "trap_id": "JT-EU-001",
                "trap_type": "standard_confusion",
                "expected_response": "Agent should NOT classify leases as operating vs finance",
                "description": (
                    "IFRS 16 uses a single lessee model. If the agent produces an "
                    "operating/finance classification, it is applying ASC 842 logic."
                ),
            },
            {
                "trap_id": "JT-EU-002",
                "trap_type": "low_value_exemption",
                "expected_response": "Agent should identify LS-EU-007 printer as potentially low-value exempt",
                "description": (
                    "The IFRS 16 low-value asset exemption (~EUR 4,500 / USD 5,000) "
                    "does not exist under ASC 842. Agent must know the IFRS 16 threshold."
                ),
            },
        ],
    )


# ── Public entry point ───────────────────────────────────────────────────────

def emit_tc04_eu(
    model: CascadeModel,
    output_dir: Path,
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    manifest: Manifest,
    **kwargs: object,
) -> None:
    """Write all TC-04-EU files to *output_dir*."""
    ctx: ScenarioContext = kwargs["ctx"]  # type: ignore[assignment]
    # Generate European leases from a dedicated RNG stream
    rng_obj = ctx.named_rng("tc04_eu_leases")
    import random as _random
    seed_bytes = rng_obj.bytes(8)
    seed_int = int.from_bytes(seed_bytes, "big")
    rng = _random.Random(seed_int)
    leases = generate_leases_eu(rng)

    _write_lease_pdfs(leases, output_dir, canaries, manifest)
    _write_partial_schedule(leases, output_dir, canaries, errors, manifest)
    _write_prompt(output_dir)
    _write_expected_behavior(output_dir)
