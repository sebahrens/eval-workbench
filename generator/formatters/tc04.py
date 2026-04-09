"""Formatter: TC-04 — Lease Extraction & ASC 842 Schedule Population (Audit, Adversarial).

Emits:
- test_cases/TC-04/input_files/leases/LS-001.pdf through LS-015.pdf
  15 lease PDFs: 10 text-native (reportlab), 3 scanned-style (fpdf2 low-res),
  2 with amendments appended as additional pages
- test_cases/TC-04/input_files/lease_schedule_partial.xlsx
  Partially populated lease schedule (8 of 15 leases filled)
- test_cases/TC-04/prompt.md
- test_cases/TC-04/expected_behavior.md
- gold_standards/TC-04_gold.json

One planted error: ERR-007 date_inconsistency — commencement date in the partial
schedule for LS-002 differs from the PDF by one month.

Uses the canonical lease model — never hardcodes numbers.
"""

from __future__ import annotations

import datetime
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any

from fpdf import FPDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from generator.canaries import (
    CanaryRegistry,
    embed_canary_pdf_fpdf2,
    embed_canary_xlsx,
)
from generator.errors import ErrorRegistry, PlantedError
from generator.golds.framework import GoldStandard, register_gold
from generator.manifest import Manifest
from generator.model.build import CascadeModel
from generator.model.leases import (
    EscalationType,
    Lease,
)

# ── Constants ────────────────────────────────────────────────────────────────

_TC = "TC-04"
_INPUT_DIR = f"test_cases/{_TC}/input_files"
_LEASES_DIR = f"{_INPUT_DIR}/leases"

# fpdf2 fixed creation date for determinism
_CREATION_DATE = datetime.datetime(2025, 1, 15, 9, 0, 0)

# ── Lease-to-rendering-type assignment ───────────────────────────────────────
# 10 text-native, 3 scanned-style, 2 with amendments.
# Amendments: LS-002 (warehouse expansion) and LS-007 (clean room conversion)
# and LS-010 (reduced footprint).  LS-002 and LS-007 are the two we render
# with amendment pages.  LS-010's amendment is the third but we render it
# as text-native (the spec says 2 with amendments *attached as additional pages*).
#
# Scanned-style: LS-003 (CNC milling), LS-008 (electron microscope),
# LS-013 (refrigerated truck) — equipment leases from smaller lessors.

_SCANNED_IDS = {"LS-003", "LS-008", "LS-013"}
_AMENDMENT_PDF_IDS = {"LS-002", "LS-007"}
# All others are text-native (10 total: 15 - 3 scanned - 2 amendment = 10)


# ── Helpers ──────────────────────────────────────────────────────────────────

def _fmt_dollars(d: Decimal | int) -> str:
    if isinstance(d, int):
        val = d
    else:
        val = int(d.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    if val < 0:
        return f"(${abs(val):,})"
    return f"${val:,}"


def _fmt_date(d: datetime.date) -> str:
    return d.strftime("%B %d, %Y")


def _escalation_text(lease: Lease) -> str:
    if lease.escalation_type == EscalationType.FIXED_PCT:
        pct = float(lease.escalation_pct) * 100
        return f"Fixed annual escalation of {pct:.1f}%"
    elif lease.escalation_type == EscalationType.CPI:
        return "Annual adjustment based on Consumer Price Index (CPI)"
    elif lease.escalation_type == EscalationType.STEPPED:
        if lease.escalation_steps:
            steps = ", ".join(_fmt_dollars(s) for s in lease.escalation_steps)
            return f"Stepped rent schedule: {steps} per month (years 1, 2, 3)"
        return "Stepped rent schedule per attached schedule"
    return "None"


def _renewal_text(lease: Lease) -> str:
    if lease.renewal_option_months <= 0:
        return "None"
    mos = lease.renewal_option_months
    increase = float(lease.renewal_rent_increase_pct) * 100
    return (
        f"Option to renew for {mos} months at {increase:.0f}% above "
        f"the then-current base rent, exercisable with 90-day prior written notice"
    )


def _purchase_text(lease: Lease) -> str:
    if not lease.purchase_option:
        return "None"
    price = _fmt_dollars(lease.purchase_option_price) if lease.purchase_option_price else "fair market value"
    return f"Purchase option at {price} at end of lease term"


def _canary_key(lease_id: str) -> str:
    num = lease_id.split("-")[1]
    return f"tc04_lease_{num}"


_SCHEDULE_KEY = "tc04_lease_schedule_partial"

# ── Which 8 leases go in the partial schedule ─────────────────────────────
# Pick the first 8 by lease_id (LS-001..LS-008).
# Some fields are left blank to test the agent's ability to fill gaps.
_PARTIAL_LEASE_IDS = {f"LS-{i:03d}" for i in range(1, 9)}
# For 3 of the 8, leave some fields blank (monthly rent, escalation)
_BLANK_FIELDS_IDS = {"LS-003", "LS-005", "LS-007"}


# ── Text-native PDF via reportlab ────────────────────────────────────────────

def _write_lease_pdf_reportlab(
    lease: Lease,
    path: Path,
    canary: str,
) -> None:
    """Write a text-native lease agreement PDF using reportlab."""
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "LeaseTitle", parent=styles["Title"],
        fontSize=14, spaceAfter=6, textColor=colors.HexColor("#1A3C6E"),
    )
    heading_style = ParagraphStyle(
        "LeaseHeading", parent=styles["Heading2"],
        fontSize=11, spaceAfter=6, textColor=colors.HexColor("#1A3C6E"),
    )
    body_style = ParagraphStyle(
        "LeaseBody", parent=styles["Normal"],
        fontSize=10, spaceAfter=6, leading=14,
    )
    small_style = ParagraphStyle(
        "LeaseSmall", parent=styles["Normal"],
        fontSize=8, spaceAfter=4, textColor=colors.HexColor("#666666"),
    )

    story: list = []

    # Title
    story.append(Paragraph("COMMERCIAL LEASE AGREEMENT", title_style))
    story.append(Spacer(1, 0.2 * inch))

    # Parties
    story.append(Paragraph("PARTIES", heading_style))
    story.append(Paragraph(
        f"This Commercial Lease Agreement (the &quot;Agreement&quot;) is entered into as of "
        f"<b>{_fmt_date(lease.commencement_date)}</b>, by and between:",
        body_style,
    ))
    story.append(Paragraph(
        f"<b>Lessor:</b> {lease.lessor}<br/>"
        f"<b>Lessee:</b> {lease.lessee}",
        body_style,
    ))
    story.append(Spacer(1, 0.15 * inch))

    # Premises / Equipment
    story.append(Paragraph("SECTION 1 — PREMISES / EQUIPMENT", heading_style))
    story.append(Paragraph(
        f"The Lessor hereby leases to the Lessee the following: "
        f"<b>{lease.description}</b>.",
        body_style,
    ))
    story.append(Spacer(1, 0.1 * inch))

    # Term
    story.append(Paragraph("SECTION 2 — LEASE TERM", heading_style))
    story.append(Paragraph(
        f"The lease term shall commence on <b>{_fmt_date(lease.commencement_date)}</b> "
        f"and shall continue for a period of <b>{lease.term_months} months</b>, "
        f"terminating on <b>{_fmt_date(lease.end_date)}</b>, unless earlier "
        f"terminated in accordance with the provisions of this Agreement.",
        body_style,
    ))
    story.append(Spacer(1, 0.1 * inch))

    # Rent
    story.append(Paragraph("SECTION 3 — BASE RENT", heading_style))
    story.append(Paragraph(
        f"The Lessee shall pay base rent of <b>{_fmt_dollars(lease.monthly_base_rent)} per month</b>, "
        f"due on the first day of each calendar month during the lease term.",
        body_style,
    ))
    story.append(Spacer(1, 0.1 * inch))

    # Escalation
    story.append(Paragraph("SECTION 4 — RENT ESCALATION", heading_style))
    story.append(Paragraph(_escalation_text(lease), body_style))
    story.append(Spacer(1, 0.1 * inch))

    # Renewal
    story.append(Paragraph("SECTION 5 — RENEWAL OPTION", heading_style))
    story.append(Paragraph(_renewal_text(lease), body_style))
    story.append(Spacer(1, 0.1 * inch))

    # Purchase option
    story.append(Paragraph("SECTION 6 — PURCHASE OPTION", heading_style))
    story.append(Paragraph(_purchase_text(lease), body_style))
    story.append(Spacer(1, 0.1 * inch))

    # Termination
    story.append(Paragraph("SECTION 7 — TERMINATION", heading_style))
    story.append(Paragraph(lease.termination_provision, body_style))
    story.append(Spacer(1, 0.1 * inch))

    # Signature block
    story.append(Paragraph("SIGNATURES", heading_style))
    story.append(Spacer(1, 0.3 * inch))

    sig_data = [
        ["Lessor:", "________________________", "Date:", "________________________"],
        ["", lease.lessor, "", _fmt_date(lease.commencement_date)],
        ["", "", "", ""],
        ["Lessee:", "________________________", "Date:", "________________________"],
        ["", lease.lessee, "", _fmt_date(lease.commencement_date)],
    ]
    sig_table = Table(sig_data, colWidths=[0.7 * inch, 2.5 * inch, 0.5 * inch, 2.5 * inch])
    sig_table.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, -1), "Helvetica", 9),
        ("FONT", (0, 0), (0, 0), "Helvetica-Bold", 9),
        ("FONT", (0, 3), (0, 3), "Helvetica-Bold", 9),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(sig_table)

    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(
        f"Lease Reference: {lease.lease_id} / CANARY: {canary}",
        small_style,
    ))

    doc = SimpleDocTemplate(
        str(path),
        pagesize=letter,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
        leftMargin=inch,
        rightMargin=inch,
        title=f"Lease Agreement — {lease.lease_id}",
        author=f"CANARY: {canary}",
        creator=lease.lessor,
        invariant=True,
    )
    doc.build(story)


def _write_amendment_page_reportlab(
    lease: Lease,
    path: Path,
    canary: str,
) -> None:
    """Write a lease PDF with the original agreement + amendment page(s)."""
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "LeaseTitle", parent=styles["Title"],
        fontSize=14, spaceAfter=6, textColor=colors.HexColor("#1A3C6E"),
    )
    heading_style = ParagraphStyle(
        "LeaseHeading", parent=styles["Heading2"],
        fontSize=11, spaceAfter=6, textColor=colors.HexColor("#1A3C6E"),
    )
    body_style = ParagraphStyle(
        "LeaseBody", parent=styles["Normal"],
        fontSize=10, spaceAfter=6, leading=14,
    )
    small_style = ParagraphStyle(
        "LeaseSmall", parent=styles["Normal"],
        fontSize=8, spaceAfter=4, textColor=colors.HexColor("#666666"),
    )

    story: list = []

    # ── Original agreement (page 1) ──
    story.append(Paragraph("COMMERCIAL LEASE AGREEMENT", title_style))
    story.append(Spacer(1, 0.2 * inch))

    story.append(Paragraph("PARTIES", heading_style))
    story.append(Paragraph(
        f"This Commercial Lease Agreement is entered into as of "
        f"<b>{_fmt_date(lease.commencement_date)}</b>, by and between:",
        body_style,
    ))
    story.append(Paragraph(
        f"<b>Lessor:</b> {lease.lessor}<br/>"
        f"<b>Lessee:</b> {lease.lessee}",
        body_style,
    ))
    story.append(Spacer(1, 0.1 * inch))

    story.append(Paragraph("SECTION 1 — PREMISES / EQUIPMENT", heading_style))
    story.append(Paragraph(f"<b>{lease.description}</b>.", body_style))

    story.append(Paragraph("SECTION 2 — LEASE TERM", heading_style))
    story.append(Paragraph(
        f"Commencement: <b>{_fmt_date(lease.commencement_date)}</b>. "
        f"Original term: <b>{lease.term_months} months</b>.",
        body_style,
    ))

    story.append(Paragraph("SECTION 3 — BASE RENT", heading_style))
    story.append(Paragraph(
        f"Base rent: <b>{_fmt_dollars(lease.monthly_base_rent)} per month</b>.",
        body_style,
    ))

    story.append(Paragraph("SECTION 4 — RENT ESCALATION", heading_style))
    story.append(Paragraph(_escalation_text(lease), body_style))

    story.append(Paragraph("SECTION 5 — RENEWAL OPTION", heading_style))
    story.append(Paragraph(_renewal_text(lease), body_style))

    story.append(Paragraph("SECTION 6 — TERMINATION", heading_style))
    story.append(Paragraph(lease.termination_provision, body_style))

    story.append(Spacer(1, 0.2 * inch))
    story.append(Paragraph(
        f"Lease Reference: {lease.lease_id} / CANARY: {canary}",
        small_style,
    ))

    # ── Amendment pages ──
    for i, amendment in enumerate(lease.amendments, start=1):
        story.append(PageBreak())
        story.append(Paragraph(
            f"AMENDMENT NO. {i} TO COMMERCIAL LEASE AGREEMENT",
            title_style,
        ))
        story.append(Spacer(1, 0.15 * inch))

        story.append(Paragraph(
            f"This Amendment No. {i} (the &quot;Amendment&quot;) to the Commercial Lease "
            f"Agreement dated {_fmt_date(lease.commencement_date)} (the &quot;Original "
            f"Agreement&quot;) is entered into effective as of "
            f"<b>{_fmt_date(amendment.effective_date)}</b>, by and between "
            f"<b>{lease.lessor}</b> (Lessor) and <b>{lease.lessee}</b> (Lessee).",
            body_style,
        ))
        story.append(Spacer(1, 0.1 * inch))

        story.append(Paragraph("RECITALS", heading_style))
        story.append(Paragraph(amendment.description, body_style))
        story.append(Spacer(1, 0.1 * inch))

        story.append(Paragraph("AMENDED TERMS", heading_style))
        amended_items: list[str] = []
        if amendment.new_monthly_rent is not None:
            amended_items.append(
                f"Section 3 (Base Rent) is amended to <b>{_fmt_dollars(amendment.new_monthly_rent)} "
                f"per month</b>, effective {_fmt_date(amendment.effective_date)}."
            )
        if amendment.new_term_months is not None:
            amended_items.append(
                f"Section 2 (Lease Term) is amended to a total term of "
                f"<b>{amendment.new_term_months} months</b> from the original "
                f"commencement date."
            )
        if amendment.new_escalation_pct is not None:
            pct = float(amendment.new_escalation_pct) * 100
            amended_items.append(
                f"Section 4 (Escalation) is amended to {pct:.1f}% annually."
            )

        for item in amended_items:
            story.append(Paragraph(item, body_style))

        story.append(Spacer(1, 0.1 * inch))
        story.append(Paragraph(
            "All other terms and conditions of the Original Agreement remain in "
            "full force and effect.",
            body_style,
        ))

        # Amendment signatures
        story.append(Spacer(1, 0.3 * inch))
        sig_data = [
            ["Lessor:", "________________________", "Date:", "________________________"],
            ["", lease.lessor, "", _fmt_date(amendment.effective_date)],
            ["", "", "", ""],
            ["Lessee:", "________________________", "Date:", "________________________"],
            ["", lease.lessee, "", _fmt_date(amendment.effective_date)],
        ]
        sig_table = Table(sig_data, colWidths=[0.7 * inch, 2.5 * inch, 0.5 * inch, 2.5 * inch])
        sig_table.setStyle(TableStyle([
            ("FONT", (0, 0), (-1, -1), "Helvetica", 9),
            ("FONT", (0, 0), (0, 0), "Helvetica-Bold", 9),
            ("FONT", (0, 3), (0, 3), "Helvetica-Bold", 9),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(sig_table)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=letter,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
        leftMargin=inch,
        rightMargin=inch,
        title=f"Lease Agreement — {lease.lease_id}",
        author=f"CANARY: {canary}",
        creator=lease.lessor,
        invariant=True,
    )
    doc.build(story)


# ── Scanned-style PDF via fpdf2 ─────────────────────────────────────────────
# Deliberately uses a different font, lower readability, and OCR-trap characters.

# OCR confusion pairs: map correct char → ambiguous rendering
_OCR_TRAPS: dict[str, str] = {
    "$": "S",   # "$18,000" → "S18,000"
    "1": "l",   # "12 months" → "l2 months"
}


def _latin1_safe(text: str) -> str:
    """Replace characters unsupported by fpdf2 core fonts (latin-1 only)."""
    return text.replace("\u2014", "--").replace("\u2013", "-").replace("\u201c", '"').replace("\u201d", '"')


def _ocr_mangle(text: str, lease_id: str) -> str:
    """Introduce deliberate OCR-like confusion into text.

    Only mangle specific instances — not every occurrence — so the document
    is readable but has a few spots where an agent should flag uncertainty.
    """
    # For the first scanned lease (LS-003): mangle a "$" to "S" in the rent line
    if lease_id == "LS-003" and "$8,500" in text:
        text = text.replace("$8,500", "S8,500", 1)
    # For LS-008: mangle "12,000" → "l2,000" once
    if lease_id == "LS-008" and "$12,000" in text:
        text = text.replace("$12,000", "$l2,000", 1)
    # For LS-013: mangle "9 months" → "9 rnonths" (m→rn OCR confusion)
    if lease_id == "LS-013" and "9 months" in text:
        text = text.replace("9 months", "9 rnonths", 1)
    return text


def _write_lease_pdf_fpdf2(
    lease: Lease,
    path: Path,
    canary: str,
) -> None:
    """Write a 'scanned-style' lease PDF using fpdf2 with OCR traps."""
    pdf = FPDF()
    pdf.set_creation_date(_CREATION_DATE)
    pdf.set_auto_page_break(auto=True, margin=15)
    embed_canary_pdf_fpdf2(pdf, canary)

    pdf.add_page()

    # Use Courier to simulate scanned document feel
    pdf.set_font("Courier", "B", 13)
    pdf.cell(0, 8, "COMMERCIAL LEASE AGREEMENT", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(4)

    pdf.set_font("Courier", "", 10)

    def _line(text: str) -> None:
        mangled = _ocr_mangle(_latin1_safe(text), lease.lease_id)
        pdf.multi_cell(0, 5, mangled, new_x="LMARGIN", new_y="NEXT")

    _line(f"Date: {_fmt_date(lease.commencement_date)}")
    _line(f"Lessor: {lease.lessor}")
    _line(f"Lessee: {lease.lessee}")
    pdf.ln(3)

    _line(f"PREMISES/EQUIPMENT: {lease.description}")
    pdf.ln(2)

    _line(f"LEASE TERM: {lease.term_months} months, commencing {_fmt_date(lease.commencement_date)}")
    pdf.ln(2)

    _line(f"BASE RENT: {_fmt_dollars(lease.monthly_base_rent)} per month")
    pdf.ln(2)

    _line(f"ESCALATION: {_escalation_text(lease)}")
    pdf.ln(2)

    _line(f"RENEWAL: {_renewal_text(lease)}")
    pdf.ln(2)

    if lease.purchase_option:
        _line(f"PURCHASE OPTION: {_purchase_text(lease)}")
        pdf.ln(2)

    _line(f"TERMINATION: {lease.termination_provision}")
    pdf.ln(4)

    pdf.set_font("Courier", "", 8)
    _line(f"Ref: {lease.lease_id} / CANARY: {canary}")

    pdf.output(str(path))


# ── Lease PDFs ───────────────────────────────────────────────────────────────

def _write_lease_pdfs(
    leases: list[Lease],
    output_dir: Path,
    canaries: CanaryRegistry,
    manifest: Manifest,
) -> None:
    """Write all 15 lease PDFs."""
    leases_dir = output_dir / _LEASES_DIR
    leases_dir.mkdir(parents=True, exist_ok=True)

    for lease in leases:
        key = _canary_key(lease.lease_id)
        canary = canaries.canary_for(key)
        path = leases_dir / f"{lease.lease_id}.pdf"
        rel_path = f"{_LEASES_DIR}/{lease.lease_id}.pdf"

        if lease.lease_id in _SCANNED_IDS:
            _write_lease_pdf_fpdf2(lease, path, canary)
        elif lease.lease_id in _AMENDMENT_PDF_IDS:
            _write_amendment_page_reportlab(lease, path, canary)
        else:
            _write_lease_pdf_reportlab(lease, path, canary)

        canaries.set_location(
            key, rel_path,
            "PDF metadata → Author (reportlab) or Subject (fpdf2); also in footer reference",
        )
        manifest.register(
            rel_path, "pdf",
            canary=canary,
            test_cases=[_TC],
        )


# ── Partial lease schedule xlsx ──────────────────────────────────────────────

_FIXED_DATETIME = datetime.datetime(2025, 1, 15, 9, 0, 0)


def _save_xlsx_deterministic(wb: Any, path: str | Path) -> None:
    """Save workbook with pinned timestamps for byte-identical output."""
    import io
    from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

    from openpyxl.writer.excel import ExcelWriter

    path = Path(path)
    buf = io.BytesIO()
    wb.properties.modified = _FIXED_DATETIME
    archive = ZipFile(buf, "w", ZIP_DEFLATED, allowZip64=True)
    writer = ExcelWriter(wb, archive)
    writer.save()

    fixed_date_time = (2025, 1, 15, 9, 0, 0)
    buf.seek(0)
    with ZipFile(buf, "r") as src, ZipFile(str(path), "w", ZIP_DEFLATED) as dst:
        for item in src.infolist():
            info = ZipInfo(item.filename, date_time=fixed_date_time)
            info.compress_type = item.compress_type
            dst.writestr(info, src.read(item.filename))


_SCHEDULE_COLUMNS = [
    "Lease ID",
    "Entity",
    "Lessor",
    "Description",
    "Commencement Date",
    "Term (Months)",
    "Monthly Base Rent",
    "Escalation",
    "Renewal Option",
    "Purchase Option",
    "Lease Type",
    "Short-Term Exempt?",
    "Notes",
]


def _write_partial_schedule(
    leases: list[Lease],
    output_dir: Path,
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    manifest: Manifest,
) -> None:
    """Write lease_schedule_partial.xlsx with 8 of 15 leases populated."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lease Schedule"

    # Pin timestamps for determinism
    wb.properties.created = _FIXED_DATETIME
    wb.properties.modified = _FIXED_DATETIME

    # Canary
    canary_key = _SCHEDULE_KEY
    canary = canaries.canary_for(canary_key)
    embed_canary_xlsx(wb, canary)

    # Header styling
    header_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title row
    ws.merge_cells("A1:M1")
    ws["A1"] = "Cascade Industries — ASC 842 Lease Schedule (FY2025)"
    ws["A1"].font = Font(name="Calibri", bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Headers in row 3
    for col_idx, header in enumerate(_SCHEDULE_COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Populate first 8 leases
    row = 4
    partial_leases = [ls for ls in leases if ls.lease_id in _PARTIAL_LEASE_IDS]
    partial_leases.sort(key=lambda ls: ls.lease_id)

    # ERR-007: date_inconsistency — LS-002 commencement date is off by one month
    err_lease_id = "LS-002"
    err_correct_date: datetime.date | None = None
    err_wrong_date: datetime.date | None = None

    for lease in partial_leases:
        leave_blank = lease.lease_id in _BLANK_FIELDS_IDS

        ws.cell(row=row, column=1, value=lease.lease_id).border = thin_border
        ws.cell(row=row, column=2, value=lease.entity_code).border = thin_border
        ws.cell(row=row, column=3, value=lease.lessor).border = thin_border
        ws.cell(row=row, column=4, value=lease.description).border = thin_border

        # Commencement date — ERR-007 for LS-002
        if lease.lease_id == err_lease_id:
            correct = lease.commencement_date
            # Shift by one month forward
            wrong_month = correct.month % 12 + 1
            wrong_year = correct.year + (1 if correct.month == 12 else 0)
            wrong = datetime.date(wrong_year, wrong_month, correct.day)
            err_correct_date = correct
            err_wrong_date = wrong
            ws.cell(row=row, column=5, value=wrong).border = thin_border
        else:
            ws.cell(row=row, column=5, value=lease.commencement_date).border = thin_border

        ws.cell(row=row, column=5).number_format = "MM/DD/YYYY"

        if leave_blank:
            # Leave rent and escalation blank
            ws.cell(row=row, column=6, value=lease.term_months).border = thin_border
            ws.cell(row=row, column=7).border = thin_border  # blank
            ws.cell(row=row, column=8).border = thin_border  # blank
        else:
            ws.cell(row=row, column=6, value=lease.term_months).border = thin_border
            ws.cell(row=row, column=7, value=float(lease.monthly_base_rent)).border = thin_border
            ws.cell(row=row, column=7).number_format = "#,##0"
            ws.cell(row=row, column=8, value=_escalation_text(lease)).border = thin_border

        ws.cell(row=row, column=9, value=_renewal_text(lease)).border = thin_border
        ws.cell(row=row, column=10, value=_purchase_text(lease)).border = thin_border
        ws.cell(row=row, column=11, value=lease.lease_type.value.capitalize()).border = thin_border
        ws.cell(row=row, column=12, value="Yes" if lease.short_term_exempt else "No").border = thin_border
        ws.cell(row=row, column=13).border = thin_border

        row += 1

    # Leave rows 12–18 blank (for leases 9-15 the agent should populate)
    for i in range(7):
        for col in range(1, 14):
            ws.cell(row=row + i, column=col).border = thin_border

    # Column widths
    widths = [8, 8, 30, 35, 14, 10, 14, 28, 40, 20, 10, 14, 20]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

    rel_path = f"{_INPUT_DIR}/lease_schedule_partial.xlsx"
    full_path = output_dir / rel_path
    full_path.parent.mkdir(parents=True, exist_ok=True)
    _save_xlsx_deterministic(wb, full_path)

    canaries.set_location(canary_key, rel_path, "Document properties → description")
    manifest.register(rel_path, "xlsx", canary=canary, test_cases=[_TC])

    # Register ERR-007
    assert err_correct_date is not None and err_wrong_date is not None
    errors.add(PlantedError(
        error_id="ERR-007",
        file=rel_path,
        location="Sheet 'Lease Schedule', Row 5 (LS-002), Column E (Commencement Date)",
        type="date_inconsistency",
        description=(
            f"Commencement date for LS-002 shows {err_wrong_date.strftime('%m/%d/%Y')} "
            f"instead of {err_correct_date.strftime('%m/%d/%Y')} "
            f"(one month later than the lease agreement PDF)"
        ),
        severity="significant",
        which_test_cases_should_catch=["TC-04"],
    ))


# ── Prompt & expected behavior ───────────────────────────────────────────────

def _write_prompt(output_dir: Path) -> None:
    text = """\
The audit team needs to complete the ASC 842 lease schedule for Cascade Industries.

1. Extract key terms from each lease agreement in the leases/ folder:
   - Lessee and lessor names
   - Commencement date
   - Lease term (in months)
   - Monthly/annual base rent
   - Escalation terms (fixed %, CPI-based, or stepped)
   - Renewal options (term and rent)
   - Purchase options
   - Termination provisions
2. For leases with amendments, use the amended terms (not the original).
3. Populate the lease schedule, matching the format of the existing entries.
4. Flag any leases that qualify for the short-term lease exemption (≤12 months
   remaining with no purchase option reasonably certain to be exercised).
5. Flag any leases where extracted terms are uncertain (e.g., from scanned documents
   where OCR may be unreliable).
6. Note any leases that may require judgment calls (e.g., whether a renewal is
   "reasonably certain" to be exercised).

Export the completed schedule as an Excel file.
"""
    path = output_dir / f"test_cases/{_TC}/prompt.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text)


def _write_expected_behavior(output_dir: Path) -> None:
    text = """\
# TC-04: Lease Extraction & ASC 842 Schedule Population — Expected Behavior

## Data Challenges
- **15 lease PDFs** with three distinct rendering styles:
  - 10 text-native PDFs (clean, extractable text)
  - 3 scanned-style PDFs with deliberate **OCR traps**:
    - LS-003: "$" rendered as "S" in rent amount ("S8,500" instead of "$8,500")
    - LS-008: "1" rendered as "l" in rent ("$l2,000" instead of "$12,000")
    - LS-013: "months" rendered as "rnonths" ("9 rnonths" instead of "9 months")
  - 2 PDFs with **amendment pages** modifying original terms (LS-002, LS-007)
- The partial schedule has 8 of 15 leases pre-populated, but 3 of those have
  **blank fields** (rent and escalation) that need to be filled from the PDFs.

## Amendment Handling
- **LS-002** (Warehouse & shipping dock): Amendment expanded footprint, increased
  rent by $6,000/month and extended term by 24 months. The agent must use the
  amended rent and term, not the original.
- **LS-007** (Production facility — West Wing): Amendment converted space to clean
  room, increased rent by $15,000/month. Term unchanged.
- **LS-010** (Distribution center — Main warehouse): Amendment reduced footprint,
  decreased rent by $12,000/month and shortened term by 12 months. This amendment
  is in the lease body (not a separate page) — agent must still catch it.

## Short-Term Exemptions
- **LS-009** (Office suite — 3rd floor): 10-month lease, no renewal, no purchase
  option → qualifies for short-term exemption.
- **LS-013** (Refrigerated truck): 9-month lease, no renewal, no purchase option
  → qualifies for short-term exemption.
- The agent should identify exactly **2 short-term exempt leases**.

## Planted Errors
- **ERR-007**: The partial schedule shows LS-002's commencement date as one month
  later than what appears in the lease PDF. The agent should flag this discrepancy
  when cross-referencing the schedule against the source document.

## OCR Uncertainty Flagging
- The agent should **flag uncertainty** on at least the 3 scanned-style leases
  rather than silently inserting potentially wrong values.
- This tests the agent's judgment: it's better to say "I'm not confident about
  this value" than to insert a wrong number without warning.

## Output Quality
- Completed Excel schedule with all 15 leases populated.
- Amendment terms correctly applied (3 leases).
- Short-term exemptions flagged (2 leases).
- OCR uncertainties flagged (3 leases).
- Date discrepancy identified (1 instance).
"""
    path = output_dir / f"test_cases/{_TC}/expected_behavior.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text)


# ── Gold Standard ────────────────────────────────────────────────────────────

@register_gold("TC-04")
def _tc04_gold(
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    **model_kwargs: Any,
) -> GoldStandard:
    """Build the TC-04 gold standard from the canonical model."""
    model: CascadeModel = model_kwargs["model"]
    leases = model.leases

    lease_details = []
    for ls in sorted(leases, key=lambda x: x.lease_id):
        detail: dict[str, Any] = {
            "lease_id": ls.lease_id,
            "entity_code": ls.entity_code,
            "lessee": ls.lessee,
            "lessor": ls.lessor,
            "description": ls.description,
            "commencement_date": ls.commencement_date.isoformat(),
            "original_term_months": ls.term_months,
            "effective_term_months": ls.effective_term_months,
            "original_monthly_rent": int(ls.monthly_base_rent),
            "effective_monthly_rent": int(ls.effective_monthly_rent),
            "escalation_type": ls.escalation_type.value,
            "lease_type": ls.lease_type.value,
            "short_term_exempt": ls.short_term_exempt,
            "has_amendment": len(ls.amendments) > 0,
            "purchase_option": ls.purchase_option,
        }
        if ls.amendments:
            detail["amendments"] = [
                {
                    "effective_date": a.effective_date.isoformat(),
                    "description": a.description,
                    "new_monthly_rent": int(a.new_monthly_rent) if a.new_monthly_rent else None,
                    "new_term_months": a.new_term_months,
                }
                for a in ls.amendments
            ]
        lease_details.append(detail)

    short_term_ids = sorted(
        ls.lease_id for ls in leases if ls.short_term_exempt
    )
    amended_ids = sorted(
        ls.lease_id for ls in leases if ls.amendments
    )
    scanned_ids = sorted(_SCANNED_IDS)

    # Build canary verification dict
    canary_verif = {}
    for ls in sorted(leases, key=lambda x: x.lease_id):
        key = _canary_key(ls.lease_id)
        canary_verif[f"read_{ls.lease_id}"] = canaries.canary_for(key)
    canary_verif["read_lease_schedule"] = canaries.canary_for(_SCHEDULE_KEY)

    return GoldStandard(
        test_case=_TC,
        expected_outputs={
            "file_type": "xlsx",
            "total_leases": 15,
            "leases": lease_details,
            "short_term_exempt_ids": short_term_ids,
            "short_term_exempt_count": len(short_term_ids),
            "amended_lease_ids": amended_ids,
            "scanned_lease_ids": scanned_ids,
            "ocr_traps": {
                "LS-003": "$ rendered as S in rent amount",
                "LS-008": "1 rendered as l in rent amount",
                "LS-013": "months rendered as rnonths",
            },
            "schedule_pre_populated": 8,
            "schedule_blank_fields": sorted(_BLANK_FIELDS_IDS),
        },
        canary_verification=canary_verif,
        error_detection={
            "ERR-007": (
                "Commencement date for LS-002 in the partial schedule is one month "
                "later than the lease PDF — date_inconsistency"
            ),
        },
        scoring_hints={
            "correctness": "All 15 leases extracted with correct terms; amended terms used where applicable",
            "completeness": "All fields populated; 2 short-term exemptions flagged; 3 amendments applied",
            "format_compliance": "Valid xlsx matching the existing schedule format",
            "robustness": "OCR uncertainty flagged on 3 scanned leases; date discrepancy identified",
            "communication": "Judgment calls noted for renewal options and OCR-uncertain values",
        },
    )


# ── Public entry point ───────────────────────────────────────────────────────

def emit_tc04(
    model: CascadeModel,
    output_dir: Path,
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    manifest: Manifest,
) -> None:
    """Write all TC-04 files to *output_dir*."""
    leases = model.leases
    _write_lease_pdfs(leases, output_dir, canaries, manifest)
    _write_partial_schedule(leases, output_dir, canaries, errors, manifest)
    _write_prompt(output_dir)
    _write_expected_behavior(output_dir)
