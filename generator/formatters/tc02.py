"""Formatter: TC-02 — Bank Reconciliation & Confirmation Matching (Audit, Complex).

Emits:
- test_cases/TC-02/input_files/bank_statement_dec2025.csv
  340-row bank CSV with cryptic descriptions, running balance
- test_cases/TC-02/input_files/cascade_gl_cash_dec2025.xlsx
  GL cash account detail for December with company-style descriptions
- test_cases/TC-02/input_files/bank_confirmation_fy2025.pdf
  Bank confirmation letter (text-native PDF) confirming balance as of 12/31/2025
- test_cases/TC-02/prompt.md
- test_cases/TC-02/expected_behavior.md
- gold_standards/TC-02_gold.json

Uses the canonical BankModel — never hardcodes numbers.
"""

from __future__ import annotations

import datetime
import io
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.writer.excel import ExcelWriter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from generator.canaries import (
    CanaryRegistry,
    embed_canary_csv_comment,
    embed_canary_xlsx,
)
from generator.errors import ErrorRegistry, PlantedError, mismatch_total, transpose_digits
from generator.golds.framework import GoldStandard, register_gold
from generator.manifest import Manifest
from generator.model.bank import (
    ADJUSTED_BALANCE,
    BANK_INTEREST,
    BANK_SERVICE_CHARGES,
    CONFIRMATION_BALANCE,
    DEPOSITS_IN_TRANSIT,
    GL_ENDING_BALANCE,
    OUTSTANDING_CHECKS,
    TOTAL_DEPOSITS_IN_TRANSIT,
    TOTAL_OUTSTANDING_CHECKS,
)
from generator.model.build import CascadeModel
from generator.noise import ExclusionZone, apply_csv_noise, make_noise_rng
from generator.scenario_context import ScenarioContext

# ── Constants ────────────────────────────────────────────────────────────────

_TC = "TC-02"
_INPUT_DIR = f"test_cases/{_TC}/input_files"

# Fixed datetime for xlsx metadata (determinism)
_FIXED_DATETIME = datetime.datetime(2025, 3, 15, 9, 0, 0)


# ── Helpers ──────────────────────────────────────────────────────────────────


def _save_xlsx_deterministic(wb: openpyxl.Workbook, path: str | Path) -> None:
    """Save workbook with pinned timestamps for byte-identical output."""
    path = Path(path)
    wb.properties.modified = _FIXED_DATETIME
    wb.properties.created = _FIXED_DATETIME

    buf = io.BytesIO()
    archive = ZipFile(buf, "w", ZIP_DEFLATED, allowZip64=True)
    writer = ExcelWriter(wb, archive)
    writer.save()

    fixed_date_time = (2025, 3, 15, 9, 0, 0)
    buf.seek(0)
    with ZipFile(buf, "r") as src, ZipFile(str(path), "w", ZIP_DEFLATED) as dst:
        for item in src.infolist():
            info = ZipInfo(item.filename, date_time=fixed_date_time)
            info.compress_type = item.compress_type
            dst.writestr(info, src.read(item.filename))


def _fmt_amount(d: Decimal) -> str:
    """Format a Decimal as a string with 2 decimal places for CSV."""
    return str(d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


# ── Bank Statement CSV ──────────────────────────────────────────────────────


def _write_bank_csv(
    model: CascadeModel,
    output_dir: Path,
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    manifest: Manifest,
) -> None:
    """Write bank_statement_dec2025.csv — 340 transactions."""
    bank = model.bank
    assert bank is not None

    file_key = "bank_statement_dec2025"
    canary_code = canaries.canary_for(file_key)
    canary_line = embed_canary_csv_comment(canary_code)

    path = output_dir / _INPUT_DIR / "bank_statement_dec2025.csv"
    path.parent.mkdir(parents=True, exist_ok=True)

    # ERR-002: transpose digits in the 5th transaction's amount
    err_txn_idx = 4  # 0-based → 5th transaction
    correct_amount = bank.bank_transactions[err_txn_idx].amount
    corrupt_amount = Decimal(str(transpose_digits(int(correct_amount))))
    errors.add(PlantedError(
        error_id="ERR-002",
        file=f"{_INPUT_DIR}/bank_statement_dec2025.csv",
        location="Row 6 (transaction 5), Amount column",
        type="transposed_digits",
        description=(
            f"Bank transaction amount shows ${int(corrupt_amount):,} "
            f"instead of ${int(correct_amount):,}"
        ),
        severity="material",
        which_test_cases_should_catch=["TC-02"],
    ))

    lines: list[str] = [canary_line]
    # Header comment with bank info
    lines.append("# First National Bank of Oregon — Account Statement\n")
    lines.append("# Account: Cascade Industries, Inc. — Operating Account #4782-0091\n")
    lines.append("# Statement Period: December 1 - 31, 2025\n")
    lines.append("# Beginning Balance: $" + f"{bank.bank_starting_balance:,.2f}\n")
    lines.append("#\n")
    lines.append("Date,Description,Amount,Running Balance\n")

    for i, txn in enumerate(bank.bank_transactions):
        date_str = txn.date.strftime("%m/%d/%Y")
        amt = corrupt_amount if i == err_txn_idx else txn.amount
        amt_str = _fmt_amount(amt)
        bal_str = _fmt_amount(txn.running_balance)
        # Escape description if it contains commas
        desc = txn.description
        if "," in desc:
            desc = f'"{desc}"'
        lines.append(f"{date_str},{desc},{amt_str},{bal_str}\n")

    # ── Controlled noise (csv_stdlib family pilot) ───────────────────
    # Protect canary line (index 0), comment lines (1-5), header (6),
    # and the ERR-002 transaction row.
    err_row_idx = 7 + err_txn_idx  # 7 = first data row
    excl = ExclusionZone(rows={0, 1, 2, 3, 4, 5, 6, err_row_idx})
    noise_rng = make_noise_rng(
        ScenarioContext(seed=42), _TC, "bank_statement_dec2025",
    )
    lines = apply_csv_noise(lines, noise_rng, excl)

    path.write_text("".join(lines))

    canaries.set_location(
        file_key,
        f"{_INPUT_DIR}/bank_statement_dec2025.csv",
        "First line comment",
    )
    manifest.register(
        f"{_INPUT_DIR}/bank_statement_dec2025.csv",
        "csv",
        canary=canary_code,
        test_cases=[_TC],
    )


# ── GL Cash Detail (xlsx) ──────────────────────────────────────────────────


def _write_gl_cash_xlsx(
    model: CascadeModel,
    output_dir: Path,
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    manifest: Manifest,
) -> None:
    """Write cascade_gl_cash_dec2025.xlsx — GL cash account detail."""
    bank = model.bank
    assert bank is not None

    file_key = "cascade_gl_cash_dec2025"
    canary_code = canaries.canary_for(file_key)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Detail - 1010"

    # Embed canary
    loc = embed_canary_xlsx(wb, canary_code)

    # ── Header ──────────────────────────────────────────────────────
    ws["A1"] = "Cascade Industries, Inc."
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "General Ledger — Cash (Account 1010)"
    ws["A2"].font = Font(italic=True, size=11)
    ws["A3"] = "December 2025"
    ws["A3"].font = Font(size=10)

    # Starting balance row
    ws["A5"] = "Beginning Balance"
    ws["A5"].font = Font(bold=True, size=10)
    ws["F5"] = int(bank.gl_starting_balance)
    ws["F5"].font = Font(bold=True, size=10)
    ws["F5"].number_format = '#,##0'

    # Column headers
    header_row = 7
    headers = ["Date", "Reference", "Description", "Debit", "Credit", "Balance"]
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A5276")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ── Data rows ───────────────────────────────────────────────────
    money_fmt = '#,##0'
    data_font = Font(size=10)
    row = header_row + 1

    for entry in bank.gl_entries:
        ws.cell(row=row, column=1, value=entry.date.strftime("%m/%d/%Y")).font = data_font
        ws.cell(row=row, column=2, value=entry.reference).font = data_font
        ws.cell(row=row, column=3, value=entry.description).font = data_font

        debit_cell = ws.cell(
            row=row, column=4,
            value=int(entry.debit) if entry.debit > 0 else None,
        )
        debit_cell.font = data_font
        debit_cell.number_format = money_fmt

        credit_cell = ws.cell(
            row=row, column=5,
            value=int(entry.credit) if entry.credit > 0 else None,
        )
        credit_cell.font = data_font
        credit_cell.number_format = money_fmt

        bal_cell = ws.cell(row=row, column=6, value=int(entry.running_balance))
        bal_cell.font = data_font
        bal_cell.number_format = money_fmt

        row += 1

    # Ending balance row
    # ERR-004: mismatched total — ending balance off by ~$3,500 (one omitted entry)
    correct_gl_ending = int(bank.gl_ending_balance)
    corrupt_gl_ending = int(mismatch_total(correct_gl_ending, -3_500))
    errors.add(PlantedError(
        error_id="ERR-004",
        file=f"{_INPUT_DIR}/cascade_gl_cash_dec2025.xlsx",
        location="Sheet 'Cash Detail - 1010', Ending Balance row, column F",
        type="mismatched_total",
        description=(
            f"GL ending balance shows ${corrupt_gl_ending:,} "
            f"instead of ${correct_gl_ending:,}"
        ),
        severity="material",
        which_test_cases_should_catch=["TC-02"],
    ))

    row += 1
    ws.cell(row=row, column=1, value="Ending Balance").font = Font(bold=True, size=10)
    end_cell = ws.cell(row=row, column=6, value=corrupt_gl_ending)
    end_cell.font = Font(bold=True, size=10)
    end_cell.number_format = money_fmt
    # Double underline
    end_cell.border = Border(bottom=Side(style="double"))

    # ── Column widths ───────────────────────────────────────────────
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18

    # ── Save ────────────────────────────────────────────────────────
    path = output_dir / _INPUT_DIR / "cascade_gl_cash_dec2025.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    _save_xlsx_deterministic(wb, path)

    canaries.set_location(
        file_key,
        f"{_INPUT_DIR}/cascade_gl_cash_dec2025.xlsx",
        loc,
    )
    manifest.register(
        f"{_INPUT_DIR}/cascade_gl_cash_dec2025.xlsx",
        "xlsx",
        canary=canary_code,
        test_cases=[_TC],
    )


# ── Bank Confirmation PDF ──────────────────────────────────────────────────


def _write_bank_confirmation_pdf(
    model: CascadeModel,
    output_dir: Path,
    canaries: CanaryRegistry,
    manifest: Manifest,
) -> None:
    """Write bank_confirmation_fy2025.pdf — bank confirmation letter."""
    bank = model.bank
    assert bank is not None

    file_key = "bank_confirmation_fy2025"
    canary_code = canaries.canary_for(file_key)

    path = output_dir / _INPUT_DIR / "bank_confirmation_fy2025.pdf"
    path.parent.mkdir(parents=True, exist_ok=True)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ConfTitle", parent=styles["Title"],
        fontSize=14, spaceAfter=6, textColor=colors.HexColor("#1A3C6E"),
    )
    heading_style = ParagraphStyle(
        "ConfHeading", parent=styles["Heading2"],
        fontSize=11, spaceAfter=6, textColor=colors.HexColor("#1A3C6E"),
    )
    body_style = ParagraphStyle(
        "ConfBody", parent=styles["Normal"],
        fontSize=10, spaceAfter=6, leading=14,
    )
    small_style = ParagraphStyle(
        "ConfSmall", parent=styles["Normal"],
        fontSize=8, spaceAfter=4, textColor=colors.HexColor("#666666"),
    )

    story: list = []

    # Bank letterhead
    story.append(Paragraph("FIRST NATIONAL BANK OF OREGON", title_style))
    story.append(Paragraph(
        "Corporate Banking Division<br/>"
        "100 SW Main Street, Suite 3200<br/>"
        "Portland, Oregon 97204",
        body_style,
    ))
    story.append(Spacer(1, 0.3 * inch))

    # Confirmation header
    story.append(Paragraph("STANDARD BANK CONFIRMATION", heading_style))
    story.append(Spacer(1, 0.2 * inch))

    # Date and addressee
    story.append(Paragraph("January 15, 2026", body_style))
    story.append(Spacer(1, 0.15 * inch))
    story.append(Paragraph(
        "Mitchell &amp; Associates LLP<br/>"
        "Attn: Audit Engagement Team<br/>"
        "1200 NW Couch Street, Suite 800<br/>"
        "Portland, Oregon 97209",
        body_style,
    ))
    story.append(Spacer(1, 0.2 * inch))

    story.append(Paragraph("Re: Cascade Industries, Inc.", body_style))
    story.append(Paragraph(
        "Dear Mitchell &amp; Associates LLP:",
        body_style,
    ))
    story.append(Spacer(1, 0.1 * inch))

    # Confirmation body
    confirmation_bal = f"${int(bank.confirmation_balance):,}"
    story.append(Paragraph(
        "In connection with your audit of the financial statements of "
        "<b>Cascade Industries, Inc.</b>, we confirm the following information "
        "as of the close of business on <b>December 31, 2025</b>:",
        body_style,
    ))
    story.append(Spacer(1, 0.15 * inch))

    # Account details table
    acct_data = [
        ["Account Name", "Account Number", "Type", "Balance"],
        [
            "Cascade Industries — Operating",
            "4782-0091",
            "Commercial Checking",
            confirmation_bal,
        ],
    ]
    acct_table = Table(acct_data, colWidths=[2.2 * inch, 1.5 * inch, 1.5 * inch, 1.3 * inch])
    acct_table.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 9),
        ("FONT", (0, 1), (-1, -1), "Helvetica", 9),
        ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#1A3C6E")),
        ("LINEBELOW", (0, -1), (-1, -1), 0.5, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EEF4")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(acct_table)
    story.append(Spacer(1, 0.2 * inch))

    # Loans and other
    story.append(Paragraph(
        "<b>Loans and Other Direct/Contingent Liabilities:</b> None",
        body_style,
    ))
    story.append(Spacer(1, 0.15 * inch))

    story.append(Paragraph(
        "This confirmation is issued for audit purposes only and should not be "
        "used for any other purpose. The information contained herein is "
        "confirmed to be correct as of the date indicated above.",
        body_style,
    ))
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph("Sincerely,", body_style))
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(
        "<b>Jennifer A. Thornton</b><br/>"
        "Senior Vice President, Corporate Banking<br/>"
        "First National Bank of Oregon",
        body_style,
    ))
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(
        f"Confirmation Reference: FNBO-CONF-2025-{canary_code}",
        small_style,
    ))

    # Build PDF
    doc = SimpleDocTemplate(
        str(path),
        pagesize=letter,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
        leftMargin=inch,
        rightMargin=inch,
        title="Bank Confirmation — Cascade Industries",
        author=f"CANARY: {canary_code}",
        creator="First National Bank of Oregon",
        invariant=True,
    )
    doc.build(story)

    canaries.set_location(
        file_key,
        f"{_INPUT_DIR}/bank_confirmation_fy2025.pdf",
        "PDF metadata → Author; also in Confirmation Reference footer",
    )
    manifest.register(
        f"{_INPUT_DIR}/bank_confirmation_fy2025.pdf",
        "pdf",
        canary=canary_code,
        test_cases=[_TC],
    )


# ── Prompt & Expected Behavior ──────────────────────────────────────────────


def _write_prompt(output_dir: Path) -> None:
    """Write test_cases/TC-02/prompt.md per spec."""
    text = """\
Perform a bank reconciliation for Cascade Industries as of December 31, 2025.

1. Match transactions between the bank statement and the general ledger.
   Use fuzzy matching on dates (allow +/-2 business days) and amounts (exact match).
2. Identify all outstanding checks (in GL but not on bank statement).
3. Identify all deposits in transit (in GL but not on bank statement).
4. Identify any bank charges or interest not recorded in the GL.
5. Prepare a standard bank reconciliation schedule showing:
   - Balance per bank statement
   - Add: deposits in transit
   - Less: outstanding checks
   - Adjusted bank balance
   - Balance per GL
   - Add/Less: adjustments needed
   - Adjusted book balance
6. Verify that the adjusted balances agree and tie to the bank confirmation letter.
7. Export as an Excel workpaper.
"""
    path = output_dir / f"test_cases/{_TC}/prompt.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text)


def _write_expected_behavior(output_dir: Path) -> None:
    """Write test_cases/TC-02/expected_behavior.md per spec."""
    text = """\
# TC-02: Bank Reconciliation & Confirmation Matching — Expected Behavior

## Data Challenges
- The bank statement uses **cryptic abbreviations** (e.g., "ACH CR", "CHK#", "WR DB")
  that must be matched to the GL's full company-style descriptions.
- Transaction dates differ by up to **2 business days** between bank and GL
  due to processing delays — the agent must use fuzzy date matching with exact
  amount matching.
- The bank statement has **340 rows** — a realistic volume requiring efficient
  matching, not manual review.

## Reconciling Items
- **4 outstanding checks** in the GL that have not yet cleared the bank.
- **2 deposits in transit** recorded in the GL but not yet credited by the bank.
- **Bank interest** credited on the bank statement but not yet recorded in the GL.
- **Bank service charges** debited on the bank statement but not yet recorded in the GL.

## Reconciliation Verification
- The adjusted bank balance and adjusted book balance must **agree at $4,287,331**.
- The bank confirmation letter confirms a balance of **$4,312,117** — this equals
  the bank statement ending balance. The agent should verify that the confirmation
  ties to the bank statement (not to the GL or adjusted balance).
- The reconciliation must be mathematically correct:
  - Bank ending ($4,312,117) + deposits in transit - outstanding checks = $4,287,331
  - GL ending + interest earned - service charges = $4,287,331

## Output Quality
- The output workpaper should contain a clear bank reconciliation schedule.
- Outstanding checks and deposits in transit should be individually listed.
- Bank-to-GL matching should be documented or summarized.
"""
    path = output_dir / f"test_cases/{_TC}/expected_behavior.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text)


# ── Gold Standard ────────────────────────────────────────────────────────────


@register_gold("TC-02")
def _tc02_gold(
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    **model_kwargs: Any,
) -> GoldStandard:
    """Build the TC-02 gold standard from the canonical model."""
    model: CascadeModel = model_kwargs["model"]
    bank = model.bank
    assert bank is not None

    outstanding_checks_detail = [
        {"description": desc, "amount": int(amt), "date": dt.isoformat()}
        for desc, amt, dt in OUTSTANDING_CHECKS
    ]
    deposits_in_transit_detail = [
        {"description": desc, "amount": int(amt), "date": dt.isoformat()}
        for desc, amt, dt in DEPOSITS_IN_TRANSIT
    ]

    return GoldStandard(
        test_case="TC-02",
        expected_outputs={
            "file_type": "xlsx",
            "reconciliation": {
                "bank_ending_balance": int(CONFIRMATION_BALANCE),
                "deposits_in_transit": int(TOTAL_DEPOSITS_IN_TRANSIT),
                "outstanding_checks": int(TOTAL_OUTSTANDING_CHECKS),
                "adjusted_bank_balance": int(ADJUSTED_BALANCE),
                "gl_ending_balance": int(GL_ENDING_BALANCE),
                "bank_interest": int(BANK_INTEREST),
                "bank_service_charges": int(BANK_SERVICE_CHARGES),
                "adjusted_book_balance": int(ADJUSTED_BALANCE),
                "balances_agree": True,
            },
            "outstanding_checks_detail": outstanding_checks_detail,
            "deposits_in_transit_detail": deposits_in_transit_detail,
            "confirmation_balance": int(CONFIRMATION_BALANCE),
            "confirmation_ties_to_bank_ending": True,
            "transaction_count": {
                "bank_statement_rows": 340,
                "outstanding_checks": 4,
                "deposits_in_transit": 2,
                "bank_only_items": 2,
            },
        },
        canary_verification={
            "read_bank_statement": canaries.canary_for("bank_statement_dec2025"),
            "read_gl_cash_detail": canaries.canary_for("cascade_gl_cash_dec2025"),
            "read_bank_confirmation": canaries.canary_for("bank_confirmation_fy2025"),
        },
        error_detection={},
        scoring_hints={
            "correctness": "Adjusted balances must agree at $4,287,331 exactly",
            "completeness": "All 4 outstanding checks and 2 deposits in transit identified",
            "format_compliance": "Valid xlsx with clear reconciliation schedule",
            "robustness": "Handled fuzzy date matching and cryptic bank descriptions",
            "communication": "Verified confirmation ties to bank statement ending balance",
        },
    )


# ── Public entry point ───────────────────────────────────────────────────────


def emit_tc02(
    model: CascadeModel,
    output_dir: Path,
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    manifest: Manifest,
    **kwargs: object,
) -> None:
    """Write all TC-02 files to *output_dir*."""
    _write_bank_csv(model, output_dir, canaries, errors, manifest)
    _write_gl_cash_xlsx(model, output_dir, canaries, errors, manifest)
    _write_bank_confirmation_pdf(model, output_dir, canaries, manifest)
    _write_prompt(output_dir)
    _write_expected_behavior(output_dir)
