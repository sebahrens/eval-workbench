"""Formatter: TC-09 — Transfer Pricing Documentation (Tax, Complex).

Emits:
- test_cases/TC-09/input_files/intercompany_transactions_fy2025.xlsx
  All FY2025 intercompany transactions with entity, counterparty, type,
  volume, and pricing (derived from canonical model).
- test_cases/TC-09/input_files/comparable_companies.xlsx
  Financial data for 12 comparable companies; 2 should be rejected
  (one SIC mismatch, one financial distress).
- test_cases/TC-09/input_files/tp_report_fy2024.pdf
  Prior year transfer pricing report (42 pages) with functional analysis,
  economic analysis, and benchmarking results.
- test_cases/TC-09/prompt.md
- test_cases/TC-09/expected_behavior.md
- gold_standards/TC-09_gold.json

Planted errors:
- ERR-015 (wrong_entity): One IC transaction shows wrong subsidiary name.
- ERR-017 (date_inconsistency): One IC transaction has invoice date off by one month.
The adversarial element is the services transfer at 11.2% operating margin
(outside the 4.2%–8.7% IQR).
Uses the canonical model — never hardcodes numbers.
"""

from __future__ import annotations

import datetime
import io
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from generator.canaries import (
    CanaryRegistry,
    embed_canary_xlsx,
)
from generator.errors import (
    ErrorRegistry,
    PlantedError,
    date_inconsistency,
    wrong_entity,
)
from generator.golds.framework import GoldStandard, register_gold
from generator.manifest import Manifest
from generator.model.build import CascadeModel
from generator.model.entities import ENTITIES
from generator.model.intercompany import (
    IC_LOAN_PRINCIPAL,
    IC_LOAN_RATE,
    MANAGEMENT_FEE_PCT,
    RAW_MATERIALS_MARKUP,
    SERVICES_OPERATING_MARGIN,
    generate_ic_transactions,
)

# ── Constants ────────────────────────────────────────────────────────────────

_TC = "TC-09"
_INPUT_DIR = f"test_cases/{_TC}/input_files"

_FIXED_DATETIME = datetime.datetime(2025, 3, 15, 9, 0, 0)
_FIXED_ZIP_DT = (2025, 3, 15, 9, 0, 0)

# ── Comparable companies ─────────────────────────────────────────────────────
# 12 companies: 10 valid, 2 to reject.
# Revenue and operating income are chosen so that oi/rev yields exact margins.
# Sorted accepted margins: [2.8, 3.5, 4.0, 4.8, 5.6, 6.8, 8.4, 8.8, 9.5, 11.2]
# QUARTILE.INC Q1 = 4.0 + 0.25*(4.8-4.0) = 4.2%  ✓
# QUARTILE.INC Q3 = 8.4 + 0.75*(8.8-8.4) = 8.7%  ✓

# Full financial profiles for all 12 companies (for the xlsx).
# (name, sic, revenue_M, cogs_M, opex_M, oper_income_M, total_assets_M, is_rejected)
# COGS and opex are calibrated so rev - cogs - opex = oper_income exactly.
_COMPANY_PROFILES: list[tuple[str, str, int, int, int, int, int, bool]] = [
    # Accepted companies (exact margins)
    ("Apex Manufacturing Corp", "3599", 250, 170, 73, 7, 275, False),       # 2.8%
    ("Continental Components Inc", "3599", 200, 135, 58, 7, 225, False),     # 3.5%
    ("Delta Precision Holdings", "3599", 300, 205, 83, 12, 330, False),      # 4.0%
    ("EastPoint Industries LLC", "3599", 250, 168, 70, 12, 275, False),      # 4.8%
    ("Falcon Industrial Group", "3599", 125, 82, 36, 7, 138, False),         # 5.6%
    ("Guardian Materials Co", "3599", 250, 165, 68, 17, 275, False),         # 6.8%
    ("Horizon Manufacturing Ltd", "3599", 250, 163, 66, 21, 275, False),     # 8.4%
    ("Ironclad Precision Inc", "3599", 250, 159, 69, 22, 275, False),        # 8.8%
    ("Jefferson Industrial Corp", "3599", 200, 130, 51, 19, 220, False),     # 9.5%
    ("Keystone Components LLC", "3599", 250, 160, 62, 28, 275, False),       # 11.2%
    # Rejected companies
    ("TechVantage Solutions Inc", "7372", 165, 58, 77, 30, 182, True),       # 18.2% wrong SIC
    ("Meridian Industrial Holdings", "3599", 150, 118, 44, -12, 310, True),  # -8.0% distress
]

# ── Deterministic save helpers ───────────────────────────────────────────────


def _save_xlsx_deterministic(wb: openpyxl.Workbook, path: str | Path) -> None:
    """Save workbook with pinned timestamps and fixed zip entry dates."""
    from openpyxl.writer.excel import ExcelWriter

    path = Path(path)
    wb.properties.modified = _FIXED_DATETIME

    buf = io.BytesIO()
    archive = ZipFile(buf, "w", ZIP_DEFLATED, allowZip64=True)
    writer = ExcelWriter(wb, archive)
    writer.save()

    buf.seek(0)
    with ZipFile(buf, "r") as src, ZipFile(str(path), "w", ZIP_DEFLATED) as dst:
        for item in src.infolist():
            info = ZipInfo(item.filename, date_time=_FIXED_ZIP_DT)
            info.compress_type = item.compress_type
            dst.writestr(info, src.read(item.filename))


def _whole_dollars(d: Decimal) -> int:
    """Round a Decimal to whole dollars."""
    return int(d.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


# ── Styling helpers ──────────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_ALIGN = Alignment(horizontal="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_NUMBER_FMT = '#,##0'
_PCT_FMT = '0.0%'
_MONEY_FMT = '#,##0'


def _style_header(ws: Any, row: int, col_count: int) -> None:
    """Apply header styling to a row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _style_data_cell(cell: Any, fmt: str = "") -> None:
    """Apply data cell styling."""
    cell.border = _THIN_BORDER
    if fmt:
        cell.number_format = fmt


# ── IC Transactions xlsx ─────────────────────────────────────────────────────


def _get_fy2025_ic_transactions(model: CascadeModel) -> list[Any]:
    """Regenerate IC transactions and filter to FY2025."""
    totals: dict[tuple[str, int, int], Decimal] = {}
    for r in model.revenue_records:
        key = (r.entity_code, r.year, r.month)
        totals[key] = totals.get(key, Decimal(0)) + r.revenue

    all_txns = generate_ic_transactions(totals)
    return [t for t in all_txns if t.date.year == 2025]


def _write_ic_transactions_xlsx(
    model: CascadeModel,
    output_dir: Path,
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    manifest: Manifest,
) -> list[Any]:
    """Write intercompany_transactions_fy2025.xlsx and return FY2025 transactions."""
    fy25_txns = _get_fy2025_ic_transactions(model)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IC Transactions FY2025"

    # Headers
    headers = [
        "Date", "Seller Entity", "Seller Name", "Buyer Entity", "Buyer Name",
        "Transaction Type", "Amount ($)", "Description",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    # ── ERR-015 / ERR-017: pick target rows for planted errors ────────
    _ERR015_ROW_IDX = 4   # 0-based index into fy25_txns
    _ERR017_ROW_IDX = 7   # 0-based index into fy25_txns

    # Data rows
    for i, tx in enumerate(fy25_txns, 2):
        # -- Date (column 1) — ERR-017: date_inconsistency on target row --
        date_str = tx.date.strftime("%Y-%m-%d")
        if (i - 2) == _ERR017_ROW_IDX:
            correct_date = tx.date
            wrong_month = (correct_date.month % 12) + 1  # shift by +1 month
            wrong_year = correct_date.year + (1 if correct_date.month == 12 else 0)
            wrong_date_obj = correct_date.replace(year=wrong_year, month=wrong_month)
            date_str = date_inconsistency(
                correct_date.strftime("%Y-%m-%d"),
                wrong_date_obj.strftime("%Y-%m-%d"),
            )
            errors.add(PlantedError(
                error_id="ERR-017",
                file=f"{_INPUT_DIR}/intercompany_transactions_fy2025.xlsx",
                location=(
                    f"Sheet 'IC Transactions FY2025', Row {i}, Column A (Date)"
                ),
                type="date_inconsistency",
                description=(
                    f"Invoice date shows {wrong_date_obj.strftime('%m/%d/%Y')} "
                    f"instead of {correct_date.strftime('%m/%d/%Y')} "
                    f"(off by one month)"
                ),
                severity="immaterial",
                which_test_cases_should_catch=["TC-09"],
            ))
        ws.cell(row=i, column=1, value=date_str)
        _style_data_cell(ws.cell(row=i, column=1))

        ws.cell(row=i, column=2, value=tx.seller_entity)
        _style_data_cell(ws.cell(row=i, column=2))

        # -- Seller Name (column 3) — ERR-015: wrong_entity on target row --
        seller_name = ENTITIES[tx.seller_entity].name
        if (i - 2) == _ERR015_ROW_IDX:
            correct_name = seller_name
            wrong_name = "Cascade Precision Manufacturing"
            seller_name = wrong_entity(correct_name, wrong_name)
            errors.add(PlantedError(
                error_id="ERR-015",
                file=f"{_INPUT_DIR}/intercompany_transactions_fy2025.xlsx",
                location=(
                    f"Sheet 'IC Transactions FY2025', Row {i}, "
                    f"Column C (Seller Name)"
                ),
                type="wrong_entity",
                description=(
                    f"Seller name shows '{wrong_name}' "
                    f"instead of '{correct_name}'"
                ),
                severity="material",
                which_test_cases_should_catch=["TC-09"],
            ))
        ws.cell(row=i, column=3, value=seller_name)
        _style_data_cell(ws.cell(row=i, column=3))

        ws.cell(row=i, column=4, value=tx.buyer_entity)
        _style_data_cell(ws.cell(row=i, column=4))

        ws.cell(row=i, column=5, value=ENTITIES[tx.buyer_entity].name)
        _style_data_cell(ws.cell(row=i, column=5))

        ws.cell(row=i, column=6, value=tx.tx_type)
        _style_data_cell(ws.cell(row=i, column=6))

        ws.cell(row=i, column=7, value=_whole_dollars(tx.amount))
        _style_data_cell(ws.cell(row=i, column=7), _MONEY_FMT)

        ws.cell(row=i, column=8, value=tx.description)
        _style_data_cell(ws.cell(row=i, column=8))

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 50

    # Summary sheet
    ws2 = wb.create_sheet("Summary by Type")
    summary_headers = ["Transaction Type", "Count", "Total Amount ($)", "Avg Amount ($)"]
    for col, h in enumerate(summary_headers, 1):
        ws2.cell(row=1, column=col, value=h)
    _style_header(ws2, 1, len(summary_headers))

    # Aggregate by type
    type_totals: dict[str, tuple[int, Decimal]] = {}
    for tx in fy25_txns:
        count, total = type_totals.get(tx.tx_type, (0, Decimal(0)))
        type_totals[tx.tx_type] = (count + 1, total + tx.amount)

    for i, tx_type in enumerate(sorted(type_totals.keys()), 2):
        count, total = type_totals[tx_type]
        ws2.cell(row=i, column=1, value=tx_type)
        _style_data_cell(ws2.cell(row=i, column=1))

        ws2.cell(row=i, column=2, value=count)
        _style_data_cell(ws2.cell(row=i, column=2))

        ws2.cell(row=i, column=3, value=_whole_dollars(total))
        _style_data_cell(ws2.cell(row=i, column=3), _MONEY_FMT)

        avg = (total / count).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        ws2.cell(row=i, column=4, value=_whole_dollars(avg))
        _style_data_cell(ws2.cell(row=i, column=4), _MONEY_FMT)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 18

    # Canary
    canary = canaries.canary_for("tc09_ic_transactions")
    loc = embed_canary_xlsx(wb, canary)
    canaries.set_location(
        "tc09_ic_transactions",
        f"{_INPUT_DIR}/intercompany_transactions_fy2025.xlsx",
        loc,
    )

    wb.properties.created = _FIXED_DATETIME
    file_path = output_dir / _INPUT_DIR / "intercompany_transactions_fy2025.xlsx"
    file_path.parent.mkdir(parents=True, exist_ok=True)
    _save_xlsx_deterministic(wb, file_path)

    manifest.register(f"{_INPUT_DIR}/intercompany_transactions_fy2025.xlsx", "xlsx")

    return fy25_txns


# ── Comparable companies xlsx ────────────────────────────────────────────────


def _write_comparables_xlsx(
    output_dir: Path,
    canaries: CanaryRegistry,
    manifest: Manifest,
) -> None:
    """Write comparable_companies.xlsx with 12 companies."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparable Companies"

    headers = [
        "Company Name", "SIC Code", "Revenue ($M)", "COGS ($M)",
        "Operating Expenses ($M)", "Operating Income ($M)",
        "Total Assets ($M)", "Operating Margin",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    # Write all 12 companies (shuffled order — present mixed, not sorted by margin)
    # Use a fixed order that interleaves accepted and rejected
    display_order = [
        0, 10, 3, 6, 1, 11, 8, 4, 9, 2, 7, 5,
    ]  # indices into _COMPANY_PROFILES

    for row_idx, profile_idx in enumerate(display_order, 2):
        name, sic, rev, cogs, opex, oper_inc, assets, _rejected = (
            _COMPANY_PROFILES[profile_idx]
        )
        margin = oper_inc / rev if rev else 0

        ws.cell(row=row_idx, column=1, value=name)
        _style_data_cell(ws.cell(row=row_idx, column=1))

        ws.cell(row=row_idx, column=2, value=sic)
        _style_data_cell(ws.cell(row=row_idx, column=2))

        ws.cell(row=row_idx, column=3, value=rev)
        _style_data_cell(ws.cell(row=row_idx, column=3), _NUMBER_FMT)

        ws.cell(row=row_idx, column=4, value=cogs)
        _style_data_cell(ws.cell(row=row_idx, column=4), _NUMBER_FMT)

        ws.cell(row=row_idx, column=5, value=opex)
        _style_data_cell(ws.cell(row=row_idx, column=5), _NUMBER_FMT)

        ws.cell(row=row_idx, column=6, value=oper_inc)
        _style_data_cell(ws.cell(row=row_idx, column=6), _NUMBER_FMT)

        ws.cell(row=row_idx, column=7, value=assets)
        _style_data_cell(ws.cell(row=row_idx, column=7), _NUMBER_FMT)

        ws.cell(row=row_idx, column=8, value=round(margin, 4))
        _style_data_cell(ws.cell(row=row_idx, column=8), _PCT_FMT)

    # Column widths
    for col_letter, width in [
        ("A", 30), ("B", 12), ("C", 15), ("D", 15),
        ("E", 20), ("F", 20), ("G", 18), ("H", 16),
    ]:
        ws.column_dimensions[col_letter].width = width

    # Notes sheet
    ws2 = wb.create_sheet("Notes")
    ws2.cell(row=1, column=1, value="Comparable Company Selection Notes")
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws2.cell(row=3, column=1, value="Source: Public filings and industry databases")
    ws2.cell(row=4, column=1, value="Period: Fiscal year ending 2024")
    ws2.cell(row=5, column=1, value="SIC Code 3599: Industrial Machinery & Equipment, NEC")
    ws2.cell(row=6, column=1, value="All figures in millions of USD")
    ws2.cell(row=8, column=1, value="Screening Criteria Applied:")
    ws2.cell(row=9, column=1, value="  - SIC code match to target company")
    ws2.cell(row=10, column=1, value="  - Revenue range: $100M - $500M")
    ws2.cell(row=11, column=1, value="  - Publicly traded with available financial data")
    ws2.cell(row=12, column=1, value="  - No financial distress or restructuring")
    ws2.cell(row=14, column=1, value=(
        "Note: Initial screen identified 12 companies. "
        "Analyst should review and exclude any that are not appropriate comparables."
    ))
    ws2.column_dimensions["A"].width = 70

    # Canary
    canary = canaries.canary_for("tc09_comparable_companies")
    loc = embed_canary_xlsx(wb, canary)
    canaries.set_location(
        "tc09_comparable_companies",
        f"{_INPUT_DIR}/comparable_companies.xlsx",
        loc,
    )

    wb.properties.created = _FIXED_DATETIME
    file_path = output_dir / _INPUT_DIR / "comparable_companies.xlsx"
    file_path.parent.mkdir(parents=True, exist_ok=True)
    _save_xlsx_deterministic(wb, file_path)

    manifest.register(f"{_INPUT_DIR}/comparable_companies.xlsx", "xlsx")


# ── Prior year TP report PDF (42 pages) ──────────────────────────────────────


def _write_tp_report_pdf(
    output_dir: Path,
    canaries: CanaryRegistry,
    manifest: Manifest,
) -> None:
    """Write a 42-page prior year TP report PDF."""
    canary = canaries.canary_for("tc09_tp_report_fy2024")
    file_path = output_dir / _INPUT_DIR / "tp_report_fy2024.pdf"
    file_path.parent.mkdir(parents=True, exist_ok=True)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=letter,
        leftMargin=inch,
        rightMargin=inch,
        topMargin=inch,
        bottomMargin=inch,
        invariant=True,
    )
    doc.title = "Transfer Pricing Documentation — Cascade Industries FY2024"
    doc.author = f"CANARY: {canary}"
    doc.subject = "Transfer Pricing Local File"
    doc.creator = "Cascade Industries Test Suite Generator"

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TPTitle", parent=styles["Title"], fontSize=18, spaceAfter=20,
    )
    heading_style = ParagraphStyle(
        "TPH1", parent=styles["Heading1"], fontSize=14, spaceBefore=12, spaceAfter=8,
    )
    heading2_style = ParagraphStyle(
        "TPH2", parent=styles["Heading2"], fontSize=12, spaceBefore=10, spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "TPBody", parent=styles["Normal"], fontSize=10, leading=14, spaceAfter=8,
    )
    small_style = ParagraphStyle(
        "TPSmall", parent=body_style, fontSize=8, textColor=colors.gray,
    )

    story: list[Any] = []

    # ── Page 1: Title page ────────────────────────────────────
    story.append(Spacer(1, 2 * inch))
    story.append(Paragraph(
        "Transfer Pricing Documentation<br/>"
        "Cascade Industries, Inc.<br/>"
        "Fiscal Year Ended December 31, 2024",
        title_style,
    ))
    story.append(Spacer(1, 0.5 * inch))
    story.append(Paragraph(
        "Prepared in accordance with IRC §482 and<br/>"
        "OECD Transfer Pricing Guidelines (2022)",
        body_style,
    ))
    story.append(Spacer(1, 0.5 * inch))
    story.append(Paragraph(
        "CONFIDENTIAL — FOR INTERNAL USE ONLY",
        ParagraphStyle("Conf", parent=body_style, fontSize=10,
                       textColor=colors.red, alignment=1),
    ))
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(f"Document ID: {canary}", small_style))
    story.append(PageBreak())

    # ── Page 2: Table of Contents ─────────────────────────────
    story.append(Paragraph("Table of Contents", heading_style))
    toc = [
        ("1. Executive Summary", "3"),
        ("2. Company Overview", "4"),
        ("3. Organizational Structure", "5"),
        ("4. Industry Analysis", "6"),
        ("5. Functional Analysis — Goods Transfers", "8"),
        ("6. Functional Analysis — Services Transfers", "10"),
        ("7. Functional Analysis — Management Fees", "12"),
        ("8. Functional Analysis — Intercompany Loan", "14"),
        ("9. Selection of Transfer Pricing Method", "16"),
        ("10. Economic Analysis — Comparable Search", "18"),
        ("11. Economic Analysis — Comparable Screening", "22"),
        ("12. Economic Analysis — Benchmarking Results", "26"),
        ("13. Economic Analysis — Results and Conclusions", "31"),
        ("14. Appendix A: Financial Statements", "35"),
        ("15. Appendix B: Comparable Company Data", "38"),
        ("16. Appendix C: Regulatory References", "41"),
    ]
    toc_data = [["Section", "Page"]]
    for section, page in toc:
        toc_data.append([section, page])
    t = Table(toc_data, colWidths=[5 * inch, 1 * inch])
    t.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#4472C4")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    story.append(PageBreak())

    # ── Page 3: Executive Summary ─────────────────────────────
    story.append(Paragraph("1. Executive Summary", heading_style))
    story.append(Paragraph(
        "This report documents the transfer pricing policies and practices of "
        "Cascade Industries, Inc. (\"Cascade\" or the \"Company\") and its "
        "wholly-owned subsidiaries for the fiscal year ended December 31, 2024. "
        "The analysis covers all material intercompany transactions conducted "
        "between Cascade Industries and its three operating subsidiaries: "
        "Cascade Precision Components LLC, Cascade Advanced Materials, Inc., "
        "and Cascade Distribution Services LLC.",
        body_style,
    ))
    story.append(Paragraph(
        "The Company's transfer pricing policies are designed to comply with "
        "the arm's-length principle as codified in IRC §482 and consistent "
        "with the OECD Transfer Pricing Guidelines. Our analysis concludes "
        "that all intercompany transactions for FY2024 fall within the arm's "
        "length range, with the exception of certain intercompany service "
        "fees that merit continued monitoring.",
        body_style,
    ))
    story.append(Paragraph(
        "Key findings for FY2024:",
        body_style,
    ))
    findings = [
        "Goods transfers (PC → AM): Cost-plus-8% markup — within arm's length range",
        "Service fees (DS → PC, AM): Operating margin of 10.8% — within range but "
        "approaching the upper boundary of the interquartile range",
        "Management fees (CI → subsidiaries): 1.5% of subsidiary revenue — "
        "consistent with industry benchmarks",
        "Intercompany loan interest (CI → AM): 5.0% annual rate on $5M principal — "
        "consistent with comparable market rates",
    ]
    for f in findings:
        story.append(Paragraph(f"&bull; {f}", body_style))
    story.append(PageBreak())

    # ── Page 4: Company Overview ──────────────────────────────
    story.append(Paragraph("2. Company Overview", heading_style))
    story.append(Paragraph(
        "Cascade Industries, Inc. is a U.S. C-Corporation headquartered in "
        "Portland, Oregon. The Company operates as a mid-market manufacturer "
        "with consolidated revenues of approximately $195 million in FY2024. "
        "The Company operates through three wholly-owned subsidiaries, each "
        "serving a distinct market segment.",
        body_style,
    ))
    story.append(Paragraph("Subsidiary Overview", heading2_style))
    co_data = [
        ["Entity", "Location", "Primary Activity", "FY2024 Revenue ($M)"],
        ["Cascade Precision Components LLC", "Portland, OR",
         "Core manufacturing — industrial parts", "~$90M"],
        ["Cascade Advanced Materials, Inc.", "Austin, TX",
         "Specialty materials R&D and manufacturing", "~$62M"],
        ["Cascade Distribution Services LLC", "Chicago, IL",
         "Warehousing and logistics", "~$38M"],
    ]
    t = Table(co_data, colWidths=[2.2 * inch, 1.2 * inch, 2 * inch, 1.2 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#D6E4F0")]),
    ]))
    story.append(t)
    story.append(PageBreak())

    # ── Page 5: Organizational Structure ──────────────────────
    story.append(Paragraph("3. Organizational Structure", heading_style))
    story.append(Paragraph(
        "Cascade Industries, Inc. is the ultimate parent entity that holds 100% "
        "of the equity interests in each of its three operating subsidiaries. "
        "The organizational structure is as follows:",
        body_style,
    ))
    story.append(Spacer(1, 0.3 * inch))
    org_data = [
        ["", "Cascade Industries, Inc. (Parent)", ""],
        ["", "Portland, OR — C-Corporation", ""],
        ["", "│", ""],
        ["┌──────────────┤", "├──────────────┐", ""],
        ["Precision Components", "Advanced Materials", "Distribution Services"],
        ["Portland, OR (LLC)", "Austin, TX (Inc.)", "Chicago, IL (LLC)"],
    ]
    t = Table(org_data, colWidths=[2.2 * inch, 2.2 * inch, 2.2 * inch])
    t.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(
        "All entities are domestic (U.S.) and file as part of a consolidated "
        "federal tax return. State filing requirements vary by entity location "
        "and nexus. The transfer pricing documentation is prepared to comply "
        "with federal documentation requirements under IRC §6662(e).",
        body_style,
    ))
    story.append(PageBreak())

    # ── Pages 6-7: Industry Analysis ──────────────────────────
    story.append(Paragraph("4. Industry Analysis", heading_style))
    story.append(Paragraph(
        "Cascade operates in the U.S. mid-market manufacturing sector, "
        "specifically in industrial machinery and equipment (SIC 3599) and "
        "specialty materials. The sector has experienced steady growth, with "
        "the overall market expanding at approximately 4-5% annually.",
        body_style,
    ))
    story.append(Paragraph("Market Characteristics", heading2_style))
    story.append(Paragraph(
        "The mid-market manufacturing sector is characterized by moderate "
        "concentration, with the top 20 firms holding approximately 35% of "
        "market share. Key competitive factors include product quality, "
        "technical capability, delivery reliability, and price.",
        body_style,
    ))
    story.append(Paragraph("Margin Benchmarks", heading2_style))
    margin_data = [
        ["Metric", "Industry Median", "Top Quartile", "Bottom Quartile"],
        ["Gross Margin", "32%", "38%", "25%"],
        ["Operating Margin", "6.5%", "9.0%", "3.5%"],
        ["EBITDA Margin", "10.2%", "14.0%", "6.0%"],
        ["SG&A / Revenue", "18%", "14%", "24%"],
    ]
    t = Table(margin_data, colWidths=[1.5 * inch, 1.5 * inch, 1.5 * inch, 1.5 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#D6E4F0")]),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(
        "The logistics and distribution subsector (relevant to DS) operates "
        "with lower gross margins (15-22%) but benefits from asset-light "
        "models and recurring contract revenue. Operating margins typically "
        "range from 5-12% depending on scale and service mix.",
        body_style,
    ))
    story.append(PageBreak())
    # Page 7 continuation
    story.append(Paragraph("4. Industry Analysis (continued)", heading_style))
    story.append(Paragraph("Intercompany Transaction Trends", heading2_style))
    story.append(Paragraph(
        "Intercompany transactions among related manufacturing entities "
        "typically involve goods transfers, shared services, management fees, "
        "and financing arrangements. Industry practice for goods transfers "
        "is predominantly cost-plus pricing, with markups ranging from 5-15% "
        "depending on the value added and risk assumed by the selling entity.",
        body_style,
    ))
    story.append(Paragraph(
        "For intercompany services, the most common methods are the comparable "
        "uncontrolled price (CUP) method and the transactional net margin "
        "method (TNMM). Management fee allocations are typically based on a "
        "percentage of revenue or headcount, with rates generally ranging "
        "from 1-3% of subsidiary revenue.",
        body_style,
    ))
    story.append(PageBreak())

    # ── Pages 8-15: Functional Analysis ───────────────────────
    # Pages 8-9: Goods transfers
    story.append(Paragraph("5. Functional Analysis — Goods Transfers", heading_style))
    story.append(Paragraph(
        "Cascade Precision Components LLC (\"PC\") supplies raw materials and "
        "precision-machined components to Cascade Advanced Materials, Inc. (\"AM\"). "
        "These transfers represent approximately 25% of AM's total cost of goods "
        "sold and are priced at cost plus an 8% markup.",
        body_style,
    ))
    story.append(Paragraph("Functions Performed", heading2_style))
    func_goods = [
        ("PC (Seller)", "Procurement of raw materials, quality control, "
         "machining and finishing, inventory management, shipping coordination"),
        ("AM (Buyer)", "Product design specifications, quality acceptance testing, "
         "demand planning, integration into final products"),
    ]
    for entity, functions in func_goods:
        story.append(Paragraph(f"<b>{entity}</b>: {functions}", body_style))
    story.append(Paragraph("Risks Assumed", heading2_style))
    risks_goods = [
        ("PC", "Inventory obsolescence risk, production risk, raw material "
         "price fluctuation risk"),
        ("AM", "Market demand risk, product liability risk, technology "
         "obsolescence risk"),
    ]
    for entity, risks in risks_goods:
        story.append(Paragraph(f"<b>{entity}</b>: {risks}", body_style))
    story.append(Paragraph("Assets Employed", heading2_style))
    story.append(Paragraph(
        "PC employs significant tangible assets including manufacturing "
        "equipment, CNC machining centers, and quality testing equipment. "
        "AM employs proprietary formulations and R&D capabilities as "
        "intangible assets.",
        body_style,
    ))
    story.append(Paragraph("Transfer Pricing Policy", heading2_style))
    story.append(Paragraph(
        "Goods are transferred at cost plus 8%. The cost base includes "
        "direct materials, direct labor, and allocated manufacturing "
        "overhead. This markup is intended to compensate PC for the "
        "functions performed and risks assumed as a contract manufacturer.",
        body_style,
    ))
    story.append(PageBreak())
    # Page 10
    story.append(Paragraph("5. Functional Analysis — Goods (continued)", heading_style))
    story.append(Paragraph("Volume and Pricing Analysis — FY2024", heading2_style))
    story.append(Paragraph(
        "Total intercompany goods transfers in FY2024 were approximately "
        "$7.2 million. Monthly volumes ranged from $540K to $680K, "
        "reflecting AM's seasonal production patterns. The 8% markup "
        "remained constant throughout the year.",
        body_style,
    ))
    story.append(Paragraph("Comparability Considerations", heading2_style))
    story.append(Paragraph(
        "The cost-plus method is the most appropriate transfer pricing "
        "method for goods transfers, given PC's role as a contract "
        "manufacturer with limited market risk. The 8% markup is "
        "within the range observed for comparable contract manufacturing "
        "arrangements in the industrial sector.",
        body_style,
    ))
    story.append(PageBreak())

    # Pages 10-11: Services transfers
    story.append(Paragraph(
        "6. Functional Analysis — Services Transfers", heading_style,
    ))
    story.append(Paragraph(
        "Cascade Distribution Services LLC (\"DS\") provides warehousing, "
        "logistics coordination, and inventory management services to both "
        "PC and AM. These intercompany services represent approximately 15% "
        "of DS's total warehousing capacity, with 60% allocated to PC and "
        "40% to AM.",
        body_style,
    ))
    story.append(Paragraph("Functions Performed", heading2_style))
    story.append(Paragraph(
        "<b>DS (Service Provider)</b>: Warehouse facility management, "
        "inventory receiving and put-away, order fulfillment, shipping "
        "coordination, inventory tracking and reporting, facility "
        "maintenance and security.",
        body_style,
    ))
    story.append(Paragraph(
        "<b>PC / AM (Service Recipients)</b>: Demand planning and "
        "forecasting, product specifications for storage requirements, "
        "shipping destination and scheduling instructions.",
        body_style,
    ))
    story.append(Paragraph("Risks Assumed", heading2_style))
    story.append(Paragraph(
        "<b>DS</b>: Facility risk (lease obligations, maintenance), "
        "labor risk, regulatory compliance risk (safety, environmental).",
        body_style,
    ))
    story.append(Paragraph("Transfer Pricing Policy", heading2_style))
    story.append(Paragraph(
        "Services are priced at market rates benchmarked to third-party "
        "warehouse service contracts that DS maintains with external "
        "customers. The fee structure is designed to approximate the "
        "margins DS earns on comparable external engagements.",
        body_style,
    ))
    story.append(Paragraph(
        "For FY2024, the intercompany services generated an operating "
        "margin of approximately 10.8% for DS, which is at the upper "
        "end of the arm's length range but within it. This should be "
        "monitored as margins may fluctuate with volume changes.",
        body_style,
    ))
    story.append(PageBreak())

    # Page 12-13: Management fees
    story.append(Paragraph(
        "7. Functional Analysis — Management Fees", heading_style,
    ))
    story.append(Paragraph(
        "Cascade Industries, Inc. (the Parent) charges each operating "
        "subsidiary a management fee equal to 1.5% of the subsidiary's "
        "annual revenue. This fee covers corporate overhead services "
        "including:",
        body_style,
    ))
    mgmt_services = [
        "Executive management and strategic planning",
        "Financial reporting, consolidation, and internal audit",
        "Legal and regulatory compliance",
        "Human resources administration and benefits management",
        "Information technology infrastructure and support",
        "Treasury and cash management",
    ]
    for s in mgmt_services:
        story.append(Paragraph(f"&bull; {s}", body_style))
    story.append(Paragraph("Benchmarking", heading2_style))
    story.append(Paragraph(
        "The 1.5% management fee rate is within the commonly observed "
        "range of 1-3% for management service allocations among "
        "mid-market manufacturing groups. The fee is assessed on a "
        "monthly basis using actual revenue.",
        body_style,
    ))
    story.append(PageBreak())
    # Page 14
    story.append(Paragraph(
        "7. Functional Analysis — Management Fees (continued)", heading_style,
    ))
    story.append(Paragraph(
        "Total management fees charged in FY2024:",
        body_style,
    ))
    mgmt_fee_data = [
        ["Subsidiary", "FY2024 Revenue ($M)", "Fee Rate", "Fee Amount ($K)"],
        ["Precision Components", "~$90M", "1.5%", "~$1,350K"],
        ["Advanced Materials", "~$62M", "1.5%", "~$930K"],
        ["Distribution Services", "~$38M", "1.5%", "~$570K"],
        ["Total", "~$190M", "", "~$2,850K"],
    ]
    t = Table(mgmt_fee_data, colWidths=[2 * inch, 1.5 * inch, 1 * inch, 1.5 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("LINEABOVE", (0, -1), (-1, -1), 1, colors.black),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    story.append(t)
    story.append(PageBreak())

    # Pages 14-15: IC Loan
    story.append(Paragraph(
        "8. Functional Analysis — Intercompany Loan", heading_style,
    ))
    story.append(Paragraph(
        "In January 2023, Cascade Industries, Inc. extended a $5,000,000 "
        "intercompany loan to Cascade Advanced Materials, Inc. to fund "
        "capital expenditures related to AM's specialty materials R&D "
        "facility expansion in Austin, TX.",
        body_style,
    ))
    story.append(Paragraph("Loan Terms", heading2_style))
    loan_data = [
        ["Parameter", "Value"],
        ["Principal", "$5,000,000"],
        ["Interest Rate", "5.0% per annum"],
        ["Term", "10 years"],
        ["Payment Structure", "Interest-only, monthly accrual"],
        ["Collateral", "None (intercompany guarantee)"],
    ]
    t = Table(loan_data, colWidths=[2 * inch, 3 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(
        "The 5.0% interest rate was established based on comparable "
        "market rates for unsecured corporate lending at the time of "
        "origination. FY2024 interest income/expense totaled $250,000 "
        "(12 monthly accruals of $20,833).",
        body_style,
    ))
    story.append(Paragraph(
        "The loan rate remains within the arm's length range based on "
        "current comparable market rates of 4.5-6.0% for similar "
        "unsecured intercompany facilities.",
        body_style,
    ))
    story.append(PageBreak())
    # Page 16
    story.append(Paragraph(
        "8. Functional Analysis — Intercompany Loan (continued)", heading_style,
    ))
    story.append(Paragraph(
        "The intercompany loan is documented via a formal loan agreement "
        "that includes standard commercial terms. The principal balance "
        "remained unchanged at $5,000,000 throughout FY2024.",
        body_style,
    ))
    story.append(PageBreak())

    # ── Pages 16-21: Selection of TP Method ───────────────────
    story.append(Paragraph(
        "9. Selection of Transfer Pricing Method", heading_style,
    ))
    story.append(Paragraph(
        "The following transfer pricing methods were considered for each "
        "category of intercompany transaction:",
        body_style,
    ))
    method_data = [
        ["Transaction Type", "Selected Method", "Rationale"],
        ["Goods (PC → AM)", "Cost Plus Method (CPM)",
         "PC functions as a contract manufacturer with limited risk. "
         "Cost-plus is the most direct and reliable method."],
        ["Services (DS → PC, AM)", "Transactional Net Margin Method (TNMM)",
         "DS provides integrated warehouse services. TNMM using "
         "operating margin as the profit level indicator is most appropriate."],
        ["Management Fees", "Cost Plus / Comparable Analysis",
         "Fees are benchmarked against industry surveys of management "
         "service allocations."],
        ["Interest", "Comparable Uncontrolled Price (CUP)",
         "Market rate comparison based on comparable lending arrangements."],
    ]
    t = Table(method_data, colWidths=[1.5 * inch, 1.8 * inch, 3.3 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(
        "For the goods and services transactions, the Transactional Net "
        "Margin Method (TNMM) is applied as the primary method, using "
        "operating margin as the profit level indicator (PLI). This "
        "approach is consistent with the OECD Transfer Pricing Guidelines "
        "and IRS regulations under §1.482-5.",
        body_style,
    ))
    story.append(PageBreak())
    # Page 18 — filler for method selection continued
    story.append(Paragraph(
        "9. Selection of Transfer Pricing Method (continued)", heading_style,
    ))
    story.append(Paragraph(
        "The comparable uncontrolled price (CUP) method was also considered "
        "but rejected for goods and services transactions due to the lack "
        "of sufficiently comparable uncontrolled transactions. The profit "
        "split method was considered but deemed unnecessary given the "
        "relatively straightforward nature of the intercompany arrangements.",
        body_style,
    ))
    story.append(PageBreak())

    # ── Pages 18-21: Economic Analysis — Comparable Search ────
    story.append(Paragraph(
        "10. Economic Analysis — Comparable Search", heading_style,
    ))
    story.append(Paragraph(
        "A comparable company search was conducted to identify independent "
        "companies whose operating margins can be used to benchmark "
        "Cascade's intercompany transactions. The search was conducted "
        "using publicly available financial databases.",
        body_style,
    ))
    story.append(Paragraph("Search Strategy", heading2_style))
    story.append(Paragraph(
        "The search targeted companies in SIC code 3599 (Industrial and "
        "Commercial Machinery and Equipment, NEC) with annual revenues "
        "between $100 million and $500 million. The geographic scope was "
        "limited to U.S.-headquartered companies.",
        body_style,
    ))
    story.append(Paragraph("Initial Search Criteria", heading2_style))
    criteria = [
        "SIC code: 3599 (primary) or related manufacturing codes",
        "Revenue range: $100M — $500M",
        "Geography: United States",
        "Public company with available financial statements",
        "Minimum 3 years of operating history",
        "No pending bankruptcy or restructuring proceedings",
    ]
    for c in criteria:
        story.append(Paragraph(f"&bull; {c}", body_style))
    story.append(Paragraph(
        "The initial search yielded 45 potential comparable companies. "
        "After applying quantitative and qualitative screens, the set "
        "was narrowed to 12 companies for detailed analysis.",
        body_style,
    ))
    story.append(PageBreak())
    # Pages 20-21 filler
    story.append(Paragraph(
        "10. Economic Analysis — Comparable Search (continued)", heading_style,
    ))
    story.append(Paragraph(
        "The screening process applied the following quantitative filters:",
        body_style,
    ))
    story.append(Paragraph("Quantitative Screening", heading2_style))
    quant_data = [
        ["Filter", "Companies Before", "Companies After"],
        ["SIC code match", "45", "38"],
        ["Revenue range ($100M-$500M)", "38", "24"],
        ["Positive operating income (3yr avg)", "24", "19"],
        ["Complete financial data available", "19", "15"],
        ["No related-party concentration >50%", "15", "12"],
    ]
    t = Table(quant_data, colWidths=[3 * inch, 1.5 * inch, 1.5 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(
        "Note: The final set of 12 companies includes all entities that "
        "passed the quantitative screens. A qualitative review of each "
        "company's business description, financial condition, and "
        "comparability is presented in Section 11.",
        body_style,
    ))
    story.append(PageBreak())
    # Additional pages for search details
    story.append(Paragraph(
        "10. Economic Analysis — Comparable Search (continued)", heading_style,
    ))
    story.append(Paragraph(
        "The following data sources were used in the comparable search:",
        body_style,
    ))
    sources = [
        "SEC EDGAR filings (10-K annual reports)",
        "S&P Capital IQ industry screening tools",
        "Bureau van Dijk / Orbis company database",
        "Industry trade association publications",
    ]
    for s in sources:
        story.append(Paragraph(f"&bull; {s}", body_style))
    story.append(Paragraph(
        "Financial data was collected for the most recent fiscal year "
        "(FY2024 where available, otherwise FY2023). All financial "
        "figures are reported in U.S. dollars.",
        body_style,
    ))
    story.append(PageBreak())

    # ── Pages 22-30: Economic Analysis — Comparable Screening ─
    story.append(Paragraph(
        "11. Economic Analysis — Comparable Screening", heading_style,
    ))
    story.append(Paragraph(
        "This section presents the qualitative review of each comparable "
        "company identified in the search process. Companies are assessed "
        "for functional comparability, risk profile, and financial condition.",
        body_style,
    ))
    # Write a brief paragraph for each of the 12 companies
    for name, sic, rev, cogs, opex, oper_inc, assets, rejected in _COMPANY_PROFILES:
        margin = oper_inc / rev if rev else 0
        story.append(Paragraph(f"<b>{name}</b> (SIC: {sic})", heading2_style))
        story.append(Paragraph(
            f"Revenue: ${rev}M | COGS: ${cogs}M | Operating Expenses: ${opex}M | "
            f"Operating Income: ${oper_inc}M | Operating Margin: {margin:.1%} | "
            f"Total Assets: ${assets}M",
            body_style,
        ))
        if rejected and sic != "3599":
            story.append(Paragraph(
                "<b>Recommendation: REJECT</b> — This company operates in SIC "
                f"code {sic} (Prepackaged Software), which is not comparable "
                "to Cascade's manufacturing operations. Different industry "
                "dynamics, margin structures, and risk profiles make this "
                "company unsuitable as a comparable.",
                body_style,
            ))
        elif rejected and oper_inc < 0:
            story.append(Paragraph(
                "<b>Recommendation: REJECT</b> — This company is currently "
                "experiencing severe financial distress with negative operating "
                "margins. The company filed for Chapter 11 bankruptcy protection "
                "in Q3 2024. Financial distress companies are not appropriate "
                "comparables under arm's length analysis.",
                body_style,
            ))
        else:
            story.append(Paragraph(
                "<b>Recommendation: ACCEPT</b> — Comparable manufacturing "
                "operations with similar functions, risks, and asset intensity.",
                body_style,
            ))
        # Add page break every 4 companies to spread across pages 22-30
        idx = _COMPANY_PROFILES.index(
            (name, sic, rev, cogs, opex, oper_inc, assets, rejected)
        )
        if idx in (3, 7, 11):
            story.append(PageBreak())

    story.append(PageBreak())

    # ── Pages 31-38: Benchmarking Results ─────────────────────
    story.append(Paragraph(
        "12. Economic Analysis — Benchmarking Results", heading_style,
    ))
    story.append(Paragraph(
        "After rejecting two companies from the initial set of 12, the "
        "benchmarking analysis was performed using the remaining 10 "
        "accepted comparable companies.",
        body_style,
    ))
    story.append(Paragraph("Accepted Comparable Set — Summary Statistics", heading2_style))
    # Show the FY2024 benchmarking results
    story.append(Paragraph(
        "The interquartile range (IQR) of operating margins for the "
        "accepted comparable set for FY2024 was:",
        body_style,
    ))
    iqr_data = [
        ["Statistic", "Value"],
        ["Minimum", "2.2%"],
        ["25th Percentile (Q1)", "3.8%"],
        ["Median", "6.2%"],
        ["75th Percentile (Q3)", "8.5%"],
        ["Maximum", "11.5%"],
    ]
    t = Table(iqr_data, colWidths=[2.5 * inch, 2 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(
        "Note: These are the FY2024 benchmarking results. The current "
        "year (FY2025) comparable data set should be updated with "
        "current year financials when available.",
        body_style,
    ))
    story.append(PageBreak())

    # Page 32-34: Application to Cascade's transactions
    story.append(Paragraph(
        "12. Benchmarking Results (continued) — Application", heading_style,
    ))
    story.append(Paragraph(
        "The benchmarking results were applied to each of Cascade's "
        "intercompany transaction categories for FY2024:",
        body_style,
    ))
    results_data = [
        ["Transaction", "Cascade Margin", "IQR Range", "Status"],
        ["Goods (PC → AM)", "8.0% (cost-plus)", "3.8% — 8.5%",
         "WITHIN RANGE"],
        ["Services (DS → PC/AM)", "10.8%", "3.8% — 8.5%",
         "MONITOR — approaching upper bound"],
        ["Management Fees", "1.5% of revenue", "1.0% — 3.0% (industry)",
         "WITHIN RANGE"],
        ["Interest (CI → AM)", "5.0%", "4.5% — 6.0% (market)",
         "WITHIN RANGE"],
    ]
    t = Table(results_data, colWidths=[1.5 * inch, 1.5 * inch, 1.5 * inch, 2 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("TEXTCOLOR", (3, 4), (3, 4), colors.HexColor("#C00000")),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(
        "<b>Services Transaction — Monitoring Note:</b> The services "
        "operating margin of 10.8% is above the 75th percentile (8.5%) "
        "of the FY2024 comparable set. While this was within the full "
        "range (2.2% — 11.5%), it is above the IQR. Management should "
        "monitor this margin and consider whether pricing adjustments "
        "are warranted if the margin continues to increase.",
        body_style,
    ))
    story.append(PageBreak())

    # Pages 35-38: More detail on benchmarking
    for page_num in range(4):
        story.append(Paragraph(
            f"12. Benchmarking Results (continued) — Detail {page_num + 1}",
            heading_style,
        ))
        if page_num == 0:
            story.append(Paragraph(
                "The following table presents the detailed operating margin "
                "analysis for each accepted comparable company:",
                body_style,
            ))
            detail_data = [["Company", "Revenue ($M)", "Oper. Income ($M)", "Margin"]]
            for name, sic, rev, cogs, opex, oi, assets, rej in _COMPANY_PROFILES:
                if not rej:
                    margin = oi / rev if rev else 0
                    detail_data.append([name, str(rev), str(oi), f"{margin:.1%}"])
            t = Table(detail_data, colWidths=[2.5 * inch, 1.2 * inch, 1.5 * inch, 1 * inch])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ]))
            story.append(t)
        else:
            story.append(Paragraph(
                "Additional sensitivity analysis and year-over-year "
                "trend data for the comparable set demonstrates "
                "consistency in the benchmarking results. "
                + ("This section provides further detail on the statistical "
                   "methodology used to compute the interquartile range and "
                   "the rationale for using operating margin as the profit "
                   "level indicator. ") * 3,
                body_style,
            ))
        story.append(PageBreak())

    # ── Pages 31-38 results section done ──────────────────────

    # ── Pages 39-40: Results and Conclusions ──────────────────
    story.append(Paragraph(
        "13. Economic Analysis — Results and Conclusions", heading_style,
    ))
    story.append(Paragraph(
        "Based on the benchmarking analysis using the TNMM with operating "
        "margin as the profit level indicator, the following conclusions "
        "are reached for FY2024:",
        body_style,
    ))
    story.append(Paragraph("Conclusions by Transaction Type", heading2_style))
    conclusions = [
        ("<b>Goods Transfers (PC → AM)</b>: The cost-plus-8% transfer price "
         "results in an operating margin of approximately 8.0% for PC on these "
         "transactions. This margin falls within the interquartile range "
         "(3.8% — 8.5%) and is therefore consistent with the arm's length "
         "principle."),
        ("<b>Services (DS → PC, AM)</b>: The intercompany service fees result "
         "in an operating margin of 10.8% for DS. This margin is above the "
         "75th percentile (8.5%) but within the full range of comparables. "
         "While no adjustment is required for FY2024, this transaction should "
         "be closely monitored and repriced if the margin exceeds the full "
         "range in future years."),
        ("<b>Management Fees (CI → Subs)</b>: The 1.5% management fee rate "
         "is within the range of 1-3% commonly observed for similar "
         "management service arrangements in the industry."),
        ("<b>Intercompany Loan Interest</b>: The 5.0% interest rate on the "
         "$5M loan from CI to AM is within the range of 4.5-6.0% observed "
         "for comparable intercompany lending arrangements."),
    ]
    for c in conclusions:
        story.append(Paragraph(c, body_style))
    story.append(PageBreak())
    # Page 40
    story.append(Paragraph(
        "13. Results and Conclusions (continued)", heading_style,
    ))
    story.append(Paragraph("Overall Assessment", heading2_style))
    story.append(Paragraph(
        "All of Cascade's intercompany transactions for FY2024 are "
        "determined to be consistent with the arm's length principle. "
        "The Company's transfer pricing policies are appropriately "
        "documented and applied consistently.",
        body_style,
    ))
    story.append(Paragraph("Recommendations", heading2_style))
    story.append(Paragraph(
        "1. Continue to monitor the services transfer pricing margin "
        "as it approached the upper boundary of the IQR in FY2024.",
        body_style,
    ))
    story.append(Paragraph(
        "2. Update the comparable company set annually to reflect "
        "current market conditions.",
        body_style,
    ))
    story.append(Paragraph(
        "3. Review the intercompany loan rate in light of current "
        "interest rate environment changes.",
        body_style,
    ))
    story.append(PageBreak())

    # ── Pages 35-38 (Appendix A): Financial Statements ────────
    story.append(Paragraph("14. Appendix A: Financial Statements", heading_style))
    story.append(Paragraph(
        "Selected consolidated financial data for Cascade Industries "
        "for FY2022–FY2024 (in millions):",
        body_style,
    ))
    fin_data = [
        ["", "FY2022", "FY2023", "FY2024"],
        ["Revenue", "$178M", "$190M", "$195M"],
        ["COGS", "$118M", "$125M", "$128M"],
        ["Gross Profit", "$60M", "$65M", "$67M"],
        ["Operating Expenses", "$38M", "$40M", "$42M"],
        ["Operating Income", "$22M", "$25M", "$25M"],
        ["Intercompany Eliminations", "($8.5M)", "($9.2M)", "($10.1M)"],
    ]
    t = Table(fin_data, colWidths=[2.5 * inch, 1.2 * inch, 1.2 * inch, 1.2 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
    ]))
    story.append(t)
    story.append(PageBreak())
    # More appendix pages
    for appendix_page in range(2):
        story.append(Paragraph(
            f"14. Appendix A (continued) — Detail {appendix_page + 1}",
            heading_style,
        ))
        story.append(Paragraph(
            "Entity-level financial statements and intercompany "
            "elimination schedules are maintained in the Company's "
            "accounting records and are available upon request. "
            "The consolidation process ensures that all intercompany "
            "transactions (9xxx accounts) net to zero on a consolidated "
            "basis.",
            body_style,
        ))
        story.append(PageBreak())

    # ── Pages 38-40 (Appendix B): Comparable Company Data ─────
    story.append(Paragraph("15. Appendix B: Comparable Company Data", heading_style))
    story.append(Paragraph(
        "Detailed financial data for all 12 comparable companies "
        "identified in the search process. See Section 11 for "
        "individual company assessments and accept/reject recommendations.",
        body_style,
    ))
    story.append(PageBreak())
    story.append(Paragraph("15. Appendix B (continued)", heading_style))
    story.append(Paragraph(
        "Full comparable company financial profiles are maintained "
        "in the transfer pricing documentation workfiles and are "
        "available for regulatory review upon request.",
        body_style,
    ))
    story.append(PageBreak())

    # ── Pages 41-42 (Appendix C): Regulatory References ──────
    story.append(Paragraph("16. Appendix C: Regulatory References", heading_style))
    regs = [
        "IRC §482 — Allocation of Income and Deductions Among Taxpayers",
        "IRC §6662(e) — Substantial and Gross Valuation Misstatements "
        "Attributable to Transfer Pricing",
        "Treas. Reg. §1.482-1 — Allocation of Income and Deductions "
        "Among Taxpayers (General Principles)",
        "Treas. Reg. §1.482-3 — Methods to Determine Taxable Income "
        "in Connection with a Transfer of Tangible Property",
        "Treas. Reg. §1.482-5 — Comparable Profits Method",
        "Treas. Reg. §1.482-9 — Methods to Determine Taxable Income "
        "in Connection with a Controlled Services Transaction",
        "OECD Transfer Pricing Guidelines for Multinational Enterprises "
        "and Tax Administrations (2022 edition)",
        "OECD Model Tax Convention, Article 9 — Associated Enterprises",
    ]
    for r in regs:
        story.append(Paragraph(f"&bull; {r}", body_style))
    story.append(PageBreak())
    # Final page
    story.append(Paragraph("16. Appendix C (continued)", heading_style))
    story.append(Paragraph(
        "This transfer pricing documentation has been prepared in "
        "compliance with the documentation requirements of IRC §6662(e) "
        "and the contemporaneous documentation standards established "
        "under Treas. Reg. §1.6662-6.",
        body_style,
    ))
    story.append(Spacer(1, inch))
    story.append(Paragraph(
        "— End of Transfer Pricing Documentation FY2024 —",
        ParagraphStyle("End", parent=body_style, alignment=1, fontSize=10,
                       textColor=colors.gray),
    ))

    # Build PDF
    doc.build(story)

    # Write deterministic PDF
    buf.seek(0)
    file_path.write_bytes(buf.getvalue())

    canaries.set_location(
        "tc09_tp_report_fy2024",
        f"{_INPUT_DIR}/tp_report_fy2024.pdf",
        "PDF metadata → Author",
    )
    manifest.register(f"{_INPUT_DIR}/tp_report_fy2024.pdf", "pdf")


# ── Prompt and expected behavior ─────────────────────────────────────────────


def _write_prompt(output_dir: Path) -> None:
    """Write test_cases/TC-09/prompt.md per spec."""
    text = """\
Update the transfer pricing analysis for Cascade Industries for FY2025.

1. Using the intercompany transaction data, calculate the actual intercompany
   margins for each transaction type (goods, services, interest).
2. Screen the comparable companies:
   - Reject any that are not appropriate comparables (explain why)
   - Compute the interquartile range of operating margins for the accepted set
3. Determine whether Cascade's intercompany margins fall within the arm's-length
   range (interquartile range of comparables).
4. Flag any transaction types that fall outside the range.
5. Draft the "Economic Analysis — Results" section of the local file, updating
   the prior year report's language with current year data and conclusions.

Export:
- Benchmarking analysis as Excel
- Updated results section as Word document
"""
    path = output_dir / f"test_cases/{_TC}/prompt.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text)


def _write_expected_behavior(output_dir: Path) -> None:
    """Write test_cases/TC-09/expected_behavior.md per spec."""
    text = """\
# TC-09: Transfer Pricing Documentation — Expected Behavior

## Key Findings the Agent Should Produce

1. **Reject two comparable companies**:
   - TechVantage Solutions Inc (SIC 7372) — wrong industry (software, not manufacturing)
   - Meridian Industrial Holdings (SIC 3599) — financial distress (negative operating
     margin, Chapter 11)

2. **Compute IQR of operating margins**: After rejecting the two companies, the
   remaining 10 companies yield:
   - Q1 (25th percentile): **4.2%**
   - Q3 (75th percentile): **8.7%**
   - IQR: 4.2% — 8.7%

3. **Assess Cascade's intercompany margins**:
   - Goods (PC → AM): Cost-plus-8% — **within range** (8% falls within 4.2%–8.7%)
   - Services (DS → PC, AM): 11.2% operating margin — **OUTSIDE range** (above Q3 of 8.7%)
   - Management fees: 1.5% of subsidiary revenue — consistent with industry benchmarks
   - Interest: 5.0% on $5M loan — consistent with market rates

4. **Flag the services transaction**: The 11.2% operating margin on intercompany
   services is outside the IQR. The agent must flag this and recommend pricing
   adjustments or documentation of the business rationale.

5. **Note the change from FY2024**: The prior year report (FY2024) showed services
   at 10.8% which was above the IQR but within the full range. For FY2025, the margin
   has increased to 11.2%, further outside the IQR.

## Data Challenges

- **Comparable screening**: The agent must correctly identify and reject the two
  inappropriate comparables — one for SIC mismatch and one for financial distress.
  Simply accepting all 12 companies would produce incorrect IQR bounds.
- **IQR computation**: The agent must use a standard quartile computation method
  (e.g., QUARTILE.INC in Excel) on the 10 accepted companies.
- **Prior year comparison**: The FY2024 TP report provides benchmarking results
  for the prior year. The agent should note the change in services margin from
  10.8% (FY2024) to 11.2% (FY2025) and the implications.
- **Multiple transaction types**: Each transaction type requires separate analysis.
  The cost-plus method for goods is different from TNMM for services.

## Expected Output Structure

### Benchmarking Analysis (Excel workbook):
- Comparable company summary (all 12 with accept/reject flags)
- Accepted set operating margins and statistical summary
- IQR computation
- Cascade's margins vs. IQR for each transaction type
- Year-over-year comparison (FY2024 vs FY2025)

### Updated Results Section (Word document):
- Updated "Economic Analysis — Results" section with FY2025 data
- Clear statement that services margin is outside the IQR
- Recommendation for services transaction repricing or documentation
- Confirmation that other transactions remain within arm's length range
"""
    path = output_dir / f"test_cases/{_TC}/expected_behavior.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text)


# ── Gold Standard ────────────────────────────────────────────────────────────


def _compute_fy2025_margins(
    model: CascadeModel,
) -> dict[str, dict[str, Any]]:
    """Compute actual FY2025 IC margins by transaction type from the model."""
    fy25_txns = _get_fy2025_ic_transactions(model)

    type_totals: dict[str, Decimal] = {}
    type_counts: dict[str, int] = {}
    for tx in fy25_txns:
        type_totals[tx.tx_type] = type_totals.get(tx.tx_type, Decimal(0)) + tx.amount
        type_counts[tx.tx_type] = type_counts.get(tx.tx_type, 0) + 1

    return {
        "goods": {
            "total": _whole_dollars(type_totals.get("goods", Decimal(0))),
            "count": type_counts.get("goods", 0),
            "markup_pct": float(RAW_MATERIALS_MARKUP * 100),
            "description": "Cost-plus-8% (PC → AM)",
        },
        "services": {
            "total": _whole_dollars(type_totals.get("services", Decimal(0))),
            "count": type_counts.get("services", 0),
            "operating_margin_pct": float(SERVICES_OPERATING_MARGIN * 100),
            "description": "DS warehouse fees to PC and AM",
        },
        "management_fee": {
            "total": _whole_dollars(type_totals.get("management_fee", Decimal(0))),
            "count": type_counts.get("management_fee", 0),
            "fee_pct": float(MANAGEMENT_FEE_PCT * 100),
            "description": "1.5% of subsidiary revenue (CI → subs)",
        },
        "interest": {
            "total": _whole_dollars(type_totals.get("interest", Decimal(0))),
            "count": type_counts.get("interest", 0),
            "rate_pct": float(IC_LOAN_RATE * 100),
            "principal": _whole_dollars(IC_LOAN_PRINCIPAL),
            "description": "5% annual on $5M loan (CI → AM)",
        },
    }


def _compute_iqr() -> dict[str, Any]:
    """Compute the IQR for the 10 accepted comparable companies."""
    # Get sorted operating margins for accepted companies (not rejected)
    margins = sorted(
        oi / rev
        for _, _, rev, _, _, oi, _, rejected in _COMPANY_PROFILES
        if not rejected and rev > 0
    )
    n = len(margins)  # 10

    # QUARTILE.INC method: Q1 at index 0.25*(n-1) = 2.25, Q3 at 0.75*(n-1) = 6.75
    q1_idx = 0.25 * (n - 1)  # 2.25
    q3_idx = 0.75 * (n - 1)  # 6.75

    q1_low = int(q1_idx)
    q1_frac = q1_idx - q1_low
    q1 = margins[q1_low] + q1_frac * (margins[q1_low + 1] - margins[q1_low])

    q3_low = int(q3_idx)
    q3_frac = q3_idx - q3_low
    q3 = margins[q3_low] + q3_frac * (margins[q3_low + 1] - margins[q3_low])

    return {
        "accepted_count": n,
        "rejected_count": 2,
        "margins_sorted_pct": [round(m * 100, 1) for m in margins],
        "q1_pct": round(q1 * 100, 1),
        "median_pct": round((margins[4] + margins[5]) / 2 * 100, 1),
        "q3_pct": round(q3 * 100, 1),
        "min_pct": round(margins[0] * 100, 1),
        "max_pct": round(margins[-1] * 100, 1),
    }


@register_gold(_TC)
def _tc09_gold(
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    **model_kwargs: Any,
) -> GoldStandard:
    """TC-09 gold standard: Transfer pricing benchmarking."""
    model: CascadeModel = model_kwargs["model"]
    margins = _compute_fy2025_margins(model)
    iqr = _compute_iqr()

    return GoldStandard(
        test_case=_TC,
        expected_outputs={
            "output_files": {
                "benchmarking_analysis": {
                    "type": "xlsx",
                    "required_sheets": [
                        "Comparable Companies",
                        "Accepted Set Analysis",
                        "IQR Computation",
                        "Cascade Margin Assessment",
                    ],
                },
                "updated_results_section": {
                    "type": "docx",
                    "required_sections": [
                        "Economic Analysis — Results",
                        "Conclusions by Transaction Type",
                        "Recommendations",
                    ],
                },
            },
            "comparable_screening": {
                "total_companies": 12,
                "accepted": iqr["accepted_count"],
                "rejected": iqr["rejected_count"],
                "rejections": [
                    {
                        "company": "TechVantage Solutions Inc",
                        "reason": "SIC code mismatch (7372 vs 3599)",
                    },
                    {
                        "company": "Meridian Industrial Holdings",
                        "reason": "Financial distress (negative operating margin, Chapter 11)",
                    },
                ],
            },
            "iqr_analysis": {
                "q1_pct": iqr["q1_pct"],
                "q3_pct": iqr["q3_pct"],
                "median_pct": iqr["median_pct"],
                "min_pct": iqr["min_pct"],
                "max_pct": iqr["max_pct"],
                "margins_sorted_pct": iqr["margins_sorted_pct"],
            },
            "cascade_margins": margins,
            "arm_length_assessment": {
                "goods": {
                    "margin_pct": float(RAW_MATERIALS_MARKUP * 100),
                    "within_iqr": True,
                    "description": "Cost-plus-8% falls within 4.2%–8.7% IQR",
                },
                "services": {
                    "margin_pct": float(SERVICES_OPERATING_MARGIN * 100),
                    "within_iqr": False,
                    "description": (
                        "11.2% operating margin exceeds Q3 of 8.7%. "
                        "Agent must flag this and recommend repricing or documentation."
                    ),
                },
                "management_fee": {
                    "rate_pct": float(MANAGEMENT_FEE_PCT * 100),
                    "within_range": True,
                    "description": "1.5% within industry range of 1-3%",
                },
                "interest": {
                    "rate_pct": float(IC_LOAN_RATE * 100),
                    "within_range": True,
                    "description": "5.0% within market range of 4.5-6.0%",
                },
            },
            "prior_year_comparison": {
                "services_fy2024_pct": 10.8,
                "services_fy2025_pct": 11.2,
                "trend": "increasing — further outside IQR",
            },
        },
        canary_verification={
            "read_ic_transactions": canaries.canary_for("tc09_ic_transactions"),
            "read_comparable_companies": canaries.canary_for("tc09_comparable_companies"),
            "read_tp_report": canaries.canary_for("tc09_tp_report_fy2024"),
        },
        error_detection={},
        scoring_hints={
            "correctness": (
                "IQR must be exactly Q1=4.2%, Q3=8.7%. "
                "Services margin must be identified as 11.2% (outside IQR). "
                "Goods margin (8.0% cost-plus) must be identified as within range."
            ),
            "completeness": (
                "All four transaction types analyzed. "
                "Both rejected companies identified with correct reasons. "
                "Prior year comparison noted."
            ),
            "format_compliance": (
                "Benchmarking analysis as Excel with IQR computation. "
                "Updated results section as Word document."
            ),
            "robustness": (
                "Agent must reject exactly 2 companies (SIC mismatch + distress). "
                "Agent must flag services margin as outside IQR, not just outside "
                "full range. Agent must note the FY2024→FY2025 services margin increase."
            ),
            "communication": (
                "Clear statement of which transactions pass/fail arm's length test. "
                "Actionable recommendation for services repricing. "
                "Professional transfer pricing language."
            ),
        },
    )


# ── Public entry point ───────────────────────────────────────────────────────


def emit_tc09(
    model: CascadeModel,
    output_dir: Path,
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    manifest: Manifest,
) -> None:
    """Emit all TC-09 files."""
    _write_ic_transactions_xlsx(model, output_dir, canaries, errors, manifest)
    _write_comparables_xlsx(output_dir, canaries, manifest)
    _write_tp_report_pdf(output_dir, canaries, manifest)
    _write_prompt(output_dir)
    _write_expected_behavior(output_dir)
