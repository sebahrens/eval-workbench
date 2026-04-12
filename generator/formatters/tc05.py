"""Formatter: TC-05 — Audit Workpaper Memo — Accounts Receivable (Routine).

Emits:
- test_cases/TC-05/input_files/ar_aging_fy2025.xlsx
  AR aging schedule by customer (current, 30, 60, 90, 120+ buckets)
- test_cases/TC-05/input_files/ar_confirmations_summary.xlsx
  Summary of AR confirmation results
- test_cases/TC-05/input_files/allowance_analysis.xlsx
  Historical bad debt write-offs and allowance calculations
- test_cases/TC-05/input_files/workpaper_memo_template.docx
  Copied from templates/workpaper_memo_template.docx
- test_cases/TC-05/prompt.md
- test_cases/TC-05/expected_behavior.md
- gold_standards/TC-05_gold.json

Planted errors:
- ERR-016 (date_inconsistency) in ar_confirmations_summary.xlsx
- ERR-020 (missing_data) in ar_aging_fy2025.xlsx
Uses the canonical model — never hardcodes numbers.
"""

from __future__ import annotations

import datetime
import random
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from generator.canaries import CanaryRegistry, embed_canary_docx, embed_canary_xlsx
from generator.errors import ErrorRegistry, PlantedError, date_inconsistency, missing_data
from generator.golds.framework import GoldStandard, register_gold
from generator.manifest import Manifest
from generator.model.ar import (
    RESERVE_RATES,
    AllowanceAnalysis,
    ARAgingEntry,
    generate_allowance,
    generate_ar_aging,
)
from generator.model.build import CascadeModel
from generator.noise import apply_docx_noise, make_noise_rng
from generator.scenario_context import ScenarioContext

# ── Constants ────────────────────────────────────────────────────────────────

_TC = "TC-05"
_INPUT_DIR = f"test_cases/{_TC}/input_files"

# Fixed datetime for xlsx metadata (determinism)
_FIXED_DATETIME = datetime.datetime(2025, 3, 15, 9, 0, 0)

# Confirmation response statuses — deterministic per customer DSO
_CONFIRMATION_STATUSES = ("Agreed", "Agreed with exceptions", "No response")


# ── Helpers ──────────────────────────────────────────────────────────────────


def _pin_xlsx_dates(wb: openpyxl.Workbook) -> None:
    """Pin created timestamp for determinism."""
    wb.properties.created = _FIXED_DATETIME


def _save_xlsx_deterministic(wb: openpyxl.Workbook, path: str | Path) -> None:
    """Save workbook with pinned timestamps and fixed zip entry dates."""
    import io
    from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

    from openpyxl.writer.excel import ExcelWriter

    path = Path(path)

    buf = io.BytesIO()
    wb.properties.modified = _FIXED_DATETIME
    archive = ZipFile(buf, "w", ZIP_DEFLATED, allowZip64=True)
    writer = ExcelWriter(wb, archive)
    writer.save()

    fixed_date_time = (2025, 3, 15, 9, 0, 0)
    buf.seek(0)
    with ZipFile(buf, "r") as src, ZipFile(str(path), "w", ZIP_DEFLATED) as dst:
        for item in src.infolist():
            info = ZipInfo(item.filename, date_time=fixed_date_time)
            info.compress_type = item.compress_type
            dst.writestr(info, src.read(item.filename))


def _save_docx_deterministic(doc: Any, path: str | Path) -> None:
    """Save a python-docx Document with fixed zip entry timestamps for determinism."""
    import io
    from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

    path = Path(path)

    # Save to in-memory buffer first
    buf = io.BytesIO()
    doc.save(buf)

    # Re-pack with fixed timestamps
    fixed_date_time = (2025, 3, 15, 9, 0, 0)
    buf.seek(0)
    with ZipFile(buf, "r") as src, ZipFile(str(path), "w", ZIP_DEFLATED) as dst:
        for item in src.infolist():
            info = ZipInfo(item.filename, date_time=fixed_date_time)
            info.compress_type = item.compress_type
            dst.writestr(info, src.read(item.filename))


def _whole_dollars(d: Decimal) -> int:
    """Round a Decimal to whole dollars."""
    return int(d.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _generate_confirmations(
    aging: list[ARAgingEntry],
    seed: int,
) -> list[dict[str, Any]]:
    """Generate deterministic AR confirmation results per customer.

    Confirmation outcomes are derived from customer DSO:
    - DSO <= 40: Agreed (clean)
    - DSO 41-60: Agreed with exceptions
    - DSO > 60: No response (for some) or Agreed with exceptions
    """
    rng = random.Random(seed + 505)  # TC-05 specific sub-seed
    results: list[dict[str, Any]] = []

    for entry in aging:
        total = _whole_dollars(entry.total)
        if total == 0:
            continue

        # Determine confirmation outcome based on DSO
        if entry.dso <= 40:
            status = "Agreed"
            confirmed_amount = total
            exception_amount = 0
        elif entry.dso <= 60:
            status = "Agreed with exceptions"
            # Small exception: 2-5% of total
            exception_pct = Decimal(str(rng.randint(2, 5))) / Decimal(100)
            exception_amount = _whole_dollars(Decimal(total) * exception_pct)
            confirmed_amount = total - exception_amount
        else:
            # High DSO customers: 50% chance of no response
            if rng.random() < 0.5:
                status = "No response"
                confirmed_amount = 0
                exception_amount = 0
            else:
                status = "Agreed with exceptions"
                exception_pct = Decimal(str(rng.randint(5, 10))) / Decimal(100)
                exception_amount = _whole_dollars(Decimal(total) * exception_pct)
                confirmed_amount = total - exception_amount

        results.append({
            "customer_id": entry.customer_id,
            "customer_name": entry.customer_name,
            "entity": entry.entity_code,
            "confirmation_sent": "Yes",
            "confirmation_amount": total,
            "status": status,
            "confirmed_amount": confirmed_amount,
            "exception_amount": exception_amount,
            "exception_notes": (
                "Timing difference — payment in transit"
                if exception_amount > 0 else ""
            ),
            "alternative_procedures": (
                "Subsequent receipts verified" if status == "No response" else ""
            ),
        })

    results.sort(key=lambda r: (r["entity"], r["customer_id"]))
    return results


# ── AR Aging Schedule ────────────────────────────────────────────────────────


def _write_ar_aging(
    aging: list[ARAgingEntry],
    output_dir: Path,
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    manifest: Manifest,
) -> None:
    """Write ar_aging_fy2025.xlsx — AR aging by customer with bucket detail."""
    wb = openpyxl.Workbook()
    _pin_xlsx_dates(wb)
    ws = wb.active
    ws.title = "AR Aging"

    # Canary
    canary_code = canaries.canary_for("ar_aging_fy2025")
    location = embed_canary_xlsx(wb, canary_code)
    canaries.set_location(
        "ar_aging_fy2025",
        f"{_INPUT_DIR}/ar_aging_fy2025.xlsx",
        location,
    )

    # Header
    ws["A1"] = "Cascade Industries, Inc."
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "Accounts Receivable Aging Schedule"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A3"] = "As of December 31, 2025"
    ws["A3"].font = Font(italic=True, size=10)

    header_row = 5
    headers = [
        "Customer ID", "Customer Name", "Entity",
        "Current (0-30)", "31-60 Days", "61-90 Days",
        "91-120 Days", "120+ Days", "Total AR",
    ]
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A3C6E")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    money_fmt = "#,##0"
    data_font = Font(size=10)
    row = header_row + 1

    # ERR-020: pick the 3rd customer row to blank out its "Total AR" cell
    err020_target_idx = 2  # 0-based index into aging list

    grand_total = {
        "current": 0, "days_30": 0, "days_60": 0,
        "days_90": 0, "days_120_plus": 0, "total": 0,
    }

    for idx, entry in enumerate(aging):
        cur = _whole_dollars(entry.current)
        d30 = _whole_dollars(entry.days_30)
        d60 = _whole_dollars(entry.days_60)
        d90 = _whole_dollars(entry.days_90)
        d120 = _whole_dollars(entry.days_120_plus)
        total = cur + d30 + d60 + d90 + d120

        ws.cell(row=row, column=1, value=entry.customer_id).font = data_font
        ws.cell(row=row, column=2, value=entry.customer_name).font = data_font
        ws.cell(row=row, column=3, value=entry.entity_code).font = data_font

        # ERR-020: blank out the Total AR cell for the target customer
        if idx == err020_target_idx:
            correct_total = total
            corrupted_total = missing_data()  # returns None → blank cell
            values = [cur, d30, d60, d90, d120, corrupted_total]
            errors.add(PlantedError(
                error_id="ERR-020",
                file=f"{_INPUT_DIR}/ar_aging_fy2025.xlsx",
                location=(
                    f"Sheet 'AR Aging', Row {row}, Column I (Total AR) "
                    f"for customer {entry.customer_name}"
                ),
                type="missing_data",
                description=(
                    f"Total AR balance is blank for {entry.customer_name} "
                    f"instead of ${correct_total:,}"
                ),
                severity="immaterial",
                which_test_cases_should_catch=["TC-05"],
            ))
        else:
            values = [cur, d30, d60, d90, d120, total]

        for c, val in enumerate(values, 4):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = data_font
            cell.number_format = money_fmt

        grand_total["current"] += cur
        grand_total["days_30"] += d30
        grand_total["days_60"] += d60
        grand_total["days_90"] += d90
        grand_total["days_120_plus"] += d120
        grand_total["total"] += total
        row += 1

    # Totals row
    ws.cell(row=row, column=2, value="TOTAL").font = Font(bold=True, size=10)
    for c, key in enumerate(
        ["current", "days_30", "days_60", "days_90", "days_120_plus", "total"], 4
    ):
        cell = ws.cell(row=row, column=c, value=grand_total[key])
        cell.font = Font(bold=True, size=10)
        cell.number_format = money_fmt

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    for col_letter in ["D", "E", "F", "G", "H", "I"]:
        ws.column_dimensions[col_letter].width = 16

    path = output_dir / _INPUT_DIR / "ar_aging_fy2025.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    _save_xlsx_deterministic(wb, path)

    manifest.register(
        f"{_INPUT_DIR}/ar_aging_fy2025.xlsx",
        "xlsx",
        canary=canary_code,
        test_cases=[_TC],
    )


# ── AR Confirmations Summary ────────────────────────────────────────────────


def _write_confirmations(
    confirmations: list[dict[str, Any]],
    output_dir: Path,
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    manifest: Manifest,
) -> None:
    """Write ar_confirmations_summary.xlsx."""
    wb = openpyxl.Workbook()
    _pin_xlsx_dates(wb)
    ws = wb.active
    ws.title = "Confirmations"

    # Canary
    canary_code = canaries.canary_for("ar_confirmations_summary")
    location = embed_canary_xlsx(wb, canary_code)
    canaries.set_location(
        "ar_confirmations_summary",
        f"{_INPUT_DIR}/ar_confirmations_summary.xlsx",
        location,
    )

    # Header
    ws["A1"] = "Cascade Industries, Inc."
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "AR Confirmation Results Summary — FY2025 Audit"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A3"] = "Confirmation date: January 15, 2026"
    ws["A3"].font = Font(italic=True, size=10)

    header_row = 5
    headers = [
        "Customer ID", "Customer Name", "Entity", "Sent",
        "Response Date", "Confirmation Amount", "Status", "Confirmed Amount",
        "Exception Amount", "Exception Notes", "Alternative Procedures",
    ]
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A3C6E")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    money_fmt = "#,##0"
    data_font = Font(size=10)
    row = header_row + 1

    totals = {"sent": 0, "conf_amt": 0, "confirmed": 0, "exception": 0}
    status_counts = {"Agreed": 0, "Agreed with exceptions": 0, "No response": 0}

    # ERR-016: pick the 2nd confirmation row to corrupt its response date
    err016_target_idx = 1  # 0-based index
    correct_response_date = "01/15/2026"
    wrong_response_date = date_inconsistency(correct_response_date, "02/15/2026")

    for conf_idx, conf in enumerate(confirmations):
        ws.cell(row=row, column=1, value=conf["customer_id"]).font = data_font
        ws.cell(row=row, column=2, value=conf["customer_name"]).font = data_font
        ws.cell(row=row, column=3, value=conf["entity"]).font = data_font
        ws.cell(row=row, column=4, value=conf["confirmation_sent"]).font = data_font

        # ERR-016: corrupt response date for one customer
        if conf_idx == err016_target_idx:
            response_date = wrong_response_date
            errors.add(PlantedError(
                error_id="ERR-016",
                file=f"{_INPUT_DIR}/ar_confirmations_summary.xlsx",
                location=(
                    f"Sheet 'Confirmations', Row {row}, Column E (Response Date) "
                    f"for customer {conf['customer_name']}"
                ),
                type="date_inconsistency",
                description=(
                    f"Confirmation response date for {conf['customer_name']} "
                    f"shows {wrong_response_date} instead of {correct_response_date} "
                    f"(inconsistent with AR aging date of 12/31/2025)"
                ),
                severity="immaterial",
                which_test_cases_should_catch=["TC-05"],
            ))
        else:
            response_date = correct_response_date

        ws.cell(row=row, column=5, value=response_date).font = data_font

        c6 = ws.cell(row=row, column=6, value=conf["confirmation_amount"])
        c6.font = data_font
        c6.number_format = money_fmt

        ws.cell(row=row, column=7, value=conf["status"]).font = data_font

        c8 = ws.cell(row=row, column=8, value=conf["confirmed_amount"])
        c8.font = data_font
        c8.number_format = money_fmt

        c9 = ws.cell(row=row, column=9, value=conf["exception_amount"])
        c9.font = data_font
        c9.number_format = money_fmt

        ws.cell(row=row, column=10, value=conf["exception_notes"]).font = data_font
        ws.cell(row=row, column=11, value=conf["alternative_procedures"]).font = data_font

        totals["sent"] += 1
        totals["conf_amt"] += conf["confirmation_amount"]
        totals["confirmed"] += conf["confirmed_amount"]
        totals["exception"] += conf["exception_amount"]
        status_counts[conf["status"]] += 1
        row += 1

    # Totals row
    ws.cell(row=row, column=2, value="TOTAL").font = Font(bold=True, size=10)
    ws.cell(row=row, column=4, value=totals["sent"]).font = Font(bold=True, size=10)
    c6t = ws.cell(row=row, column=6, value=totals["conf_amt"])
    c6t.font = Font(bold=True, size=10)
    c6t.number_format = money_fmt
    c8t = ws.cell(row=row, column=8, value=totals["confirmed"])
    c8t.font = Font(bold=True, size=10)
    c8t.number_format = money_fmt
    c9t = ws.cell(row=row, column=9, value=totals["exception"])
    c9t.font = Font(bold=True, size=10)
    c9t.number_format = money_fmt

    # Summary row
    row += 2
    ws.cell(row=row, column=1, value="Summary").font = Font(bold=True, size=11)
    row += 1
    ws.cell(row=row, column=1, value="Confirmations sent:").font = data_font
    ws.cell(row=row, column=2, value=totals["sent"]).font = data_font
    row += 1
    ws.cell(row=row, column=1, value="Agreed:").font = data_font
    ws.cell(row=row, column=2, value=status_counts["Agreed"]).font = data_font
    row += 1
    ws.cell(row=row, column=1, value="Agreed with exceptions:").font = data_font
    ws.cell(row=row, column=2, value=status_counts["Agreed with exceptions"]).font = data_font
    row += 1
    ws.cell(row=row, column=1, value="No response:").font = data_font
    ws.cell(row=row, column=2, value=status_counts["No response"]).font = data_font
    row += 1
    ws.cell(row=row, column=1, value="Response rate:").font = data_font
    responded = status_counts["Agreed"] + status_counts["Agreed with exceptions"]
    rate = (responded / totals["sent"] * 100) if totals["sent"] > 0 else 0
    ws.cell(row=row, column=2, value=f"{rate:.1f}%").font = data_font

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 24
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["J"].width = 35
    ws.column_dimensions["K"].width = 30

    path = output_dir / _INPUT_DIR / "ar_confirmations_summary.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    _save_xlsx_deterministic(wb, path)

    manifest.register(
        f"{_INPUT_DIR}/ar_confirmations_summary.xlsx",
        "xlsx",
        canary=canary_code,
        test_cases=[_TC],
    )

    return status_counts, totals


# ── Allowance Analysis ──────────────────────────────────────────────────────


def _write_allowance(
    aging: list[ARAgingEntry],
    allowance_data: list[AllowanceAnalysis],
    output_dir: Path,
    canaries: CanaryRegistry,
    manifest: Manifest,
) -> dict[str, Any]:
    """Write allowance_analysis.xlsx — historical bad debt and allowance.

    Returns summary data for gold standard.
    """
    wb = openpyxl.Workbook()
    _pin_xlsx_dates(wb)

    # Canary
    canary_code = canaries.canary_for("allowance_analysis")
    location = embed_canary_xlsx(wb, canary_code)
    canaries.set_location(
        "allowance_analysis",
        f"{_INPUT_DIR}/allowance_analysis.xlsx",
        location,
    )

    # ── Sheet 1: Allowance Rollforward ──────────────────────────────
    ws1 = wb.active
    ws1.title = "Allowance Rollforward"

    ws1["A1"] = "Cascade Industries, Inc."
    ws1["A1"].font = Font(bold=True, size=12)
    ws1["A2"] = "Allowance for Doubtful Accounts — Rollforward"
    ws1["A2"].font = Font(bold=True, size=11)

    header_row = 4
    headers = ["Entity", "Year", "Beginning Balance", "Provision", "Ending Balance"]
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A3C6E")
    money_fmt = "#,##0"
    data_font = Font(size=10)

    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row = header_row + 1
    for rec in allowance_data:
        ws1.cell(row=row, column=1, value=rec.entity_code).font = data_font
        ws1.cell(row=row, column=2, value=rec.year).font = data_font
        for c, val in enumerate(
            [rec.beginning_balance, rec.provision, rec.ending_balance], 3
        ):
            cell = ws1.cell(row=row, column=c, value=_whole_dollars(val))
            cell.font = data_font
            cell.number_format = money_fmt
        row += 1

    # FY2025 consolidated totals
    fy2025_allowance = [r for r in allowance_data if r.year == 2025]
    total_ending = sum(_whole_dollars(r.ending_balance) for r in fy2025_allowance)
    total_provision = sum(_whole_dollars(r.provision) for r in fy2025_allowance)

    row += 1
    ws1.cell(row=row, column=1, value="CONSOLIDATED FY2025").font = Font(
        bold=True, size=10
    )
    c4 = ws1.cell(row=row, column=4, value=total_provision)
    c4.font = Font(bold=True, size=10)
    c4.number_format = money_fmt
    c5 = ws1.cell(row=row, column=5, value=total_ending)
    c5.font = Font(bold=True, size=10)
    c5.number_format = money_fmt

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 10
    ws1.column_dimensions["C"].width = 20
    ws1.column_dimensions["D"].width = 16
    ws1.column_dimensions["E"].width = 18

    # ── Sheet 2: Reserve Rate Analysis ──────────────────────────────
    ws2 = wb.create_sheet("Reserve Rates")

    ws2["A1"] = "Bad Debt Reserve Rates by Aging Bucket"
    ws2["A1"].font = Font(bold=True, size=11)

    headers2 = [
        "Aging Bucket", "Reserve Rate", "FY2025 AR Balance",
        "FY2025 Reserve",
    ]
    header_row2 = 3
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=header_row2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    bucket_labels = {
        "current": "Current (0-30 days)",
        "days_30": "31-60 Days",
        "days_60": "61-90 Days",
        "days_90": "91-120 Days",
        "days_120_plus": "120+ Days",
    }

    row2 = header_row2 + 1
    reserve_detail: dict[str, dict[str, int]] = {}

    for bucket_key, label in bucket_labels.items():
        rate = RESERVE_RATES[bucket_key]
        bucket_total = sum(
            _whole_dollars(getattr(e, bucket_key)) for e in aging
        )
        reserve = _whole_dollars(Decimal(bucket_total) * rate)

        ws2.cell(row=row2, column=1, value=label).font = data_font
        ws2.cell(row=row2, column=2, value=f"{float(rate)*100:.0f}%").font = data_font
        c3 = ws2.cell(row=row2, column=3, value=bucket_total)
        c3.font = data_font
        c3.number_format = money_fmt
        c4 = ws2.cell(row=row2, column=4, value=reserve)
        c4.font = data_font
        c4.number_format = money_fmt

        reserve_detail[bucket_key] = {
            "ar_balance": bucket_total,
            "rate_pct": float(rate) * 100,
            "reserve": reserve,
        }
        row2 += 1

    # Total row
    total_ar_aging = sum(d["ar_balance"] for d in reserve_detail.values())
    total_reserve = sum(d["reserve"] for d in reserve_detail.values())
    ws2.cell(row=row2, column=1, value="TOTAL").font = Font(bold=True, size=10)
    c3t = ws2.cell(row=row2, column=3, value=total_ar_aging)
    c3t.font = Font(bold=True, size=10)
    c3t.number_format = money_fmt
    c4t = ws2.cell(row=row2, column=4, value=total_reserve)
    c4t.font = Font(bold=True, size=10)
    c4t.number_format = money_fmt

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 18

    path = output_dir / _INPUT_DIR / "allowance_analysis.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    _save_xlsx_deterministic(wb, path)

    manifest.register(
        f"{_INPUT_DIR}/allowance_analysis.xlsx",
        "xlsx",
        canary=canary_code,
        test_cases=[_TC],
    )

    return {
        "total_ending_allowance": total_ending,
        "total_provision_fy2025": total_provision,
        "total_ar_aging": total_ar_aging,
        "total_reserve_calculated": total_reserve,
        "reserve_detail": reserve_detail,
    }


# ── Copy template ───────────────────────────────────────────────────────────


def _copy_template(
    output_dir: Path,
    canaries: CanaryRegistry,
    manifest: Manifest,
) -> None:
    """Copy workpaper_memo_template.docx into TC-05 input_files and embed canary.

    Reads the static template from the repo's templates/ directory (next to
    the project root), embeds a canary, and writes to the output tree.
    """
    from docx import Document

    canary_code = canaries.canary_for("workpaper_memo_template")

    # Locate the repo-root templates/ directory
    repo_root = Path(__file__).resolve().parent.parent.parent
    template_path = repo_root / "templates" / "workpaper_memo_template.docx"

    doc = Document(str(template_path))
    location = embed_canary_docx(doc, canary_code)
    canaries.set_location(
        "workpaper_memo_template",
        f"{_INPUT_DIR}/workpaper_memo_template.docx",
        location,
    )

    # Pin core-properties dates for determinism
    doc.core_properties.created = _FIXED_DATETIME
    doc.core_properties.modified = _FIXED_DATETIME

    # ── Controlled noise (docx_python_docx family pilot) ────────────
    # No paragraphs need exclusion — no planted errors in the template.
    # Canary lives in core_properties.comments, which noise never touches.
    noise_rng = make_noise_rng(
        ScenarioContext(seed=42), _TC, "workpaper_memo_template",
    )
    apply_docx_noise(doc, noise_rng)

    dest = output_dir / _INPUT_DIR / "workpaper_memo_template.docx"
    dest.parent.mkdir(parents=True, exist_ok=True)
    _save_docx_deterministic(doc, dest)

    manifest.register(
        f"{_INPUT_DIR}/workpaper_memo_template.docx",
        "docx",
        canary=canary_code,
        test_cases=[_TC],
    )


# ── Prompt & Expected Behavior ──────────────────────────────────────────────


def _write_prompt(output_dir: Path) -> None:
    """Write test_cases/TC-05/prompt.md per spec."""
    text = """\
Draft the accounts receivable workpaper memo for the FY2025 audit of Cascade Industries.

Use the firm template provided. The memo should document:
- The audit objective for AR
- Procedures performed (confirmation, aging analysis, subsequent receipts,
  allowance assessment)
- Key findings from the aging, confirmation results, and allowance analysis
- Your conclusion on whether the AR balance is fairly stated

Reference specific data from the provided workpapers. Keep the tone professional
and consistent with Big 4 audit documentation standards.

Save as a Word document.
"""
    path = output_dir / f"test_cases/{_TC}/prompt.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text)


def _write_expected_behavior(output_dir: Path) -> None:
    """Write test_cases/TC-05/expected_behavior.md per spec."""
    text = """\
# TC-05: Audit Workpaper Memo — Accounts Receivable — Expected Behavior

## Template Usage
- The agent should use the firm template (workpaper_memo_template.docx) as a
  starting point, filling in all placeholder sections.
- All metadata fields should be populated: Client, Engagement, Fiscal Year,
  Prepared By, Reviewed By.

## Required Sections
The memo must include all template sections properly completed:
1. **Objective**: State the audit objective for AR (existence, completeness,
   valuation, presentation).
2. **Scope**: Define the population tested, materiality thresholds, and period.
3. **Procedures Performed**: Document each procedure:
   - AR confirmation procedures and results
   - Aging analysis and concentration review
   - Subsequent receipts testing
   - Allowance for doubtful accounts assessment
4. **Findings**: Reference specific data from the workpapers:
   - Total AR balance and aging distribution
   - Confirmation response rate and exceptions
   - Customer concentration (top customer percentage)
   - Allowance adequacy assessment
5. **Conclusion**: Professional audit conclusion using appropriate language
   ("we obtained sufficient appropriate audit evidence...").

## Data References
- The memo should reference specific numbers from the AR aging, confirmation
  summary, and allowance analysis workpapers.
- Customer concentration risk should be identified (top customer ~18% of
  consolidated revenue).
- Any aged receivables (90+ days) should be discussed.

## Professional Standards
- Tone should be consistent with Big 4 audit documentation.
- Use audit-specific language ("procedures performed", "sufficient appropriate
  audit evidence", "fairly stated in all material respects").
- Logical flow from objective → procedures → findings → conclusion.

## Output Quality
- Output must be a valid .docx file that opens without errors.
- Template formatting should be preserved.
"""
    path = output_dir / f"test_cases/{_TC}/expected_behavior.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text)


# ── Gold Standard ────────────────────────────────────────────────────────────


@register_gold("TC-05")
def _tc05_gold(
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    **model_kwargs: Any,
) -> GoldStandard:
    """Build the TC-05 gold standard from the canonical model."""
    model: CascadeModel = model_kwargs["model"]

    aging = generate_ar_aging(model.revenue_records, year=2025)
    allowance_data = generate_allowance(model.revenue_records)
    confirmations = _generate_confirmations(aging, seed=42)

    # Compute totals from aging
    total_ar = sum(_whole_dollars(e.total) for e in aging)
    current_bucket = sum(_whole_dollars(e.current) for e in aging)
    aged_90_plus = sum(
        _whole_dollars(e.days_90) + _whole_dollars(e.days_120_plus)
        for e in aging
    )

    # Top customer concentration
    top_customer = max(aging, key=lambda e: e.total)
    top_customer_pct = (
        float(top_customer.total / Decimal(total_ar) * 100)
        if total_ar > 0 else 0
    )

    # Confirmation stats
    total_sent = len(confirmations)
    agreed = sum(1 for c in confirmations if c["status"] == "Agreed")
    agreed_with_exceptions = sum(
        1 for c in confirmations if c["status"] == "Agreed with exceptions"
    )
    no_response = sum(
        1 for c in confirmations if c["status"] == "No response"
    )
    response_rate = (
        (agreed + agreed_with_exceptions) / total_sent * 100
        if total_sent > 0 else 0
    )

    # Allowance
    fy2025_allowance = [r for r in allowance_data if r.year == 2025]
    total_allowance = sum(
        _whole_dollars(r.ending_balance) for r in fy2025_allowance
    )

    return GoldStandard(
        test_case="TC-05",
        expected_outputs={
            "file_type": "docx",
            "required_sections": [
                "Objective", "Scope", "Procedures Performed",
                "Findings", "Conclusion",
            ],
            "ar_data": {
                "total_ar_balance": total_ar,
                "current_bucket": current_bucket,
                "aged_90_plus": aged_90_plus,
                "customer_count": len(aging),
            },
            "concentration": {
                "top_customer_name": top_customer.customer_name,
                "top_customer_ar": _whole_dollars(top_customer.total),
                "top_customer_pct_of_ar": round(top_customer_pct, 1),
            },
            "confirmations": {
                "total_sent": total_sent,
                "agreed": agreed,
                "agreed_with_exceptions": agreed_with_exceptions,
                "no_response": no_response,
                "response_rate_pct": round(response_rate, 1),
            },
            "allowance": {
                "total_ending_allowance": total_allowance,
            },
        },
        canary_verification={
            "read_aging": canaries.canary_for("ar_aging_fy2025"),
            "read_confirmations": canaries.canary_for("ar_confirmations_summary"),
            "read_allowance": canaries.canary_for("allowance_analysis"),
            "read_template": canaries.canary_for("workpaper_memo_template"),
        },
        error_detection={
            "ERR-016": "date_inconsistency in ar_confirmations_summary.xlsx",
            "ERR-020": "missing_data in ar_aging_fy2025.xlsx",
        },
        scoring_hints={
            "correctness": "Data references must match gold standard values from workpapers",
            "completeness": "All 5 template sections filled; all data sources referenced",
            "format_compliance": "Valid docx using the firm template; formatting preserved",
            "communication": (
                "Professional audit language; logical flow; identified concentration "
                "risk and aged receivables"
            ),
        },
    )


# ── Public entry point ──────────────────────────────────────────────────────


def emit_tc05(
    model: CascadeModel,
    output_dir: Path,
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    manifest: Manifest,
    **kwargs: object,
) -> None:
    """Write all TC-05 files to *output_dir*."""
    aging = generate_ar_aging(model.revenue_records, year=2025)
    allowance_data = generate_allowance(model.revenue_records)
    confirmations = _generate_confirmations(aging, seed=42)

    _write_ar_aging(aging, output_dir, canaries, errors, manifest)
    _write_confirmations(confirmations, output_dir, canaries, errors, manifest)
    _write_allowance(aging, allowance_data, output_dir, canaries, manifest)
    _copy_template(output_dir, canaries, manifest)
    _write_prompt(output_dir)
    _write_expected_behavior(output_dir)
