"""Formatter: TC-15 — DCF Valuation (Advisory, Complex).

Emits:
- test_cases/TC-15/input_files/historical_financials_3yr.xlsx
  3 years of income statement, balance sheet, cash flow (from canonical model)
- test_cases/TC-15/input_files/management_projections.xlsx
  5-year revenue and EBITDA projections with assumptions
- test_cases/TC-15/input_files/comparable_companies_trading.xlsx
  10 comparable public companies with trading multiples
- test_cases/TC-15/input_files/industry_overview.pdf
  15-page industry report with growth forecasts and risk factors
- test_cases/TC-15/prompt.md
- test_cases/TC-15/expected_behavior.md
- gold_standards/TC-15_gold.json

Planted errors:
- ERR-022 (missing_data): One comparable company has a blank EV/EBITDA field
- ERR-025 (mismatched_total): FY2024 total revenue doesn't match component sum
Uses the canonical model — never hardcodes numbers.
"""

from __future__ import annotations

import datetime
import io
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from generator.canaries import CanaryRegistry, embed_canary_xlsx
from generator.errors import ErrorRegistry, PlantedError, mismatch_total, missing_data
from generator.golds.framework import GoldStandard, register_gold
from generator.manifest import Manifest
from generator.model.build import CascadeModel
from generator.model.consolidation import (
    build_balance_sheet,
    build_cash_flow,
    build_income_statement,
)

# ── Constants ────────────────────────────────────────────────────────────────

_TC = "TC-15"
_INPUT_DIR = f"test_cases/{_TC}/input_files"

_FIXED_DATETIME = datetime.datetime(2025, 3, 15, 9, 0, 0)
_FIXED_ZIP_DT = (2025, 3, 15, 9, 0, 0)

# DCF parameters from prompt.md TC-15
_RISK_FREE_RATE = Decimal("0.042")     # 4.2%
_EQUITY_RISK_PREMIUM = Decimal("0.055")  # 5.5%
_SIZE_PREMIUM = Decimal("0.020")       # 2.0%
_PERPETUITY_GROWTH = Decimal("0.025")  # 2.5%

# Projection years
_PROJECTION_YEARS = [2026, 2027, 2028, 2029, 2030]

# Comparable companies — synthetic data for 10 public companies.
# Designed so median D/E ≈ 0.25, median EV/EBITDA ≈ 8.5x, median beta ≈ 1.15.
# Format: (name, ticker, market_cap_M, ev_M, revenue_M, ebitda_M, net_income_M, beta)
_COMPARABLE_COMPANIES: list[tuple[str, str, int, int, int, int, int, float]] = [
    ("Precision Manufacturing Corp", "PMC", 450, 585, 406, 65, 26, 1.15),
    ("Advanced Industrial Holdings", "AIH", 680, 830, 582, 99, 42, 1.08),
    ("MidWest Materials Group", "MMG", 320, 432, 307, 43, 16, 1.25),
    ("Pacific Components Inc", "PCI", 510, 612, 469, 75, 32, 1.12),
    ("National Parts & Supply", "NPS", 290, 406, 300, 39, 14, 1.30),
    ("Summit Manufacturing Ltd", "SML", 420, 538, 394, 63, 25, 1.18),
    ("Continental Industrial Corp", "CIC", 750, 840, 622, 112, 50, 1.05),
    ("Delta Precision Systems", "DPS", 380, 502, 353, 53, 20, 1.22),
    ("Atlas Materials Technology", "AMT", 560, 644, 488, 83, 36, 1.10),
    ("Cascade Peer Industries", "CPI", 480, 528, 469, 75, 32, 1.14),
]

# Management projection assumptions — growth rates and margin targets.
# Revenue grows from historical ~9% trending down to 5% by Year 5.
# EBITDA margin starts at a target base (management's view of normalized
# profitability after cost optimization) and expands 200bps over 5 years —
# the aggressive assumption the agent should flag.
_PROJ_REVENUE_GROWTH = [
    Decimal("0.080"),  # Y1: 8%
    Decimal("0.070"),  # Y2: 7%
    Decimal("0.060"),  # Y3: 6%
    Decimal("0.055"),  # Y4: 5.5%
    Decimal("0.050"),  # Y5: 5%
]

# Base EBITDA margin for Year 1 projections.  Management presents this
# as the "normalized" margin after planned cost restructuring.
_PROJ_EBITDA_MARGIN_BASE = Decimal("0.150")   # 15.0%
_PROJ_EBITDA_MARGIN_TARGET = Decimal("0.170")  # 17.0% = base + 200bps

# Capex as % of revenue
_PROJ_CAPEX_PCT = [
    Decimal("0.040"),  # Y1: 4.0%
    Decimal("0.038"),  # Y2: 3.8%
    Decimal("0.035"),  # Y3: 3.5%
    Decimal("0.035"),  # Y4: 3.5%
    Decimal("0.035"),  # Y5: 3.5%
]


# ── Deterministic save helpers ───────────────────────────────────────────────


def _save_xlsx_deterministic(wb: openpyxl.Workbook, path: str | Path) -> None:
    """Save workbook with pinned timestamps and fixed zip entry dates."""
    from openpyxl.writer.excel import ExcelWriter

    path = Path(path)
    wb.properties.modified = _FIXED_DATETIME

    buf = io.BytesIO()
    archive = ZipFile(buf, "w", ZIP_DEFLATED, allowZip64=True)
    writer = ExcelWriter(wb, archive)
    writer.save()

    buf.seek(0)
    with ZipFile(buf, "r") as src, ZipFile(str(path), "w", ZIP_DEFLATED) as dst:
        for item in src.infolist():
            info = ZipInfo(item.filename, date_time=_FIXED_ZIP_DT)
            info.compress_type = item.compress_type
            dst.writestr(info, src.read(item.filename))


def _whole_dollars(d: Decimal) -> int:
    """Round a Decimal to whole dollars."""
    return int(d.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _millions(d: Decimal) -> float:
    """Convert to millions with 1 decimal."""
    return float((d / Decimal("1000000")).quantize(
        Decimal("0.1"), rounding=ROUND_HALF_UP
    ))


# ── Style helpers ────────────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_NUM_FMT = '#,##0'
_PCT_FMT = '0.0%'


def _style_header_row(ws: Any, row: int, cols: int) -> None:
    """Apply header styling to a row."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT_WHITE
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center")


def _style_data_cell(cell: Any, fmt: str = _NUM_FMT) -> None:
    """Apply data styling to a cell."""
    cell.border = _THIN_BORDER
    cell.number_format = fmt


# ── Historical financials ────────────────────────────────────────────────────


def _get_historical_data(
    model: CascadeModel,
) -> dict[str, Any]:
    """Extract 3 years of historical financial data from the canonical model."""
    data: dict[str, Any] = {}

    for year in [2023, 2024, 2025]:
        is_stmt = build_income_statement(model.ledger, year)
        bs = build_balance_sheet(
            model.ledger, datetime.date(year, 12, 31),
        )
        cf = build_cash_flow(model.ledger, year)

        # D&A from IS detail (accounts 6210, 6220, 6330)
        da = Decimal(0)
        for acct in ["6210", "6220", "6330"]:
            da += is_stmt.detail.get(acct, Decimal(0))

        # EBITDA = operating income + D&A
        ebitda = is_stmt.operating_income + da

        # Debt from balance sheet
        debt_accounts = ["2100", "2110", "2200", "2210"]
        total_debt = sum(bs.liabilities.get(a, Decimal(0)) for a in debt_accounts)

        # Cash
        cash_accounts = ["1010", "1020", "1030", "1050"]
        total_cash = sum(bs.assets.get(a, Decimal(0)) for a in cash_accounts)

        # Working capital
        wc_asset_accounts = [
            "1100", "1150", "1200", "1210", "1220", "1230",
            "1300", "1310", "1320", "1350",
        ]
        wc_liab_accounts = [
            "2010", "2020", "2030", "2040", "2050", "2060",
            "2070", "2075", "2080", "2120", "2130", "2150",
        ]
        wc_assets = sum(bs.assets.get(a, Decimal(0)) for a in wc_asset_accounts)
        wc_liabs = sum(bs.liabilities.get(a, Decimal(0)) for a in wc_liab_accounts)
        nwc = wc_assets - wc_liabs

        # Interest expense from IS detail (account 7010 or similar)
        interest_expense = Decimal(0)
        for acct, bal in is_stmt.detail.items():
            if acct.startswith("7"):
                interest_expense += bal

        data[year] = {
            "revenue": is_stmt.total_revenue,
            "cogs": is_stmt.total_cogs,
            "gross_profit": is_stmt.gross_profit,
            "opex": is_stmt.total_opex,
            "operating_income": is_stmt.operating_income,
            "da": da,
            "ebitda": ebitda,
            "interest_expense": interest_expense,
            "other": is_stmt.total_other,
            "pre_tax_income": is_stmt.pre_tax_income,
            "tax": is_stmt.total_tax,
            "net_income": is_stmt.net_income,
            "total_assets": bs.total_assets,
            "total_liabilities": bs.total_liabilities,
            "total_equity": bs.total_equity,
            "total_debt": total_debt,
            "total_cash": total_cash,
            "net_debt": total_debt - total_cash,
            "nwc": nwc,
            "capex": cf.capex,  # negative = outflow
            "cf_operations": cf.cash_from_operations,
            "cf_investing": cf.cash_from_investing,
            "cf_financing": cf.cash_from_financing,
            "effective_tax_rate": (
                model.tax_provisions[year].effective_tax_rate
                if year in model.tax_provisions
                else Decimal("0.25")
            ),
        }

    return data


def _write_historical_xlsx(
    model: CascadeModel,
    output_dir: Path,
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    manifest: Manifest,
) -> dict[str, Any]:
    """Write historical_financials_3yr.xlsx and return the historical data dict."""
    hist = _get_historical_data(model)

    # ERR-025: corrupt FY2024 total revenue by a small delta ($3,247)
    _err025_delta = Decimal("3247")
    _err025_correct = hist[2024]["revenue"]
    _err025_wrong = mismatch_total(_whole_dollars(_err025_correct), _whole_dollars(_err025_delta))
    # We'll write the corrupted value; store it for the IS sheet below
    _err025_corrupted_revenue = _err025_wrong

    errors.add(PlantedError(
        error_id="ERR-025",
        file=f"{_INPUT_DIR}/historical_financials_3yr.xlsx",
        location="Sheet 'Income Statement', Row 2, Column C (FY2024 Revenue)",
        type="mismatched_total",
        description=(
            f"FY2024 total revenue shows ${_err025_wrong:,} "
            f"instead of ${_whole_dollars(_err025_correct):,}"
        ),
        severity="material",
        which_test_cases_should_catch=["TC-15"],
    ))

    wb = openpyxl.Workbook()
    wb.properties.created = _FIXED_DATETIME

    # ── Income Statement sheet ──────────────────────────────
    ws = wb.active
    ws.title = "Income Statement"
    ws.column_dimensions["A"].width = 30
    for col_letter in ["B", "C", "D"]:
        ws.column_dimensions[col_letter].width = 18

    headers = ["", "FY2023", "FY2024", "FY2025"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, len(headers))

    is_rows = [
        ("Revenue", "revenue"),
        ("Cost of Goods Sold", "cogs"),
        ("Gross Profit", "gross_profit"),
        ("Operating Expenses", "opex"),
        ("  Depreciation & Amortization", "da"),
        ("Operating Income", "operating_income"),
        ("Other Income / (Expense)", "other"),
        ("Pre-Tax Income", "pre_tax_income"),
        ("Income Tax Provision", "tax"),
        ("Net Income", "net_income"),
        ("", None),
        ("EBITDA", "ebitda"),
    ]

    for r, (label, key) in enumerate(is_rows, 2):
        ws.cell(row=r, column=1, value=label).font = Font(
            bold=label in ("Gross Profit", "Operating Income", "Net Income", "EBITDA"),
        )
        if key:
            for c, year in enumerate([2023, 2024, 2025], 2):
                val = _whole_dollars(hist[year][key])
                # ERR-025: corrupt FY2024 revenue
                if key == "revenue" and year == 2024:
                    val = _err025_corrupted_revenue
                cell = ws.cell(row=r, column=c, value=val)
                _style_data_cell(cell)

    # ── Balance Sheet sheet ─────────────────────────────────
    ws_bs = wb.create_sheet("Balance Sheet")
    ws_bs.column_dimensions["A"].width = 30
    for col_letter in ["B", "C", "D"]:
        ws_bs.column_dimensions[col_letter].width = 18

    for c, h in enumerate(headers, 1):
        ws_bs.cell(row=1, column=c, value=h)
    _style_header_row(ws_bs, 1, len(headers))

    bs_rows = [
        ("Total Assets", "total_assets"),
        ("Total Liabilities", "total_liabilities"),
        ("Total Equity", "total_equity"),
        ("", None),
        ("Total Debt", "total_debt"),
        ("Cash & Equivalents", "total_cash"),
        ("Net Debt", "net_debt"),
        ("Net Working Capital", "nwc"),
    ]

    for r, (label, key) in enumerate(bs_rows, 2):
        ws_bs.cell(row=r, column=1, value=label).font = Font(
            bold=label in ("Total Assets", "Total Liabilities", "Total Equity"),
        )
        if key:
            for c, year in enumerate([2023, 2024, 2025], 2):
                cell = ws_bs.cell(row=r, column=c, value=_whole_dollars(hist[year][key]))
                _style_data_cell(cell)

    # ── Cash Flow sheet ─────────────────────────────────────
    ws_cf = wb.create_sheet("Cash Flow")
    ws_cf.column_dimensions["A"].width = 30
    for col_letter in ["B", "C", "D"]:
        ws_cf.column_dimensions[col_letter].width = 18

    for c, h in enumerate(headers, 1):
        ws_cf.cell(row=1, column=c, value=h)
    _style_header_row(ws_cf, 1, len(headers))

    cf_rows = [
        ("Cash from Operations", "cf_operations"),
        ("Capital Expenditures", "capex"),
        ("Cash from Investing", "cf_investing"),
        ("Cash from Financing", "cf_financing"),
    ]

    for r, (label, key) in enumerate(cf_rows, 2):
        ws_cf.cell(row=r, column=1, value=label)
        for c, year in enumerate([2023, 2024, 2025], 2):
            cell = ws_cf.cell(row=r, column=c, value=_whole_dollars(hist[year][key]))
            _style_data_cell(cell)

    # ── Key Ratios sheet ────────────────────────────────────
    ws_r = wb.create_sheet("Key Ratios")
    ws_r.column_dimensions["A"].width = 30
    for col_letter in ["B", "C", "D"]:
        ws_r.column_dimensions[col_letter].width = 18

    for c, h in enumerate(headers, 1):
        ws_r.cell(row=1, column=c, value=h)
    _style_header_row(ws_r, 1, len(headers))

    ratio_rows = [
        "Gross Margin",
        "EBITDA Margin",
        "Operating Margin",
        "Net Margin",
        "Effective Tax Rate",
        "NWC / Revenue",
        "CapEx / Revenue",
    ]

    for r, label in enumerate(ratio_rows, 2):
        ws_r.cell(row=r, column=1, value=label)
        for c, year in enumerate([2023, 2024, 2025], 2):
            h = hist[year]
            rev = h["revenue"]
            if rev == 0:
                continue
            if label == "Gross Margin":
                val = h["gross_profit"] / rev
            elif label == "EBITDA Margin":
                val = h["ebitda"] / rev
            elif label == "Operating Margin":
                val = h["operating_income"] / rev
            elif label == "Net Margin":
                val = h["net_income"] / rev
            elif label == "Effective Tax Rate":
                val = h["effective_tax_rate"]
            elif label == "NWC / Revenue":
                val = h["nwc"] / rev
            elif label == "CapEx / Revenue":
                val = abs(h["capex"]) / rev
            else:
                val = Decimal(0)
            cell = ws_r.cell(row=r, column=c, value=float(val))
            _style_data_cell(cell, _PCT_FMT)

    # Embed canary and save
    canary_code = canaries.canary_for("tc15_historical_financials")
    loc = embed_canary_xlsx(wb, canary_code)
    canaries.set_location(
        "tc15_historical_financials",
        f"{_INPUT_DIR}/historical_financials_3yr.xlsx",
        loc,
    )
    file_path = output_dir / _INPUT_DIR / "historical_financials_3yr.xlsx"
    file_path.parent.mkdir(parents=True, exist_ok=True)
    _save_xlsx_deterministic(wb, file_path)

    manifest.register(f"{_INPUT_DIR}/historical_financials_3yr.xlsx", "xlsx")
    return hist


# ── Management projections ───────────────────────────────────────────────────


def _compute_projections(
    hist: dict[str, Any],
) -> dict[int, dict[str, Any]]:
    """Compute 5-year management projections from historical base.

    Management assumptions:
    - Revenue growth declines from 8% to 5%
    - EBITDA margin starts at 15.5% (management's "normalized" target after
      planned cost restructuring) and expands 200bps to 17.5% by Year 5
    - CapEx decreases as % of revenue
    - NWC/Revenue ratio stays at FY2025 level

    The margin base is management's forward view, NOT a mechanical extrapolation
    from historical EBITDA.  This is typical in advisory/deal contexts where
    management projects operational improvements.
    """
    base_rev = hist[2025]["revenue"]
    margin_step = (_PROJ_EBITDA_MARGIN_TARGET - _PROJ_EBITDA_MARGIN_BASE) / Decimal(5)

    nwc_rev_ratio = hist[2025]["nwc"] / hist[2025]["revenue"]

    projections: dict[int, dict[str, Any]] = {}
    prev_rev = base_rev
    prev_nwc = hist[2025]["nwc"]

    for i, year in enumerate(_PROJECTION_YEARS):
        growth = _PROJ_REVENUE_GROWTH[i]
        rev = (prev_rev * (1 + growth)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        margin = _PROJ_EBITDA_MARGIN_BASE + margin_step * Decimal(i + 1)
        ebitda = (rev * margin).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        capex_pct = _PROJ_CAPEX_PCT[i]
        capex = (rev * capex_pct).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        nwc = (rev * nwc_rev_ratio).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        delta_nwc = nwc - prev_nwc

        projections[year] = {
            "revenue": rev,
            "growth_rate": growth,
            "ebitda": ebitda,
            "ebitda_margin": margin,
            "capex": capex,
            "capex_pct": capex_pct,
            "nwc": nwc,
            "delta_nwc": delta_nwc,
        }

        prev_rev = rev
        prev_nwc = nwc

    return projections


def _write_projections_xlsx(
    hist: dict[str, Any],
    output_dir: Path,
    canaries: CanaryRegistry,
    manifest: Manifest,
) -> dict[int, dict[str, Any]]:
    """Write management_projections.xlsx and return the projections dict."""
    proj = _compute_projections(hist)
    wb = openpyxl.Workbook()
    wb.properties.created = _FIXED_DATETIME

    # ── Revenue Projections sheet ───────────────────────────
    ws = wb.active
    ws.title = "Revenue Projections"
    ws.column_dimensions["A"].width = 28

    years = _PROJECTION_YEARS
    headers = ["Metric"] + [str(y) for y in years]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 16
    _style_header_row(ws, 1, len(headers))

    row = 2
    ws.cell(row=row, column=1, value="Revenue ($)")
    for c, year in enumerate(years, 2):
        cell = ws.cell(row=row, column=c, value=_whole_dollars(proj[year]["revenue"]))
        _style_data_cell(cell)

    row = 3
    ws.cell(row=row, column=1, value="Revenue Growth (%)")
    for c, year in enumerate(years, 2):
        cell = ws.cell(row=row, column=c, value=float(proj[year]["growth_rate"]))
        _style_data_cell(cell, _PCT_FMT)

    # ── EBITDA Projections sheet ────────────────────────────
    ws_e = wb.create_sheet("EBITDA Projections")
    ws_e.column_dimensions["A"].width = 28

    for c, h in enumerate(headers, 1):
        ws_e.cell(row=1, column=c, value=h)
        ws_e.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 16
    _style_header_row(ws_e, 1, len(headers))

    rows_data = [
        ("EBITDA ($)", "ebitda", _NUM_FMT),
        ("EBITDA Margin (%)", "ebitda_margin", _PCT_FMT),
    ]
    for r, (label, key, fmt) in enumerate(rows_data, 2):
        ws_e.cell(row=r, column=1, value=label)
        for c, year in enumerate(years, 2):
            val = proj[year][key]
            if fmt == _NUM_FMT:
                val = _whole_dollars(val)
            else:
                val = float(val)
            cell = ws_e.cell(row=r, column=c, value=val)
            _style_data_cell(cell, fmt)

    # ── Assumptions sheet ───────────────────────────────────
    ws_a = wb.create_sheet("Assumptions")
    ws_a.column_dimensions["A"].width = 35

    for c, h in enumerate(headers, 1):
        ws_a.cell(row=1, column=c, value=h)
        ws_a.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 16
    _style_header_row(ws_a, 1, len(headers))

    assumption_rows = [
        ("Revenue Growth Rate", "growth_rate", _PCT_FMT),
        ("EBITDA Margin (target)", "ebitda_margin", _PCT_FMT),
        ("CapEx (% of Revenue)", "capex_pct", _PCT_FMT),
        ("CapEx ($)", "capex", _NUM_FMT),
        ("Change in NWC ($)", "delta_nwc", _NUM_FMT),
    ]
    for r, (label, key, fmt) in enumerate(assumption_rows, 2):
        ws_a.cell(row=r, column=1, value=label)
        for c, year in enumerate(years, 2):
            val = proj[year][key]
            if fmt == _NUM_FMT:
                val = _whole_dollars(val)
            else:
                val = float(val)
            cell = ws_a.cell(row=r, column=c, value=val)
            _style_data_cell(cell, fmt)

    # Note about margin expansion
    note_row = len(assumption_rows) + 3
    ws_a.cell(
        row=note_row, column=1,
        value="Note: EBITDA margin assumes 200bps expansion over 5 years "
        "driven by operational efficiencies and scale benefits.",
    ).font = Font(italic=True, size=10)

    canary_code = canaries.canary_for("tc15_management_projections")
    loc = embed_canary_xlsx(wb, canary_code)
    canaries.set_location(
        "tc15_management_projections",
        f"{_INPUT_DIR}/management_projections.xlsx",
        loc,
    )
    file_path = output_dir / _INPUT_DIR / "management_projections.xlsx"
    file_path.parent.mkdir(parents=True, exist_ok=True)
    _save_xlsx_deterministic(wb, file_path)

    manifest.register(f"{_INPUT_DIR}/management_projections.xlsx", "xlsx")
    return proj


# ── Comparable companies ─────────────────────────────────────────────────────


def _write_comps_xlsx(
    output_dir: Path,
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    manifest: Manifest,
) -> None:
    """Write comparable_companies_trading.xlsx."""
    wb = openpyxl.Workbook()
    wb.properties.created = _FIXED_DATETIME

    ws = wb.active
    ws.title = "Trading Comparables"

    headers = [
        "Company", "Ticker", "Market Cap ($M)", "Enterprise Value ($M)",
        "Revenue ($M)", "EBITDA ($M)", "Net Income ($M)",
        "EV/Revenue", "EV/EBITDA", "P/E", "Beta",
    ]
    ws.column_dimensions["A"].width = 30
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
        if c > 1:
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 16
    _style_header_row(ws, 1, len(headers))

    # ERR-022: blank EV/EBITDA for MidWest Materials Group (index 2)
    _ERR022_COMP_IDX = 2  # MidWest Materials Group
    _err022_correct_ev_ebitda: float | None = None

    for r, comp in enumerate(_COMPARABLE_COMPANIES, 2):
        name, ticker, mcap, ev, rev, ebitda, ni, beta = comp
        ev_rev = round(ev / rev, 1)
        ev_ebitda = round(ev / ebitda, 1)
        pe = round(mcap / ni, 1)

        # Plant ERR-022: blank EV/EBITDA for one company
        if r - 2 == _ERR022_COMP_IDX:
            _err022_correct_ev_ebitda = ev_ebitda
            ev_ebitda = missing_data()  # writes None → blank cell

        row_data = [name, ticker, mcap, ev, rev, ebitda, ni, ev_rev, ev_ebitda, pe, beta]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c >= 3:
                _style_data_cell(cell, "0.0" if c >= 8 else _NUM_FMT)

    errors.add(PlantedError(
        error_id="ERR-022",
        file=f"{_INPUT_DIR}/comparable_companies_trading.xlsx",
        location="Sheet 'Trading Comparables', Row 4, Column I (EV/EBITDA)",
        type="missing_data",
        description=(
            f"EV/EBITDA multiple is blank for MidWest Materials Group "
            f"instead of {_err022_correct_ev_ebitda}x"
        ),
        severity="immaterial",
        which_test_cases_should_catch=["TC-15"],
    ))

    # Summary statistics row
    summary_row = len(_COMPARABLE_COMPANIES) + 3
    ws.cell(row=summary_row, column=1, value="Median").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=1, value="Mean").font = Font(bold=True)

    # Compute medians and means for key columns
    ev_rev_list = sorted(
        round(comp[3] / comp[4], 1) for comp in _COMPARABLE_COMPANIES
    )
    ev_ebitda_list = sorted(
        round(comp[3] / comp[5], 1) for comp in _COMPARABLE_COMPANIES
    )
    beta_list = sorted(comp[7] for comp in _COMPARABLE_COMPANIES)
    de_list = sorted(
        round((comp[3] - comp[2]) / comp[2], 2) for comp in _COMPARABLE_COMPANIES
    )

    n = len(_COMPARABLE_COMPANIES)

    def _median(lst: list[float]) -> float:
        mid = n // 2
        if n % 2 == 0:
            return round((lst[mid - 1] + lst[mid]) / 2, 2)
        return lst[mid]

    def _mean(lst: list[float]) -> float:
        return round(sum(lst) / n, 2)

    # EV/Revenue median (col 8)
    ws.cell(row=summary_row, column=8, value=_median(ev_rev_list)).font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=8, value=_mean(ev_rev_list)).font = Font(bold=True)

    # EV/EBITDA median (col 9)
    ws.cell(row=summary_row, column=9, value=_median(ev_ebitda_list)).font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=9, value=_mean(ev_ebitda_list)).font = Font(bold=True)

    # Beta median (col 11)
    ws.cell(row=summary_row, column=11, value=_median(beta_list)).font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=11, value=_mean(beta_list)).font = Font(bold=True)

    # D/E ratios sheet
    ws_de = wb.create_sheet("Capital Structure")
    de_headers = [
        "Company", "Ticker", "Market Cap ($M)", "Enterprise Value ($M)",
        "Net Debt ($M)", "Debt/Equity", "Debt/Capital",
    ]
    ws_de.column_dimensions["A"].width = 30
    for c, h in enumerate(de_headers, 1):
        ws_de.cell(row=1, column=c, value=h)
        if c > 1:
            ws_de.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 16
    _style_header_row(ws_de, 1, len(de_headers))

    for r, comp in enumerate(_COMPARABLE_COMPANIES, 2):
        name, ticker, mcap, ev, rev, ebitda, ni, beta = comp
        net_debt = ev - mcap
        de_ratio = round(net_debt / mcap, 2)
        dc_ratio = round(net_debt / ev, 2)

        row_data = [name, ticker, mcap, ev, net_debt, de_ratio, dc_ratio]
        for c, val in enumerate(row_data, 1):
            cell = ws_de.cell(row=r, column=c, value=val)
            if c >= 3:
                _style_data_cell(cell, "0.00" if c >= 6 else _NUM_FMT)

    # Summary
    summary_de_row = len(_COMPARABLE_COMPANIES) + 3
    ws_de.cell(row=summary_de_row, column=1, value="Median").font = Font(bold=True)
    ws_de.cell(row=summary_de_row + 1, column=1, value="Mean").font = Font(bold=True)

    ws_de.cell(row=summary_de_row, column=6, value=_median(de_list)).font = Font(bold=True)
    ws_de.cell(row=summary_de_row + 1, column=6, value=_mean(de_list)).font = Font(bold=True)

    canary_code = canaries.canary_for("tc15_comparable_companies")
    loc = embed_canary_xlsx(wb, canary_code)
    canaries.set_location(
        "tc15_comparable_companies",
        f"{_INPUT_DIR}/comparable_companies_trading.xlsx",
        loc,
    )
    file_path = output_dir / _INPUT_DIR / "comparable_companies_trading.xlsx"
    file_path.parent.mkdir(parents=True, exist_ok=True)
    _save_xlsx_deterministic(wb, file_path)

    manifest.register(f"{_INPUT_DIR}/comparable_companies_trading.xlsx", "xlsx")


# ── Industry overview PDF ────────────────────────────────────────────────────


def _write_industry_pdf(
    output_dir: Path,
    canaries: CanaryRegistry,
    manifest: Manifest,
) -> None:
    """Write a 15-page synthetic industry overview PDF."""
    canary = canaries.canary_for("tc15_industry_overview")
    file_path = output_dir / _INPUT_DIR / "industry_overview.pdf"
    file_path.parent.mkdir(parents=True, exist_ok=True)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=letter,
        leftMargin=inch,
        rightMargin=inch,
        topMargin=inch,
        bottomMargin=inch,
        invariant=True,
    )
    doc.title = "U.S. Mid-Market Manufacturing Industry Overview 2025"
    doc.author = "Cascade Research Associates"
    doc.subject = f"CANARY: {canary}"
    doc.creator = "Cascade Industries Test Suite Generator"

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle", parent=styles["Title"],
        fontSize=18, spaceAfter=20,
    )
    heading_style = ParagraphStyle(
        "CustomH1", parent=styles["Heading1"],
        fontSize=14, spaceBefore=12, spaceAfter=8,
    )
    heading2_style = ParagraphStyle(
        "CustomH2", parent=styles["Heading2"],
        fontSize=12, spaceBefore=10, spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "CustomBody", parent=styles["Normal"],
        fontSize=10, leading=14, spaceAfter=8,
    )

    story: list[Any] = []

    # ── Page 1: Title page ──────────────────────────────────
    story.append(Spacer(1, 2 * inch))
    story.append(Paragraph(
        "U.S. Mid-Market Manufacturing<br/>Industry Overview 2025",
        title_style,
    ))
    story.append(Spacer(1, 0.5 * inch))
    story.append(Paragraph(
        "Prepared by Cascade Research Associates<br/>"
        "March 2025",
        body_style,
    ))
    story.append(Spacer(1, 0.5 * inch))
    story.append(Paragraph(
        f"Document ID: {canary}",
        ParagraphStyle("Canary", parent=body_style, fontSize=8, textColor=colors.gray),
    ))
    story.append(PageBreak())

    # ── Page 2: Table of Contents ───────────────────────────
    story.append(Paragraph("Table of Contents", heading_style))
    toc_items = [
        "1. Executive Summary",
        "2. Market Size and Growth",
        "3. Sector Analysis: Industrial Manufacturing",
        "4. Sector Analysis: Advanced Materials & Composites",
        "5. Sector Analysis: Logistics & Distribution",
        "6. Competitive Dynamics",
        "7. M&A Activity and Valuation Trends",
        "8. Regulatory and Macroeconomic Factors",
        "9. Technology and Innovation Trends",
        "10. Risk Factors",
        "11. Five-Year Outlook",
        "12. Appendix: Data Sources and Methodology",
    ]
    for item in toc_items:
        story.append(Paragraph(item, body_style))
    story.append(PageBreak())

    # ── Page 3: Executive Summary ───────────────────────────
    story.append(Paragraph("1. Executive Summary", heading_style))
    story.append(Paragraph(
        "The U.S. mid-market manufacturing sector ($50B-$500M revenue) continues "
        "to demonstrate resilient growth despite persistent supply chain challenges "
        "and rising input costs. The sector generated approximately $1.8 trillion "
        "in combined revenue in 2024, representing 5.3% year-over-year growth.",
        body_style,
    ))
    story.append(Paragraph(
        "Key findings from our analysis of 250+ mid-market manufacturers:",
        body_style,
    ))
    exec_summary_points = [
        "Overall sector growth of 5.3% in 2024, expected to moderate to 4.5-5.0% in 2025",
        "Advanced materials subsector outperforming at 12.8% growth",
        "Logistics segment showing steady 7.5% growth driven by e-commerce",
        "M&A multiples averaging 8.5-10.5x EV/EBITDA for quality assets",
        "Margin pressure from labor costs partially offset by automation investments",
    ]
    for point in exec_summary_points:
        story.append(Paragraph(f"&bull; {point}", body_style))
    story.append(PageBreak())

    # ── Page 4: Market Size and Growth ──────────────────────
    story.append(Paragraph("2. Market Size and Growth", heading_style))
    story.append(Paragraph(
        "The mid-market manufacturing sector encompasses companies with annual "
        "revenues between $50 million and $500 million. Our analysis covers "
        "three primary subsectors: industrial manufacturing, advanced materials "
        "and composites, and logistics and distribution.",
        body_style,
    ))
    story.append(Paragraph("Historical Growth Rates by Subsector", heading2_style))

    growth_data = [
        ["Subsector", "2022", "2023", "2024", "2025E"],
        ["Industrial Manufacturing", "3.8%", "3.5%", "4.2%", "4.0-4.5%"],
        ["Advanced Materials", "8.5%", "10.2%", "12.8%", "11.0-13.0%"],
        ["Logistics & Distribution", "6.2%", "6.8%", "7.5%", "7.0-8.0%"],
        ["Overall Sector", "4.8%", "5.0%", "5.3%", "4.5-5.0%"],
    ]
    t = Table(growth_data, colWidths=[2.2 * inch, 1 * inch, 1 * inch, 1 * inch, 1.2 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#D6E4F0")]),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(
        "The industrial manufacturing subsector, which includes precision "
        "components and custom machining, grew 4.2% in 2024. This growth was "
        "primarily driven by increased defense spending and reshoring initiatives. "
        "We expect growth to moderate to 4.0-4.5% in 2025 as inventory "
        "restocking cycles normalize.",
        body_style,
    ))
    story.append(PageBreak())

    # ── Page 5: Sector Analysis: Industrial Manufacturing ───
    story.append(Paragraph("3. Sector Analysis: Industrial Manufacturing", heading_style))
    story.append(Paragraph(
        "The industrial manufacturing subsector remains the backbone of the "
        "mid-market, accounting for approximately 55% of sector revenue. "
        "Key characteristics include:",
        body_style,
    ))
    ind_points = [
        "Average gross margins of 30-38% (higher for precision/custom work)",
        "Capital intensity ratio of 3.5-4.5% of revenue",
        "Workforce challenges driving automation investments",
        "Growing demand for precision components in aerospace and defense",
    ]
    for p in ind_points:
        story.append(Paragraph(f"&bull; {p}", body_style))
    story.append(Paragraph("Valuation Benchmarks", heading2_style))
    story.append(Paragraph(
        "Industrial manufacturing companies in the mid-market are currently "
        "trading at 7.5-9.5x EV/EBITDA, with a median of approximately 8.5x. "
        "Companies with higher gross margins and recurring revenue streams "
        "command premium multiples of 10-12x.",
        body_style,
    ))
    story.append(PageBreak())

    # ── Page 6: Advanced Materials sector ───────────────────
    story.append(Paragraph("4. Sector Analysis: Advanced Materials & Composites", heading_style))
    story.append(Paragraph(
        "The advanced materials subsector is the fastest-growing segment, "
        "driven by demand for lightweight composites, specialty coatings, "
        "and high-performance polymers. The subsector grew 12.8% in 2024, "
        "significantly outpacing overall manufacturing growth.",
        body_style,
    ))
    story.append(Paragraph("Growth Drivers", heading2_style))
    am_drivers = [
        "Electric vehicle components and battery materials",
        "Aerospace composite structures and thermal management",
        "Semiconductor packaging materials",
        "Sustainable and recyclable material alternatives",
    ]
    for d in am_drivers:
        story.append(Paragraph(f"&bull; {d}", body_style))
    story.append(Paragraph("Margin Profile", heading2_style))
    story.append(Paragraph(
        "Advanced materials companies typically exhibit gross margins of "
        "45-55%, driven by proprietary formulations and high switching costs. "
        "R&D spending averages 8-12% of revenue. EBITDA margins range from "
        "18-28%, with scale being a significant driver of profitability.",
        body_style,
    ))
    story.append(PageBreak())

    # ── Page 7: Logistics & Distribution ────────────────────
    story.append(Paragraph("5. Sector Analysis: Logistics & Distribution", heading_style))
    story.append(Paragraph(
        "The logistics and distribution subsector grew 7.5% in 2024, "
        "benefiting from continued e-commerce growth and supply chain "
        "diversification strategies. Warehouse utilization rates remain "
        "above 90% in key markets.",
        body_style,
    ))
    story.append(Paragraph("Key Metrics", heading2_style))
    logistics_data = [
        ["Metric", "Industry Avg", "Top Quartile"],
        ["Revenue Growth (2024)", "7.5%", "10.2%"],
        ["Gross Margin", "15-22%", "22-28%"],
        ["EBITDA Margin", "8-14%", "14-18%"],
        ["CapEx / Revenue", "4-6%", "3-4%"],
        ["Warehouse Utilization", "88-92%", "94-97%"],
    ]
    lt = Table(logistics_data, colWidths=[2.5 * inch, 1.5 * inch, 1.5 * inch])
    lt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
    ]))
    story.append(lt)
    story.append(PageBreak())

    # ── Page 8: Competitive Dynamics ────────────────────────
    story.append(Paragraph("6. Competitive Dynamics", heading_style))
    story.append(Paragraph(
        "The mid-market manufacturing landscape remains highly fragmented, "
        "with the top 50 companies accounting for less than 15% of total "
        "sector revenue. This fragmentation creates significant M&A "
        "consolidation opportunities.",
        body_style,
    ))
    story.append(Paragraph(
        "Competitive advantages in the sector are primarily derived from: "
        "proprietary technology, customer relationships, geographic proximity "
        "to end markets, and operational efficiency. Companies with recurring "
        "revenue streams (service contracts, long-term supply agreements) "
        "command premium valuations.",
        body_style,
    ))
    story.append(Paragraph("Barriers to Entry", heading2_style))
    barriers = [
        "Capital requirements ($10-50M for competitive manufacturing capability)",
        "Regulatory compliance (ISO, AS9100 for aerospace, IATF for automotive)",
        "Workforce training and retention (skilled labor shortage persists)",
        "Customer qualification cycles (6-18 months for major OEMs)",
    ]
    for b in barriers:
        story.append(Paragraph(f"&bull; {b}", body_style))
    story.append(PageBreak())

    # ── Page 9: M&A Activity ────────────────────────────────
    story.append(Paragraph("7. M&A Activity and Valuation Trends", heading_style))
    story.append(Paragraph(
        "M&A activity in mid-market manufacturing remained robust in 2024, "
        "with 180+ transactions completed. Private equity sponsors accounted "
        "for approximately 45% of deal volume. Key valuation trends:",
        body_style,
    ))
    ma_data = [
        ["Category", "EV/Revenue", "EV/EBITDA"],
        ["Industrial Manufacturing", "1.2-1.8x", "7.5-9.5x"],
        ["Advanced Materials", "2.0-3.0x", "10.0-14.0x"],
        ["Logistics & Distribution", "1.0-1.5x", "8.0-10.0x"],
        ["Overall Mid-Market", "1.5-2.0x", "8.5-10.5x"],
    ]
    mat = Table(ma_data, colWidths=[2.5 * inch, 1.5 * inch, 1.5 * inch])
    mat.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
    ]))
    story.append(mat)
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(
        "Premium multiples (above the ranges shown) were paid for targets with: "
        "proprietary technology, strong IP portfolios, above-market growth, "
        "high customer retention rates (>95%), and EBITDA margins above the "
        "sector median.",
        body_style,
    ))
    story.append(PageBreak())

    # ── Page 10: Regulatory & Macro ─────────────────────────
    story.append(Paragraph("8. Regulatory and Macroeconomic Factors", heading_style))
    story.append(Paragraph(
        "Several macroeconomic factors are shaping the mid-market manufacturing "
        "environment in 2025:",
        body_style,
    ))
    macro_points = [
        "Interest rates: Federal funds rate at 4.25-4.50%, with expectations "
        "of 1-2 cuts in 2025. Higher rates increasing cost of capital for "
        "leveraged acquisitions.",
        "Trade policy: Ongoing tariff uncertainties creating both risks and "
        "reshoring opportunities for domestic manufacturers.",
        "Labor market: Unemployment at 3.9%, with manufacturing wages growing "
        "4.2% annually. Skilled labor shortages remain acute.",
        "Input costs: Steel and aluminum prices stabilized after 2023 volatility. "
        "Specialty materials costs continue to rise 3-5% annually.",
        "ESG requirements: Growing customer demands for sustainability reporting "
        "and carbon footprint reduction driving capex in green technologies.",
    ]
    for p in macro_points:
        story.append(Paragraph(f"&bull; {p}", body_style))
    story.append(PageBreak())

    # ── Page 11: Technology Trends ──────────────────────────
    story.append(Paragraph("9. Technology and Innovation Trends", heading_style))
    story.append(Paragraph(
        "Digital transformation is accelerating across mid-market manufacturing, "
        "with key technology investments including:",
        body_style,
    ))
    tech_points = [
        "Industry 4.0 / Smart Factory: IoT sensors, real-time monitoring, "
        "predictive maintenance. Average ROI of 15-25% within 3 years.",
        "Additive Manufacturing: 3D printing adoption growing 20% annually "
        "for prototyping and short-run production.",
        "AI/ML Applications: Quality inspection, demand forecasting, and "
        "supply chain optimization. Early adopters reporting 10-15% "
        "reduction in scrap rates.",
        "ERP Modernization: 40% of mid-market manufacturers still on legacy "
        "systems, creating significant upgrade cycle opportunity.",
    ]
    for p in tech_points:
        story.append(Paragraph(f"&bull; {p}", body_style))
    story.append(PageBreak())

    # ── Page 12: Risk Factors ───────────────────────────────
    story.append(Paragraph("10. Risk Factors", heading_style))
    story.append(Paragraph(
        "Key risks facing mid-market manufacturers in the current environment:",
        body_style,
    ))
    risks = [
        ("Supply Chain Disruption",
         "Geopolitical tensions and logistics bottlenecks continue to pose "
         "risks. Companies with single-source suppliers are particularly vulnerable."),
        ("Margin Compression",
         "Rising labor and input costs may compress margins, particularly "
         "for companies unable to pass through cost increases. Margin expansion "
         "assumptions of more than 100-150bps should be scrutinized carefully "
         "unless supported by specific operational initiatives."),
        ("Technology Obsolescence",
         "Rapid technological change creates risk of stranded assets and "
         "competitive disadvantage for slow adopters."),
        ("Concentration Risk",
         "Companies with >20% revenue from a single customer face significant "
         "customer concentration risk."),
        ("Regulatory Compliance",
         "Evolving environmental and safety regulations may require significant "
         "capital investment for compliance."),
    ]
    for title, desc in risks:
        story.append(Paragraph(f"<b>{title}</b>: {desc}", body_style))
    story.append(PageBreak())

    # ── Page 13: Five-Year Outlook ──────────────────────────
    story.append(Paragraph("11. Five-Year Outlook", heading_style))
    story.append(Paragraph(
        "Our five-year outlook for mid-market manufacturing is cautiously "
        "optimistic. We project overall sector growth of 4.0-5.5% annually "
        "through 2029, with advanced materials maintaining above-average "
        "growth of 10-12%.",
        body_style,
    ))
    story.append(Paragraph("Projected Growth Rates (2025-2029)", heading2_style))

    outlook_data = [
        ["Subsector", "2025E", "2026E", "2027E", "2028E", "2029E"],
        ["Industrial Manufacturing", "4.0%", "3.8%", "3.5%", "3.5%", "3.5%"],
        ["Advanced Materials", "12.0%", "11.0%", "10.5%", "10.0%", "10.0%"],
        ["Logistics & Distribution", "7.0%", "6.5%", "6.0%", "5.5%", "5.5%"],
        ["Overall Sector", "5.0%", "4.5%", "4.2%", "4.0%", "4.0%"],
    ]
    ot = Table(outlook_data, colWidths=[2.2 * inch] + [0.9 * inch] * 5)
    ot.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
    ]))
    story.append(ot)
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(
        "Critical assumptions underlying this outlook include: stable trade "
        "policy, gradual interest rate normalization, continued reshoring "
        "activity, and no major recessionary events. Downside scenarios could "
        "reduce growth by 200-300bps across all subsectors.",
        body_style,
    ))
    story.append(PageBreak())

    # ── Page 14: Appendix ───────────────────────────────────
    story.append(Paragraph("12. Appendix: Data Sources and Methodology", heading_style))
    story.append(Paragraph(
        "This report is based on analysis of publicly available data, "
        "proprietary databases, and interviews with industry executives.",
        body_style,
    ))
    story.append(Paragraph("Data Sources", heading2_style))
    sources = [
        "U.S. Census Bureau — Annual Survey of Manufactures",
        "Bureau of Economic Analysis — GDP by Industry",
        "Federal Reserve Economic Data (FRED)",
        "S&P Capital IQ — Public company financial data",
        "PitchBook — Private company and M&A transaction data",
        "Industry trade associations (NAM, MAPI, MHI)",
        "Executive interviews (45 companies surveyed)",
    ]
    for s in sources:
        story.append(Paragraph(f"&bull; {s}", body_style))
    story.append(Paragraph("Methodology", heading2_style))
    story.append(Paragraph(
        "Growth rates are computed on a revenue-weighted basis across our "
        "coverage universe of 250+ companies. Valuation multiples are based "
        "on trailing twelve-month financials as of December 2024. All financial "
        "data has been normalized for non-recurring items.",
        body_style,
    ))
    story.append(PageBreak())

    # ── Page 15: Disclaimer ─────────────────────────────────
    story.append(Paragraph("Disclaimer", heading_style))
    story.append(Paragraph(
        "This report is provided for informational purposes only and does not "
        "constitute investment advice. Cascade Research Associates makes no "
        "representations or warranties regarding the accuracy or completeness "
        "of the information contained herein.",
        body_style,
    ))
    story.append(Spacer(1, 0.5 * inch))
    story.append(Paragraph(
        "© 2025 Cascade Research Associates. All rights reserved.",
        body_style,
    ))
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(
        f"Document Reference: {canary}",
        ParagraphStyle("CanaryFooter", parent=body_style, fontSize=8, textColor=colors.gray),
    ))

    doc.build(story)

    with open(file_path, "wb") as f:
        f.write(buf.getvalue())

    canaries.set_location(
        "tc15_industry_overview",
        f"{_INPUT_DIR}/industry_overview.pdf",
        "PDF metadata field 'subject' and title page footer",
    )
    manifest.register(f"{_INPUT_DIR}/industry_overview.pdf", "pdf")


# ── Prompt and expected behavior ─────────────────────────────────────────────


def _write_prompt(output_dir: Path) -> None:
    """Write test_cases/TC-15/prompt.md per spec."""
    text = """\
Prepare a DCF valuation of Cascade Industries.

1. Derive unlevered free cash flow (UFCF) from management's projections:
   - EBITDA
   - Less: taxes (use effective rate from historical data)
   - Less: capex (from projections)
   - Less: changes in net working capital (derive from historical NWC/revenue ratios)
2. Compute WACC:
   - Cost of equity: risk-free rate (use 4.2%), equity risk premium (5.5%),
     beta (derive from comparable companies), size premium (2.0%)
   - Cost of debt: derive from the company's interest expense / average debt
   - Capital structure: use comparable companies' average debt/equity ratio
3. Compute terminal value using both:
   - Gordon Growth Model (2.5% perpetuity growth rate)
   - Exit multiple method (use comparable companies' median EV/EBITDA)
4. Discount to present value and compute implied enterprise value range.
5. Compute implied equity value (subtract net debt).
6. Perform sensitivity analysis:
   - WACC ± 1% vs. terminal growth rate ± 0.5% (for Gordon Growth)
   - WACC ± 1% vs. exit multiple ± 1x (for exit multiple method)
7. Draft a one-page valuation summary with the range and key assumptions.

Export:
- DCF model as Excel workbook (with clearly labeled assumptions, calculations,
  and sensitivity tables)
- Valuation summary as Word document
"""
    path = output_dir / f"test_cases/{_TC}/prompt.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text)


def _write_expected_behavior(output_dir: Path) -> None:
    """Write test_cases/TC-15/expected_behavior.md per spec."""
    text = """\
# TC-15: DCF Valuation — Expected Behavior

## Key Findings the Agent Should Produce

1. **WACC approximately 10.8%**: Derived from:
   - Cost of equity ~14.2% (risk-free 4.2% + beta × 5.5% ERP + 2.0% size premium)
   - Cost of debt from historical interest expense / average debt
   - Capital structure weights from comparable companies

2. **Enterprise value ranges**:
   - Gordon Growth Model: $245M - $305M
   - Exit Multiple Method: $260M - $320M

3. **Implied equity value approximately $255M**: Enterprise value minus net debt

4. **Flag aggressive margin assumption**: Management's projections assume 200bps
   EBITDA margin expansion by Year 5. The industry overview notes that margin
   expansion assumptions above 100-150bps should be "scrutinized carefully unless
   supported by specific operational initiatives." The agent should flag this.

## Data Challenges

- **Historical data extraction**: The agent must correctly derive EBITDA from
  historical financials (operating income + D&A), not accept a pre-labeled EBITDA line.
- **Effective tax rate**: Must compute from historical data, not assume a standard rate.
- **Comparable company beta**: Must unlever and relever betas using the target's
  capital structure.
- **NWC/Revenue ratio**: Must compute from historical balance sheets and apply
  consistently to projected revenue.
- **Terminal value**: Both methods must be computed and compared. A significant
  divergence should be noted.
- **Industry report depth**: The margin expansion warning is on page 12 of the
  15-page PDF (Risk Factors section). An agent that only reads the executive
  summary will miss this critical finding.

## Expected Output Structure

### DCF Model (Excel workbook):
- Assumptions tab (WACC inputs, growth rates, margin targets)
- Historical financials summary (3 years)
- Projected UFCF (5 years)
- Terminal value calculations (both methods)
- DCF bridge: UFCF → PV → Enterprise Value → Equity Value
- Sensitivity tables:
  - WACC vs. terminal growth rate
  - WACC vs. exit multiple

### Valuation Summary (Word document):
- Key assumptions and sources
- Valuation range with both methods
- Comparison of methods
- Key risk factors and caveats
- Recommendation or conclusion
"""
    path = output_dir / f"test_cases/{_TC}/expected_behavior.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text)


# ── Gold Standard ────────────────────────────────────────────────────────────


def _compute_dcf(
    hist: dict[str, Any],
    proj: dict[int, dict[str, Any]],
) -> dict[str, Any]:
    """Pre-compute the DCF model values for the gold standard.

    Returns a dict with all key DCF outputs.
    """
    # ── Effective tax rate (average of historical) ──────────
    etr_sum = sum(hist[y]["effective_tax_rate"] for y in [2023, 2024, 2025])
    avg_etr = (etr_sum / Decimal(3)).quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)

    # ── Beta from comparables ───────────────────────────────
    betas = [comp[7] for comp in _COMPARABLE_COMPANIES]
    de_ratios = [(comp[3] - comp[2]) / comp[2] for comp in _COMPARABLE_COMPANIES]

    # Unlever each beta: beta_u = beta_l / (1 + (1-t) * D/E)
    # Use avg ETR as tax rate
    t = float(avg_etr)
    unlevered_betas = [
        b / (1 + (1 - t) * de) for b, de in zip(betas, de_ratios)
    ]
    avg_unlevered_beta = sum(unlevered_betas) / len(unlevered_betas)

    # Target D/E from comparable median
    sorted_de = sorted(de_ratios)
    n = len(sorted_de)
    median_de = (sorted_de[n // 2 - 1] + sorted_de[n // 2]) / 2

    # Relever: beta_l = beta_u * (1 + (1-t) * D/E)
    relevered_beta = avg_unlevered_beta * (1 + (1 - t) * median_de)

    # ── Cost of equity ──────────────────────────────────────
    rf = float(_RISK_FREE_RATE)
    erp = float(_EQUITY_RISK_PREMIUM)
    sp = float(_SIZE_PREMIUM)
    cost_of_equity = rf + relevered_beta * erp + sp

    # ── Cost of debt ────────────────────────────────────────
    # From historical interest expense / average debt
    interest_2025 = float(hist[2025]["interest_expense"])
    debt_2024 = float(hist[2024]["total_debt"])
    debt_2025 = float(hist[2025]["total_debt"])
    avg_debt = (debt_2024 + debt_2025) / 2
    if avg_debt > 0:
        pre_tax_cost_of_debt = abs(interest_2025) / avg_debt
    else:
        pre_tax_cost_of_debt = 0.05  # fallback
    after_tax_cost_of_debt = pre_tax_cost_of_debt * (1 - t)

    # ── WACC ────────────────────────────────────────────────
    equity_weight = 1 / (1 + median_de)
    debt_weight = median_de / (1 + median_de)
    wacc = equity_weight * cost_of_equity + debt_weight * after_tax_cost_of_debt

    # ── Projected UFCF ──────────────────────────────────────
    ufcf_by_year: dict[int, float] = {}
    # NWC/Revenue ratio from FY2025
    nwc_rev_ratio = float(hist[2025]["nwc"] / hist[2025]["revenue"])

    prev_nwc = float(hist[2025]["nwc"])
    for year in _PROJECTION_YEARS:
        p = proj[year]
        ebitda = float(p["ebitda"])
        taxes = ebitda * t  # Taxes on EBIT (approximate via EBITDA × tax rate)
        capex = float(p["capex"])
        nwc = float(p["revenue"]) * nwc_rev_ratio
        delta_nwc = nwc - prev_nwc
        ufcf = ebitda - taxes - capex - delta_nwc
        ufcf_by_year[year] = ufcf
        prev_nwc = nwc

    # ── Terminal value (Gordon Growth) ──────────────────────
    terminal_ufcf = ufcf_by_year[2030] * (1 + float(_PERPETUITY_GROWTH))
    tv_gordon = terminal_ufcf / (wacc - float(_PERPETUITY_GROWTH))

    # ── Terminal value (Exit Multiple) ──────────────────────
    ev_ebitda_list = sorted(comp[3] / comp[5] for comp in _COMPARABLE_COMPANIES)
    median_ev_ebitda = (ev_ebitda_list[n // 2 - 1] + ev_ebitda_list[n // 2]) / 2
    terminal_ebitda = float(proj[2030]["ebitda"])
    tv_exit = terminal_ebitda * median_ev_ebitda

    # ── Discount to PV ──────────────────────────────────────
    pv_ufcf_gordon = 0.0
    pv_ufcf_exit = 0.0
    for i, year in enumerate(_PROJECTION_YEARS):
        discount = (1 + wacc) ** (i + 1)
        pv_ufcf_gordon += ufcf_by_year[year] / discount
        pv_ufcf_exit += ufcf_by_year[year] / discount

    n_years = len(_PROJECTION_YEARS)
    terminal_discount = (1 + wacc) ** n_years

    pv_tv_gordon = tv_gordon / terminal_discount
    pv_tv_exit = tv_exit / terminal_discount

    ev_gordon = pv_ufcf_gordon + pv_tv_gordon
    ev_exit = pv_ufcf_exit + pv_tv_exit

    # ── Equity value ────────────────────────────────────────
    net_debt = float(hist[2025]["net_debt"])
    equity_gordon = ev_gordon - net_debt
    equity_exit = ev_exit - net_debt
    equity_midpoint = (equity_gordon + equity_exit) / 2

    return {
        "wacc": round(wacc, 4),
        "wacc_pct": round(wacc * 100, 1),
        "cost_of_equity": round(cost_of_equity, 4),
        "cost_of_equity_pct": round(cost_of_equity * 100, 1),
        "relevered_beta": round(relevered_beta, 2),
        "median_de_ratio": round(median_de, 2),
        "avg_etr": round(float(avg_etr), 3),
        "pre_tax_cost_of_debt": round(pre_tax_cost_of_debt, 4),
        "after_tax_cost_of_debt": round(after_tax_cost_of_debt, 4),
        "ufcf_by_year": {str(y): round(v / 1e6, 1) for y, v in ufcf_by_year.items()},
        "tv_gordon_M": round(tv_gordon / 1e6, 0),
        "tv_exit_M": round(tv_exit / 1e6, 0),
        "ev_gordon_M": round(ev_gordon / 1e6, 0),
        "ev_exit_M": round(ev_exit / 1e6, 0),
        "net_debt_M": round(net_debt / 1e6, 0),
        "equity_gordon_M": round(equity_gordon / 1e6, 0),
        "equity_exit_M": round(equity_exit / 1e6, 0),
        "equity_midpoint_M": round(equity_midpoint / 1e6, 0),
        "median_ev_ebitda": round(median_ev_ebitda, 1),
        "perpetuity_growth": float(_PERPETUITY_GROWTH),
    }


@register_gold("TC-15")
def _tc15_gold(
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    **model_kwargs: Any,
) -> GoldStandard:
    """TC-15 gold standard: DCF valuation."""
    model: CascadeModel = model_kwargs["model"]
    hist = _get_historical_data(model)
    proj = _compute_projections(hist)
    dcf = _compute_dcf(hist, proj)

    return GoldStandard(
        test_case="TC-15",
        expected_outputs={
            "output_files": {
                "dcf_model": {
                    "type": "xlsx",
                    "required_sheets": [
                        "Assumptions",
                        "Historical Summary",
                        "Projected UFCF",
                        "Terminal Value",
                        "DCF Bridge",
                        "Sensitivity - Gordon Growth",
                        "Sensitivity - Exit Multiple",
                    ],
                },
                "valuation_summary": {
                    "type": "docx",
                    "required_sections": [
                        "Key Assumptions",
                        "Valuation Range",
                        "Risk Factors",
                    ],
                },
            },
            "wacc": {
                "value_pct": dcf["wacc_pct"],
                "cost_of_equity_pct": dcf["cost_of_equity_pct"],
                "relevered_beta": dcf["relevered_beta"],
                "pre_tax_cost_of_debt": round(dcf["pre_tax_cost_of_debt"] * 100, 1),
                "target_de_ratio": dcf["median_de_ratio"],
            },
            "enterprise_value_gordon_growth": {
                "range_low_M": 245,
                "range_high_M": 305,
                "computed_M": dcf["ev_gordon_M"],
            },
            "enterprise_value_exit_multiple": {
                "range_low_M": 260,
                "range_high_M": 320,
                "computed_M": dcf["ev_exit_M"],
            },
            "implied_equity_value_M": dcf["equity_midpoint_M"],
            "margin_expansion_flag": {
                "assumption_bps": 200,
                "industry_warning_threshold_bps": "100-150",
                "should_flag": True,
                "description": (
                    "Management projects 200bps EBITDA margin expansion over 5 years. "
                    "Industry report (Risk Factors, p.12) warns that margin expansion "
                    "above 100-150bps should be scrutinized unless supported by specific "
                    "operational initiatives."
                ),
            },
            "projected_ufcf_M": dcf["ufcf_by_year"],
            "terminal_value": {
                "gordon_growth_M": dcf["tv_gordon_M"],
                "exit_multiple_M": dcf["tv_exit_M"],
                "median_ev_ebitda_multiple": dcf["median_ev_ebitda"],
                "perpetuity_growth_rate": dcf["perpetuity_growth"],
            },
        },
        canary_verification={
            "read_historical_financials": canaries.canary_for("tc15_historical_financials"),
            "read_management_projections": canaries.canary_for("tc15_management_projections"),
            "read_comparable_companies": canaries.canary_for("tc15_comparable_companies"),
            "read_industry_overview": canaries.canary_for("tc15_industry_overview"),
        },
        error_detection={
            "ERR-022": (
                "EV/EBITDA multiple is blank for MidWest Materials Group — "
                "missing_data in comparable_companies_trading.xlsx"
            ),
            "ERR-025": (
                "FY2024 total revenue does not match sum of components — "
                "mismatched_total in historical_financials_3yr.xlsx"
            ),
        },
        scoring_hints={
            "correctness": (
                "WACC must be within 0.5pp of gold standard (~10.8%). "
                "EV range must overlap with gold standard ranges. "
                "UFCF derivation must be methodologically sound."
            ),
            "completeness": (
                "Both terminal value methods computed and compared. "
                "Sensitivity tables for both WACC ± 1% vs TGR ± 0.5% "
                "and WACC ± 1% vs exit multiple ± 1x."
            ),
            "format_compliance": (
                "DCF model as Excel with required sheets. "
                "Valuation summary as Word document."
            ),
            "robustness": (
                "Agent must flag the 200bps margin expansion as aggressive, "
                "citing the industry report's risk factor warning. "
                "Agent must derive effective tax rate from historical data, "
                "not assume 21% or 25%."
            ),
            "communication": (
                "Clear one-page summary with range and key assumptions. "
                "Explicit comparison of Gordon Growth vs Exit Multiple results."
            ),
        },
    )


# ── Public entry point ───────────────────────────────────────────────────────


def emit_tc15(
    model: CascadeModel,
    output_dir: Path,
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    manifest: Manifest,
) -> None:
    """Emit all TC-15 files."""
    hist = _write_historical_xlsx(model, output_dir, canaries, errors, manifest)
    _write_projections_xlsx(hist, output_dir, canaries, manifest)
    _write_comps_xlsx(output_dir, canaries, errors, manifest)
    _write_industry_pdf(output_dir, canaries, manifest)
    _write_prompt(output_dir)
    _write_expected_behavior(output_dir)
