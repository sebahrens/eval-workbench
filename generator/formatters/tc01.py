"""Formatter: TC-01 — Trial Balance Reconciliation (Audit, Complex).

Emits:
- test_cases/TC-01/input_files/cascade_tb_fy2025.xlsx
  Messy client TB: merged header, inconsistent names, ERR-001 (transposed digit)
- test_cases/TC-01/input_files/cascade_tb_fy2024_workpaper.xlsx
  Clean prior year workpaper with standardized names and lead schedule mappings
- test_cases/TC-01/input_files/cascade_financials_fy2024_signed.pdf
  Text-native PDF of signed FY2024 financial statements (BS, IS, CF, notes)
- test_cases/TC-01/prompt.md
- test_cases/TC-01/expected_behavior.md
- gold_standards/TC-01_gold.json

Uses the canonical model (Ledger) — never hardcodes numbers.
"""

from __future__ import annotations

import datetime
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from generator.canaries import CanaryRegistry, embed_canary_xlsx
from generator.errors import ErrorRegistry, PlantedError, transpose_digits
from generator.golds.framework import GoldStandard, register_gold
from generator.manifest import Manifest
from generator.model.build import CascadeModel
from generator.model.coa import ACCOUNTS_BY_NUMBER, Account
from generator.model.consolidation import (
    build_balance_sheet,
    build_income_statement,
    consolidated_trial_balance_eliminated,
)
from generator.noise import ExclusionZone, apply_xlsx_noise, make_noise_rng
from generator.scenario_context import ScenarioContext

# ── Constants ────────────────────────────────────────────────────────────────

_TC = "TC-01"
_INPUT_DIR = f"test_cases/{_TC}/input_files"

# Two accounts renamed from prior year (spec: "two accounts renamed from prior
# year without mapping notes").  The agent must recognize these.
_RENAMED_ACCOUNTS: dict[str, str] = {
    # prior_year_name → current_year_name
    "1100": "Trade Receivables — Net",       # was "Accounts Receivable — Trade"
    "6150": "Business Travel",               # was "Travel & Entertainment"
}

# Accounts that are NEW in FY2025 (not in FY2024 workpaper)
_NEW_FY2025_ACCOUNTS = ["1350", "6260", "6270"]

# One account MISSING from FY2025 that was in FY2024
_MISSING_FROM_FY2025 = "1310"  # "Prepaid Rent" — absorbed into lease accounting

# Name abbreviations for the messy client TB
_ABBREVIATIONS: dict[str, str] = {
    "1100": "Accts Recv — Trade",
    "2010": "A/P Trade",
    "2020": "A/P Accrued Exp",
    "6030": "Emp Benefits",
    "6210": "Depr Expense",
    "6220": "Amort Expense",
    "1150": "Allow Doubtful Accts",
    "5010": "COGS — Dir Matl",
    "5020": "COGS — Dir Labor",
}

# ERR-001: Transposed digits in Accounts Receivable (1100) FY2025 balance.
# The error is in the FY2025 TB — the balance has two digits swapped,
# and the agent must catch it when tying out to the prior year signed financials.
_ERR_001_ACCOUNT = "1100"

# Fixed datetime for xlsx metadata (determinism)
_FIXED_DATETIME = datetime.datetime(2025, 3, 15, 9, 0, 0)


# ── Helpers ──────────────────────────────────────────────────────────────────


def _pin_xlsx_dates(wb: openpyxl.Workbook) -> None:
    """Pin created timestamp for determinism. Modified is set at save time."""
    wb.properties.created = _FIXED_DATETIME


def _save_xlsx_deterministic(wb: openpyxl.Workbook, path: str | Path) -> None:
    """Save workbook with pinned timestamps (openpyxl overrides modified on save).

    Also re-packs the zip with fixed entry dates so the file is byte-identical
    across runs regardless of wall-clock time.
    """
    import io
    from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

    from openpyxl.writer.excel import ExcelWriter

    path = Path(path)

    # Step 1: Write to an in-memory buffer (openpyxl will set modified=now)
    buf = io.BytesIO()
    wb.properties.modified = _FIXED_DATETIME
    archive = ZipFile(buf, "w", ZIP_DEFLATED, allowZip64=True)
    writer = ExcelWriter(wb, archive)
    writer.save()

    # Step 2: Re-pack with fixed timestamps
    fixed_date_time = (2025, 3, 15, 9, 0, 0)
    buf.seek(0)
    with ZipFile(buf, "r") as src, ZipFile(str(path), "w", ZIP_DEFLATED) as dst:
        for item in src.infolist():
            info = ZipInfo(item.filename, date_time=fixed_date_time)
            info.compress_type = item.compress_type
            dst.writestr(info, src.read(item.filename))


def _account_name_for_messy_tb(acct: Account) -> str:
    """Return a deliberately messy account name for the client TB."""
    num = acct.number
    # Renamed accounts get the new name
    if num in _RENAMED_ACCOUNTS:
        return _RENAMED_ACCOUNTS[num]
    # Some accounts get abbreviations
    if num in _ABBREVIATIONS:
        return _ABBREVIATIONS[num]
    # Add inconsistent formatting: some have leading spaces
    if num in ("1200", "1210", "1220", "2030", "6040"):
        return "  " + acct.name  # leading spaces
    return acct.name


def _whole_dollars(d: Decimal) -> int:
    """Round a Decimal to whole dollars (integer)."""
    return int(d.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


# ── FY2025 Messy Client Trial Balance ────────────────────────────────────────


def _write_messy_tb(
    model: CascadeModel,
    output_dir: Path,
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    manifest: Manifest,
    *,
    ctx: ScenarioContext | None = None,
) -> dict[str, int]:
    """Write cascade_tb_fy2025.xlsx — the messy client trial balance.

    Returns a dict of account_number → balance (whole dollars) *after* error
    injection, for use by the gold standard.
    """
    wb = openpyxl.Workbook()
    _pin_xlsx_dates(wb)
    ws = wb.active
    ws.title = "Trial Balance"

    # ── Canary ───────────────────────────────────────────────────────
    canary_code = canaries.canary_for("cascade_tb_fy2025")
    location = embed_canary_xlsx(wb, canary_code)
    canaries.set_location(
        "cascade_tb_fy2025",
        f"{_INPUT_DIR}/cascade_tb_fy2025.xlsx",
        location,
    )

    # ── Merged header row (deliberately messy) ───────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"] = "Cascade Industries, Inc. — Trial Balance — FY2025 (Unaudited)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = "As of December 31, 2025"
    ws["A2"].font = Font(italic=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Blank row 3
    header_row = 4

    # ── Column headers ───────────────────────────────────────────────
    headers = ["Account #", "Account Name", "Debit", "Credit", "Net Balance", "Prior Year Ref"]
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A3C6E")
    header_border = Border(bottom=Side(style="thin"))

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center")

    # ── Get consolidated TB (IC-eliminated) for FY2025 ───────────────
    eoy_2025 = datetime.date(2025, 12, 31)
    tb_2025 = consolidated_trial_balance_eliminated(model.ledger, eoy_2025)

    # Build the account list: include COA accounts that have a non-zero balance,
    # but skip the missing account and include new accounts
    tb_accounts: list[str] = sorted(
        acct for acct in tb_2025
        if acct != _MISSING_FROM_FY2025
    )
    # Ensure new FY2025 accounts are present (they may already be if they have balances)
    for acct_num in _NEW_FY2025_ACCOUNTS:
        if acct_num not in tb_accounts and acct_num in ACCOUNTS_BY_NUMBER:
            tb_accounts.append(acct_num)
    tb_accounts = sorted(tb_accounts)

    # ── ERR-001: Transposed digits in A/R ────────────────────────────
    correct_ar_balance = _whole_dollars(tb_2025.get(_ERR_001_ACCOUNT, Decimal(0)))
    if correct_ar_balance == 0:
        # A/R should always have a balance; if not, use a placeholder
        correct_ar_balance = 18_423_109
    err_ar_balance = transpose_digits(correct_ar_balance)

    # Register the error
    err = PlantedError(
        error_id="ERR-001",
        file=f"{_INPUT_DIR}/cascade_tb_fy2025.xlsx",
        location="Sheet 'Trial Balance', Account 1100 Net Balance",
        type="transposed_digits",
        description=(
            f"Accounts Receivable balance shows ${err_ar_balance:,} "
            f"instead of ${correct_ar_balance:,}"
        ),
        severity="material",
        which_test_cases_should_catch=["TC-01", "TC-02"],
    )
    errors.add(err)

    # ── Write account rows ───────────────────────────────────────────
    balances_written: dict[str, int] = {}
    data_font = Font(size=10)
    money_fmt = '#,##0'
    row = header_row + 1

    for acct_num in tb_accounts:
        acct_obj = ACCOUNTS_BY_NUMBER.get(acct_num)
        if acct_obj is None:
            continue

        raw_bal = tb_2025.get(acct_num, Decimal(0))
        bal = _whole_dollars(raw_bal)

        # Inject ERR-001
        if acct_num == _ERR_001_ACCOUNT:
            bal = err_ar_balance

        # Debit/Credit presentation
        if bal >= 0:
            debit_val = bal
            credit_val = None
        else:
            debit_val = None
            credit_val = abs(bal)

        name = _account_name_for_messy_tb(acct_obj)

        ws.cell(row=row, column=1, value=acct_num).font = data_font
        ws.cell(row=row, column=2, value=name).font = data_font
        debit_cell = ws.cell(row=row, column=3, value=debit_val)
        debit_cell.font = data_font
        debit_cell.number_format = money_fmt
        credit_cell = ws.cell(row=row, column=4, value=credit_val)
        credit_cell.font = data_font
        credit_cell.number_format = money_fmt
        net_cell = ws.cell(row=row, column=5, value=bal)
        net_cell.font = data_font
        net_cell.number_format = money_fmt
        # Prior Year Ref column — left blank (messy: no consistent mapping)
        ws.cell(row=row, column=6, value="").font = data_font

        balances_written[acct_num] = bal
        row += 1

    # ── Totals row ───────────────────────────────────────────────────
    total_debit = sum(v for v in balances_written.values() if v > 0)
    total_credit = sum(abs(v) for v in balances_written.values() if v < 0)

    ws.cell(row=row, column=2, value="TOTALS").font = Font(bold=True, size=10)
    ws.cell(row=row, column=3, value=total_debit).font = Font(bold=True, size=10)
    ws.cell(row=row, column=3).number_format = money_fmt
    ws.cell(row=row, column=4, value=total_credit).font = Font(bold=True, size=10)
    ws.cell(row=row, column=4).number_format = money_fmt

    # ── Column widths ────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    # ── Controlled noise (xlsx_openpyxl family pilot) ────────────────
    # Protect merged title rows, ERR-001 balance cell, and all numeric
    # balance cells (model fact preservation).  Canary lives in document
    # properties, not in a cell, so no cell exclusion needed for it.
    excl = ExclusionZone()
    # Protect merged title area (rows 1-3)
    for r in range(1, 4):
        for c in range(1, 7):
            excl.cells.add((ws.title, r, c))
    # Protect all numeric balance cells (columns 3-5) — model facts
    for r in range(header_row + 1, row + 1):
        for c in (3, 4, 5):
            excl.cells.add((ws.title, r, c))
    noise_ctx = ctx if ctx is not None else ScenarioContext(seed=42)
    noise_rng = make_noise_rng(noise_ctx, _TC, "cascade_tb_fy2025")
    apply_xlsx_noise(wb, noise_rng, excl)

    # ── Save ─────────────────────────────────────────────────────────
    path = output_dir / _INPUT_DIR / "cascade_tb_fy2025.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    _save_xlsx_deterministic(wb, path)

    manifest.register(
        f"{_INPUT_DIR}/cascade_tb_fy2025.xlsx",
        "xlsx",
        canary=canary_code,
        test_cases=[_TC],
    )

    return balances_written


# ── FY2024 Clean Workpaper ───────────────────────────────────────────────────


def _write_prior_year_workpaper(
    model: CascadeModel,
    output_dir: Path,
    canaries: CanaryRegistry,
    manifest: Manifest,
) -> dict[str, int]:
    """Write cascade_tb_fy2024_workpaper.xlsx — clean prior year workpaper.

    Returns a dict of account_number → balance (whole dollars) for gold standard.
    """
    wb = openpyxl.Workbook()
    _pin_xlsx_dates(wb)
    ws = wb.active
    ws.title = "Trial Balance"

    # ── Canary ───────────────────────────────────────────────────────
    canary_code = canaries.canary_for("cascade_tb_fy2024_workpaper")
    location = embed_canary_xlsx(wb, canary_code)
    canaries.set_location(
        "cascade_tb_fy2024_workpaper",
        f"{_INPUT_DIR}/cascade_tb_fy2024_workpaper.xlsx",
        location,
    )

    # ── Header ───────────────────────────────────────────────────────
    ws["A1"] = "Cascade Industries, Inc."
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "Audited Trial Balance — FY2024"
    ws["A2"].font = Font(italic=True, size=11)
    ws["A3"] = "As of December 31, 2024"
    ws["A3"].font = Font(size=10)

    header_row = 5
    headers = ["Account #", "Account Name", "Lead Schedule", "Debit", "Credit", "Net Balance", "Notes"]
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E5090")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ── Get consolidated TB (IC-eliminated) for FY2024 ───────────────
    eoy_2024 = datetime.date(2024, 12, 31)
    tb_2024 = consolidated_trial_balance_eliminated(model.ledger, eoy_2024)

    # Lead schedule mapping (first digit of account number)
    _lead_schedule: dict[str, str] = {
        "1": "A — Assets",
        "2": "B — Liabilities",
        "3": "C — Equity",
        "4": "D — Revenue",
        "5": "E — Cost of Sales",
        "6": "F — Operating Expenses",
        "7": "G — Other Income/Expense",
        "8": "H — Income Tax",
    }

    tb_accounts_2024: list[str] = sorted(
        acct for acct in tb_2024 if not acct.startswith("9")
    )
    # Include the missing-from-2025 account (it exists in 2024)
    if _MISSING_FROM_FY2025 not in tb_accounts_2024:
        tb_accounts_2024.append(_MISSING_FROM_FY2025)
        tb_accounts_2024 = sorted(tb_accounts_2024)

    balances_2024: dict[str, int] = {}
    data_font = Font(size=10)
    money_fmt = '#,##0'
    row = header_row + 1

    for acct_num in tb_accounts_2024:
        acct_obj = ACCOUNTS_BY_NUMBER.get(acct_num)
        if acct_obj is None:
            continue

        raw_bal = tb_2024.get(acct_num, Decimal(0))
        bal = _whole_dollars(raw_bal)

        if bal >= 0:
            debit_val = bal
            credit_val = None
        else:
            debit_val = None
            credit_val = abs(bal)

        lead = _lead_schedule.get(acct_num[0], "")

        ws.cell(row=row, column=1, value=acct_num).font = data_font
        ws.cell(row=row, column=2, value=acct_obj.name).font = data_font  # Clean names
        ws.cell(row=row, column=3, value=lead).font = data_font
        debit_cell = ws.cell(row=row, column=4, value=debit_val)
        debit_cell.font = data_font
        debit_cell.number_format = money_fmt
        credit_cell = ws.cell(row=row, column=5, value=credit_val)
        credit_cell.font = data_font
        credit_cell.number_format = money_fmt
        net_cell = ws.cell(row=row, column=6, value=bal)
        net_cell.font = data_font
        net_cell.number_format = money_fmt
        ws.cell(row=row, column=7, value="").font = data_font

        balances_2024[acct_num] = bal
        row += 1

    # ── Column widths ────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 30

    # ── Save ─────────────────────────────────────────────────────────
    path = output_dir / _INPUT_DIR / "cascade_tb_fy2024_workpaper.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    _save_xlsx_deterministic(wb, path)

    manifest.register(
        f"{_INPUT_DIR}/cascade_tb_fy2024_workpaper.xlsx",
        "xlsx",
        canary=canary_code,
        test_cases=[_TC],
    )

    return balances_2024


# ── FY2024 Signed Financial Statements (PDF) ────────────────────────────────


def _write_signed_financials_pdf(
    model: CascadeModel,
    output_dir: Path,
    canaries: CanaryRegistry,
    manifest: Manifest,
) -> None:
    """Write cascade_financials_fy2024_signed.pdf — 8-page signed financials."""
    path = output_dir / _INPUT_DIR / "cascade_financials_fy2024_signed.pdf"
    path.parent.mkdir(parents=True, exist_ok=True)

    canary_code = canaries.canary_for("cascade_financials_fy2024_signed")

    eoy_2024 = datetime.date(2024, 12, 31)
    eoy_2023 = datetime.date(2023, 12, 31)

    bs_2024 = build_balance_sheet(model.ledger, eoy_2024)
    bs_2023 = build_balance_sheet(model.ledger, eoy_2023)
    is_2024 = build_income_statement(model.ledger, 2024)
    is_2023 = build_income_statement(model.ledger, 2023)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "FinTitle", parent=styles["Title"],
        fontSize=16, spaceAfter=6, textColor=colors.HexColor("#1A3C6E"),
    )
    heading_style = ParagraphStyle(
        "FinHeading", parent=styles["Heading2"],
        fontSize=12, spaceAfter=4, textColor=colors.HexColor("#1A3C6E"),
    )
    body_style = ParagraphStyle(
        "FinBody", parent=styles["Normal"], fontSize=10, spaceAfter=4,
    )
    note_style = ParagraphStyle(
        "FinNote", parent=styles["Normal"], fontSize=9, spaceAfter=3,
        textColor=colors.HexColor("#333333"),
    )

    story: list = []

    # ── Page 1: Cover ────────────────────────────────────────────────
    story.append(Spacer(1, 2 * inch))
    story.append(Paragraph("Cascade Industries, Inc.", title_style))
    story.append(Paragraph("Consolidated Financial Statements", heading_style))
    story.append(Paragraph("For the Year Ended December 31, 2024", body_style))
    story.append(Spacer(1, inch))
    story.append(Paragraph(
        "<i>Audited by Mitchell & Associates LLP</i>", body_style
    ))
    story.append(Paragraph(
        "<i>Report Date: March 15, 2025</i>", body_style
    ))
    story.append(PageBreak())

    # ── Page 2: Auditor's Report ─────────────────────────────────────
    story.append(Paragraph("Independent Auditor's Report", title_style))
    story.append(Spacer(1, 12))
    story.append(Paragraph("To the Board of Directors and Shareholders", body_style))
    story.append(Paragraph("Cascade Industries, Inc.", body_style))
    story.append(Spacer(1, 8))
    story.append(Paragraph(
        "We have audited the accompanying consolidated financial statements of "
        "Cascade Industries, Inc. and subsidiaries, which comprise the consolidated "
        "balance sheet as of December 31, 2024, and the related consolidated "
        "statements of income, comprehensive income, stockholders' equity, and cash "
        "flows for the year then ended, and the related notes to the financial "
        "statements.", body_style
    ))
    story.append(Spacer(1, 8))
    story.append(Paragraph(
        "<b>Opinion</b>", body_style
    ))
    story.append(Paragraph(
        "In our opinion, the financial statements referred to above present fairly, "
        "in all material respects, the financial position of Cascade Industries, Inc. "
        "and subsidiaries as of December 31, 2024, and the results of their operations "
        "and their cash flows for the year then ended in accordance with accounting "
        "principles generally accepted in the United States of America.", body_style
    ))
    story.append(Spacer(1, 24))
    story.append(Paragraph("Mitchell & Associates LLP", body_style))
    story.append(Paragraph("Portland, Oregon", body_style))
    story.append(Paragraph("March 15, 2025", body_style))
    story.append(PageBreak())

    # ── Page 3–4: Balance Sheet ──────────────────────────────────────
    story.append(Paragraph(
        "Consolidated Balance Sheet", title_style
    ))
    story.append(Paragraph(
        "As of December 31, 2024 and 2023 (in whole dollars)", body_style
    ))
    story.append(Spacer(1, 12))

    def _fmt(val: Decimal | int) -> str:
        v = int(val) if isinstance(val, Decimal) else val
        if v < 0:
            return f"({abs(v):,})"
        return f"{v:,}"

    # Balance sheet table — uses aggregate totals from the BalanceSheet dataclass
    bs_data = [
        ["", "2024", "2023"],
        ["ASSETS", "", ""],
        ["  Total Assets", _fmt(bs_2024.total_assets), _fmt(bs_2023.total_assets)],
        ["", "", ""],
        ["LIABILITIES", "", ""],
        ["  Total Liabilities", _fmt(bs_2024.total_liabilities), _fmt(bs_2023.total_liabilities)],
        ["", "", ""],
        ["EQUITY", "", ""],
        ["  Total Stockholders' Equity", _fmt(bs_2024.total_equity), _fmt(bs_2023.total_equity)],
        ["", "", ""],
        [
            "TOTAL LIABILITIES & EQUITY",
            _fmt(bs_2024.total_liabilities + bs_2024.total_equity),
            _fmt(bs_2023.total_liabilities + bs_2023.total_equity),
        ],
    ]

    bs_table = Table(bs_data, colWidths=[3.5 * inch, 1.5 * inch, 1.5 * inch])
    bs_table.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 10),
        ("FONT", (0, 1), (-1, -1), "Helvetica", 9),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#1A3C6E")),
        ("LINEBELOW", (0, 2), (-1, 2), 0.5, colors.black),
        ("LINEBELOW", (0, 7), (-1, 7), 0.5, colors.black),
        ("LINEBELOW", (0, -1), (-1, -1), 1.5, colors.black),
        ("FONT", (0, 2), (0, 2), "Helvetica-Bold", 9),
        ("FONT", (0, 7), (0, 7), "Helvetica-Bold", 9),
        ("FONT", (0, -1), (0, -1), "Helvetica-Bold", 9),
    ]))
    story.append(bs_table)
    story.append(PageBreak())

    # ── Page 5–6: Income Statement ───────────────────────────────────
    story.append(Paragraph(
        "Consolidated Statement of Income", title_style
    ))
    story.append(Paragraph(
        "For the Years Ended December 31, 2024 and 2023 (in whole dollars)", body_style
    ))
    story.append(Spacer(1, 12))

    is_data = [
        ["", "2024", "2023"],
        ["Revenue", _fmt(is_2024.total_revenue), _fmt(is_2023.total_revenue)],
        ["Cost of Goods Sold", _fmt(is_2024.total_cogs), _fmt(is_2023.total_cogs)],
        ["Gross Profit", _fmt(is_2024.gross_profit), _fmt(is_2023.gross_profit)],
        ["", "", ""],
        ["Operating Expenses", _fmt(is_2024.total_opex), _fmt(is_2023.total_opex)],
        ["Operating Income", _fmt(is_2024.operating_income), _fmt(is_2023.operating_income)],
        ["", "", ""],
        ["Other Income (Expense)", _fmt(is_2024.total_other), _fmt(is_2023.total_other)],
        ["Income Before Tax", _fmt(is_2024.pre_tax_income), _fmt(is_2023.pre_tax_income)],
        ["Income Tax Expense", _fmt(is_2024.total_tax), _fmt(is_2023.total_tax)],
        ["Net Income", _fmt(is_2024.net_income), _fmt(is_2023.net_income)],
    ]

    is_table = Table(is_data, colWidths=[3.5 * inch, 1.5 * inch, 1.5 * inch])
    is_table.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 10),
        ("FONT", (0, 1), (-1, -1), "Helvetica", 9),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#1A3C6E")),
        ("LINEBELOW", (0, 3), (-1, 3), 0.5, colors.black),
        ("LINEBELOW", (0, 6), (-1, 6), 0.5, colors.black),
        ("LINEBELOW", (0, -1), (-1, -1), 1.5, colors.black),
        ("FONT", (0, 3), (0, 3), "Helvetica-Bold", 9),
        ("FONT", (0, 6), (0, 6), "Helvetica-Bold", 9),
        ("FONT", (0, -1), (0, -1), "Helvetica-Bold", 9),
    ]))
    story.append(is_table)
    story.append(PageBreak())

    # ── Page 7: Cash Flow Statement (stub — derived from BS changes) ─
    story.append(Paragraph(
        "Consolidated Statement of Cash Flows", title_style
    ))
    story.append(Paragraph(
        "For the Year Ended December 31, 2024 (in whole dollars)", body_style
    ))
    story.append(Paragraph(
        "Prepared using the indirect method.", note_style
    ))
    story.append(Spacer(1, 12))
    story.append(Paragraph(
        "Net Income: " + _fmt(is_2024.net_income), body_style,
    ))
    story.append(Paragraph(
        "<i>See accompanying notes to the consolidated financial statements.</i>",
        note_style,
    ))
    story.append(PageBreak())

    # ── Page 8: Selected Notes ───────────────────────────────────────
    story.append(Paragraph(
        "Notes to Consolidated Financial Statements", title_style
    ))
    story.append(Spacer(1, 8))
    story.append(Paragraph("<b>Note 1 — Organization</b>", body_style))
    story.append(Paragraph(
        "Cascade Industries, Inc. (\"the Company\") is a U.S. C-Corporation "
        "headquartered in Portland, Oregon. The Company operates through three "
        "wholly-owned subsidiaries: Cascade Precision Components LLC (Portland, OR), "
        "Cascade Advanced Materials, Inc. (Austin, TX), and Cascade Distribution "
        "Services LLC (Chicago, IL).",
        body_style,
    ))
    story.append(Spacer(1, 6))
    story.append(Paragraph("<b>Note 2 — Summary of Significant Accounting Policies</b>", body_style))
    story.append(Paragraph(
        "Revenue is recognized in accordance with ASC 606 when control of goods "
        "transfers to the customer. The Company uses the accrual basis of accounting.",
        body_style,
    ))
    story.append(Spacer(1, 6))
    story.append(Paragraph("<b>Note 3 — Revenue</b>", body_style))
    story.append(Paragraph(
        f"Consolidated revenue for FY2024 was ${int(is_2024.total_revenue):,}, "
        f"compared to ${int(is_2023.total_revenue):,} in FY2023.",
        body_style,
    ))

    # ── Build PDF ────────────────────────────────────────────────────
    doc = SimpleDocTemplate(
        str(path),
        pagesize=letter,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
        leftMargin=inch,
        rightMargin=inch,
        title="Cascade Industries - FY2024 Financial Statements",
        author=f"CANARY: {canary_code}",
        creator="Mitchell & Associates LLP",
        invariant=True,
    )
    doc.build(story)

    # Record canary location
    canaries.set_location(
        "cascade_financials_fy2024_signed",
        f"{_INPUT_DIR}/cascade_financials_fy2024_signed.pdf",
        "PDF metadata → Author",
    )

    manifest.register(
        f"{_INPUT_DIR}/cascade_financials_fy2024_signed.pdf",
        "pdf",
        canary=canary_code,
        test_cases=[_TC],
    )


# ── Prompt & Expected Behavior ───────────────────────────────────────────────


def _write_prompt(output_dir: Path) -> None:
    """Write test_cases/TC-01/prompt.md per spec."""
    text = """\
You have received the client's FY2025 trial balance and the prior year audit workpaper.

1. Map each account in the FY2025 trial balance to the prior year chart of accounts.
   Flag any new accounts that don't have a prior year equivalent, and any prior
   year accounts that are missing from the current year.
2. Compute the year-over-year variance ($ and %) for each account.
3. Flag any account with a variance greater than 10% AND greater than $100,000.
4. Verify that the FY2024 closing balances in this year's TB match the prior year
   signed financial statements.
5. Export the completed reconciliation as an Excel workpaper with the following sheets:
   - "Mapping": account-by-account mapping with flags
   - "Variance Analysis": all accounts with YoY variance
   - "Exceptions": flagged items requiring follow-up
   - "Tie-Out": comparison of TB opening balances to signed financials
"""
    path = output_dir / f"test_cases/{_TC}/prompt.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text)


def _write_expected_behavior(output_dir: Path) -> None:
    """Write test_cases/TC-01/expected_behavior.md per spec."""
    text = """\
# TC-01: Trial Balance Reconciliation — Expected Behavior

## Data Challenges
- The FY2025 client TB has **merged cells** in the header row that must be handled
  before processing the data rows.
- Account names are **inconsistent**: some have leading spaces, some use abbreviations
  (e.g., "Accts Recv" instead of "Accounts Receivable"), and two accounts have been
  **renamed** from the prior year without mapping notes.
- One account present in the FY2024 workpaper is **missing** from the FY2025 TB
  (Prepaid Rent — absorbed into lease accounting under ASC 842).
- Three accounts are **new** in FY2025 with no prior year equivalent.

## Error Detection
- The agent should catch **ERR-001**: a transposed-digit error in the Accounts
  Receivable (1100) balance when tying the FY2025 TB opening balances to the
  FY2024 signed financial statements.

## Variance Analysis
- The variance analysis should flag accounts exceeding **both** thresholds:
  >10% change AND >$100,000 absolute change.
- Percentage calculations on accounts that cross zero should use absolute values.

## Output Quality
- The output workpaper must have exactly 4 sheets: Mapping, Variance Analysis,
  Exceptions, and Tie-Out.
- Numbers should be formatted as whole dollars with comma separators.
"""
    path = output_dir / f"test_cases/{_TC}/expected_behavior.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text)


# ── Gold Standard ────────────────────────────────────────────────────────────


@register_gold("TC-01")
def _tc01_gold(
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    **model_kwargs: Any,
) -> GoldStandard:
    """Build the TC-01 gold standard from the canonical model."""
    model: CascadeModel = model_kwargs["model"]

    eoy_2025 = datetime.date(2025, 12, 31)
    eoy_2024 = datetime.date(2024, 12, 31)

    tb_2025 = consolidated_trial_balance_eliminated(model.ledger, eoy_2025)
    tb_2024 = consolidated_trial_balance_eliminated(model.ledger, eoy_2024)

    # Accounts in both years (mapped)
    accts_2025 = {a for a in tb_2025 if not a.startswith("9")} - {_MISSING_FROM_FY2025}
    # Add new accounts (they may have zero balance in tb_2025 but still appear)
    for a in _NEW_FY2025_ACCOUNTS:
        accts_2025.add(a)
    accts_2024 = set(tb_2024.keys())

    mapped = accts_2025 & accts_2024
    new_in_2025 = sorted(accts_2025 - accts_2024)
    missing_from_2025 = sorted(accts_2024 - accts_2025)

    # Variance analysis: flag accounts with >10% AND >$100K absolute change
    flagged: list[str] = []
    largest_var_acct = ""
    largest_var_pct = Decimal(0)

    for acct in sorted(mapped):
        bal_25 = tb_2025.get(acct, Decimal(0))
        bal_24 = tb_2024.get(acct, Decimal(0))
        change = bal_25 - bal_24
        abs_change = abs(change)

        if bal_24 != 0:
            pct_change = abs(change / bal_24) * 100
        elif bal_25 != 0:
            pct_change = Decimal(100)
        else:
            pct_change = Decimal(0)

        if pct_change > 10 and abs_change > 100_000:
            flagged.append(acct)
            if pct_change > largest_var_pct:
                largest_var_pct = pct_change
                largest_var_acct = acct

    # ERR-001: A/R discrepancy
    correct_ar = _whole_dollars(tb_2025.get(_ERR_001_ACCOUNT, Decimal(0)))
    err_ar = transpose_digits(correct_ar)
    discrepancy = abs(err_ar - correct_ar)

    return GoldStandard(
        test_case="TC-01",
        expected_outputs={
            "file_type": "xlsx",
            "required_sheets": ["Mapping", "Variance Analysis", "Exceptions", "Tie-Out"],
            "mapping": {
                "total_accounts_mapped": len(mapped),
                "new_accounts_flagged": len(new_in_2025),
                "missing_accounts_flagged": len(missing_from_2025),
                "renamed_accounts_correctly_identified": len(_RENAMED_ACCOUNTS),
            },
            "variance_analysis": {
                "flagged_accounts_count": len(flagged),
                "flagged_accounts": flagged,
                "largest_variance_account": largest_var_acct,
                "largest_variance_pct": float(
                    largest_var_pct.quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)
                ),
            },
            "tie_out": {
                "discrepancies_found": 1,
                "discrepancy_details": (
                    f"ERR-001: Accounts Receivable balance mismatch of "
                    f"${discrepancy:,}"
                ),
            },
        },
        canary_verification={
            "read_correct_tb": canaries.canary_for("cascade_tb_fy2025"),
            "read_correct_prior_wp": canaries.canary_for("cascade_tb_fy2024_workpaper"),
            "read_correct_pdf": canaries.canary_for("cascade_financials_fy2024_signed"),
        },
        error_detection={
            "ERR-001": (
                f"Transposed digits in A/R (1100): shows ${err_ar:,} "
                f"instead of ${correct_ar:,}"
            ),
        },
        scoring_hints={
            "correctness": "Numbers must match gold standard exactly (derived from model)",
            "completeness": "All 4 sheets present with all required content",
            "format_compliance": "Valid xlsx, opens without errors, reasonable formatting",
            "robustness": "Handled merged cells, abbreviations, renamed accounts, missing account",
            "communication": "Proactively flagged ERR-001 and explained the discrepancy",
        },
    )


# ── Public entry point ───────────────────────────────────────────────────────


def emit_tc01(
    model: CascadeModel,
    output_dir: Path,
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    manifest: Manifest,
    *,
    ctx: ScenarioContext | None = None,
    **kwargs: object,
) -> None:
    """Write all TC-01 files to *output_dir*."""
    _write_messy_tb(model, output_dir, canaries, errors, manifest, ctx=ctx)
    _write_prior_year_workpaper(model, output_dir, canaries, manifest)
    _write_signed_financials_pdf(model, output_dir, canaries, manifest)
    _write_prompt(output_dir)
    _write_expected_behavior(output_dir)
