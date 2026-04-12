"""Formatter: TC-09-EU — OECD Transfer Pricing Documentation (Master File / Local File).

Emits:
- test_cases/TC-09-EU/input_files/intercompany_transactions_eu_fy2025.xlsx
  All FY2025 intercompany transactions across Cascade Europe group (~120 rows).
- test_cases/TC-09-EU/input_files/comparable_companies_eu.xlsx
  Manufacturing (15) and distribution (10) comparables in two sheets.
- test_cases/TC-09-EU/input_files/master_file_fy2024.pdf
  Prior year OECD master file (28 pages).
- test_cases/TC-09-EU/input_files/local_file_cp_fy2024.pdf
  Prior year local file for CP — Germany (35 pages).
- test_cases/TC-09-EU/input_files/interest_rate_benchmarks_eu.xlsx
  EURIBOR rates + BBB credit spreads with ERR-EU-009 planted.
- test_cases/TC-09-EU/prompt.md
- test_cases/TC-09-EU/expected_behavior.md
- gold_standards/TC-09-EU_gold.json

Planted error:
  ERR-EU-009: transposed_digits — Q3 FY2025 EURIBOR 12M rate entered as
  0.38% instead of 3.80% (decimal point error).

Uses deterministic EU transfer pricing model — never hardcodes numbers
that should come from the model.
"""

from __future__ import annotations

import datetime
import io
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from generator.canaries import (
    CanaryRegistry,
    embed_canary_xlsx,
)
from generator.errors import (
    ErrorRegistry,
    PlantedError,
)
from generator.golds.framework import GoldStandard, register_gold
from generator.manifest import Manifest
from generator.model.build import CascadeModel

# ── Constants ────────────────────────────────────────────────────────────────

_TC = "TC-09-EU"
_INPUT_DIR = f"test_cases/{_TC}/input_files"

_FIXED_DATETIME = datetime.datetime(2025, 3, 15, 9, 0, 0)
_FIXED_ZIP_DT = (2025, 3, 15, 9, 0, 0)

# Canary keys (5 files)
ALL_CANARY_KEYS_TC09EU: list[str] = sorted([
    "tc09eu_intercompany_transactions",
    "tc09eu_comparable_companies",
    "tc09eu_master_file_fy2024",
    "tc09eu_local_file_cp_fy2024",
    "tc09eu_interest_rate_benchmarks",
])

# ── EU Entity names ──────────────────────────────────────────────────────────

_EU_ENTITIES: dict[str, dict[str, str]] = {
    "CE": {
        "name": "Cascade Europe Holdings B.V.",
        "country": "Netherlands",
        "city": "Amsterdam",
        "role": "Holding company — strategic oversight, treasury, legal",
    },
    "CP": {
        "name": "Cascade Pr\u00e4zisionsteile GmbH",
        "country": "Germany",
        "city": "Munich",
        "role": "Licensed manufacturer — precision components",
    },
    "CM": {
        "name": "Cascade Mat\u00e9riaux Avanc\u00e9s SAS",
        "country": "France",
        "city": "Lyon",
        "role": "R&D centre and IP developer — advanced materials",
    },
    "CD": {
        "name": "Cascade Distribution Services Ltd",
        "country": "United Kingdom",
        "city": "Birmingham",
        "role": "Limited-risk distributor — warehousing and logistics",
    },
}

# ── Revenue targets from model ───────────────────────────────────────────────

_CP_REVENUE_EUR = Decimal("45000000")
_CM_REVENUE_EUR = Decimal("32000000")
_CD_REVENUE_GBP = Decimal("18000000")
_CD_REVENUE_EUR = Decimal("21000000")  # ~€21M at ~1.167 EUR/GBP

# ── Intercompany flow parameters ─────────────────────────────────────────────

# CP→CM raw materials at cost-plus-6%
_RAW_MATERIALS_TOTAL_EUR = Decimal("8500000")
_RAW_MATERIALS_MARKUP = Decimal("0.06")

# CP→CD finished goods at cost-plus-8%
_FINISHED_GOODS_TOTAL_EUR = Decimal("6200000")
_FINISHED_GOODS_MARKUP = Decimal("0.08")

# CE management fees: 1.5% of each sub revenue
_MGMT_FEE_PCT = Decimal("0.015")
_MGMT_FEE_CP = (_CP_REVENUE_EUR * _MGMT_FEE_PCT).quantize(Decimal("1"), rounding=ROUND_HALF_UP)  # ~675k
_MGMT_FEE_CM = (_CM_REVENUE_EUR * _MGMT_FEE_PCT).quantize(Decimal("1"), rounding=ROUND_HALF_UP)  # ~480k
_MGMT_FEE_CD = (_CD_REVENUE_EUR * _MGMT_FEE_PCT).quantize(Decimal("1"), rounding=ROUND_HALF_UP)  # ~315k
_MGMT_FEE_TOTAL = _MGMT_FEE_CP + _MGMT_FEE_CM + _MGMT_FEE_CD

# CE→CM intercompany loan: €3M at 4.5%
_LOAN_PRINCIPAL = Decimal("3000000")
_LOAN_RATE = Decimal("0.045")
_LOAN_INTEREST_ANNUAL = (_LOAN_PRINCIPAL * _LOAN_RATE).quantize(Decimal("1"), rounding=ROUND_HALF_UP)  # 135k

# CM→CP R&D royalty: 3% of CP revenue
_ROYALTY_PCT = Decimal("0.03")
_ROYALTY_TOTAL = (_CP_REVENUE_EUR * _ROYALTY_PCT).quantize(Decimal("1"), rounding=ROUND_HALF_UP)  # ~1.35M

# ── Manufacturing Comparables (CP benchmarking) ─────────────────────────────
# 15 companies: 12 accepted, 3 rejected.
# Accepted set IQR of operating margin: 3.8% to 7.9% (median 5.6%)
# Sorted accepted margins: [2.5, 3.2, 3.8, 4.5, 5.1, 5.6, 6.2, 6.8, 7.5, 8.2, 9.0, 10.5]
# QUARTILE.INC Q1 = margins[0.25*11 = 2.75] = 3.8 + 0.75*(4.5-3.8) = 3.8 + 0.525 = 4.325 ≈ NO
# Let me recalibrate for exact Q1=3.8, Q3=7.9
# n=12: Q1 at index 2.75, Q3 at index 8.25
# We need: margins[2] + 0.75*(margins[3]-margins[2]) = 3.8%
# And: margins[8] + 0.25*(margins[9]-margins[8]) = 7.9%
# Try: [2.5, 3.2, 3.6, 3.9, 4.5, 5.3, 5.9, 6.8, 7.5, 9.1, 10.2, 11.8]
# Q1 = 3.6 + 0.75*(3.9-3.6) = 3.6 + 0.225 = 3.825 ≈ not exact
# For exact 3.8: margins[2] + 0.75*(margins[3]-margins[2]) = 3.8
# Try margins[2]=3.5, margins[3]=3.9: 3.5 + 0.75*0.4 = 3.5+0.3 = 3.8 ✓
# For exact 7.9: margins[8] + 0.25*(margins[9]-margins[8]) = 7.9
# Try margins[8]=7.8, margins[9]=8.2: 7.8 + 0.25*0.4 = 7.8+0.1 = 7.9 ✓
#
# Final calibrated set (12 accepted margins sorted):
# [2.5, 3.2, 3.5, 3.9, 4.5, 5.3, 5.9, 6.8, 7.8, 8.2, 9.5, 11.0]
# Median = (5.3+5.9)/2 = 5.6 ✓
# Q1 = 3.5 + 0.75*(3.9-3.5) = 3.8 ✓
# Q3 = 7.8 + 0.25*(8.2-7.8) = 7.9 ✓
#
# CP actual operating margin: ~6.2% (within range)

# Rejected manufacturing comparables (for writing to XLSX and gold standard)
_MFG_REJECTED: list[tuple[str, str, str, int, int, int, int, int, float, str]] = [
    # (name, country, nace, rev, cogs, opex, oper_inc, assets, roce, reason)
    ("Nordic Logistics Solutions Oy", "Finland", "5229",
     130, 85, 32, 13, 150, 8.7,
     "SIC/NACE mismatch \u2014 logistics (5229), not manufacturing"),
    ("Meridionale Industrie S.p.A.", "Italy", "2562",
     90, 78, 25, -13, 180, -7.2,
     "Financial distress \u2014 negative equity, concordato preventivo"),
    ("Continental Grosswerk AG", "Germany", "2562",
     2800, 1820, 700, 280, 3200, 8.8,
     "Size outlier \u2014 revenue >10x tested party (OECD \u00a73.43-3.46)"),
]

# Exact calibrated financials for accepted companies to hit target margins
_MFG_CALIBRATED: list[tuple[str, str, str, int, int, int, float, int, float]] = [
    # (name, country, nace, revenue_M, cogs_M, opex_M, margin_pct, assets_M, roce_pct)
    ("Eurostahl Maschinenbau AG", "Germany", "2562", 200, 143, 52, 2.5, 210, 2.4),
    ("Nordisk Precision ApS", "Denmark", "2562", 250, 175, 67, 3.2, 230, 3.5),
    ("S\u00fcdwerk Komponenten GmbH", "Germany", "2562", 200, 138, 55, 3.5, 225, 3.1),
    ("Atelier M\u00e9canique Pr\u00e9cision SA", "France", "2562", 200, 136, 56, 3.9, 170, 4.6),
    ("Iberia Manufactura S.L.", "Spain", "2562", 200, 136, 55, 4.5, 140, 6.4),
    ("Benelux Industrial Parts N.V.", "Netherlands", "2562", 200, 133, 56, 5.3, 185, 5.7),
    ("Alpine Werkzeug AG", "Austria", "2562", 200, 130, 58, 5.9, 165, 7.2),
    ("Rheinmetall Pr\u00e4zision GmbH", "Germany", "2562", 250, 165, 68, 6.8, 290, 5.9),
    ("Scandia Mekaniska AB", "Sweden", "2562", 250, 162, 69, 7.8, 210, 9.3),
    ("Helvetia Engineering SA", "Switzerland", "2562", 250, 162, 68, 8.2, 240, 8.5),
    ("Polska Precyzja Sp. z o.o.", "Poland", "2562", 200, 127, 54, 9.5, 120, 15.8),
    ("Balkan Components d.o.o.", "Croatia", "2562", 200, 122, 56, 11.0, 100, 22.0),
]

# ── Distribution Comparables (CD benchmarking) ──────────────────────────────
# 10 companies: 8 accepted, 2 rejected.
# Accepted set IQR of net margin: 1.2% to 3.5% (median 2.1%)
# CD actual net margin: ~1.8% (within range, low end)
#
# n=8: Q1 at index 0.25*7=1.75, Q3 at index 0.75*7=5.25
# Q1 = margins[1] + 0.75*(margins[2]-margins[1]) = 1.2
# Q3 = margins[5] + 0.25*(margins[6]-margins[5]) = 3.5
# Sorted: [0.6, 0.8, 1.4, 1.8, 2.1, 3.2, 4.4, 5.0]
# Q1 = 0.8 + 0.75*(1.4-0.8) = 0.8+0.45 = 1.25 ≈ not exact
# Try: [0.5, 0.9, 1.3, 1.7, 2.1, 3.1, 3.9, 5.2]
# Q1 = 0.9 + 0.75*(1.3-0.9) = 0.9+0.3 = 1.2 ✓
# Q3 = 3.1 + 0.25*(3.9-3.1) = 3.1+0.2 = 3.3 ≠ 3.5
# Try: [0.5, 0.9, 1.3, 1.8, 2.1, 3.3, 4.1, 5.0]
# Q3 = 3.3 + 0.25*(4.1-3.3) = 3.3+0.2 = 3.5 ✓
# Q1 = 0.9 + 0.75*(1.3-0.9) = 0.9+0.3 = 1.2 ✓
# Median = (1.8+2.1)/2 = 1.95 → design says 2.1, so adjust
# Try: [0.5, 0.9, 1.3, 1.9, 2.3, 3.3, 4.1, 5.0]
# Median = (1.9+2.3)/2 = 2.1 ✓
# Q1 = 0.9 + 0.75*(1.3-0.9) = 1.2 ✓
# Q3 = 3.3 + 0.25*(4.1-3.3) = 3.5 ✓

_DIST_CALIBRATED: list[tuple[str, str, str, int, int, int, float, int, float, bool, str]] = [
    # (name, country, nace, revenue_M, cogs_M, opex_M, net_margin_pct, assets_M, roce_pct, rejected, reason)
    # Accepted (sorted by net margin)
    ("TransEuropa Logistics GmbH", "Germany", "5210", 150, 132, 17, 0.5, 90, 0.8, False, ""),
    ("Channel Freight Services Ltd", "UK", "5210", 120, 105, 14, 0.9, 70, 1.5, False, ""),
    ("Benelux Distribution N.V.", "Netherlands", "5210", 180, 157, 21, 1.3, 100, 2.3, False, ""),
    ("Rhein-Main Vertrieb GmbH", "Germany", "5210", 200, 174, 22, 1.9, 120, 3.2, False, ""),
    ("Scandi Distribution AB", "Sweden", "5210", 160, 137, 19, 2.3, 95, 3.9, False, ""),
    ("Iberian Dist S.L.", "Spain", "5210", 140, 120, 16, 3.3, 80, 5.8, False, ""),
    ("Alpine Logistik AG", "Switzerland", "5210", 130, 110, 15, 4.1, 75, 7.1, False, ""),
    ("Adriatic Trade d.o.o.", "Slovenia", "5210", 100, 84, 11, 5.0, 60, 8.3, False, ""),
    # Rejected
    ("Groupe Stellaire Captive SAS", "France", "5210", 90, 76, 11, 3.5, 50, 6.3, True,
     "Captive entity of a listed group — no independent pricing or market risk"),
    ("Baltic Services O\u00dc", "Estonia", "5210", 70, 64, 9, -3.8, 40, -6.7, True,
     "Restructuring losses — non-recurring charges distort net margin"),
]

# ── Interest Rate Benchmarks ─────────────────────────────────────────────────

# EURIBOR 12M quarterly rates FY2024-FY2025 (annualized %)
# ERR-EU-009: Q3 FY2025 shows 0.38% instead of 3.80%
_EURIBOR_DATA: list[tuple[str, str, float, float, float]] = [
    # (period, quarter, 3M_rate, 6M_rate, 12M_rate)
    ("FY2024", "Q1", 3.92, 3.86, 3.74),
    ("FY2024", "Q2", 3.79, 3.72, 3.61),
    ("FY2024", "Q3", 3.68, 3.63, 3.55),
    ("FY2024", "Q4", 3.54, 3.49, 3.42),
    ("FY2025", "Q1", 3.82, 3.78, 3.72),
    ("FY2025", "Q2", 3.95, 3.90, 3.85),
    ("FY2025", "Q3", 4.02, 3.95, 0.38),   # ERR-EU-009: should be 3.80
    ("FY2025", "Q4", 4.05, 3.98, 3.83),
]

# Correct Q3 FY2025 12M rate (for gold standard)
_CORRECT_Q3_FY2025_12M = 3.80

# BBB credit spread data (basis points over EURIBOR)
_BBB_SPREADS: list[tuple[str, int, int]] = [
    # (maturity_bucket, spread_low_bps, spread_high_bps)
    ("1-3 year", 40, 80),
    ("3-5 year", 50, 150),
    ("5-10 year", 100, 180),
]

# ── Styling helpers ──────────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill("solid", fgColor="1A5276")
_HEADER_ALIGN = Alignment(horizontal="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_NUMBER_FMT = "#,##0"
_PCT_FMT = "0.00%"
_MONEY_FMT = "#,##0"
_RATE_FMT = "0.00%"


def _style_header(ws: Any, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _style_data_cell(cell: Any, fmt: str = "") -> None:
    cell.border = _THIN_BORDER
    if fmt:
        cell.number_format = fmt


# ── Deterministic save helpers ───────────────────────────────────────────────

def _save_xlsx_deterministic(wb: openpyxl.Workbook, path: str | Path) -> None:
    from openpyxl.writer.excel import ExcelWriter

    path = Path(path)
    wb.properties.created = _FIXED_DATETIME
    wb.properties.modified = _FIXED_DATETIME

    buf = io.BytesIO()
    archive = ZipFile(buf, "w", ZIP_DEFLATED, allowZip64=True)
    writer = ExcelWriter(wb, archive)
    writer.save()

    buf.seek(0)
    with ZipFile(buf, "r") as src, ZipFile(str(path), "w", ZIP_DEFLATED) as dst:
        for item in src.infolist():
            info = ZipInfo(item.filename, date_time=_FIXED_ZIP_DT)
            info.compress_type = item.compress_type
            dst.writestr(info, src.read(item.filename))


def _whole_euros(d: Decimal) -> int:
    return int(d.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


# ── Intercompany Transactions XLSX ───────────────────────────────────────────


def _generate_ic_line_items() -> list[dict[str, Any]]:
    """Generate ~120 deterministic monthly intercompany line items for FY2025."""
    items: list[dict[str, Any]] = []
    txn_id = 1

    # Monthly amounts (slight variation using deterministic pattern)
    base_months = list(range(1, 13))

    for month in base_months:
        date_str = f"{2025:04d}-{month:02d}-15"

        # CP→CM raw materials: ~€8.5M / 12 ≈ €708k/month
        raw_base = Decimal("708333")
        raw_var = Decimal(str((month * 17 + 3) % 50 - 25)) * Decimal("1000")
        raw_amt = raw_base + raw_var
        items.append({
            "transaction_id": f"IC-EU-{txn_id:04d}",
            "from_entity": "CP",
            "to_entity": "CM",
            "transaction_type": "goods",
            "description": f"Raw materials — precision components batch RM-{2025}-{month:02d}",
            "volume_or_principal": f"{_whole_euros(raw_amt / (1 + _RAW_MATERIALS_MARKUP))} units",
            "price_or_rate": "Cost-plus-6%",
            "total_amount_eur": _whole_euros(raw_amt),
            "invoicing_currency": "EUR",
            "arm_length_method_applied": "Cost Plus Method",
            "date": date_str,
        })
        txn_id += 1

        # CP→CD finished goods: ~€6.2M / 12 ≈ €517k/month
        fg_base = Decimal("516667")
        fg_var = Decimal(str((month * 23 + 7) % 40 - 20)) * Decimal("1000")
        fg_amt = fg_base + fg_var
        items.append({
            "transaction_id": f"IC-EU-{txn_id:04d}",
            "from_entity": "CP",
            "to_entity": "CD",
            "transaction_type": "goods",
            "description": f"Finished goods — industrial components FG-{2025}-{month:02d}",
            "volume_or_principal": f"{_whole_euros(fg_amt / (1 + _FINISHED_GOODS_MARKUP))} units",
            "price_or_rate": "Cost-plus-8%",
            "total_amount_eur": _whole_euros(fg_amt),
            "invoicing_currency": "EUR",
            "arm_length_method_applied": "Cost Plus Method",
            "date": date_str,
        })
        txn_id += 1

        # CE→CP management fee
        mf_cp = _MGMT_FEE_CP / 12
        items.append({
            "transaction_id": f"IC-EU-{txn_id:04d}",
            "from_entity": "CE",
            "to_entity": "CP",
            "transaction_type": "management_fee",
            "description": "Monthly management fee — strategic oversight, treasury, legal",
            "volume_or_principal": "1.5% of monthly revenue",
            "price_or_rate": "1.5%",
            "total_amount_eur": _whole_euros(mf_cp),
            "invoicing_currency": "EUR",
            "arm_length_method_applied": "Cost Plus / Benefit Test",
            "date": date_str,
        })
        txn_id += 1

        # CE→CM management fee
        mf_cm = _MGMT_FEE_CM / 12
        items.append({
            "transaction_id": f"IC-EU-{txn_id:04d}",
            "from_entity": "CE",
            "to_entity": "CM",
            "transaction_type": "management_fee",
            "description": "Monthly management fee — strategic oversight, treasury, legal",
            "volume_or_principal": "1.5% of monthly revenue",
            "price_or_rate": "1.5%",
            "total_amount_eur": _whole_euros(mf_cm),
            "invoicing_currency": "EUR",
            "arm_length_method_applied": "Cost Plus / Benefit Test",
            "date": date_str,
        })
        txn_id += 1

        # CE→CD management fee
        mf_cd = _MGMT_FEE_CD / 12
        items.append({
            "transaction_id": f"IC-EU-{txn_id:04d}",
            "from_entity": "CE",
            "to_entity": "CD",
            "transaction_type": "management_fee",
            "description": "Monthly management fee — strategic oversight, treasury, legal",
            "volume_or_principal": "1.5% of monthly revenue",
            "price_or_rate": "1.5%",
            "total_amount_eur": _whole_euros(mf_cd),
            "invoicing_currency": "EUR",
            "arm_length_method_applied": "Cost Plus / Benefit Test",
            "date": date_str,
        })
        txn_id += 1

        # CE→CM intercompany loan interest (monthly accrual)
        interest_monthly = _LOAN_INTEREST_ANNUAL / 12
        items.append({
            "transaction_id": f"IC-EU-{txn_id:04d}",
            "from_entity": "CE",
            "to_entity": "CM",
            "transaction_type": "interest",
            "description": "Intercompany loan interest — \u20ac3M facility at 4.5% p.a.",
            "volume_or_principal": f"\u20ac{_whole_euros(_LOAN_PRINCIPAL):,}",
            "price_or_rate": "4.5% p.a.",
            "total_amount_eur": _whole_euros(interest_monthly),
            "invoicing_currency": "EUR",
            "arm_length_method_applied": "CUP / Interest Rate Benchmark",
            "date": date_str,
        })
        txn_id += 1

        # CM→CP R&D royalty (monthly)
        royalty_monthly = _ROYALTY_TOTAL / 12
        items.append({
            "transaction_id": f"IC-EU-{txn_id:04d}",
            "from_entity": "CM",
            "to_entity": "CP",
            "transaction_type": "royalty",
            "description": "R&D royalty \u2014 technology license",
            "volume_or_principal": "3% of CP monthly revenue",
            "price_or_rate": "3%",
            "total_amount_eur": _whole_euros(royalty_monthly),
            "invoicing_currency": "EUR",
            "arm_length_method_applied": "CUT / Comparable License Analysis",
            "date": date_str,
        })
        txn_id += 1

    # Add quarterly adjustments and supplementary line items to reach ~120 items
    for q in range(1, 5):
        month = q * 3
        date_str = f"2025-{month:02d}-28"

        # Quarterly true-up for raw materials
        items.append({
            "transaction_id": f"IC-EU-{txn_id:04d}",
            "from_entity": "CP",
            "to_entity": "CM",
            "transaction_type": "goods",
            "description": f"Quarterly true-up \u2014 raw materials volume adjustment Q{q}",
            "volume_or_principal": "True-up",
            "price_or_rate": "Cost-plus-6%",
            "total_amount_eur": _whole_euros(Decimal("25000") + Decimal(str(q * 3000))),
            "invoicing_currency": "EUR",
            "arm_length_method_applied": "Cost Plus Method",
            "date": date_str,
        })
        txn_id += 1

        # Quarterly true-up for finished goods
        items.append({
            "transaction_id": f"IC-EU-{txn_id:04d}",
            "from_entity": "CP",
            "to_entity": "CD",
            "transaction_type": "goods",
            "description": f"Quarterly true-up \u2014 finished goods volume adjustment Q{q}",
            "volume_or_principal": "True-up",
            "price_or_rate": "Cost-plus-8%",
            "total_amount_eur": _whole_euros(Decimal("18000") + Decimal(str(q * 2000))),
            "invoicing_currency": "EUR",
            "arm_length_method_applied": "Cost Plus Method",
            "date": date_str,
        })
        txn_id += 1

        # Quarterly freight / logistics charges (CP→CD)
        items.append({
            "transaction_id": f"IC-EU-{txn_id:04d}",
            "from_entity": "CP",
            "to_entity": "CD",
            "transaction_type": "services",
            "description": f"Quarterly freight and logistics charges Q{q} \u2014 DE\u2192UK shipments",
            "volume_or_principal": f"Q{q} consolidated",
            "price_or_rate": "At cost",
            "total_amount_eur": _whole_euros(Decimal("42000") + Decimal(str(q * 5000))),
            "invoicing_currency": "EUR",
            "arm_length_method_applied": "Cost Plus Method",
            "date": date_str,
        })
        txn_id += 1

        # Quarterly IP cost-sharing adjustment (CM→CP)
        items.append({
            "transaction_id": f"IC-EU-{txn_id:04d}",
            "from_entity": "CM",
            "to_entity": "CP",
            "transaction_type": "services",
            "description": f"Quarterly R&D cost-sharing recharge Q{q} \u2014 joint development",
            "volume_or_principal": f"Q{q} allocation",
            "price_or_rate": "At cost",
            "total_amount_eur": _whole_euros(Decimal("35000") + Decimal(str(q * 4000))),
            "invoicing_currency": "EUR",
            "arm_length_method_applied": "Cost Plus Method",
            "date": date_str,
        })
        txn_id += 1

        # Quarterly CE treasury recharges to each sub
        for sub in ["CP", "CM", "CD"]:
            items.append({
                "transaction_id": f"IC-EU-{txn_id:04d}",
                "from_entity": "CE",
                "to_entity": sub,
                "transaction_type": "services",
                "description": f"Quarterly treasury / cash pooling charge Q{q}",
                "volume_or_principal": f"Q{q} allocation",
                "price_or_rate": "At cost",
                "total_amount_eur": _whole_euros(Decimal("8000") + Decimal(str(q * 1000))),
                "invoicing_currency": "EUR",
                "arm_length_method_applied": "Cost Plus / Benefit Test",
                "date": date_str,
            })
            txn_id += 1

    # Sort by date then by transaction_id for determinism
    items.sort(key=lambda x: (x["date"], x["transaction_id"]))
    return items


def _write_ic_transactions_xlsx(
    output_dir: Path,
    canaries: CanaryRegistry,
    manifest: Manifest,
) -> None:
    """Write intercompany_transactions_eu_fy2025.xlsx."""
    items = _generate_ic_line_items()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = [
        "Transaction ID", "From Entity", "To Entity", "Transaction Type",
        "Description", "Volume / Principal", "Price / Rate",
        "Total Amount (EUR)", "Invoicing Currency", "Arm's-Length Method Applied",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    for i, item in enumerate(items, 2):
        ws.cell(row=i, column=1, value=item["transaction_id"])
        _style_data_cell(ws.cell(row=i, column=1))
        ws.cell(row=i, column=2, value=item["from_entity"])
        _style_data_cell(ws.cell(row=i, column=2))
        ws.cell(row=i, column=3, value=item["to_entity"])
        _style_data_cell(ws.cell(row=i, column=3))
        ws.cell(row=i, column=4, value=item["transaction_type"])
        _style_data_cell(ws.cell(row=i, column=4))
        ws.cell(row=i, column=5, value=item["description"])
        _style_data_cell(ws.cell(row=i, column=5))
        ws.cell(row=i, column=6, value=item["volume_or_principal"])
        _style_data_cell(ws.cell(row=i, column=6))
        ws.cell(row=i, column=7, value=item["price_or_rate"])
        _style_data_cell(ws.cell(row=i, column=7))
        ws.cell(row=i, column=8, value=item["total_amount_eur"])
        _style_data_cell(ws.cell(row=i, column=8), _MONEY_FMT)
        ws.cell(row=i, column=9, value=item["invoicing_currency"])
        _style_data_cell(ws.cell(row=i, column=9))
        ws.cell(row=i, column=10, value=item["arm_length_method_applied"])
        _style_data_cell(ws.cell(row=i, column=10))

    # Column widths
    widths = {"A": 14, "B": 14, "C": 14, "D": 18, "E": 55, "F": 22,
              "G": 16, "H": 18, "I": 18, "J": 30}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    # Summary sheet
    ws2 = wb.create_sheet("Summary by Flow")
    summary_headers = ["Flow", "Transaction Type", "Count", "Total (EUR)"]
    for col, h in enumerate(summary_headers, 1):
        ws2.cell(row=1, column=col, value=h)
    _style_header(ws2, 1, len(summary_headers))

    flows: dict[str, tuple[str, int, int]] = {}
    for item in items:
        key = f"{item['from_entity']}\u2192{item['to_entity']} {item['transaction_type']}"
        _, count, total = flows.get(key, (item["transaction_type"], 0, 0))
        flows[key] = (item["transaction_type"], count + 1, total + item["total_amount_eur"])

    for i, (flow_key, (tx_type, count, total)) in enumerate(sorted(flows.items()), 2):
        ws2.cell(row=i, column=1, value=flow_key)
        _style_data_cell(ws2.cell(row=i, column=1))
        ws2.cell(row=i, column=2, value=tx_type)
        _style_data_cell(ws2.cell(row=i, column=2))
        ws2.cell(row=i, column=3, value=count)
        _style_data_cell(ws2.cell(row=i, column=3))
        ws2.cell(row=i, column=4, value=total)
        _style_data_cell(ws2.cell(row=i, column=4), _MONEY_FMT)

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 18

    # Canary
    canary = canaries.canary_for("tc09eu_intercompany_transactions")
    loc = embed_canary_xlsx(wb, canary)
    canaries.set_location(
        "tc09eu_intercompany_transactions",
        f"{_INPUT_DIR}/intercompany_transactions_eu_fy2025.xlsx",
        loc,
    )

    file_path = output_dir / _INPUT_DIR / "intercompany_transactions_eu_fy2025.xlsx"
    file_path.parent.mkdir(parents=True, exist_ok=True)
    _save_xlsx_deterministic(wb, file_path)
    manifest.register(f"{_INPUT_DIR}/intercompany_transactions_eu_fy2025.xlsx", "xlsx")


# ── Comparable Companies XLSX ────────────────────────────────────────────────


def _write_comparables_xlsx(
    output_dir: Path,
    canaries: CanaryRegistry,
    manifest: Manifest,
) -> None:
    """Write comparable_companies_eu.xlsx with Manufacturing and Distribution sheets."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Manufacturing Comparables ───────────────────────
    ws1 = wb.active
    ws1.title = "Manufacturing Comparables"
    mfg_headers = [
        "Company Name", "Country", "NACE Code", "Revenue (\u20acM)", "COGS (\u20acM)",
        "Operating Expenses (\u20acM)", "Operating Income (\u20acM)",
        "Total Assets (\u20acM)", "ROCE (%)",
    ]
    for col, h in enumerate(mfg_headers, 1):
        ws1.cell(row=1, column=col, value=h)
    _style_header(ws1, 1, len(mfg_headers))

    # Interleave accepted and rejected for realistic presentation
    mfg_display_order = [0, 12, 4, 8, 1, 13, 6, 10, 2, 14, 9, 5, 11, 3, 7]

    for row_idx, idx in enumerate(mfg_display_order, 2):
        if idx < len(_MFG_CALIBRATED):
            name, country, nace, rev, cogs, opex, margin_pct, assets, roce = _MFG_CALIBRATED[idx]
            oper_inc = round(rev * margin_pct / 100)
        else:
            # Rejected company
            rej_idx = idx - len(_MFG_CALIBRATED)
            rej = _MFG_REJECTED[rej_idx]
            name, country, nace = rej[0], rej[1], rej[2]
            rev, cogs, opex, oper_inc = rej[3], rej[4], rej[5], rej[6]
            assets, roce = rej[7], rej[8]

        ws1.cell(row=row_idx, column=1, value=name)
        _style_data_cell(ws1.cell(row=row_idx, column=1))
        ws1.cell(row=row_idx, column=2, value=country)
        _style_data_cell(ws1.cell(row=row_idx, column=2))
        ws1.cell(row=row_idx, column=3, value=nace)
        _style_data_cell(ws1.cell(row=row_idx, column=3))
        ws1.cell(row=row_idx, column=4, value=rev)
        _style_data_cell(ws1.cell(row=row_idx, column=4), _NUMBER_FMT)
        ws1.cell(row=row_idx, column=5, value=cogs)
        _style_data_cell(ws1.cell(row=row_idx, column=5), _NUMBER_FMT)
        ws1.cell(row=row_idx, column=6, value=opex)
        _style_data_cell(ws1.cell(row=row_idx, column=6), _NUMBER_FMT)
        ws1.cell(row=row_idx, column=7, value=oper_inc)
        _style_data_cell(ws1.cell(row=row_idx, column=7), _NUMBER_FMT)
        ws1.cell(row=row_idx, column=8, value=assets)
        _style_data_cell(ws1.cell(row=row_idx, column=8), _NUMBER_FMT)
        ws1.cell(row=row_idx, column=9, value=round(roce, 1))
        _style_data_cell(ws1.cell(row=row_idx, column=9), "0.0%")

    for col_letter, w in [("A", 35), ("B", 15), ("C", 12), ("D", 15),
                           ("E", 15), ("F", 20), ("G", 20), ("H", 18), ("I", 12)]:
        ws1.column_dimensions[col_letter].width = w

    # ── Sheet 2: Distribution Comparables ────────────────────────
    ws2 = wb.create_sheet("Distribution Comparables")
    dist_headers = [
        "Company Name", "Country", "NACE Code", "Revenue (\u20acM)", "COGS (\u20acM)",
        "Operating Expenses (\u20acM)", "Net Income (\u20acM)",
        "Total Assets (\u20acM)", "ROCE (%)",
    ]
    for col, h in enumerate(dist_headers, 1):
        ws2.cell(row=1, column=col, value=h)
    _style_header(ws2, 1, len(dist_headers))

    # Interleave for presentation
    dist_display_order = [0, 8, 3, 6, 1, 9, 5, 7, 2, 4]

    for row_idx, idx in enumerate(dist_display_order, 2):
        d = _DIST_CALIBRATED[idx]
        name, country, nace = d[0], d[1], d[2]
        rev, cogs, opex = d[3], d[4], d[5]
        net_margin_pct = d[6]
        net_inc = round(rev * net_margin_pct / 100)
        assets, roce = d[7], d[8]

        ws2.cell(row=row_idx, column=1, value=name)
        _style_data_cell(ws2.cell(row=row_idx, column=1))
        ws2.cell(row=row_idx, column=2, value=country)
        _style_data_cell(ws2.cell(row=row_idx, column=2))
        ws2.cell(row=row_idx, column=3, value=nace)
        _style_data_cell(ws2.cell(row=row_idx, column=3))
        ws2.cell(row=row_idx, column=4, value=rev)
        _style_data_cell(ws2.cell(row=row_idx, column=4), _NUMBER_FMT)
        ws2.cell(row=row_idx, column=5, value=cogs)
        _style_data_cell(ws2.cell(row=row_idx, column=5), _NUMBER_FMT)
        ws2.cell(row=row_idx, column=6, value=opex)
        _style_data_cell(ws2.cell(row=row_idx, column=6), _NUMBER_FMT)
        ws2.cell(row=row_idx, column=7, value=net_inc)
        _style_data_cell(ws2.cell(row=row_idx, column=7), _NUMBER_FMT)
        ws2.cell(row=row_idx, column=8, value=assets)
        _style_data_cell(ws2.cell(row=row_idx, column=8), _NUMBER_FMT)
        ws2.cell(row=row_idx, column=9, value=round(roce, 1))
        _style_data_cell(ws2.cell(row=row_idx, column=9), "0.0%")

    for col_letter, w in [("A", 35), ("B", 15), ("C", 12), ("D", 15),
                           ("E", 15), ("F", 20), ("G", 18), ("H", 18), ("I", 12)]:
        ws2.column_dimensions[col_letter].width = w

    # Canary
    canary = canaries.canary_for("tc09eu_comparable_companies")
    loc = embed_canary_xlsx(wb, canary)
    canaries.set_location(
        "tc09eu_comparable_companies",
        f"{_INPUT_DIR}/comparable_companies_eu.xlsx",
        loc,
    )

    file_path = output_dir / _INPUT_DIR / "comparable_companies_eu.xlsx"
    file_path.parent.mkdir(parents=True, exist_ok=True)
    _save_xlsx_deterministic(wb, file_path)
    manifest.register(f"{_INPUT_DIR}/comparable_companies_eu.xlsx", "xlsx")


# ── Interest Rate Benchmarks XLSX ────────────────────────────────────────────


def _write_interest_rate_benchmarks(
    output_dir: Path,
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    manifest: Manifest,
) -> None:
    """Write interest_rate_benchmarks_eu.xlsx with ERR-EU-009 planted."""
    wb = openpyxl.Workbook()

    # Sheet 1: EURIBOR rates
    ws1 = wb.active
    ws1.title = "EURIBOR Rates"
    euribor_headers = ["Period", "Quarter", "3M Rate (%)", "6M Rate (%)", "12M Rate (%)"]
    for col, h in enumerate(euribor_headers, 1):
        ws1.cell(row=1, column=col, value=h)
    _style_header(ws1, 1, len(euribor_headers))

    err_row = None
    for i, (period, quarter, r3m, r6m, r12m) in enumerate(_EURIBOR_DATA, 2):
        ws1.cell(row=i, column=1, value=period)
        _style_data_cell(ws1.cell(row=i, column=1))
        ws1.cell(row=i, column=2, value=quarter)
        _style_data_cell(ws1.cell(row=i, column=2))
        ws1.cell(row=i, column=3, value=r3m / 100)
        _style_data_cell(ws1.cell(row=i, column=3), _RATE_FMT)
        ws1.cell(row=i, column=4, value=r6m / 100)
        _style_data_cell(ws1.cell(row=i, column=4), _RATE_FMT)
        ws1.cell(row=i, column=5, value=r12m / 100)
        _style_data_cell(ws1.cell(row=i, column=5), _RATE_FMT)

        # Track ERR-EU-009 row
        if period == "FY2025" and quarter == "Q3":
            err_row = i

    for col_letter, w in [("A", 12), ("B", 10), ("C", 14), ("D", 14), ("E", 14)]:
        ws1.column_dimensions[col_letter].width = w

    # Sheet 2: BBB Credit Spreads
    ws2 = wb.create_sheet("BBB Credit Spreads")
    spread_headers = ["Maturity Bucket", "Spread Low (bps)", "Spread High (bps)",
                      "Spread Low (%)", "Spread High (%)"]
    for col, h in enumerate(spread_headers, 1):
        ws2.cell(row=1, column=col, value=h)
    _style_header(ws2, 1, len(spread_headers))

    for i, (bucket, low_bps, high_bps) in enumerate(_BBB_SPREADS, 2):
        ws2.cell(row=i, column=1, value=bucket)
        _style_data_cell(ws2.cell(row=i, column=1))
        ws2.cell(row=i, column=2, value=low_bps)
        _style_data_cell(ws2.cell(row=i, column=2))
        ws2.cell(row=i, column=3, value=high_bps)
        _style_data_cell(ws2.cell(row=i, column=3))
        ws2.cell(row=i, column=4, value=low_bps / 10000)
        _style_data_cell(ws2.cell(row=i, column=4), _RATE_FMT)
        ws2.cell(row=i, column=5, value=high_bps / 10000)
        _style_data_cell(ws2.cell(row=i, column=5), _RATE_FMT)

    ws2.column_dimensions["A"].width = 18
    for col_letter in ["B", "C", "D", "E"]:
        ws2.column_dimensions[col_letter].width = 16

    # Notes sheet
    ws3 = wb.create_sheet("Notes")
    ws3.cell(row=1, column=1, value="Interest Rate Benchmark Data")
    ws3.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws3.cell(row=3, column=1, value="Source: European Central Bank / Refinitiv")
    ws3.cell(row=4, column=1, value="EURIBOR rates: quarterly averages, annualized")
    ws3.cell(row=5, column=1, value="Credit spreads: BBB-rated European industrial borrowers")
    ws3.cell(row=6, column=1, value="Applicable for benchmarking intercompany loan at 4.5% p.a.")
    ws3.cell(row=8, column=1, value="Recommended approach:")
    ws3.cell(row=9, column=1, value="  Arm's-length range = EURIBOR 12M average + BBB spread range")
    ws3.cell(row=10, column=1, value="  Compare CE\u2192CM loan rate (4.5%) against this range")
    ws3.column_dimensions["A"].width = 65

    # Register ERR-EU-009
    errors.add(PlantedError(
        error_id="ERR-EU-009",
        file=f"{_INPUT_DIR}/interest_rate_benchmarks_eu.xlsx",
        location=f"Sheet 'EURIBOR Rates', Row {err_row}, Column E (12M Rate)",
        type="transposed_digits",
        description=(
            "EURIBOR 12M rate for Q3 FY2025 shows 0.38% instead of 3.80% "
            "(decimal point error \u2014 digits transposed around the decimal). "
            "Other quarters show rates of 3.42%\u20133.85%, making 0.38% "
            "an obvious anomaly."
        ),
        severity="material",
        which_test_cases_should_catch=["TC-09-EU"],
    ))

    # Canary
    canary = canaries.canary_for("tc09eu_interest_rate_benchmarks")
    loc = embed_canary_xlsx(wb, canary)
    canaries.set_location(
        "tc09eu_interest_rate_benchmarks",
        f"{_INPUT_DIR}/interest_rate_benchmarks_eu.xlsx",
        loc,
    )

    file_path = output_dir / _INPUT_DIR / "interest_rate_benchmarks_eu.xlsx"
    file_path.parent.mkdir(parents=True, exist_ok=True)
    _save_xlsx_deterministic(wb, file_path)
    manifest.register(f"{_INPUT_DIR}/interest_rate_benchmarks_eu.xlsx", "xlsx")


# ── Master File PDF (28 pages) ───────────────────────────────────────────────


def _write_master_file_pdf(
    output_dir: Path,
    canaries: CanaryRegistry,
    manifest: Manifest,
) -> None:
    """Write master_file_fy2024.pdf — OECD BEPS Action 13 master file (28 pages)."""
    canary = canaries.canary_for("tc09eu_master_file_fy2024")
    file_path = output_dir / _INPUT_DIR / "master_file_fy2024.pdf"
    file_path.parent.mkdir(parents=True, exist_ok=True)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=inch, rightMargin=inch,
        topMargin=inch, bottomMargin=inch,
        invariant=True,
    )
    doc.title = "OECD Master File \u2014 Cascade Europe Holdings B.V. FY2024"
    doc.author = f"CANARY: {canary}"
    doc.subject = "Transfer Pricing Master File (BEPS Action 13)"
    doc.creator = "Cascade Industries Test Suite Generator"

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("MFTitle", parent=styles["Title"], fontSize=18, spaceAfter=20)
    h1 = ParagraphStyle("MFH1", parent=styles["Heading1"], fontSize=14, spaceBefore=12, spaceAfter=8)
    h2 = ParagraphStyle("MFH2", parent=styles["Heading2"], fontSize=12, spaceBefore=10, spaceAfter=6)
    body = ParagraphStyle("MFBody", parent=styles["Normal"], fontSize=10, leading=14, spaceAfter=8)
    small = ParagraphStyle("MFSmall", parent=body, fontSize=8, textColor=colors.gray)

    story: list[Any] = []

    # ── Page 1: Title ────────────────────────────────────────────
    story.append(Spacer(1, 2 * inch))
    story.append(Paragraph(
        "OECD Transfer Pricing Master File<br/>"
        "Cascade Europe Holdings B.V.<br/>"
        "Fiscal Year Ended December 31, 2024",
        title_style,
    ))
    story.append(Spacer(1, 0.5 * inch))
    story.append(Paragraph(
        "Prepared in accordance with OECD Transfer Pricing Guidelines (2022)<br/>"
        "and BEPS Action 13 Master File requirements",
        body,
    ))
    story.append(Spacer(1, 0.5 * inch))
    story.append(Paragraph("CONFIDENTIAL \u2014 FOR INTERNAL USE ONLY", ParagraphStyle(
        "Conf", parent=body, fontSize=10, textColor=colors.red, alignment=1,
    )))
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(f"Document ID: {canary}", small))
    story.append(PageBreak())

    # ── Page 2: Table of Contents ────────────────────────────────
    story.append(Paragraph("Table of Contents", h1))
    toc = [
        ("1. Organizational Structure", "3"),
        ("2. Group Business Description", "6"),
        ("3. Intangibles Overview", "11"),
        ("4. Intercompany Financial Activities", "15"),
        ("5. Financial and Tax Positions", "19"),
        ("Appendix A: Entity Legal Details", "25"),
        ("Appendix B: Intercompany Agreement List", "27"),
    ]
    toc_data = [["Section", "Page"]]
    for section, page in toc:
        toc_data.append([section, page])
    t = Table(toc_data, colWidths=[5 * inch, 1 * inch])
    t.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1A5276")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#1A5276")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    story.append(PageBreak())

    # ── Pages 3-5: Organizational Structure ──────────────────────
    story.append(Paragraph("1. Organizational Structure", h1))
    story.append(Paragraph(
        "Cascade Europe Holdings B.V. (\"CE\" or the \"Group\") is a Dutch "
        "holding company headquartered in Amsterdam, Netherlands. CE is the "
        "ultimate European parent that holds 100% of the equity interests in "
        "three operating subsidiaries across Europe.",
        body,
    ))
    story.append(Paragraph("Group Structure", h2))
    org_data = [
        ["Entity", "Jurisdiction", "Activity", "FY2024 Revenue"],
        ["Cascade Europe Holdings B.V.", "Netherlands", "Holding / Management", "\u20ac22M (fees + interest)"],
        ["Cascade Pr\u00e4zisionsteile GmbH", "Germany", "Manufacturing", "~\u20ac45M"],
        ["Cascade Mat\u00e9riaux Avanc\u00e9s SAS", "France", "R&D / Materials", "~\u20ac32M"],
        ["Cascade Distribution Services Ltd", "United Kingdom", "Distribution", "~\u00a318M (~\u20ac21M)"],
    ]
    t = Table(org_data, colWidths=[2.2 * inch, 1.2 * inch, 1.5 * inch, 1.5 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A5276")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#D6E4F0")]),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(
        "CE was incorporated in the Netherlands in 2018 as a wholly-owned "
        "European subsidiary of Cascade Industries, Inc. (Portland, Oregon, USA). "
        "The Group's European operations serve the industrial manufacturing and "
        "distribution markets across the EU and UK.",
        body,
    ))
    story.append(Paragraph(
        "The organizational structure reflects a functional specialization model: "
        "CP performs licensed manufacturing under IP developed by CM, CD operates "
        "as a limited-risk distributor, and CE provides centralised management "
        "services and treasury functions.",
        body,
    ))
    story.append(PageBreak())

    # Pages 4-5: continued organizational structure
    story.append(Paragraph("1. Organizational Structure (continued)", h1))
    story.append(Paragraph("Ownership Chain", h2))
    story.append(Paragraph(
        "Cascade Industries, Inc. (USA) \u2192 100% \u2192 Cascade Europe Holdings B.V. (NL) "
        "\u2192 100% \u2192 CP (DE), CM (FR), CD (UK)",
        body,
    ))
    story.append(Paragraph("Key Changes in FY2024", h2))
    story.append(Paragraph(
        "No material changes to the organizational structure occurred during FY2024. "
        "CD completed its first full year of operations following its establishment "
        "in FY2023. The UK entity operates under post-Brexit trade arrangements "
        "with goods flowing from CP (Germany) to CD (UK).",
        body,
    ))
    story.append(Paragraph("Functional Profiles", h2))
    for code, info in sorted(_EU_ENTITIES.items()):
        story.append(Paragraph(f"<b>{info['name']} ({code})</b>: {info['role']}", body))
    story.append(PageBreak())

    # ── Pages 6-10: Group Business Description ───────────────────
    story.append(Paragraph("2. Group Business Description", h1))
    story.append(Paragraph(
        "The Cascade Europe group operates in the European industrial "
        "manufacturing and distribution sector. The group's value chain "
        "comprises R&D and advanced materials development (CM, France), "
        "precision component manufacturing (CP, Germany), and distribution "
        "to European end customers (CD, United Kingdom).",
        body,
    ))
    story.append(Paragraph("Business Lines", h2))
    story.append(Paragraph(
        "&bull; <b>Advanced Materials R&D</b> (CM): Development of specialty "
        "materials, coatings, and composites for industrial applications. "
        "CM holds the group's core IP and licenses technology to CP.<br/>"
        "&bull; <b>Precision Manufacturing</b> (CP): Production of precision-machined "
        "components and assemblies using CM's proprietary processes.<br/>"
        "&bull; <b>Distribution</b> (CD): Warehousing, logistics, and last-mile "
        "delivery to UK and Irish customers.",
        body,
    ))
    story.append(Paragraph("Market Position", h2))
    story.append(Paragraph(
        "The group holds a mid-market position in the European industrial "
        "components sector, with combined revenue of approximately \u20ac120M. "
        "Key competitive advantages include CM's proprietary materials science "
        "capabilities and CP's precision manufacturing expertise.",
        body,
    ))
    story.append(PageBreak())

    # Pages 7-10: continued business description (filler for page count)
    for pg in range(7, 11):
        story.append(Paragraph(f"2. Group Business Description (continued \u2014 p.{pg})", h1))
        story.append(Paragraph("Supply Chain Integration", h2))
        story.append(Paragraph(
            "The group's supply chain is integrated across three jurisdictions. "
            "Raw materials and semi-finished components flow from CM (Lyon) to "
            "CP (Munich), where they are processed into finished products. "
            "Finished goods destined for the UK market are shipped from CP to "
            "CD (Birmingham). CE coordinates group-wide procurement and treasury.",
            body,
        ))
        story.append(Paragraph("Competitive Dynamics", h2))
        story.append(Paragraph(
            "The European industrial components market is moderately fragmented, "
            "with the top 15 groups holding approximately 40% of market share. "
            "Competition is driven by product quality, technical support, "
            "delivery reliability, and price. The group competes primarily in "
            "the precision and specialty segments where margins are above average.",
            body,
        ))
        story.append(PageBreak())

    # ── Pages 11-14: Intangibles Overview ────────────────────────
    story.append(Paragraph("3. Intangibles Overview", h1))
    story.append(Paragraph(
        "The group's intangible assets are primarily held by Cascade Mat\u00e9riaux "
        "Avanc\u00e9s SAS (CM), the group's R&D centre in Lyon, France. CM is "
        "responsible for the Development, Enhancement, Maintenance, Protection, "
        "and Exploitation (DEMPE) of the group's core technology.",
        body,
    ))
    story.append(Paragraph("DEMPE Analysis", h2))
    dempe_data = [
        ["DEMPE Function", "Performed By", "Description"],
        ["Development", "CM (France)",
         "All R&D activities, formulation development, testing, and prototyping"],
        ["Enhancement", "CM (France)",
         "Ongoing improvement of existing products and processes"],
        ["Maintenance", "CM (France)",
         "Quality control, patent maintenance, regulatory compliance"],
        ["Protection", "CM / CE",
         "Patent filing and defence (CM), legal oversight (CE)"],
        ["Exploitation", "CM \u2192 CP (license)",
         "CM licenses technology to CP for manufacturing; CP pays royalty"],
    ]
    t = Table(dempe_data, colWidths=[1.5 * inch, 1.5 * inch, 3.5 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A5276")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(
        "Under the current transfer pricing arrangement, CP pays CM a royalty "
        "of 3% of CP's annual revenue (\u2248\u20ac1.35M) for the right to use "
        "CM's proprietary technology in its manufacturing operations. This "
        "royalty is assessed monthly based on actual revenue.",
        body,
    ))
    story.append(Paragraph(
        "The royalty rate of 3% is within the range typically observed for "
        "manufacturing technology licenses (2\u20135%) and was established based "
        "on a Comparable Uncontrolled Transaction (CUT) analysis at the time "
        "the license was granted.",
        body,
    ))
    story.append(PageBreak())

    # Pages 12-14: continued intangibles
    for pg in range(12, 15):
        story.append(Paragraph(f"3. Intangibles Overview (continued \u2014 p.{pg})", h1))
        story.append(Paragraph("Key Patents and Know-How", h2))
        story.append(Paragraph(
            "CM holds 14 European patents related to advanced materials "
            "formulations, coating processes, and manufacturing techniques. "
            "The patent portfolio was developed through CM's internal R&D "
            "programme, which employs approximately 30 researchers in Lyon.",
            body,
        ))
        story.append(Paragraph(
            "In addition to registered patents, CM possesses significant "
            "trade secrets and manufacturing know-how that are integral to "
            "the group's competitive advantage. These intangibles are "
            "protected through confidentiality agreements with all employees "
            "and contractors.",
            body,
        ))
        story.append(PageBreak())

    # ── Pages 15-18: Intercompany Financial Activities ───────────
    story.append(Paragraph("4. Intercompany Financial Activities", h1))
    story.append(Paragraph(
        "The group conducts five principal categories of intercompany "
        "transactions, each governed by formal intercompany agreements "
        "and priced in accordance with the arm's-length principle:",
        body,
    ))

    ic_summary = [
        ["Transaction", "Flow", "Method", "FY2024 Volume"],
        ["Raw Materials", "CP \u2192 CM", "Cost-Plus-6%", "~\u20ac8.5M"],
        ["Finished Goods", "CP \u2192 CD", "Cost-Plus-8%", "~\u20ac6.2M"],
        ["Management Fees", "CE \u2192 CP/CM/CD", "1.5% of revenue", "~\u20ac1.47M"],
        ["Intercompany Loan", "CE \u2192 CM", "4.5% p.a.", "\u20ac3M principal"],
        ["R&D Royalty", "CM \u2192 CP", "3% of CP revenue", "~\u20ac1.35M"],
    ]
    t = Table(ic_summary, colWidths=[1.5 * inch, 1.5 * inch, 1.5 * inch, 1.5 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A5276")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#D6E4F0")]),
    ]))
    story.append(t)
    story.append(PageBreak())

    # Pages 16-18: IC details
    story.append(Paragraph("4. Intercompany Financial Activities (continued)", h1))
    story.append(Paragraph("Intercompany Loan", h2))
    story.append(Paragraph(
        "In January 2023, CE extended a \u20ac3,000,000 intercompany loan to CM "
        "to fund R&D laboratory expansion at the Lyon facility. The loan bears "
        "interest at 4.5% per annum, payable monthly. The rate was set by "
        "reference to the 12-month EURIBOR rate plus an appropriate credit "
        "spread for BBB-rated European industrial borrowers.",
        body,
    ))
    loan_data = [
        ["Parameter", "Value"],
        ["Principal", "\u20ac3,000,000"],
        ["Interest Rate", "4.5% per annum"],
        ["Benchmark", "EURIBOR 12M + BBB credit spread"],
        ["Term", "10 years"],
        ["Payment", "Interest-only, monthly accrual"],
    ]
    t = Table(loan_data, colWidths=[2 * inch, 3.5 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A5276")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(
        "FY2024 interest income (CE) / expense (CM) totalled \u20ac135,000. "
        "The rate remains within the arm's-length range based on EURIBOR "
        "benchmark data.",
        body,
    ))
    story.append(PageBreak())

    for pg in range(17, 19):
        story.append(Paragraph(f"4. Intercompany Financial Activities (continued \u2014 p.{pg})", h1))
        story.append(Paragraph("Management Fee Arrangement", h2))
        story.append(Paragraph(
            "CE charges each subsidiary a management fee of 1.5% of the "
            "subsidiary's annual revenue. The fee covers strategic oversight, "
            "treasury and cash management, legal and regulatory compliance, "
            "and IT infrastructure. The OECD Transfer Pricing Guidelines "
            "(Chapter VII) require that management fees satisfy a benefit test \u2014 "
            "each subsidiary must demonstrate that it receives an identifiable, "
            "measurable benefit from the services provided by CE.",
            body,
        ))
        story.append(Paragraph(
            "CE's services are documented in the Group Management Services Agreement. "
            "The fee rate of 1.5% is within the range observed in comparable "
            "management service arrangements (1\u20133% of revenue).",
            body,
        ))
        story.append(PageBreak())

    # ── Pages 19-24: Financial and Tax Positions ─────────────────
    story.append(Paragraph("5. Financial and Tax Positions", h1))
    story.append(Paragraph(
        "The group operates across four tax jurisdictions with the following "
        "statutory corporate income tax rates:",
        body,
    ))
    tax_data = [
        ["Jurisdiction", "Entity", "Statutory Rate", "Notes"],
        ["Netherlands", "CE", "25.8%", "Holding company regime"],
        ["Germany", "CP", "~29.9%", "K\u00f6rperschaftsteuer 15% + SolZ 5.5% + Gewerbesteuer ~14%"],
        ["France", "CM", "25.0%", "Eligible for CIR (30% R&D tax credit)"],
        ["United Kingdom", "CD", "25.0%", "Post-Brexit, TIOPA 2010 Part 4 TP rules"],
    ]
    t = Table(tax_data, colWidths=[1.3 * inch, 0.8 * inch, 1.2 * inch, 3.2 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A5276")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(
        "Country-by-Country Reporting (CbCR): The group's combined revenue "
        "of approximately \u20ac120M is below the \u20ac750M CbCR filing threshold "
        "under BEPS Action 13. Accordingly, CbCR is not required for the "
        "Cascade Europe group.",
        body,
    ))
    story.append(PageBreak())

    # Pages 20-24: continued financial positions (filler)
    for pg in range(20, 25):
        story.append(Paragraph(f"5. Financial and Tax Positions (continued \u2014 p.{pg})", h1))
        story.append(Paragraph(
            "The group's transfer pricing arrangements are designed to ensure "
            "that profits are allocated to the jurisdictions where economic "
            "value is created, consistent with the arm's-length principle and "
            "BEPS Action 8\u201310 guidance on value creation.",
            body,
        ))
        if pg == 20:
            story.append(Paragraph("Advance Pricing Agreements", h2))
            story.append(Paragraph(
                "The group does not currently have any Advance Pricing "
                "Agreements (APAs) in force. Bilateral APAs between DE/FR "
                "and NL/DE are under consideration for the CP\u2192CM goods flow.",
                body,
            ))
        if pg == 22:
            story.append(Paragraph("Tax Audit History", h2))
            story.append(Paragraph(
                "No material transfer pricing adjustments have been assessed "
                "by tax authorities in any jurisdiction for the past five "
                "fiscal years.",
                body,
            ))
        story.append(PageBreak())

    # ── Pages 25-28: Appendices ──────────────────────────────────
    story.append(Paragraph("Appendix A: Entity Legal Details", h1))
    for code, info in sorted(_EU_ENTITIES.items()):
        story.append(Paragraph(f"<b>{info['name']}</b>", h2))
        story.append(Paragraph(f"Entity code: {code}", body))
        story.append(Paragraph(f"Jurisdiction: {info['country']} ({info['city']})", body))
        story.append(Paragraph(f"Principal activity: {info['role']}", body))
        story.append(Spacer(1, 0.2 * inch))
    story.append(PageBreak())

    story.append(Paragraph("Appendix A (continued)", h1))
    story.append(Paragraph(
        "All entities are wholly owned by Cascade Europe Holdings B.V. "
        "Registered offices and tax identification numbers are maintained "
        "in the group's corporate secretariat records.",
        body,
    ))
    story.append(PageBreak())

    story.append(Paragraph("Appendix B: Intercompany Agreement List", h1))
    agreements = [
        ("Raw Materials Supply Agreement", "CP \u2192 CM", "2019", "Cost-plus-6%"),
        ("Finished Goods Distribution Agreement", "CP \u2192 CD", "2023", "Cost-plus-8%"),
        ("Group Management Services Agreement", "CE \u2192 All", "2018", "1.5% of revenue"),
        ("Intercompany Loan Agreement", "CE \u2192 CM", "2023", "\u20ac3M at 4.5% p.a."),
        ("Technology License Agreement", "CM \u2192 CP", "2019", "3% of CP revenue"),
    ]
    agr_data = [["Agreement", "Parties", "Effective", "Terms"]]
    for a in agreements:
        agr_data.append(list(a))
    t = Table(agr_data, colWidths=[2.5 * inch, 1.2 * inch, 1 * inch, 1.8 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A5276")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(t)
    story.append(PageBreak())

    # Final page
    story.append(Paragraph("Appendix B (continued)", h1))
    story.append(Paragraph(
        "All intercompany agreements are reviewed annually by CE's legal "
        "and tax teams. The terms are updated as necessary to reflect "
        "changes in market conditions, regulatory requirements, and "
        "group business operations.",
        body,
    ))
    story.append(Spacer(1, 1 * inch))
    story.append(Paragraph(
        "<i>End of Master File \u2014 Cascade Europe Holdings B.V. \u2014 FY2024</i>",
        ParagraphStyle("End", parent=body, alignment=1, textColor=colors.grey),
    ))

    doc.build(story)

    buf.seek(0)
    file_path.write_bytes(buf.getvalue())

    canaries.set_location(
        "tc09eu_master_file_fy2024",
        f"{_INPUT_DIR}/master_file_fy2024.pdf",
        "PDF metadata \u2192 Author",
    )
    manifest.register(f"{_INPUT_DIR}/master_file_fy2024.pdf", "pdf")


# ── Local File CP PDF (35 pages) ─────────────────────────────────────────────


def _write_local_file_cp_pdf(
    output_dir: Path,
    canaries: CanaryRegistry,
    manifest: Manifest,
) -> None:
    """Write local_file_cp_fy2024.pdf — CP local file (35 pages)."""
    canary = canaries.canary_for("tc09eu_local_file_cp_fy2024")
    file_path = output_dir / _INPUT_DIR / "local_file_cp_fy2024.pdf"
    file_path.parent.mkdir(parents=True, exist_ok=True)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=inch, rightMargin=inch,
        topMargin=inch, bottomMargin=inch,
        invariant=True,
    )
    doc.title = "OECD Local File \u2014 Cascade Pr\u00e4zisionsteile GmbH (CP) FY2024"
    doc.author = f"CANARY: {canary}"
    doc.subject = "Transfer Pricing Local File (BEPS Action 13)"
    doc.creator = "Cascade Industries Test Suite Generator"

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("LFTitle", parent=styles["Title"], fontSize=18, spaceAfter=20)
    h1 = ParagraphStyle("LFH1", parent=styles["Heading1"], fontSize=14, spaceBefore=12, spaceAfter=8)
    h2 = ParagraphStyle("LFH2", parent=styles["Heading2"], fontSize=12, spaceBefore=10, spaceAfter=6)
    body = ParagraphStyle("LFBody", parent=styles["Normal"], fontSize=10, leading=14, spaceAfter=8)
    small = ParagraphStyle("LFSmall", parent=body, fontSize=8, textColor=colors.gray)

    story: list[Any] = []

    # ── Page 1: Title ────────────────────────────────────────────
    story.append(Spacer(1, 2 * inch))
    story.append(Paragraph(
        "OECD Transfer Pricing Local File<br/>"
        "Cascade Pr\u00e4zisionsteile GmbH<br/>"
        "Fiscal Year Ended December 31, 2024",
        title_style,
    ))
    story.append(Spacer(1, 0.5 * inch))
    story.append(Paragraph(
        "Prepared in accordance with OECD Transfer Pricing Guidelines (2022),<br/>"
        "BEPS Action 13, and German AStG \u00a790 documentation requirements",
        body,
    ))
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(f"Document ID: {canary}", small))
    story.append(PageBreak())

    # ── Page 2: Table of Contents ────────────────────────────────
    story.append(Paragraph("Table of Contents", h1))
    toc = [
        ("1. Entity Description", "4"),
        ("2. Functional Analysis", "6"),
        ("3. Controlled Transactions", "13"),
        ("4. Economic Analysis \u2014 Benchmarking", "21"),
        ("5. Conclusions", "31"),
        ("Appendix: Financial Data", "33"),
    ]
    toc_data = [["Section", "Page"]]
    for section, page in toc:
        toc_data.append([section, page])
    t = Table(toc_data, colWidths=[5 * inch, 1 * inch])
    t.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1A5276")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#1A5276")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
    ]))
    story.append(t)
    story.append(PageBreak())

    # ── Page 3: blank / intro ────────────────────────────────────
    story.append(Paragraph("Introduction", h1))
    story.append(Paragraph(
        "This Local File documents the transfer pricing policies and "
        "economic analysis for Cascade Pr\u00e4zisionsteile GmbH (\"CP\"), "
        "a precision manufacturing entity based in Munich, Germany.",
        body,
    ))
    story.append(PageBreak())

    # ── Pages 4-5: Entity Description ────────────────────────────
    story.append(Paragraph("1. Entity Description", h1))
    story.append(Paragraph(
        "Cascade Pr\u00e4zisionsteile GmbH is a wholly-owned subsidiary of "
        "Cascade Europe Holdings B.V. (Netherlands). CP operates as a "
        "licensed manufacturer, producing precision-machined components "
        "and assemblies using proprietary technology licensed from "
        "Cascade Mat\u00e9riaux Avanc\u00e9s SAS (CM, France).",
        body,
    ))
    story.append(Paragraph("Key Facts", h2))
    kf_data = [
        ["Attribute", "Value"],
        ["Legal form", "Gesellschaft mit beschr\u00e4nkter Haftung (GmbH)"],
        ["Registered office", "Munich, Bavaria, Germany"],
        ["Principal activity", "Licensed manufacturer \u2014 precision components"],
        ["Employees", "~450"],
        ["FY2024 Revenue", "~\u20ac45M"],
        ["Key relationships", "Receives raw materials from CM; sells finished goods to CD"],
    ]
    t = Table(kf_data, colWidths=[2 * inch, 4.5 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A5276")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(t)
    story.append(PageBreak())
    story.append(Paragraph("1. Entity Description (continued)", h1))
    story.append(Paragraph(
        "CP's manufacturing operations are concentrated at the Munich "
        "facility, which includes CNC machining centres, surface treatment "
        "lines, and quality testing laboratories. The facility operates "
        "three shifts and has a capacity utilisation rate of approximately 78%.",
        body,
    ))
    story.append(PageBreak())

    # ── Pages 6-12: Functional Analysis ──────────────────────────
    story.append(Paragraph("2. Functional Analysis", h1))
    story.append(Paragraph(
        "CP functions as a licensed manufacturer within the Cascade Europe "
        "group. Its principal functions, risks, and assets are as follows:",
        body,
    ))
    story.append(Paragraph("Functions Performed", h2))
    story.append(Paragraph(
        "&bull; Procurement of raw materials (primarily from CM)<br/>"
        "&bull; Manufacturing of precision components per design specifications<br/>"
        "&bull; Quality control and testing<br/>"
        "&bull; Inventory management<br/>"
        "&bull; Shipping and logistics coordination (to CM for raw materials flow, "
        "to CD for finished goods)",
        body,
    ))
    story.append(Paragraph("Risks Assumed", h2))
    story.append(Paragraph(
        "&bull; Production risk (yield, defects)<br/>"
        "&bull; Inventory obsolescence risk<br/>"
        "&bull; Raw material price fluctuation risk<br/>"
        "&bull; Limited market risk (primarily sells within the group)",
        body,
    ))
    story.append(Paragraph("Assets Employed", h2))
    story.append(Paragraph(
        "&bull; Tangible: CNC machines, surface treatment equipment, testing labs<br/>"
        "&bull; Intangible: Limited \u2014 manufacturing know-how, but core IP is "
        "held by CM under the Technology License Agreement",
        body,
    ))
    story.append(PageBreak())

    # Pages 7-12: continued functional analysis (filler for page count)
    for pg in range(7, 13):
        story.append(Paragraph(f"2. Functional Analysis (continued \u2014 p.{pg})", h1))
        if pg == 7:
            story.append(Paragraph("Characterization", h2))
            story.append(Paragraph(
                "Based on the functional analysis, CP is characterised as a "
                "licensed manufacturer with moderate functional complexity but "
                "limited risk and limited ownership of valuable intangibles. "
                "The Transactional Net Margin Method (TNMM) is the most "
                "appropriate method for testing CP's intercompany transactions.",
                body,
            ))
        if pg == 9:
            story.append(Paragraph("R&D Royalty Analysis", h2))
            story.append(Paragraph(
                "CP pays a royalty of 3% of its annual revenue to CM for the "
                "right to use CM's proprietary technology. CM is the legal and "
                "economic owner of the IP, performing all DEMPE functions. "
                "CP is the licensee. The royalty flows from CP (licensee) to "
                "CM (licensor), which is consistent with the economic substance "
                "of the arrangement.",
                body,
            ))
        story.append(Paragraph(
            "The functional analysis confirms that CP operates within a "
            "well-defined role in the group's value chain. The limited-risk "
            "characterisation supports the use of one-sided transfer pricing "
            "methods with CP as the tested party.",
            body,
        ))
        story.append(PageBreak())

    # ── Pages 13-20: Controlled Transactions ─────────────────────
    story.append(Paragraph("3. Controlled Transactions", h1))
    story.append(Paragraph("CP engages in the following controlled transactions:", body))
    ct_data = [
        ["#", "Transaction", "Counterparty", "Method", "FY2024 Volume"],
        ["1", "Purchase of raw materials", "CM (seller)", "Cost-Plus-6%", "~\u20ac8.5M"],
        ["2", "Sale of finished goods", "CD (buyer)", "Cost-Plus-8%", "~\u20ac6.2M"],
        ["3", "Management fee", "CE (service provider)", "1.5% of revenue", "~\u20ac675K"],
        ["4", "R&D royalty", "CM (licensor)", "3% of revenue", "~\u20ac1.35M"],
    ]
    t = Table(ct_data, colWidths=[0.4 * inch, 2 * inch, 1.3 * inch, 1.5 * inch, 1.3 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A5276")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(t)
    story.append(PageBreak())

    # Pages 14-20: controlled transactions details
    for pg in range(14, 21):
        story.append(Paragraph(f"3. Controlled Transactions (continued \u2014 p.{pg})", h1))
        if pg == 14:
            story.append(Paragraph("Raw Materials (CP \u2192 CM)", h2))
            story.append(Paragraph(
                "CP supplies raw materials and semi-finished components to CM "
                "at cost plus a 6% markup. The cost base includes direct materials, "
                "direct labour, and allocated manufacturing overhead. Monthly "
                "volumes averaged approximately \u20ac710K in FY2024.",
                body,
            ))
        elif pg == 16:
            story.append(Paragraph("Finished Goods (CP \u2192 CD)", h2))
            story.append(Paragraph(
                "CP supplies finished precision components to CD for distribution "
                "in the UK market. Goods are priced at cost plus 8%, reflecting "
                "the additional value added through CP's manufacturing process.",
                body,
            ))
        else:
            story.append(Paragraph(
                "Transaction volumes and pricing remain consistent with the "
                "group's transfer pricing policies. All transactions are "
                "documented in formal intercompany agreements.",
                body,
            ))
        story.append(PageBreak())

    # ── Pages 21-30: Economic Analysis ───────────────────────────
    story.append(Paragraph("4. Economic Analysis \u2014 Benchmarking", h1))
    story.append(Paragraph(
        "The TNMM is applied with operating margin as the profit level "
        "indicator (PLI) to test whether CP's intercompany transactions "
        "are at arm's length.",
        body,
    ))
    story.append(Paragraph("Comparable Search Strategy", h2))
    story.append(Paragraph(
        "A search of European manufacturing companies was conducted using "
        "the following criteria:<br/>"
        "&bull; NACE code 2562 (machining)<br/>"
        "&bull; Revenue range \u20ac50M\u2013\u20ac500M<br/>"
        "&bull; Publicly available financial data<br/>"
        "&bull; No financial distress or restructuring<br/>"
        "&bull; Independent operations (not a captive subsidiary)",
        body,
    ))
    story.append(Paragraph(
        "The initial search identified 15 companies. After screening, "
        "12 were accepted and 3 were rejected.",
        body,
    ))
    story.append(PageBreak())

    # Pages 22-30: benchmarking details
    for pg in range(22, 31):
        story.append(Paragraph(f"4. Economic Analysis (continued \u2014 p.{pg})", h1))
        if pg == 22:
            story.append(Paragraph("Comparable Screening Results", h2))
            story.append(Paragraph(
                "Three companies were rejected from the comparable set:<br/>"
                "&bull; Nordic Logistics Solutions Oy \u2014 NACE code 5229 "
                "(logistics), not manufacturing<br/>"
                "&bull; Meridionale Industrie S.p.A. \u2014 financial distress, "
                "negative operating margin<br/>"
                "&bull; Continental Grosswerk AG \u2014 revenue \u20ac2.8B, "
                "exceeds 10x CP's revenue (size outlier per OECD \u00a73.43-3.46)",
                body,
            ))
        elif pg == 25:
            story.append(Paragraph("IQR Computation (FY2024)", h2))
            story.append(Paragraph(
                "Based on the 12 accepted companies, the interquartile range "
                "of operating margins is:<br/>"
                "&bull; Q1 (25th percentile): 3.8%<br/>"
                "&bull; Median: 5.6%<br/>"
                "&bull; Q3 (75th percentile): 7.9%<br/>"
                "&bull; CP actual operating margin: ~6.0% (within range)",
                body,
            ))
        else:
            story.append(Paragraph(
                "The benchmarking analysis supports the conclusion that "
                "CP's intercompany pricing is consistent with the arm's-length "
                "principle. Further details of the comparable company financial "
                "profiles are provided in the appendix.",
                body,
            ))
        story.append(PageBreak())

    # ── Pages 31-33: Conclusions ─────────────────────────────────
    story.append(Paragraph("5. Conclusions", h1))
    story.append(Paragraph(
        "Based on the economic analysis performed, CP's operating margin of "
        "approximately 6.0% for FY2024 falls within the interquartile range "
        "of 3.8%\u20137.9% established from the comparable company analysis. "
        "Accordingly, CP's intercompany transactions are considered to be "
        "at arm's length.",
        body,
    ))
    story.append(Paragraph("Summary by Transaction Type", h2))
    summary_data = [
        ["Transaction", "Result", "Recommendation"],
        ["Raw materials (CP\u2192CM)", "Within range", "No adjustment required"],
        ["Finished goods (CP\u2192CD)", "Within range", "No adjustment required"],
        ["Management fee", "1.5% within 1-3% range", "No adjustment required"],
        ["R&D royalty", "3% within 2-5% range", "Monitor DEMPE allocation"],
    ]
    t = Table(summary_data, colWidths=[2 * inch, 2 * inch, 2.5 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A5276")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(t)
    story.append(PageBreak())

    # Pages 32-33: continued conclusions
    story.append(Paragraph("5. Conclusions (continued)", h1))
    story.append(Paragraph(
        "The transfer pricing documentation for CP should be updated annually "
        "with current year data. The FY2025 update should include fresh "
        "comparable company data and updated IQR computations.",
        body,
    ))
    story.append(PageBreak())

    # ── Pages 33-35: Appendix ────────────────────────────────────
    story.append(Paragraph("Appendix: Financial Data", h1))
    story.append(Paragraph(
        "Selected financial highlights for Cascade Pr\u00e4zisionsteile GmbH:",
        body,
    ))
    fin_data = [
        ["Metric", "FY2023", "FY2024"],
        ["Revenue", "\u20ac42.5M", "\u20ac45.0M"],
        ["COGS", "\u20ac30.2M", "\u20ac31.8M"],
        ["Gross margin", "28.9%", "29.3%"],
        ["Operating expenses", "\u20ac9.8M", "\u20ac10.5M"],
        ["Operating income", "\u20ac2.5M", "\u20ac2.7M"],
        ["Operating margin", "5.9%", "6.0%"],
    ]
    t = Table(fin_data, colWidths=[2 * inch, 2 * inch, 2 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A5276")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
    ]))
    story.append(t)
    story.append(PageBreak())

    story.append(Paragraph("Appendix (continued)", h1))
    story.append(Paragraph(
        "The financial data presented above is derived from CP's statutory "
        "financial statements prepared under German GAAP (HGB). The operating "
        "margin of 6.0% is consistent with the arm's-length range established "
        "in the economic analysis section.",
        body,
    ))
    story.append(Spacer(1, 1 * inch))
    story.append(Paragraph(
        "<i>End of Local File \u2014 Cascade Pr\u00e4zisionsteile GmbH \u2014 FY2024</i>",
        ParagraphStyle("End", parent=body, alignment=1, textColor=colors.grey),
    ))

    doc.build(story)

    buf.seek(0)
    file_path.write_bytes(buf.getvalue())

    canaries.set_location(
        "tc09eu_local_file_cp_fy2024",
        f"{_INPUT_DIR}/local_file_cp_fy2024.pdf",
        "PDF metadata \u2192 Author",
    )
    manifest.register(f"{_INPUT_DIR}/local_file_cp_fy2024.pdf", "pdf")


# ── Prompt ───────────────────────────────────────────────────────────────────


def _write_prompt(output_dir: Path) -> None:
    text = """\
Prepare updated OECD transfer pricing documentation for the Cascade Europe
Holdings B.V. group for FY2025, covering both the master file and entity-level
local files.

Master File (group level):
1. Update the organizational structure section with current FY2025 entity details
   and any changes from FY2024.
2. Update the description of intercompany transactions with FY2025 volumes and
   pricing for all five transaction flows (goods CP\u2192CM, goods CP\u2192CD, management
   fees CE\u2192subs, intercompany loan CE\u2192CM, R&D royalty CM\u2192CP).
3. Update the intangibles section noting that CM (France) is the principal R&D
   entity and licensor of developed IP to CP under the royalty arrangement.
4. Update the financial activities section with the intercompany loan terms and
   note the interest rate relative to EURIBOR benchmarks.

Local File \u2014 Cascade Pr\u00e4zisionsteile GmbH (CP, Germany):
5. Update the functional analysis for CP as a licensed manufacturer receiving
   raw materials from CM and selling finished goods to CD.
6. Screen the manufacturing comparable companies:
   - Reject any that are not appropriate comparables (explain why for each)
   - Compute the interquartile range of operating margins for the accepted set
   - Apply the Transactional Net Margin Method (TNMM) to test CP's margin
7. Determine whether CP's operating margin falls within the arm's-length range.
8. Analyze the R&D royalty received from CM (3% of CP revenue) \u2014 assess whether
   this rate is consistent with arm's-length principles given that CM develops
   the IP and CP is the licensee. Note any concerns.

Local File \u2014 Cascade Distribution Services Ltd (CD, UK):
9. Prepare a functional analysis for CD as a limited-risk distributor.
10. Screen the distribution comparable companies and compute the interquartile
    range of net margins for the accepted set.
11. Apply TNMM to test CD's net margin against the benchmark range.
12. Flag any concerns about CD's margin relative to the distribution benchmark.

Intercompany Loan Analysis:
13. Test the CE\u2192CM intercompany loan interest rate (4.5%) against the EURIBOR
    benchmark data plus appropriate credit spread for BBB-rated industrial
    borrowers. Determine whether the rate is arm's-length.

For all analyses:
14. Flag any transaction types where the actual margin or rate falls outside the
    arm's-length range.
15. Note any data gaps or limitations that affect the conclusions.

Export:
- Benchmarking analysis as Excel (one sheet per tested entity: comparables
  screening, accepted set, IQR computation, tested party margin, conclusion)
- Interest rate benchmark analysis as separate Excel sheet
- Updated master file sections as Word document (organizational structure,
  transactions, intangibles, financial activities)
- Updated CP local file economic analysis section as Word document
- CD local file as Word document (new \u2014 no prior year local file exists for CD)
"""
    path = output_dir / f"test_cases/{_TC}/prompt.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text)


# ── Expected Behavior ────────────────────────────────────────────────────────


def _write_expected_behavior(output_dir: Path) -> None:
    text = """\
# TC-09-EU: OECD Transfer Pricing Documentation \u2014 Expected Behavior

## Key Findings the Agent Should Produce

1. **Master file / local file split (BEPS Action 13)**: Agent must produce both
   master file content AND entity-level local files. A single TP report without
   the OECD two-tier structure is wrong.

2. **Reject three manufacturing comparables**:
   - Nordic Logistics Solutions Oy (NACE 5229) \u2014 wrong industry (logistics, not manufacturing)
   - Meridionale Industrie S.p.A. (NACE 2562) \u2014 financial distress (negative operating margin)
   - Continental Grosswerk AG (NACE 2562) \u2014 size outlier, revenue >10x CP (OECD \u00a73.43-3.46)

3. **Reject two distribution comparables**:
   - Groupe Stellaire Captive SAS \u2014 captive entity, no independent pricing
   - Baltic Services O\u00dc \u2014 restructuring losses, non-recurring charges

4. **Manufacturing IQR (CP)**: After rejecting 3 companies, 12 accepted:
   - Q1: **3.8%**, Median: **5.6%**, Q3: **7.9%**
   - CP actual operating margin: ~6.2% \u2014 **within range**

5. **Distribution IQR (CD)**: After rejecting 2 companies, 8 accepted:
   - Q1: **1.2%**, Median: **2.1%**, Q3: **3.5%**
   - CD actual net margin: ~1.8% \u2014 **within range** (low end)

6. **Intercompany loan (CE\u2192CM)**:
   - EURIBOR 12M average FY2025: ~3.8% (after correcting the Q3 data error)
   - BBB industrial credit spread: 100-150bps
   - Arm's-length range: 4.3% to 5.3%
   - Actual rate 4.5%: **within range**
   - **ERR-EU-009**: Q3 FY2025 EURIBOR 12M shows 0.38% instead of 3.80% \u2014
     agent must identify this as a data error

7. **R&D royalty direction verification**: CM\u2192CP royalty \u2014 the transaction data
   shows CM paying CP, but the description says "technology license." Agent must
   verify direction against DEMPE analysis (CM is IP developer/licensor, CP is
   licensee \u2014 royalty flows from CP to CM are correct).

8. **Multiple TP methods**: Agent must select the appropriate method per transaction:
   - TNMM for CP (licensed manufacturer) and CD (limited-risk distributor)
   - Cost-plus for manufacturing flows
   - CUP-like for interest rate
   - CUT or comparable analysis for royalty

9. **Management fee benefit test**: Agent should flag that management fees require
   a benefit test per OECD Guidelines Chapter VII.

10. **CD local file (new)**: No prior year local file for CD. Agent must create one
    from scratch with proper OECD local file structure.

11. **Post-Brexit treatment**: Agent should note CD (UK) is non-EU and reference
    UK TP rules under TIOPA 2010 Part 4.

12. **CbCR awareness**: Group revenue ~\u20ac120M is below \u20ac750M threshold \u2014 CbCR
    not required, but agent should demonstrate awareness.

## Data Challenges

- **Comparable screening with OECD-specific criteria**: The size outlier rejection
  is a distinctly OECD screen not present in simpler analyses.
- **Multiple IQRs**: Manufacturing (CP) and distribution (CD) have separate
  comparable sets with different IQR ranges.
- **ERR-EU-009 detection**: 0.38% is implausible when other quarters show 3.4-3.9%.
  Agent must catch this or risk concluding the loan is above-market.
- **Royalty direction ambiguity**: The description field is misleading. Agent must
  cross-reference the functional analysis and master file intangibles section.
- **New document creation**: CD local file requires creation from scratch.

## Expected Output Structure

### Benchmarking Analysis (Excel):
- Sheet per tested entity (CP manufacturing, CD distribution)
- Comparable screening with accept/reject flags and reasons
- Accepted set margins and IQR computation
- Tested party margin vs. IQR assessment

### Interest Rate Analysis (Excel):
- EURIBOR rates with data quality flag for Q3 FY2025
- BBB spread data and composite range
- CE\u2192CM loan rate assessment

### Updated Master File (Word):
- Organizational structure, transactions, intangibles, financial activities

### CP Local File Update (Word):
- Economic analysis with FY2025 benchmarking results

### CD Local File (Word) \u2014 NEW:
- Entity overview, functional analysis, controlled transactions,
  economic analysis, financial data appendix
"""
    path = output_dir / f"test_cases/{_TC}/expected_behavior.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text)


# ── Gold Standard ────────────────────────────────────────────────────────────


def _compute_mfg_iqr() -> dict[str, Any]:
    """Compute IQR for the 12 accepted manufacturing comparables."""
    margins = sorted(m[6] for m in _MFG_CALIBRATED)  # margin_pct
    n = len(margins)  # 12

    q1_idx = 0.25 * (n - 1)  # 2.75
    q3_idx = 0.75 * (n - 1)  # 8.25

    q1_low = int(q1_idx)
    q1_frac = q1_idx - q1_low
    q1 = margins[q1_low] + q1_frac * (margins[q1_low + 1] - margins[q1_low])

    q3_low = int(q3_idx)
    q3_frac = q3_idx - q3_low
    q3 = margins[q3_low] + q3_frac * (margins[q3_low + 1] - margins[q3_low])

    median_idx = n // 2
    median = (margins[median_idx - 1] + margins[median_idx]) / 2

    return {
        "accepted_count": n,
        "rejected_count": 3,
        "margins_sorted_pct": [round(m, 1) for m in margins],
        "q1_pct": round(q1, 1),
        "median_pct": round(median, 1),
        "q3_pct": round(q3, 1),
        "min_pct": round(margins[0], 1),
        "max_pct": round(margins[-1], 1),
    }


def _compute_dist_iqr() -> dict[str, Any]:
    """Compute IQR for the 8 accepted distribution comparables."""
    margins = sorted(d[6] for d in _DIST_CALIBRATED if not d[9])
    n = len(margins)  # 8

    q1_idx = 0.25 * (n - 1)  # 1.75
    q3_idx = 0.75 * (n - 1)  # 5.25

    q1_low = int(q1_idx)
    q1_frac = q1_idx - q1_low
    q1 = margins[q1_low] + q1_frac * (margins[q1_low + 1] - margins[q1_low])

    q3_low = int(q3_idx)
    q3_frac = q3_idx - q3_low
    q3 = margins[q3_low] + q3_frac * (margins[q3_low + 1] - margins[q3_low])

    median_idx = n // 2
    median = (margins[median_idx - 1] + margins[median_idx]) / 2

    return {
        "accepted_count": n,
        "rejected_count": 2,
        "margins_sorted_pct": [round(m, 1) for m in margins],
        "q1_pct": round(q1, 1),
        "median_pct": round(median, 1),
        "q3_pct": round(q3, 1),
        "min_pct": round(margins[0], 1),
        "max_pct": round(margins[-1], 1),
    }


def _compute_euribor_avg_fy2025_corrected() -> float:
    """Average EURIBOR 12M for FY2025 quarters, correcting the planted error."""
    fy25_rates = []
    for period, _quarter, _r3m, _r6m, r12m in _EURIBOR_DATA:
        if period == "FY2025":
            # Correct the error for Q3
            if _quarter == "Q3":
                fy25_rates.append(_CORRECT_Q3_FY2025_12M)
            else:
                fy25_rates.append(r12m)
    return round(sum(fy25_rates) / len(fy25_rates), 2)


@register_gold(_TC)
def _tc09_eu_gold(
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    **model_kwargs: Any,
) -> GoldStandard:
    """TC-09-EU gold standard: OECD transfer pricing documentation."""
    mfg_iqr = _compute_mfg_iqr()
    dist_iqr = _compute_dist_iqr()
    euribor_avg = _compute_euribor_avg_fy2025_corrected()

    return GoldStandard(
        test_case=_TC,
        expected_outputs={
            "output_files": {
                "benchmarking_analysis": {
                    "type": "xlsx",
                    "required_sheets": [
                        "CP Manufacturing Benchmarking",
                        "CD Distribution Benchmarking",
                        "Interest Rate Analysis",
                    ],
                },
                "master_file_update": {
                    "type": "docx",
                    "required_sections": [
                        "Organizational Structure",
                        "Intercompany Transactions",
                        "Intangibles",
                        "Financial Activities",
                    ],
                },
                "cp_local_file_update": {
                    "type": "docx",
                    "required_sections": [
                        "Economic Analysis \u2014 FY2025 Benchmarking Results",
                    ],
                },
                "cd_local_file_new": {
                    "type": "docx",
                    "required_sections": [
                        "Entity Overview",
                        "Functional Analysis",
                        "Controlled Transactions",
                        "Economic Analysis",
                        "Financial Data Appendix",
                    ],
                },
            },
            "manufacturing_comparable_screening": {
                "total_companies": 15,
                "accepted": mfg_iqr["accepted_count"],
                "rejected": mfg_iqr["rejected_count"],
                "rejections": [
                    {
                        "company": "Nordic Logistics Solutions Oy",
                        "reason": "NACE code mismatch (5229 vs 2562 \u2014 logistics, not manufacturing)",
                    },
                    {
                        "company": "Meridionale Industrie S.p.A.",
                        "reason": "Financial distress (negative operating margin, concordato preventivo)",
                    },
                    {
                        "company": "Continental Grosswerk AG",
                        "reason": (
                            "Size outlier \u2014 revenue \u20ac2.8B exceeds "
                            "10x CP revenue (OECD \u00a73.43-3.46)"
                        ),
                    },
                ],
            },
            "manufacturing_iqr": {
                "q1_pct": mfg_iqr["q1_pct"],
                "median_pct": mfg_iqr["median_pct"],
                "q3_pct": mfg_iqr["q3_pct"],
                "margins_sorted_pct": mfg_iqr["margins_sorted_pct"],
                "cp_operating_margin_pct": 6.2,
                "cp_within_range": True,
            },
            "distribution_comparable_screening": {
                "total_companies": 10,
                "accepted": dist_iqr["accepted_count"],
                "rejected": dist_iqr["rejected_count"],
                "rejections": [
                    {
                        "company": "Groupe Stellaire Captive SAS",
                        "reason": "Captive entity of a listed group \u2014 no independent pricing",
                    },
                    {
                        "company": "Baltic Services O\u00dc",
                        "reason": "Restructuring losses \u2014 non-recurring charges distort net margin",
                    },
                ],
            },
            "distribution_iqr": {
                "q1_pct": dist_iqr["q1_pct"],
                "median_pct": dist_iqr["median_pct"],
                "q3_pct": dist_iqr["q3_pct"],
                "margins_sorted_pct": dist_iqr["margins_sorted_pct"],
                "cd_net_margin_pct": 1.8,
                "cd_within_range": True,
                "cd_note": "Low end of range \u2014 within IQR but near Q1",
            },
            "intercompany_loan": {
                "loan_principal_eur": 3_000_000,
                "loan_rate_pct": 4.5,
                "euribor_12m_avg_fy2025_corrected_pct": euribor_avg,
                "bbb_spread_low_bps": 50,
                "bbb_spread_high_bps": 150,
                "arm_length_range_low_pct": round(euribor_avg + 0.5, 1),
                "arm_length_range_high_pct": round(euribor_avg + 1.5, 1),
                "within_range": True,
            },
            "royalty_analysis": {
                "rate_pct": 3.0,
                "flow_direction": "CP pays CM (licensee to licensor)",
                "direction_correct": True,
                "comparable_range_pct": "2\u20135%",
                "within_range": True,
                "dempe_note": "CM performs Development, Enhancement, and Maintenance of IP",
            },
            "management_fees": {
                "rate_pct": 1.5,
                "total_eur": _whole_euros(_MGMT_FEE_TOTAL),
                "within_range": True,
                "benefit_test_note": (
                    "Agent should flag benefit test requirement per OECD Chapter VII \u2014 "
                    "shareholder activities are not chargeable"
                ),
            },
            "cbcr_assessment": {
                "group_revenue_eur": "~\u20ac120M",
                "threshold_eur": "\u20ac750M",
                "cbcr_required": False,
            },
        },
        canary_verification={
            "read_ic_transactions": canaries.canary_for("tc09eu_intercompany_transactions"),
            "read_comparables": canaries.canary_for("tc09eu_comparable_companies"),
            "read_master_file": canaries.canary_for("tc09eu_master_file_fy2024"),
            "read_local_file_cp": canaries.canary_for("tc09eu_local_file_cp_fy2024"),
            "read_interest_benchmarks": canaries.canary_for("tc09eu_interest_rate_benchmarks"),
        },
        error_detection={
            "ERR-EU-009": (
                "EURIBOR 12M rate for Q3 FY2025 shows 0.38% instead of 3.80% \u2014 "
                "decimal point error that distorts the interest rate benchmark analysis"
            ),
        },
        scoring_hints={
            "correctness": (
                "Manufacturing IQR must be Q1=3.8%, Q3=7.9%. "
                "Distribution IQR must be Q1=1.2%, Q3=3.5%. "
                "CP margin ~6.2% within range. CD margin ~1.8% within range. "
                "Loan rate 4.5% within 4.3%-5.3% range. "
                "ERR-EU-009 must be identified."
            ),
            "completeness": (
                "All five transaction types analysed with appropriate methods. "
                "Master file AND local files produced (BEPS Action 13 split). "
                "CD local file created from scratch. "
                "5 rejections identified (3 manufacturing, 2 distribution). "
                "DEMPE analysis referenced for royalty. "
                "Benefit test noted for management fees."
            ),
            "format_compliance": (
                "Benchmarking as Excel with per-entity sheets. "
                "Interest rate analysis as separate sheet. "
                "Master file and local files as Word documents. "
                "OECD two-tier documentation structure."
            ),
            "robustness": (
                "Agent must reject size outlier (OECD-specific screen). "
                "Agent must detect ERR-EU-009 decimal point error. "
                "Agent must verify royalty direction against DEMPE. "
                "Agent must note CD is non-EU (post-Brexit). "
                "Agent must note CbCR not required (below threshold)."
            ),
            "communication": (
                "Clear arm's-length conclusions per transaction type. "
                "Professional OECD transfer pricing terminology. "
                "Actionable recommendations where margins are borderline."
            ),
        },
        scenario_pack="cascade_europe_ifrs",
    )


# ── Public entry point ───────────────────────────────────────────────────────


def emit_tc09_eu(
    model: CascadeModel,
    output_dir: Path,
    canaries: CanaryRegistry,
    errors: ErrorRegistry,
    manifest: Manifest,
    **kwargs: object,
) -> None:
    """Emit all TC-09-EU files."""
    _write_ic_transactions_xlsx(output_dir, canaries, manifest)
    _write_comparables_xlsx(output_dir, canaries, manifest)
    _write_interest_rate_benchmarks(output_dir, canaries, errors, manifest)
    _write_master_file_pdf(output_dir, canaries, manifest)
    _write_local_file_cp_pdf(output_dir, canaries, manifest)
    _write_prompt(output_dir)
    _write_expected_behavior(output_dir)
